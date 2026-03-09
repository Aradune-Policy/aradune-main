#!/usr/bin/env python3
"""
build_lake_snap_tanf.py — Ingest SNAP and TANF enrollment data into the lake.

Reads from: data/raw/snap_tanf/
  - archive/FY{00-26}.xlsx  — SNAP state-level monthly participation & cost
  - tanf_caseload_fy2024.xlsx — TANF families & recipients by state and month
  - tanf_caseload_fy2023.xlsx — TANF FY2023

Writes to: data/lake/
  fact_snap_enrollment   — SNAP monthly participation and cost by state
  fact_tanf_enrollment   — TANF monthly families and recipients by state

Usage:
  python3 scripts/build_lake_snap_tanf.py
  python3 scripts/build_lake_snap_tanf.py --dry-run
  python3 scripts/build_lake_snap_tanf.py --only fact_snap_enrollment
"""

import argparse
import json
import uuid
from datetime import date, datetime
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "snap_tanf"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

# State name → code mapping
STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Guam": "GU", "Puerto Rico": "PR", "Virgin Islands": "VI",
    "U.S. Virgin Islands": "VI", "American Samoa": "AS",
    "Northern Mariana Islands": "MP", "N. Mariana Islands": "MP",
}

# Month abbreviation → number
MONTH_ABBR = {
    "Oct": 10, "Nov": 11, "Dec": 12, "Jan": 1, "Feb": 2, "Mar": 3,
    "Apr": 4, "May": 5, "Jun": 6, "Jul": 7, "Aug": 8, "Sep": 9,
}


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


# ---------------------------------------------------------------------------
# SNAP Enrollment — Parse state-level monthly data from regional Excel files
# ---------------------------------------------------------------------------

def _parse_snap_file(xlsx_path: Path) -> list[dict]:
    """Parse one SNAP FY file (FY00-FY26) into flat records."""
    rows = []
    # Determine file type (.xls vs .xlsx)
    suffix = xlsx_path.suffix.lower()
    if suffix == ".xls":
        # Old .xls files — use xlrd
        try:
            import xlrd
        except ImportError:
            print(f"  SKIPPED {xlsx_path.name} — xlrd not installed (pip install xlrd)")
            return []
        wb = xlrd.open_workbook(str(xlsx_path))
        for sheet_name in wb.sheet_names():
            if sheet_name == "US Summary":
                continue
            ws = wb.sheet_by_name(sheet_name)
            current_state = None
            for row_idx in range(ws.nrows):
                cell0 = ws.cell_value(row_idx, 0)
                if not cell0 or not isinstance(cell0, str):
                    continue
                cell0 = cell0.strip()
                # Check if this is a state name
                if cell0 in STATE_NAME_TO_CODE:
                    current_state = STATE_NAME_TO_CODE[cell0]
                    continue
                # Check if this is a month row like "Oct 2023"
                if current_state and len(cell0) > 4 and cell0[:3] in MONTH_ABBR:
                    parts = cell0.split()
                    if len(parts) == 2:
                        try:
                            month_num = MONTH_ABBR[parts[0]]
                            year = int(parts[1])
                            households = ws.cell_value(row_idx, 1)
                            persons = ws.cell_value(row_idx, 2)
                            cost = ws.cell_value(row_idx, 3)
                            if isinstance(households, (int, float)) and households > 0:
                                rows.append({
                                    "state_code": current_state,
                                    "year": year,
                                    "month": month_num,
                                    "households": int(households),
                                    "persons": int(persons) if isinstance(persons, (int, float)) else None,
                                    "benefit_cost": round(float(cost), 2) if isinstance(cost, (int, float)) else None,
                                })
                        except (ValueError, KeyError):
                            pass
                # "Total" row resets state
                if cell0 == "Total":
                    current_state = None
        return rows

    # .xlsx files
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name == "US Summary":
            continue
        ws = wb[sheet_name]
        current_state = None
        for row in ws.iter_rows(values_only=True):
            cell0 = row[0]
            if not cell0 or not isinstance(cell0, str):
                continue
            cell0 = cell0.strip()
            # Check if this is a state name
            if cell0 in STATE_NAME_TO_CODE:
                current_state = STATE_NAME_TO_CODE[cell0]
                continue
            # Check if this is a month row like "Oct 2023"
            if current_state and len(cell0) > 4 and cell0[:3] in MONTH_ABBR:
                parts = cell0.split()
                if len(parts) == 2:
                    try:
                        month_num = MONTH_ABBR[parts[0]]
                        year = int(parts[1])
                        households = row[1]
                        persons = row[2]
                        cost = row[3]
                        if isinstance(households, (int, float)) and households > 0:
                            rows.append({
                                "state_code": current_state,
                                "year": year,
                                "month": month_num,
                                "households": int(households),
                                "persons": int(persons) if isinstance(persons, (int, float)) else None,
                                "benefit_cost": round(float(cost), 2) if isinstance(cost, (int, float)) else None,
                            })
                    except (ValueError, KeyError):
                        pass
            # "Total" row resets state
            if cell0 == "Total":
                current_state = None
    wb.close()
    return rows


def build_fact_snap_enrollment(con, dry_run: bool) -> int:
    print("Building fact_snap_enrollment...")
    archive_dir = RAW_DIR / "archive"
    if not archive_dir.exists():
        print(f"  SKIPPED — {archive_dir} not found (unzip snap-archive.zip first)")
        return 0

    all_rows = []
    # Parse recent .xlsx files (FY16+)
    for xlsx in sorted(archive_dir.glob("FY*.xlsx")):
        print(f"  Parsing {xlsx.name}...")
        file_rows = _parse_snap_file(xlsx)
        print(f"    {len(file_rows)} records")
        all_rows.extend(file_rows)

    # Optionally parse older .xls files
    for xls in sorted(archive_dir.glob("FY*.xls")):
        if xls.suffix == ".xls":
            print(f"  Parsing {xls.name}...")
            file_rows = _parse_snap_file(xls)
            print(f"    {len(file_rows)} records")
            all_rows.extend(file_rows)

    if not all_rows:
        print("  No SNAP data parsed!")
        return 0

    print(f"  Total raw records: {len(all_rows):,}")

    # Deduplicate: keep latest if same state+year+month appears in multiple files
    seen = {}
    for r in all_rows:
        key = (r["state_code"], r["year"], r["month"])
        seen[key] = r
    deduped = list(seen.values())
    print(f"  After dedup: {len(deduped):,}")

    # Load into DuckDB
    con.execute("DROP TABLE IF EXISTS _fact_snap")
    con.execute("""
        CREATE TABLE _fact_snap (
            state_code VARCHAR, year INTEGER, month INTEGER,
            households INTEGER, persons INTEGER,
            benefit_cost DOUBLE,
            source VARCHAR, snapshot_date DATE
        )
    """)
    for r in deduped:
        con.execute("""
            INSERT INTO _fact_snap VALUES (?, ?, ?, ?, ?, ?, 'fns.usda.gov', ?)
        """, [r["state_code"], r["year"], r["month"],
              r["households"], r["persons"], r["benefit_cost"],
              SNAPSHOT_DATE])

    count = write_parquet(con, "_fact_snap", _snapshot_path("snap_enrollment"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_snap").fetchone()[0]
    years = con.execute("SELECT MIN(year), MAX(year) FROM _fact_snap").fetchone()
    print(f"  {count:,} rows, {states} states, {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_snap")
    return count


# ---------------------------------------------------------------------------
# TANF Enrollment — Parse state-level monthly data from ACF Excel files
# ---------------------------------------------------------------------------

def _parse_tanf_sheet(wb, sheet_name: str, measure: str) -> list[dict]:
    """Parse one TANF summary sheet (TFam, TRec, Adults, Children)."""
    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(values_only=True))

    # Find header row with month dates (datetime objects)
    header_row = None
    header_idx = None
    for idx, row in enumerate(rows_data):
        if row[0] == "State":
            header_row = row
            header_idx = idx
            break

    if header_row is None:
        return []

    # Extract month columns — handle both datetime and string formats
    month_cols = []
    for col_idx in range(1, min(len(header_row), 14)):
        val = header_row[col_idx]
        if isinstance(val, datetime):
            month_cols.append((col_idx, val.year, val.month))
        elif isinstance(val, str):
            if "Average" in val:
                break
            # Handle "Oct-22", "Nov-22" format
            for abbr, mnum in MONTH_ABBR.items():
                if val.startswith(abbr):
                    parts = val.split("-")
                    if len(parts) == 2:
                        try:
                            yr = int(parts[1])
                            yr = yr + 2000 if yr < 100 else yr
                            month_cols.append((col_idx, yr, mnum))
                        except ValueError:
                            pass
                    break

    records = []
    # Parse state rows (start after header, skip U.S. Totals)
    for row in rows_data[header_idx + 1:]:
        state_name = row[0]
        if not state_name or not isinstance(state_name, str):
            continue
        state_name = state_name.strip()
        if state_name in ("U.S. Totals", "U.S. Total"):
            continue
        state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            continue

        for col_idx, year, month in month_cols:
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, (int, float)) and val >= 0:
                records.append({
                    "state_code": state_code,
                    "year": year,
                    "month": month,
                    "measure": measure,
                    "value": int(val),
                })

    return records


def build_fact_tanf_enrollment(con, dry_run: bool) -> int:
    print("Building fact_tanf_enrollment...")

    all_records = []

    # Parse each TANF fiscal year file
    tanf_files = sorted(RAW_DIR.glob("tanf_caseload_fy*.xlsx"))
    if not tanf_files:
        print("  SKIPPED — no tanf_caseload_fy*.xlsx files found")
        return 0

    sheet_measures = [
        ("TFam", "total_families"),
        ("Two-par", "two_parent_families"),
        ("One-par", "one_parent_families"),
        ("Zero-par", "no_parent_families"),
        ("TRec", "total_recipients"),
        ("Adults", "adult_recipients"),
        ("Children", "child_recipients"),
    ]

    for fpath in tanf_files:
        print(f"  Parsing {fpath.name}...")
        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        file_total = 0
        for sheet_name, measure in sheet_measures:
            if sheet_name in wb.sheetnames:
                records = _parse_tanf_sheet(wb, sheet_name, measure)
                file_total += len(records)
                all_records.extend(records)
        wb.close()
        print(f"    {file_total:,} records")

    if not all_records:
        print("  No TANF data parsed!")
        return 0

    # Deduplicate
    seen = {}
    for r in all_records:
        key = (r["state_code"], r["year"], r["month"], r["measure"])
        seen[key] = r
    deduped = list(seen.values())
    print(f"  Total: {len(deduped):,} records (deduped from {len(all_records):,})")

    # Load into DuckDB
    con.execute("DROP TABLE IF EXISTS _fact_tanf")
    con.execute("""
        CREATE TABLE _fact_tanf (
            state_code VARCHAR, year INTEGER, month INTEGER,
            measure VARCHAR, value INTEGER,
            source VARCHAR, snapshot_date DATE
        )
    """)
    for r in deduped:
        con.execute("""
            INSERT INTO _fact_tanf VALUES (?, ?, ?, ?, ?, 'acf.gov', ?)
        """, [r["state_code"], r["year"], r["month"],
              r["measure"], r["value"], SNAPSHOT_DATE])

    count = write_parquet(con, "_fact_tanf", _snapshot_path("tanf_enrollment"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_tanf").fetchone()[0]
    measures = con.execute("SELECT DISTINCT measure FROM _fact_tanf ORDER BY measure").fetchall()
    print(f"  {count:,} rows, {states} states")
    print(f"  Measures: {', '.join(m[0] for m in measures)}")
    con.execute("DROP TABLE IF EXISTS _fact_tanf")
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

ALL_TABLES = {
    "fact_snap_enrollment": build_fact_snap_enrollment,
    "fact_tanf_enrollment": build_fact_tanf_enrollment,
}


def main():
    parser = argparse.ArgumentParser(description="Ingest SNAP/TANF data into Aradune lake")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only", type=str, default=None)
    args = parser.parse_args()

    tables = list(ALL_TABLES.keys())
    if args.only:
        tables = [t.strip() for t in args.only.split(",")]

    print(f"Snapshot: {SNAPSHOT_DATE}")
    print(f"Run ID:   {RUN_ID}")
    print(f"Building: {', '.join(tables)}")
    print()

    con = duckdb.connect()
    totals = {}
    for name in tables:
        totals[name] = ALL_TABLES[name](con, args.dry_run)
        print()

    con.close()

    print("=" * 60)
    print("SNAP/TANF LAKE INGESTION COMPLETE")
    print("=" * 60)
    total_rows = sum(totals.values())
    for name, count in totals.items():
        status = "written" if not args.dry_run else "dry-run"
        print(f"  {name:35s} {count:>12,} rows  [{status}]")
    print(f"  {'TOTAL':35s} {total_rows:>12,} rows")

    if not args.dry_run and total_rows > 0:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "snapshot_date": SNAPSHOT_DATE,
            "pipeline_run_id": RUN_ID,
            "created_at": datetime.now().isoformat() + "Z",
            "tables": {name: {"rows": count} for name, count in totals.items()},
            "total_rows": total_rows,
        }
        manifest_file = META_DIR / f"manifest_snap_tanf_{SNAPSHOT_DATE}.json"
        with open(manifest_file, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"\n  Manifest: {manifest_file}")

    print("\nDone.")


if __name__ == "__main__":
    main()
