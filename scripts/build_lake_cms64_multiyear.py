#!/usr/bin/env python3
"""
build_lake_cms64_multiyear.py — Parse CMS-64 FMR Excel workbooks (FY2018-2024) into a
single multi-year fact table for the Aradune data lake.

Reads from:  data/raw/fmr_historical/fmr_fyYYYY.zip
             Each ZIP contains "FY YYYY FMR NET EXPENDITURES.xlsx" (Medicaid) and a CHIP variant
Writes to:   data/lake/fact/cms64_multiyear/snapshot={DATE}/data.parquet

Output schema (extends existing fact_financial_mgmt):
  state_code VARCHAR(2)          -- 2-letter code mapped from sheet state name
  program VARCHAR                -- "Medical Assistance Program" or "Administration"
  service_category VARCHAR       -- Row-level service category label
  total_computable DOUBLE
  federal_share DOUBLE
  federal_share_medicaid DOUBLE  -- NULL for ADM sheets
  federal_share_arra DOUBLE      -- NULL for ADM sheets
  federal_share_covid DOUBLE     -- Column was "BIPP" pre-FY2020, "COVID" FY2020+
  state_share DOUBLE
  fiscal_year INTEGER
  source VARCHAR                 -- 'medicaid.gov/fmr'
  snapshot_date DATE
  snapshot DATE

Usage:
  python3 scripts/build_lake_cms64_multiyear.py
  python3 scripts/build_lake_cms64_multiyear.py --dry-run
  python3 scripts/build_lake_cms64_multiyear.py --years 2022 2023 2024
  python3 scripts/build_lake_cms64_multiyear.py --years 2024 --dry-run
"""

import argparse
import io
import json
import sys
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path
from urllib.request import urlretrieve

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "fmr_historical"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

ALL_YEARS = list(range(2018, 2025))  # FY2018-2024

FMR_DOWNLOAD_URL = (
    "https://www.medicaid.gov/medicaid/financial-management/downloads/"
    "financial-management-report-fy{year}.zip"
)

# ──────────────────────────────────────────────────────────────────────
# State/territory name -> 2-letter code (56 entries: 50 states + DC + 5 territories)
# Names must match EXACTLY how they appear in the FMR Excel sheet names
# ──────────────────────────────────────────────────────────────────────
STATE_MAP = {
    # 50 states
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    # DC
    "District of Columbia": "DC",
    "Dist. Of Col.": "DC",
    # Territories
    "Puerto Rico": "PR",
    "Guam": "GU",
    "Virgin Islands": "VI",
    "American Samoa": "AS",
    "Amer. Samoa": "AS",
    "Northern Mariana Islands": "MP",
    "N. Mariana Islands": "MP",
    # Aggregate
    "National Totals": "US",
}

# Rows to skip: summary/total rows, metadata rows
SKIP_PREFIXES = (
    "Service Category",
    "Balance",
    "Collections",
    "Total Net Expenditures",
    "Total Newly Eligible",
    "Total Not Newly",
    "Total VIII Group",
    "Total COVID",
    "Created On:",
)


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(
            f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)"
        )
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def ensure_zip(year: int) -> Path:
    """Download FMR ZIP if not already present in raw dir."""
    zip_path = RAW_DIR / f"fmr_fy{year}.zip"
    if zip_path.exists():
        return zip_path

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    url = FMR_DOWNLOAD_URL.format(year=year)
    print(f"  Downloading FY{year} FMR from {url} ...")
    try:
        urlretrieve(url, zip_path)
        size_mb = zip_path.stat().st_size / (1024 * 1024)
        print(f"  Downloaded {zip_path.name} ({size_mb:.1f} MB)")
    except Exception as e:
        print(f"  ERROR downloading FY{year}: {e}")
        if zip_path.exists():
            zip_path.unlink()
        return None
    return zip_path


def _find_medicaid_xlsx(zf: zipfile.ZipFile, year: int) -> str | None:
    """Find the Medicaid (non-CHIP) FMR Excel filename inside the ZIP."""
    for name in zf.namelist():
        lower = name.lower()
        # Match "FY YYYY FMR NET EXPENDITURES.xlsx" but NOT the CHIP variant
        if (
            str(year) in lower
            and "fmr" in lower
            and "expenditures" in lower
            and name.lower().endswith(".xlsx")
            and "chip" not in lower
        ):
            return name
    return None


def _resolve_state_code(sheet_name: str) -> tuple[str | None, str | None]:
    """Extract state name from sheet name and map to 2-letter code.

    Sheet names are like "MAP - Alabama" or "ADM - Dist. Of Col."
    Returns (state_code, state_name_from_sheet).
    """
    parts = sheet_name.split(" - ", 1)
    if len(parts) != 2:
        return None, None
    state_name = parts[1].strip()
    code = STATE_MAP.get(state_name)
    return code, state_name


def _should_skip_row(service_category: str | None) -> bool:
    """Return True if this row is a header, blank, or summary row to skip."""
    if not service_category or not isinstance(service_category, str):
        return True
    stripped = service_category.strip()
    if not stripped:
        return True
    for prefix in SKIP_PREFIXES:
        if stripped.startswith(prefix):
            return True
    return False


def _safe_float(val) -> float | None:
    """Convert cell value to float, returning None for non-numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_year(year: int, zip_path: Path) -> list[dict]:
    """Parse all sheets from one FMR Excel workbook into a list of row dicts."""
    rows = []

    zf = zipfile.ZipFile(zip_path)
    xlsx_name = _find_medicaid_xlsx(zf, year)
    if not xlsx_name:
        print(f"  WARNING: No Medicaid FMR Excel found in {zip_path.name}")
        zf.close()
        return rows

    print(f"  Parsing {xlsx_name} ...")
    wb = openpyxl.load_workbook(
        io.BytesIO(zf.read(xlsx_name)), data_only=True, read_only=True
    )

    for sheet_name in wb.sheetnames:
        # Determine sheet type: MAP or ADM
        upper = sheet_name.upper()
        if upper.startswith("MAP"):
            sheet_type = "MAP"
        elif upper.startswith("ADM"):
            sheet_type = "ADM"
        else:
            continue

        state_code, state_label = _resolve_state_code(sheet_name)
        if state_code is None:
            print(f"    SKIPPED sheet {sheet_name!r} (unmapped state)")
            continue

        # Skip National Totals -- we keep state-level data only
        if state_code == "US":
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(max_col=8, values_only=True))

        # Detect program type from row 5
        program = None
        if len(all_rows) > 5 and all_rows[5][0]:
            program = str(all_rows[5][0]).strip()
        if not program:
            program = (
                "Medical Assistance Program" if sheet_type == "MAP" else "Administration"
            )

        # Detect header row to understand column layout
        # MAP sheets: Service Category | Total Computable | Federal Share |
        #             Federal Share Medicaid | Federal Share ARRA |
        #             Federal Share BIPP/COVID | State Share
        # ADM sheets: Service Category | Total Computable | Federal Share | State Share
        #
        # Data starts at row 7 (index 7)

        for row_data in all_rows[7:]:
            service_category = row_data[0]
            if _should_skip_row(service_category):
                continue

            service_category = str(service_category).strip()

            if sheet_type == "MAP":
                # 7 columns: svc, total, fed, fed_medicaid, fed_arra, fed_covid/bipp, state
                rec = {
                    "state_code": state_code,
                    "program": program,
                    "service_category": service_category,
                    "total_computable": _safe_float(row_data[1]),
                    "federal_share": _safe_float(row_data[2]),
                    "federal_share_medicaid": _safe_float(row_data[3]),
                    "federal_share_arra": _safe_float(row_data[4]),
                    "federal_share_covid": _safe_float(row_data[5]),
                    "state_share": _safe_float(row_data[6]),
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }
            else:
                # ADM: 4 columns: svc, total, federal, state
                rec = {
                    "state_code": state_code,
                    "program": program,
                    "service_category": service_category,
                    "total_computable": _safe_float(row_data[1]),
                    "federal_share": _safe_float(row_data[2]),
                    "federal_share_medicaid": None,
                    "federal_share_arra": None,
                    "federal_share_covid": None,
                    "state_share": _safe_float(row_data[3]),
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }

            rows.append(rec)

    wb.close()
    zf.close()
    return rows


def main():
    parser = argparse.ArgumentParser(
        description="CMS-64 FMR multi-year ETL: FY2018-2024 Excel -> Parquet"
    )
    parser.add_argument(
        "--years",
        nargs="+",
        type=int,
        default=ALL_YEARS,
        help=f"Fiscal years to process (default: {ALL_YEARS[0]}-{ALL_YEARS[-1]})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate but do not write files",
    )
    args = parser.parse_args()

    print(f"=== CMS-64 FMR Multi-Year ETL ===")
    print(f"Years: {args.years}")
    print(f"Snapshot: {SNAPSHOT_DATE}")
    print(f"Run ID: {RUN_ID}")
    if args.dry_run:
        print("[DRY RUN]")
    print()

    all_rows = []
    year_stats = {}

    for year in sorted(args.years):
        print(f"--- FY{year} ---")

        # Step 1: Ensure ZIP exists (download if missing)
        zip_path = ensure_zip(year)
        if zip_path is None:
            print(f"  SKIPPED FY{year} (download failed)")
            continue

        # Step 2: Parse the Medicaid FMR Excel
        year_rows = parse_year(year, zip_path)
        print(f"  Parsed {len(year_rows):,} rows from FY{year}")

        if year_rows:
            states = set(r["state_code"] for r in year_rows)
            programs = set(r["program"] for r in year_rows)
            total_comp = sum(r["total_computable"] or 0 for r in year_rows)
            year_stats[year] = {
                "rows": len(year_rows),
                "states": len(states),
                "programs": sorted(programs),
                "total_computable": total_comp,
            }
            print(f"  {len(states)} states/territories, {sorted(programs)}")
            print(f"  Total computable: ${total_comp:,.0f}")

        all_rows.extend(year_rows)
        print()

    if not all_rows:
        print("No data parsed. Exiting.")
        sys.exit(1)

    # Step 3: Load into DuckDB and write Parquet
    print(f"=== Combined: {len(all_rows):,} rows across {len(year_stats)} years ===")

    con = duckdb.connect()

    # Create table from Python list of dicts
    con.execute("""
        CREATE TABLE _cms64_multiyear (
            state_code VARCHAR,
            program VARCHAR,
            service_category VARCHAR,
            total_computable DOUBLE,
            federal_share DOUBLE,
            federal_share_medicaid DOUBLE,
            federal_share_arra DOUBLE,
            federal_share_covid DOUBLE,
            state_share DOUBLE,
            fiscal_year INTEGER,
            source VARCHAR,
            snapshot_date DATE,
            snapshot DATE
        )
    """)

    # Insert in batches for efficiency
    BATCH_SIZE = 5000
    for i in range(0, len(all_rows), BATCH_SIZE):
        batch = all_rows[i : i + BATCH_SIZE]
        values_list = []
        for r in batch:
            values_list.append((
                r["state_code"],
                r["program"],
                r["service_category"],
                r["total_computable"],
                r["federal_share"],
                r["federal_share_medicaid"],
                r["federal_share_arra"],
                r["federal_share_covid"],
                r["state_share"],
                r["fiscal_year"],
                r["source"],
                r["snapshot_date"],
                r["snapshot"],
            ))
        con.executemany(
            """INSERT INTO _cms64_multiyear VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?::DATE, ?::DATE
            )""",
            values_list,
        )

    # Validation stats
    print()
    print("=== Validation ===")

    stats = con.execute("""
        SELECT
            COUNT(*) AS total_rows,
            COUNT(DISTINCT state_code) AS state_count,
            COUNT(DISTINCT fiscal_year) AS year_count,
            MIN(fiscal_year) AS min_year,
            MAX(fiscal_year) AS max_year,
            COUNT(DISTINCT service_category) AS category_count,
            SUM(total_computable) AS total_spending,
            SUM(federal_share) AS total_federal,
            SUM(state_share) AS total_state
        FROM _cms64_multiyear
    """).fetchone()

    print(f"  Total rows:        {stats[0]:,}")
    print(f"  States/territories: {stats[1]}")
    print(f"  Years:              {stats[2]} ({stats[3]}-{stats[4]})")
    print(f"  Service categories: {stats[5]}")
    print(f"  Total computable:   ${stats[6]:,.0f}")
    print(f"  Federal share:      ${stats[7]:,.0f}")
    print(f"  State share:        ${stats[8]:,.0f}")

    # Per-year spending
    print()
    print("  Per-year total computable:")
    year_totals = con.execute("""
        SELECT fiscal_year, COUNT(*) AS rows,
               COUNT(DISTINCT state_code) AS states,
               SUM(total_computable) AS total
        FROM _cms64_multiyear
        GROUP BY fiscal_year ORDER BY fiscal_year
    """).fetchall()
    for yr, rows, states, total in year_totals:
        print(f"    FY{yr}: {rows:,} rows, {states} states, ${total:,.0f}")

    # Per-program breakdown
    print()
    print("  Per-program breakdown:")
    prog_totals = con.execute("""
        SELECT program, COUNT(*) AS rows, SUM(total_computable) AS total
        FROM _cms64_multiyear
        GROUP BY program ORDER BY total DESC
    """).fetchall()
    for prog, rows, total in prog_totals:
        print(f"    {prog}: {rows:,} rows, ${total:,.0f}")

    # Top 5 states by total computable (most recent year)
    print()
    max_yr = stats[4]
    print(f"  Top 5 states by total computable (FY{max_yr}):")
    top_states = con.execute(f"""
        SELECT state_code, SUM(total_computable) AS total
        FROM _cms64_multiyear
        WHERE fiscal_year = {max_yr}
        GROUP BY state_code ORDER BY total DESC LIMIT 5
    """).fetchall()
    for sc, total in top_states:
        print(f"    {sc}: ${total:,.0f}")

    # Write Parquet
    out_path = _snapshot_path("cms64_multiyear")
    total_written = write_parquet(con, "_cms64_multiyear", out_path, args.dry_run)

    # Write manifest
    manifest = {
        "snapshot_date": SNAPSHOT_DATE,
        "pipeline_run_id": RUN_ID,
        "created_at": datetime.utcnow().isoformat() + "Z",
        "source_files": [
            str(RAW_DIR / f"fmr_fy{y}.zip") for y in sorted(year_stats.keys())
        ],
        "source_url": "https://www.medicaid.gov/medicaid/financial-management/state-expenditure-reports",
        "tables": {
            "fact_cms64_multiyear": {
                "rows": total_written,
                "years": sorted(year_stats.keys()),
                "states": stats[1],
                "per_year": {
                    str(yr): {"rows": ys["rows"], "states": ys["states"]}
                    for yr, ys in year_stats.items()
                },
            }
        },
        "total_rows": total_written,
        "notes": (
            "CMS-64 Financial Management Report (FMR) net expenditures by state, "
            "service category, and program (MAP/ADM). Multi-year FY2018-2024. "
            "Parsed from Excel workbooks in FMR historical ZIPs. "
            "Column 'federal_share_covid' was 'BIPP' in FY2016-2019, renamed 'COVID' FY2020+. "
            "ADM sheets have fewer columns (no Medicaid/ARRA/COVID breakdown). "
            "National Totals excluded. Source: medicaid.gov/fmr."
        ),
    }

    if not args.dry_run:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest_path = META_DIR / f"manifest_cms64_multiyear_{SNAPSHOT_DATE}.json"
        manifest_path.write_text(json.dumps(manifest, indent=2))
        print(f"\n  Manifest: {manifest_path.relative_to(LAKE_DIR)}")

    con.execute("DROP TABLE IF EXISTS _cms64_multiyear")
    con.close()

    print()
    print(f"Done. {total_written:,} rows written to fact_cms64_multiyear.")


if __name__ == "__main__":
    main()
