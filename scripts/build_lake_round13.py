"""
Round 13 data ingestion: ACO/MSSP, Part D prescribing, NHSC workforce,
FQHC quality, ACO beneficiaries by county.

Produces:
  - fact_mssp_aco — MSSP ACO organizations with track, assignment, service area
  - fact_mssp_participants — ACO participant TINs/NPIs (15K providers)
  - fact_aco_beneficiaries_county — ACO assigned beneficiaries by county (135K rows)
  - fact_aco_reach_results — ACO REACH financial & quality results
  - fact_part_d_geo — Part D prescribing by state/drug (116K rows)
  - fact_part_d_quarterly_spending — Part D quarterly drug spending (28K rows)
  - fact_nhsc_field_strength — NHSC clinician counts by state & discipline (FY2025)
  - fact_fqhc_hypertension — FQHC hypertension control rates (UDS 2019-2023)
  - fact_fqhc_quality_badges — FQHC quality recognition badges (2021-2025)
"""

import csv
import duckdb
import json
import openpyxl
import re
from pathlib import Path
from datetime import date

LAKE = Path(__file__).resolve().parent.parent / "data" / "lake"
RAW  = Path(__file__).resolve().parent.parent / "data" / "raw"
SNAP = str(date.today())


def _num(val):
    if val is None or val == '' or val == '.':
        return None
    if isinstance(val, (int, float)):
        return val
    val = str(val).strip().replace(',', '')
    try:
        return float(val)
    except ValueError:
        return None


# ── 1. MSSP ACO Organizations ───────────────────────────────────────────

def ingest_mssp_orgs(con):
    path = RAW / "mssp_aco_organizations_2026.csv"
    if not path.exists():
        print("  MSSP ACOs: file not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _mssp_aco")
    con.execute(f"""
        CREATE TABLE _mssp_aco AS
        SELECT * FROM read_csv('{path}', header=true, auto_detect=true, ignore_errors=true)
    """)

    cnt = con.execute("SELECT COUNT(*) FROM _mssp_aco").fetchone()[0]
    out_dir = LAKE / "fact" / "mssp_aco" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _mssp_aco TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  mssp_aco: {cnt:,} organizations")
    return cnt


# ── 2. MSSP ACO Participants ────────────────────────────────────────────

def ingest_mssp_participants(con):
    path = RAW / "mssp_aco_participants_2026.csv"
    if not path.exists():
        print("  MSSP participants: file not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _mssp_part")
    con.execute(f"""
        CREATE TABLE _mssp_part AS
        SELECT * FROM read_csv('{path}', header=true, auto_detect=true, ignore_errors=true)
    """)

    cnt = con.execute("SELECT COUNT(*) FROM _mssp_part").fetchone()[0]
    out_dir = LAKE / "fact" / "mssp_participants" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _mssp_part TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  mssp_participants: {cnt:,} rows")
    return cnt


# ── 3. ACO Beneficiaries by County ──────────────────────────────────────

def ingest_aco_bene_county(con):
    path = RAW / "aco_beneficiaries_by_county_2024.csv"
    if not path.exists():
        print("  ACO beneficiaries: file not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _aco_bene")
    con.execute(f"""
        CREATE TABLE _aco_bene AS
        SELECT * FROM read_csv('{path}', header=true, auto_detect=true, ignore_errors=true)
    """)

    cnt = con.execute("SELECT COUNT(*) FROM _aco_bene").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT State_Name) FROM _aco_bene").fetchone()[0]
    out_dir = LAKE / "fact" / "aco_beneficiaries_county" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _aco_bene TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  aco_beneficiaries_county: {cnt:,} rows ({states} states)")
    return cnt


# ── 4. ACO REACH Financial & Quality Results ────────────────────────────

def ingest_aco_reach(con):
    path = RAW / "aco_reach_financial_quality.csv"
    if not path.exists():
        print("  ACO REACH: file not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _aco_reach")
    con.execute(f"""
        CREATE TABLE _aco_reach AS
        SELECT * FROM read_csv('{path}', header=true, auto_detect=true, ignore_errors=true)
    """)

    cnt = con.execute("SELECT COUNT(*) FROM _aco_reach").fetchone()[0]
    out_dir = LAKE / "fact" / "aco_reach_results" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _aco_reach TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  aco_reach_results: {cnt:,} rows")
    return cnt


# ── 5. Part D Prescribing by Geography ──────────────────────────────────

def ingest_part_d_geo(con):
    path = RAW / "part_d_prescriber_geo_2023.csv"
    if not path.exists():
        print("  Part D geo: file not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _partd_geo")
    con.execute(f"""
        CREATE TABLE _partd_geo AS
        SELECT * FROM read_csv('{path}', header=true, auto_detect=true, ignore_errors=true)
    """)

    cnt = con.execute("SELECT COUNT(*) FROM _partd_geo").fetchone()[0]
    out_dir = LAKE / "fact" / "part_d_geo" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _partd_geo TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  part_d_geo: {cnt:,} rows")
    return cnt


# ── 6. Part D Quarterly Spending ────────────────────────────────────────

def ingest_part_d_spending(con):
    path = RAW / "part_d_quarterly_spending.csv"
    if not path.exists():
        print("  Part D spending: file not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _partd_spend")
    con.execute(f"""
        CREATE TABLE _partd_spend AS
        SELECT * FROM read_csv('{path}', header=true, auto_detect=true, ignore_errors=true)
    """)

    cnt = con.execute("SELECT COUNT(*) FROM _partd_spend").fetchone()[0]
    out_dir = LAKE / "fact" / "part_d_quarterly_spending" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _partd_spend TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  part_d_quarterly_spending: {cnt:,} rows")
    return cnt


# ── 7. NHSC Field Strength ──────────────────────────────────────────────

def ingest_nhsc(con):
    path = RAW / "hrsa_nhsc_field_strength_2025.xlsx"
    if not path.exists():
        print("  NHSC: file not found")
        return 0

    wb = openpyxl.load_workbook(str(path))
    rows = []

    # Main summary sheet
    ws = wb['2025 NHSC Field Strength']
    headers = [str(c.value).strip() if c.value else '' for c in ws[3]]

    for row in ws.iter_rows(min_row=4, values_only=True):
        state = str(row[0]).strip() if row[0] else ''
        if not state or state in ('', 'None', 'NATIONAL HEALTH SERVICE CORPS'):
            continue
        if 'total' in state.lower() and state != 'Total':
            continue

        rows.append({
            'state_name': state,
            'discipline': 'All',
            'fiscal_year': 2025,
            'total_clinicians': _num(row[1]),
            'nhsc_lrp': _num(row[2]),
            'nhsc_sud_lrp': _num(row[3]),
            'nhsc_rc_lrp': _num(row[4]),
            'nhsc_sp': _num(row[5]),
            's2s_lrp': _num(row[6]),
            'slrp': _num(row[7]),
            'non_rural': _num(row[8]),
            'rural': _num(row[9]),
        })

    # Primary Care sheet
    if 'Primary Care FS' in wb.sheetnames:
        ws_pc = wb['Primary Care FS']
        for row in ws_pc.iter_rows(min_row=4, values_only=True):
            state = str(row[0]).strip() if row[0] else ''
            if not state or state in ('', 'None'):
                continue
            if 'total' in state.lower() and state != 'Total':
                continue
            rows.append({
                'state_name': state,
                'discipline': 'Primary Care',
                'fiscal_year': 2025,
                'total_clinicians': _num(row[1]),
                'nhsc_lrp': _num(row[2]),
                'nhsc_sud_lrp': None,
                'nhsc_rc_lrp': None,
                'nhsc_sp': _num(row[3]),
                's2s_lrp': _num(row[4]),
                'slrp': _num(row[5]),
                'non_rural': _num(row[12]) if len(row) > 12 else None,
                'rural': _num(row[13]) if len(row) > 13 else None,
            })

    # Mental Health sheet
    if 'Mental Health FS' in wb.sheetnames:
        ws_mh = wb['Mental Health FS']
        for row in ws_mh.iter_rows(min_row=4, values_only=True):
            state = str(row[0]).strip() if row[0] else ''
            if not state or state in ('', 'None'):
                continue
            if 'total' in state.lower() and state != 'Total':
                continue
            rows.append({
                'state_name': state,
                'discipline': 'Mental Health',
                'fiscal_year': 2025,
                'total_clinicians': _num(row[1]),
                'nhsc_lrp': _num(row[2]),
                'nhsc_sud_lrp': _num(row[3]),
                'nhsc_rc_lrp': _num(row[4]),
                'nhsc_sp': _num(row[5]),
                's2s_lrp': _num(row[6]),
                'slrp': _num(row[7]),
                'non_rural': _num(row[21]) if len(row) > 21 else None,
                'rural': _num(row[22]) if len(row) > 22 else None,
            })

    if not rows:
        return 0

    con.execute("DROP TABLE IF EXISTS _nhsc")
    con.execute("""
        CREATE TABLE _nhsc (
            state_name VARCHAR, discipline VARCHAR, fiscal_year INTEGER,
            total_clinicians INTEGER, nhsc_lrp INTEGER, nhsc_sud_lrp INTEGER,
            nhsc_rc_lrp INTEGER, nhsc_sp INTEGER, s2s_lrp INTEGER, slrp INTEGER,
            non_rural INTEGER, rural INTEGER
        )
    """)
    con.executemany(
        "INSERT INTO _nhsc VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        [(r['state_name'], r['discipline'], r['fiscal_year'],
          int(r['total_clinicians']) if r['total_clinicians'] else None,
          int(r['nhsc_lrp']) if r['nhsc_lrp'] else None,
          int(r['nhsc_sud_lrp']) if r['nhsc_sud_lrp'] else None,
          int(r['nhsc_rc_lrp']) if r['nhsc_rc_lrp'] else None,
          int(r['nhsc_sp']) if r['nhsc_sp'] else None,
          int(r['s2s_lrp']) if r['s2s_lrp'] else None,
          int(r['slrp']) if r['slrp'] else None,
          int(r['non_rural']) if r['non_rural'] else None,
          int(r['rural']) if r['rural'] else None)
         for r in rows]
    )

    out_dir = LAKE / "fact" / "nhsc_field_strength" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _nhsc TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  nhsc_field_strength: {len(rows):,} rows ({len(set(r['state_name'] for r in rows))} states, {len(set(r['discipline'] for r in rows))} disciplines)")
    return len(rows)


# ── 8. FQHC Hypertension Control ────────────────────────────────────────

def ingest_fqhc_hypertension(con):
    path = RAW / "hrsa_hypertension_uds.xlsx"
    if not path.exists():
        print("  FQHC hypertension: file not found")
        return 0

    wb = openpyxl.load_workbook(str(path))
    rows = []
    for sheet_name in wb.sheetnames:
        year = int(sheet_name) if sheet_name.isdigit() else None
        if not year:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            grant = str(row[0]).strip() if row[0] else ''
            if not grant:
                continue
            rows.append({
                'year': year,
                'grant_number': grant,
                'health_center_name': str(row[1]).strip() if row[1] else '',
                'nhci_awardee': str(row[2]).strip() if row[2] else '',
                'state': str(row[3]).strip() if row[3] else '',
                'hypertension_control_pct': _num(row[4]),
            })

    if not rows:
        return 0

    con.execute("DROP TABLE IF EXISTS _fqhc_htn")
    con.execute("""
        CREATE TABLE _fqhc_htn (
            year INTEGER, grant_number VARCHAR, health_center_name VARCHAR,
            nhci_awardee VARCHAR, state VARCHAR, hypertension_control_pct DOUBLE
        )
    """)
    con.executemany(
        "INSERT INTO _fqhc_htn VALUES (?,?,?,?,?,?)",
        [(r['year'], r['grant_number'], r['health_center_name'],
          r['nhci_awardee'], r['state'], r['hypertension_control_pct'])
         for r in rows]
    )

    out_dir = LAKE / "fact" / "fqhc_hypertension" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _fqhc_htn TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  fqhc_hypertension: {len(rows):,} rows (2019-2023)")
    return len(rows)


# ── 9. FQHC Quality Recognition Badges ──────────────────────────────────

def ingest_fqhc_badges(con):
    path = RAW / "hrsa_chqr_badges.xlsx"
    if not path.exists():
        print("  FQHC badges: file not found")
        return 0

    wb = openpyxl.load_workbook(str(path))
    rows = []
    for sheet_name in wb.sheetnames:
        match = re.search(r'(\d{4})', sheet_name)
        if not match:
            continue
        year = int(match.group(1))
        ws = wb[sheet_name]

        # Get headers from row 1
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            grant = str(row[1]).strip() if row[1] else ''
            if not grant or grant == 'None':
                continue
            state = str(row[3]).strip() if row[3] else ''
            if not state:
                continue

            # Count badges earned
            badge_count = sum(1 for i in range(4, len(row)) if row[i] and str(row[i]).strip().lower() == 'yes')

            rows.append({
                'year': year,
                'hc_type': str(row[0]).strip() if row[0] else '',
                'grant_number': grant,
                'grantee_name': str(row[2]).strip() if row[2] else '',
                'state': state,
                'badge_count': badge_count,
                'advancing_hit': 1 if len(row) > 4 and row[4] and str(row[4]).strip().lower() == 'yes' else 0,
                'quality_leader_gold': 1 if len(row) > 5 and row[5] and str(row[5]).strip().lower() == 'yes' else 0,
                'quality_leader_silver': 1 if len(row) > 6 and row[6] and str(row[6]).strip().lower() == 'yes' else 0,
                'quality_leader_bronze': 1 if len(row) > 7 and row[7] and str(row[7]).strip().lower() == 'yes' else 0,
            })

    if not rows:
        return 0

    con.execute("DROP TABLE IF EXISTS _fqhc_badges")
    con.execute("""
        CREATE TABLE _fqhc_badges (
            year INTEGER, hc_type VARCHAR, grant_number VARCHAR,
            grantee_name VARCHAR, state VARCHAR, badge_count INTEGER,
            advancing_hit INTEGER, quality_leader_gold INTEGER,
            quality_leader_silver INTEGER, quality_leader_bronze INTEGER
        )
    """)
    con.executemany(
        "INSERT INTO _fqhc_badges VALUES (?,?,?,?,?,?,?,?,?,?)",
        [(r['year'], r['hc_type'], r['grant_number'], r['grantee_name'],
          r['state'], r['badge_count'], r['advancing_hit'],
          r['quality_leader_gold'], r['quality_leader_silver'],
          r['quality_leader_bronze']) for r in rows]
    )

    out_dir = LAKE / "fact" / "fqhc_quality_badges" / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY _fqhc_badges TO '{out_dir / 'data.parquet'}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  fqhc_quality_badges: {len(rows):,} rows (2021-2025)")
    return len(rows)


# ── Main ────────────────────────────────────────────────────────────────

def main():
    con = duckdb.connect()
    total = 0

    print("── Round 13 Data Ingestion ──\n")

    total += ingest_mssp_orgs(con)
    total += ingest_mssp_participants(con)
    total += ingest_aco_bene_county(con)
    total += ingest_aco_reach(con)
    total += ingest_part_d_geo(con)
    total += ingest_part_d_spending(con)
    total += ingest_nhsc(con)
    total += ingest_fqhc_hypertension(con)
    total += ingest_fqhc_badges(con)

    con.close()

    print(f"\n── Total: {total:,} rows across round 13 tables ──")

    manifest = {
        "pipeline_run": "round13",
        "snapshot_date": SNAP,
        "total_rows": total,
        "tables": [
            "fact_mssp_aco", "fact_mssp_participants",
            "fact_aco_beneficiaries_county", "fact_aco_reach_results",
            "fact_part_d_geo", "fact_part_d_quarterly_spending",
            "fact_nhsc_field_strength", "fact_fqhc_hypertension",
            "fact_fqhc_quality_badges",
        ],
    }
    manifest_path = LAKE / "metadata" / f"manifest_round13_{SNAP}.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"  Manifest: {manifest_path}")


if __name__ == "__main__":
    main()
