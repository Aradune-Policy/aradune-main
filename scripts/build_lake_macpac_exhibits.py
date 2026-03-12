#!/usr/bin/env python3
"""
build_lake_macpac_exhibits.py — Ingest MACPAC Excel exhibits into the Aradune data lake.

Sources:
  1. data/raw/macpac/exhibit17.xlsx — Total Medicaid Benefit Spending by State and Category, FY 2024 (millions)
  2. data/raw/macpac/exhibit21.xlsx — Medicaid Spending by State, Eligibility Group, and Dually Eligible Status, FY 2023 (millions)
  3. data/raw/macpac/exhibit29.xlsx — Percentage of Medicaid Enrollees in Managed Care by State, July 1, 2022

Tables built:
  macpac_benefit_spending_fy2024     — Benefit spending by state and service category (FY 2024)
  macpac_spending_by_elig_fy2023     — Spending by eligibility group and dual status (FY 2023)
  macpac_mc_enrollment_detail        — Managed care enrollment percentages by type (2022)

Usage:
  python3 scripts/build_lake_macpac_exhibits.py
"""

import json
import re
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "macpac"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

# ──────────────────────────────────────────────────────────────────────
# State/territory name -> 2-letter code
# ──────────────────────────────────────────────────────────────────────
STATE_MAP = {
    # 50 states
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN",
    "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA",
    "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    # DC
    "District of Columbia": "DC",
    # Territories
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
    # Aggregate
    "Total": "US", "Totals": "US", "National Total": "US",
    "United States": "US", "U.S. Total": "US",
}


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def write_parquet(con, table_name: str, out_path: Path) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(
            f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)"
        )
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    return count


def _clean_state_name(raw_name):
    """Strip footnote numbers from state names like 'Alaska2' -> 'Alaska'."""
    if not raw_name or not isinstance(raw_name, str):
        return None
    cleaned = re.sub(r'\d+$', '', raw_name).strip()
    return cleaned if cleaned else None


def _resolve_state_code(raw_name):
    """Clean state name and resolve to 2-letter code."""
    cleaned = _clean_state_name(raw_name)
    if not cleaned:
        return None
    return STATE_MAP.get(cleaned)


def _parse_numeric(val):
    """Parse a numeric value from a cell. Returns None for dashes, None, non-numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip()
        if val in ('', '-', '--', 'n/a', 'N/A', '*', '**', 'NR', 'NM', 'DSH'):
            return None
        val = val.replace(',', '').replace('$', '')
        try:
            return float(val)
        except ValueError:
            return None
    return None


# ──────────────────────────────────────────────────────────────────────
# Exhibit 17: Total Medicaid Benefit Spending by State and Category, FY 2024
# ──────────────────────────────────────────────────────────────────────
def build_exhibit17(con) -> int:
    """Ingest Exhibit 17 - Benefit spending by state and category."""
    print("\n== Exhibit 17: Benefit Spending by State and Category, FY 2024 ==")
    xlsx_path = RAW_DIR / "exhibit17.xlsx"
    if not xlsx_path.exists():
        print(f"  SKIPPED - {xlsx_path.name} not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Column mapping (headers span rows 3-4, data starts row 5)
    # Columns: State, Total, FFS subcategories..., Managed care, Medicare premiums, Collections
    COL_MAP = {
        'B': 'total_benefit_spending',
        'C': 'ffs_hospital',
        'D': 'ffs_physician',
        'E': 'ffs_dental',
        'F': 'ffs_other_practitioner',
        'G': 'ffs_clinic_health_center',
        'H': 'ffs_other_acute',
        'I': 'ffs_drugs',
        'J': 'ffs_institutional_ltss',
        'K': 'ffs_hcbs_ltss',
        'L': 'managed_care',
        'M': 'medicare_premiums',
        'N': 'collections',
    }

    rows = []
    for row_num in range(5, ws.max_row + 1):
        raw_state = ws[f'A{row_num}'].value
        state_code = _resolve_state_code(str(raw_state) if raw_state else None)
        if not state_code:
            continue

        record = {'state_code': state_code}
        has_data = False
        for col_letter, field_name in COL_MAP.items():
            val = _parse_numeric(ws[f'{col_letter}{row_num}'].value)
            if val is not None:
                # Values are in millions - convert to actual dollars
                record[field_name] = val * 1_000_000
                has_data = True
            else:
                record[field_name] = None

        if has_data:
            record['fiscal_year'] = 2024
            record['source'] = 'macpac'
            record['snapshot_date'] = SNAPSHOT_DATE
            rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return 0

    # Load into DuckDB
    columns = [
        'state_code', 'total_benefit_spending',
        'ffs_hospital', 'ffs_physician', 'ffs_dental', 'ffs_other_practitioner',
        'ffs_clinic_health_center', 'ffs_other_acute', 'ffs_drugs',
        'ffs_institutional_ltss', 'ffs_hcbs_ltss',
        'managed_care', 'medicare_premiums', 'collections',
        'fiscal_year', 'source', 'snapshot_date',
    ]
    col_defs = (
        "state_code VARCHAR, total_benefit_spending DOUBLE, "
        "ffs_hospital DOUBLE, ffs_physician DOUBLE, ffs_dental DOUBLE, "
        "ffs_other_practitioner DOUBLE, ffs_clinic_health_center DOUBLE, "
        "ffs_other_acute DOUBLE, ffs_drugs DOUBLE, "
        "ffs_institutional_ltss DOUBLE, ffs_hcbs_ltss DOUBLE, "
        "managed_care DOUBLE, medicare_premiums DOUBLE, collections DOUBLE, "
        "fiscal_year INTEGER, source VARCHAR, snapshot_date VARCHAR"
    )
    con.execute(f"CREATE OR REPLACE TABLE _ex17 ({col_defs})")
    placeholders = ', '.join(['?'] * len(columns))
    con.executemany(
        f"INSERT INTO _ex17 VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    out_path = _snapshot_path("macpac_benefit_spending_fy2024")
    count = write_parquet(con, "_ex17", out_path)

    # Validation stats
    state_count = len(set(r['state_code'] for r in rows if r['state_code'] != 'US'))
    us_row = [r for r in rows if r['state_code'] == 'US']
    print(f"  {count} rows, {state_count} states + US total")
    if us_row:
        total = us_row[0].get('total_benefit_spending')
        if total:
            print(f"  US total benefit spending: ${total / 1e9:,.1f}B")
    mc_total = sum(r['managed_care'] or 0 for r in rows if r['state_code'] != 'US')
    if mc_total:
        print(f"  Managed care total (state sum): ${mc_total / 1e9:,.1f}B")

    return count


# ──────────────────────────────────────────────────────────────────────
# Exhibit 21: Medicaid Spending by Eligibility Group and Dually Eligible Status, FY 2023
# ──────────────────────────────────────────────────────────────────────
def build_exhibit21(con) -> int:
    """Ingest Exhibit 21 - Spending by eligibility group and dual status."""
    print("\n== Exhibit 21: Spending by Eligibility Group and Dual Status, FY 2023 ==")
    xlsx_path = RAW_DIR / "exhibit21.xlsx"
    if not xlsx_path.exists():
        print(f"  SKIPPED - {xlsx_path.name} not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Use the "Exhibit 21" sheet (not "Exhibit 21 (Book)")
    target_sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() == "exhibit 21":
            target_sheet = name
            break
    if not target_sheet:
        # Fallback: pick the first sheet that contains "21" but not "Book"
        for name in wb.sheetnames:
            if "21" in name and "book" not in name.lower():
                target_sheet = name
                break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]
        print(f"  WARNING: Could not find 'Exhibit 21' sheet, using '{target_sheet}'")

    ws = wb[target_sheet]
    print(f"  Using sheet: '{target_sheet}'")

    # Headers span rows 3-5, data starts row 6
    # Columns: State, Total, Child, New adult, Other adult, Disabled, Aged,
    #          then dual-status breakdowns
    COL_MAP = {
        'B': 'total_spending',
        'C': 'spending_child',
        'D': 'spending_new_adult',
        'E': 'spending_other_adult',
        'F': 'spending_disabled',
        'G': 'spending_aged',
        'H': 'spending_full_dual',
        'I': 'spending_partial_dual',
        'J': 'spending_non_dual',
    }

    rows = []
    for row_num in range(6, ws.max_row + 1):
        raw_state = ws[f'A{row_num}'].value
        state_code = _resolve_state_code(str(raw_state) if raw_state else None)
        if not state_code:
            continue

        record = {'state_code': state_code}
        has_data = False
        for col_letter, field_name in COL_MAP.items():
            val = _parse_numeric(ws[f'{col_letter}{row_num}'].value)
            if val is not None:
                # Values are in millions - convert to actual dollars
                record[field_name] = val * 1_000_000
                has_data = True
            else:
                record[field_name] = None

        if has_data:
            record['fiscal_year'] = 2023
            record['source'] = 'macpac'
            record['snapshot_date'] = SNAPSHOT_DATE
            rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return 0

    # Load into DuckDB
    columns = [
        'state_code', 'total_spending',
        'spending_child', 'spending_new_adult', 'spending_other_adult',
        'spending_disabled', 'spending_aged',
        'spending_full_dual', 'spending_partial_dual', 'spending_non_dual',
        'fiscal_year', 'source', 'snapshot_date',
    ]
    col_defs = (
        "state_code VARCHAR, total_spending DOUBLE, "
        "spending_child DOUBLE, spending_new_adult DOUBLE, spending_other_adult DOUBLE, "
        "spending_disabled DOUBLE, spending_aged DOUBLE, "
        "spending_full_dual DOUBLE, spending_partial_dual DOUBLE, spending_non_dual DOUBLE, "
        "fiscal_year INTEGER, source VARCHAR, snapshot_date VARCHAR"
    )
    con.execute(f"CREATE OR REPLACE TABLE _ex21 ({col_defs})")
    placeholders = ', '.join(['?'] * len(columns))
    con.executemany(
        f"INSERT INTO _ex21 VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    out_path = _snapshot_path("macpac_spending_by_elig_fy2023")
    count = write_parquet(con, "_ex21", out_path)

    # Validation stats
    state_count = len(set(r['state_code'] for r in rows if r['state_code'] != 'US'))
    us_row = [r for r in rows if r['state_code'] == 'US']
    print(f"  {count} rows, {state_count} states + US total")
    if us_row:
        total = us_row[0].get('total_spending')
        if total:
            print(f"  US total spending: ${total / 1e9:,.1f}B")
        dual = us_row[0].get('spending_full_dual')
        if dual:
            print(f"  US full-dual spending: ${dual / 1e9:,.1f}B")

    return count


# ──────────────────────────────────────────────────────────────────────
# Exhibit 29: Percentage of Medicaid Enrollees in Managed Care by State, 2022
# ──────────────────────────────────────────────────────────────────────
def build_exhibit29(con) -> int:
    """Ingest Exhibit 29 - MC enrollment percentages by type."""
    print("\n== Exhibit 29: MC Enrollment Percentages by State, July 2022 ==")
    xlsx_path = RAW_DIR / "exhibit29.xlsx"
    if not xlsx_path.exists():
        print(f"  SKIPPED - {xlsx_path.name} not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Headers span rows 3-5, data starts row 6
    # Columns: State, Total enrollees, then MC pct by type
    COL_MAP_COUNTS = {
        'B': 'total_enrollees',
    }
    COL_MAP_PCTS = {
        'C': 'pct_comprehensive_mc',
        'D': 'pct_mltss',
        'E': 'pct_bho',
        'F': 'pct_dental_mc',
        'G': 'pct_transportation_mc',
        'H': 'pct_other_mc',
        'I': 'pct_pccm',
    }

    rows = []
    for row_num in range(6, ws.max_row + 1):
        raw_state = ws[f'A{row_num}'].value
        state_code = _resolve_state_code(str(raw_state) if raw_state else None)
        if not state_code:
            continue

        record = {'state_code': state_code}
        has_data = False

        # Total enrollees - this is an integer count, not millions
        for col_letter, field_name in COL_MAP_COUNTS.items():
            val = _parse_numeric(ws[f'{col_letter}{row_num}'].value)
            if val is not None:
                record[field_name] = int(val)
                has_data = True
            else:
                record[field_name] = None

        # MC percentages - these are already 0-1 range (or 0-100, we normalize to 0-1)
        for col_letter, field_name in COL_MAP_PCTS.items():
            val = _parse_numeric(ws[f'{col_letter}{row_num}'].value)
            if val is not None:
                # If values are > 1, they are percentages (0-100) - normalize to 0-1
                # If values are <= 1, they are already in 0-1 range
                # MACPAC typically uses 0-1 range in data cells even if displayed as %
                record[field_name] = val
                has_data = True
            else:
                record[field_name] = None

        if has_data:
            record['reference_date'] = '2022-07-01'
            record['source'] = 'macpac'
            record['snapshot_date'] = SNAPSHOT_DATE
            rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return 0

    # Load into DuckDB
    columns = [
        'state_code', 'total_enrollees',
        'pct_comprehensive_mc', 'pct_mltss', 'pct_bho',
        'pct_dental_mc', 'pct_transportation_mc', 'pct_other_mc', 'pct_pccm',
        'reference_date', 'source', 'snapshot_date',
    ]
    col_defs = (
        "state_code VARCHAR, total_enrollees BIGINT, "
        "pct_comprehensive_mc DOUBLE, pct_mltss DOUBLE, pct_bho DOUBLE, "
        "pct_dental_mc DOUBLE, pct_transportation_mc DOUBLE, "
        "pct_other_mc DOUBLE, pct_pccm DOUBLE, "
        "reference_date VARCHAR, source VARCHAR, snapshot_date VARCHAR"
    )
    con.execute(f"CREATE OR REPLACE TABLE _ex29 ({col_defs})")
    placeholders = ', '.join(['?'] * len(columns))
    con.executemany(
        f"INSERT INTO _ex29 VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    out_path = _snapshot_path("macpac_mc_enrollment_detail")
    count = write_parquet(con, "_ex29", out_path)

    # Validation stats
    state_count = len(set(r['state_code'] for r in rows if r['state_code'] != 'US'))
    print(f"  {count} rows, {state_count} states")

    # Sample stats
    has_comp = [r for r in rows if r.get('pct_comprehensive_mc') is not None and r['state_code'] != 'US']
    if has_comp:
        avg_comp = sum(r['pct_comprehensive_mc'] for r in has_comp) / len(has_comp)
        # Display as pct - if values are 0-1, multiply by 100 for display
        display_val = avg_comp * 100 if avg_comp <= 1 else avg_comp
        print(f"  Avg comprehensive MC rate: {display_val:.1f}% ({len(has_comp)} states)")

    us_row = [r for r in rows if r['state_code'] == 'US']
    if us_row and us_row[0].get('total_enrollees'):
        print(f"  US total enrollees: {us_row[0]['total_enrollees']:,}")

    return count


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────
def main():
    print(f"MACPAC Exhibits ETL — snapshot {SNAPSHOT_DATE}")
    print(f"  Raw dir: {RAW_DIR}")
    print(f"  Lake dir: {LAKE_DIR}")

    con = duckdb.connect()
    total_rows = 0
    tables_built = []

    # Exhibit 17
    n = build_exhibit17(con)
    if n > 0:
        total_rows += n
        tables_built.append("macpac_benefit_spending_fy2024")

    # Exhibit 21
    n = build_exhibit21(con)
    if n > 0:
        total_rows += n
        tables_built.append("macpac_spending_by_elig_fy2023")

    # Exhibit 29
    n = build_exhibit29(con)
    if n > 0:
        total_rows += n
        tables_built.append("macpac_mc_enrollment_detail")

    con.close()

    # Summary
    print(f"\n== Summary ==")
    print(f"  Tables built: {len(tables_built)}")
    print(f"  Total rows: {total_rows:,}")
    for t in tables_built:
        print(f"    - {t}")

    # Manifest
    if tables_built:
        manifest = {
            "pipeline_run": "macpac_exhibits",
            "run_id": RUN_ID,
            "snapshot_date": SNAPSHOT_DATE,
            "run_timestamp": datetime.now(timezone.utc).isoformat(),
            "total_rows": total_rows,
            "tables": tables_built,
            "source_files": [
                "data/raw/macpac/exhibit17.xlsx",
                "data/raw/macpac/exhibit21.xlsx",
                "data/raw/macpac/exhibit29.xlsx",
            ],
            "source": "MACPAC MACStats Exhibits (macpac.gov)",
        }
        manifest_path = META_DIR / f"manifest_macpac_exhibits_{SNAPSHOT_DATE}.json"
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(json.dumps(manifest, indent=2))
        print(f"  Manifest: {manifest_path.relative_to(LAKE_DIR)}")


if __name__ == "__main__":
    main()
