#!/usr/bin/env python3
"""
build_lake_cms64_historical.py — Parse CMS-64 FMR Excel workbooks (FY1997-2017) into a
single historical fact table for the Aradune data lake.

Handles 3 distinct file formats across 4 eras:
  Format A (FY2013-2017): Per-state MAP/ADM sheets, 7 MAP cols (incl. Medicaid/ARRA/BIPP)
  Format B (FY2012):      Per-state single sheet, MAP + ADM side-by-side, same columns as A
  Format C (FY1997-2011): One sheet per year, all states stacked vertically,
                          4 MAP cols (total/fed/state), 4 ADM cols (total/fed/state)

Downloads from:
  https://www.medicaid.gov/medicaid/downloads/financial-management-report-fyXXXX.zip
  https://www.medicaid.gov/medicaid/downloads/fy02throughfy11netexpenditure.zip
  https://www.medicaid.gov/medicaid/downloads/financial-management-report-fy1997-2001.zip

Reads from:  data/raw/fmr_historical/
Writes to:   data/lake/fact/cms64_historical/snapshot={DATE}/data.parquet

Output schema (matches cms64_multiyear for union compatibility):
  state_code VARCHAR(2)
  program VARCHAR                -- "Medical Assistance Program" or "Administration"
  service_category VARCHAR
  total_computable DOUBLE
  federal_share DOUBLE
  federal_share_medicaid DOUBLE  -- NULL for FY1997-2011 and ADM sheets
  federal_share_arra DOUBLE      -- NULL for FY1997-2011 and ADM sheets
  federal_share_covid DOUBLE     -- "BIPP" column; NULL for FY1997-2011 and ADM sheets
  state_share DOUBLE
  fiscal_year INTEGER
  source VARCHAR                 -- 'medicaid.gov/fmr'
  snapshot_date DATE
  snapshot DATE

Usage:
  python3 scripts/build_lake_cms64_historical.py
  python3 scripts/build_lake_cms64_historical.py --dry-run
  python3 scripts/build_lake_cms64_historical.py --years 2014 2015 2016 2017
"""

import argparse
import io
import json
import subprocess
import sys
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "fmr_historical"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

ALL_YEARS = list(range(1997, 2018))  # FY1997-2017

# Download URLs vary by era
DOWNLOAD_URLS = {
    # FY2014-2017: individual year ZIPs (different base path than FY2018+)
    2017: "https://www.medicaid.gov/medicaid/downloads/financial-management-report-fy2017.zip",
    2016: "https://www.medicaid.gov/medicaid/downloads/financial-management-report-fy2016.zip",
    2015: "https://www.medicaid.gov/medicaid/downloads/financial-management-report-fy2015.zip",
    2014: "https://www.medicaid.gov/medicaid/downloads/financial-management-report-fy2014.zip",
    # FY2012-2013: bundled ZIP
    "2012-13": "https://www.medicaid.gov/medicaid/downloads/financial-management-report-fy2012-13.zip",
    # FY2002-2011: single file with 10 sheets
    "2002-11": "https://www.medicaid.gov/medicaid/downloads/fy02throughfy11netexpenditure.zip",
    # FY1997-2001: single file with 5 sheets
    "1997-2001": "https://www.medicaid.gov/medicaid/downloads/financial-management-report-fy1997-2001.zip",
}

# Map download key to local filename
ZIP_FILENAMES = {
    2017: "fmr_fy2017.zip",
    2016: "fmr_fy2016.zip",
    2015: "fmr_fy2015.zip",
    2014: "fmr_fy2014.zip",
    "2012-13": "fmr_fy2012-13.zip",
    "2002-11": "fmr_fy2002-11.zip",
    "1997-2001": "fmr_fy1997-2001.zip",
}

# Which years come from which ZIP
YEAR_TO_ZIP_KEY = {}
for y in range(2014, 2018):
    YEAR_TO_ZIP_KEY[y] = y
for y in range(2012, 2014):
    YEAR_TO_ZIP_KEY[y] = "2012-13"
for y in range(2002, 2012):
    YEAR_TO_ZIP_KEY[y] = "2002-11"
for y in range(1997, 2002):
    YEAR_TO_ZIP_KEY[y] = "1997-2001"

# Excel filenames inside each ZIP (keyed by year)
XLSX_NAMES = {
    2017: "FY 2017 FMR NET EXPENDITURES.xlsx",
    2016: "FY 2016 FMR NET EXPENDITURES.xlsx",
    2015: "FY 2015 NET EXPENDITURES.xlsx",
    2014: "FMR Net Expenditures FY14.xlsx",
    2013: "FMR Net Expenditures FY13.xlsx",
    2012: "FMR Net Expenditures FY12.xlsx",
    # FY2002-2011: single file, one sheet per year
    "2002-11": "NetExpenditure02through11.xlsx",
    # FY1997-2001: single file, one sheet per year
    "1997-2001": "FMR1997through2001.xlsx",
}

# Sheet names in multi-year files
MULTI_YEAR_SHEET_NAMES = {
    2002: "2002", 2003: "2003", 2004: "2004", 2005: "2005", 2006: "2006",
    2007: "2007", 2008: "2008", 2009: "2009", 2010: "2010", 2011: "2011",
    1997: "FMR1997", 1998: "FMR1998", 1999: "FMR1999", 2000: "FMR2000", 2001: "FMR2001",
}

# ──────────────────────────────────────────────────────────────────────
# State/territory name -> 2-letter code
# Names must match EXACTLY how they appear in the FMR Excel sheet names
# ──────────────────────────────────────────────────────────────────────
STATE_MAP = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN",
    "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA",
    "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    # DC
    "District of Columbia": "DC", "Dist. Of Col.": "DC",
    # Territories
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "American Samoa": "AS", "Amer. Samoa": "AS",
    "Northern Mariana Islands": "MP", "N. Mariana Islands": "MP",
    # Historical oddities
    "Mass. Blind": "MA",  # Massachusetts Commission for the Blind (FY1997-~2001)
    # Aggregate
    "National Totals": "US",
}

# Rows to skip: summary/total rows, metadata rows
SKIP_PREFIXES = (
    "Service Category",
    "Balance",
    "Collections",
    "Total Net Expenditures",
    "Total Net",
    "Total Newly Eligible",
    "Total Not Newly",
    "Total VIII Group",
    "Total COVID",
    "Created On:",
    "Created on:",
    "Medical Assistance Program",
    "Administration",
    "Medicaid Financial",
    "Medicaid Finanacial",
    "FY ",
    "Net Services",
    "National",
)

# CHIP row prefix (skip these - we want Medicaid only)
CHIP_PREFIX = "C-"


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(
            f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)"
        )
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def ensure_zip(zip_key) -> Path | None:
    """Download FMR ZIP if not already present in raw dir."""
    filename = ZIP_FILENAMES[zip_key]
    zip_path = RAW_DIR / filename
    if zip_path.exists() and zip_path.stat().st_size > 1000:
        return zip_path

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    url = DOWNLOAD_URLS[zip_key]
    print(f"  Downloading {filename} from {url} ...")
    try:
        result = subprocess.run(
            ["curl", "-sL", "-o", str(zip_path), url],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            print(f"  ERROR downloading: {result.stderr}")
            if zip_path.exists():
                zip_path.unlink()
            return None
        size_mb = zip_path.stat().st_size / (1024 * 1024)
        print(f"  Downloaded {filename} ({size_mb:.1f} MB)")
    except Exception as e:
        print(f"  ERROR downloading: {e}")
        if zip_path.exists():
            zip_path.unlink()
        return None
    return zip_path


def _safe_float(val) -> float | None:
    """Convert cell value to float, returning None for non-numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _should_skip_row(service_category: str | None) -> bool:
    """Return True if this row is a header, blank, summary, or CHIP row to skip."""
    if not service_category or not isinstance(service_category, str):
        return True
    stripped = service_category.strip()
    if not stripped:
        return True
    # Skip CHIP rows (prefixed with "C-")
    if stripped.startswith(CHIP_PREFIX):
        return True
    for prefix in SKIP_PREFIXES:
        if stripped.startswith(prefix):
            return True
    return False


def _resolve_state_code(name: str) -> str | None:
    """Map state name to 2-letter code."""
    name = name.strip()
    code = STATE_MAP.get(name)
    return code


# ──────────────────────────────────────────────────────────────────────
# Format A: FY2013-2017 (per-state MAP/ADM sheets)
# Same structure as the existing build_lake_cms64_multiyear.py
# ──────────────────────────────────────────────────────────────────────

def parse_format_a(year: int, zip_path: Path) -> list[dict]:
    """Parse FY2013-2017 workbooks (per-state MAP/ADM sheets)."""
    rows = []
    zf = zipfile.ZipFile(zip_path)
    xlsx_name = XLSX_NAMES.get(year)
    if xlsx_name not in zf.namelist():
        # Try to find it
        for name in zf.namelist():
            if name.lower().endswith(".xlsx") and "chip" not in name.lower():
                xlsx_name = name
                break
    if not xlsx_name or xlsx_name not in zf.namelist():
        print(f"  WARNING: No Medicaid FMR Excel found in {zip_path.name}")
        zf.close()
        return rows

    print(f"  Parsing {xlsx_name} ...")
    wb = openpyxl.load_workbook(
        io.BytesIO(zf.read(xlsx_name)), data_only=True, read_only=True
    )

    for sheet_name in wb.sheetnames:
        upper = sheet_name.upper()
        if upper.startswith("MAP"):
            sheet_type = "MAP"
        elif upper.startswith("ADM"):
            sheet_type = "ADM"
        else:
            continue

        # Extract state from "MAP - Alabama" or "ADM - Dist. Of Col."
        parts = sheet_name.split(" - ", 1)
        if len(parts) != 2:
            continue
        state_name = parts[1].strip()
        state_code = _resolve_state_code(state_name)
        if state_code is None:
            print(f"    SKIPPED sheet {sheet_name!r} (unmapped state)")
            continue
        if state_code == "US":
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(max_col=8, values_only=True))

        # Detect program type
        program = None
        if len(all_rows) > 5 and all_rows[5] and all_rows[5][0]:
            program = str(all_rows[5][0]).strip()
        if not program:
            program = "Medical Assistance Program" if sheet_type == "MAP" else "Administration"

        # Data starts at row 7 (index 7)
        for row_data in all_rows[7:]:
            service_category = row_data[0]
            if _should_skip_row(service_category):
                continue
            service_category = str(service_category).strip()

            if sheet_type == "MAP":
                rec = {
                    "state_code": state_code,
                    "program": program,
                    "service_category": service_category,
                    "total_computable": _safe_float(row_data[1]),
                    "federal_share": _safe_float(row_data[2]),
                    "federal_share_medicaid": _safe_float(row_data[3]),
                    "federal_share_arra": _safe_float(row_data[4]),
                    "federal_share_covid": _safe_float(row_data[5]),
                    "state_share": _safe_float(row_data[6]),
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }
            else:
                rec = {
                    "state_code": state_code,
                    "program": program,
                    "service_category": service_category,
                    "total_computable": _safe_float(row_data[1]),
                    "federal_share": _safe_float(row_data[2]),
                    "federal_share_medicaid": None,
                    "federal_share_arra": None,
                    "federal_share_covid": None,
                    "state_share": _safe_float(row_data[3]),
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }
            rows.append(rec)

    wb.close()
    zf.close()
    return rows


# ──────────────────────────────────────────────────────────────────────
# Format B: FY2012 (per-state single sheet, MAP + ADM side-by-side)
# ──────────────────────────────────────────────────────────────────────

def parse_format_b(year: int, zip_path: Path) -> list[dict]:
    """Parse FY2012 workbook (per-state sheets, MAP + ADM side-by-side)."""
    rows = []
    zf = zipfile.ZipFile(zip_path)
    xlsx_name = XLSX_NAMES.get(year)
    if xlsx_name not in zf.namelist():
        for name in zf.namelist():
            if name.lower().endswith(".xlsx") and "chip" not in name.lower() and str(year)[-2:] in name:
                xlsx_name = name
                break

    if not xlsx_name or xlsx_name not in zf.namelist():
        print(f"  WARNING: No Medicaid FMR Excel for FY{year} in {zip_path.name}")
        zf.close()
        return rows

    print(f"  Parsing {xlsx_name} ...")
    wb = openpyxl.load_workbook(
        io.BytesIO(zf.read(xlsx_name)), data_only=True, read_only=True
    )

    for sheet_name in wb.sheetnames:
        state_code = _resolve_state_code(sheet_name)
        if state_code is None:
            print(f"    SKIPPED sheet {sheet_name!r} (unmapped state)")
            continue
        if state_code == "US":
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(max_col=12, values_only=True))

        # Row 5: "Medical Assistance Program" ... "Administration"
        # Row 6: headers
        # Row 7+: data
        # MAP columns: 0=svc, 1=total, 2=fed, 3=fed_medicaid, 4=fed_arra, 5=fed_bipp, 6=state
        # ADM columns: 7=svc, 8=total, 9=fed, 10=state

        for row_data in all_rows[7:]:
            # MAP side
            map_svc = row_data[0] if len(row_data) > 0 else None
            if not _should_skip_row(map_svc):
                map_svc = str(map_svc).strip()
                rec = {
                    "state_code": state_code,
                    "program": "Medical Assistance Program",
                    "service_category": map_svc,
                    "total_computable": _safe_float(row_data[1]),
                    "federal_share": _safe_float(row_data[2]),
                    "federal_share_medicaid": _safe_float(row_data[3]),
                    "federal_share_arra": _safe_float(row_data[4]),
                    "federal_share_covid": _safe_float(row_data[5]),
                    "state_share": _safe_float(row_data[6]),
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }
                rows.append(rec)

            # ADM side
            adm_svc = row_data[7] if len(row_data) > 7 else None
            if not _should_skip_row(adm_svc):
                adm_svc = str(adm_svc).strip()
                rec = {
                    "state_code": state_code,
                    "program": "Administration",
                    "service_category": adm_svc,
                    "total_computable": _safe_float(row_data[8]),
                    "federal_share": _safe_float(row_data[9]),
                    "federal_share_medicaid": None,
                    "federal_share_arra": None,
                    "federal_share_covid": None,
                    "state_share": _safe_float(row_data[10]),
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }
                rows.append(rec)

    wb.close()
    zf.close()
    return rows


# ──────────────────────────────────────────────────────────────────────
# Format C: FY1997-2011 (one sheet per year, all states stacked)
# ──────────────────────────────────────────────────────────────────────

def parse_format_c(year: int, zip_path: Path) -> list[dict]:
    """Parse FY1997-2011 workbooks (one sheet per year, states stacked vertically).

    Two sub-layouts detected from header row:
      FY1997-2008: MAP cols 0-3 (svc, total, fed, state), ADM cols 4-7
      FY2009-2011: MAP cols 0-5 (svc, total, fed, fed_medicaid, fed_arra, state), ADM cols 6-9
    """
    rows = []
    zip_key = YEAR_TO_ZIP_KEY[year]
    zf = zipfile.ZipFile(zip_path)

    # Determine which Excel file and sheet
    if year >= 2002:
        xlsx_name = XLSX_NAMES["2002-11"]
        sheet_name = MULTI_YEAR_SHEET_NAMES[year]
    else:
        xlsx_name = XLSX_NAMES["1997-2001"]
        sheet_name = MULTI_YEAR_SHEET_NAMES[year]

    if xlsx_name not in zf.namelist():
        print(f"  WARNING: {xlsx_name} not found in {zip_path.name}")
        zf.close()
        return rows

    print(f"  Parsing {xlsx_name} / sheet '{sheet_name}' ...")
    wb = openpyxl.load_workbook(
        io.BytesIO(zf.read(xlsx_name)), data_only=True, read_only=True
    )

    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found in {xlsx_name}")
        wb.close()
        zf.close()
        return rows

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(max_col=12, values_only=True))
    print(f"    {len(all_rows)} rows in sheet")

    # Detect column layout from the header row (first "Service Category" row)
    # FY1997-2008: Service Category, Total Computable, Federal Share, State Share,
    #              Service Category, Total Computable, Federal Share, State Share
    # FY2009-2011: Service Category, Total Computable, Federal Share, Federal Share Medicaid,
    #              Federal Share ARRA, State Share,
    #              Service Category, Total Computable, Federal Share, State Share
    has_arra = False
    for r in all_rows[:20]:
        if r[0] and isinstance(r[0], str) and r[0].strip() == "Service Category":
            # Check if col 3 or 4 mentions "ARRA"
            for ci in range(3, min(6, len(r))):
                if r[ci] and isinstance(r[ci], str) and "ARRA" in r[ci].upper():
                    has_arra = True
                    break
            break

    if has_arra:
        # MAP: 0=svc, 1=total, 2=fed, 3=fed_medicaid, 4=fed_arra, 5=state
        # ADM: 6=svc, 7=total, 8=fed, 9=state
        map_cols = {"svc": 0, "total": 1, "fed": 2, "fed_medicaid": 3, "fed_arra": 4, "state": 5}
        adm_cols = {"svc": 6, "total": 7, "fed": 8, "state": 9}
        print(f"    Detected layout: 6-col MAP (with Medicaid/ARRA) + 4-col ADM")
    else:
        # MAP: 0=svc, 1=total, 2=fed, 3=state
        # ADM: 4=svc, 5=total, 6=fed, 7=state
        map_cols = {"svc": 0, "total": 1, "fed": 2, "state": 3}
        adm_cols = {"svc": 4, "total": 5, "fed": 6, "state": 7}
        print(f"    Detected layout: 4-col MAP + 4-col ADM")

    # Find all state start positions
    state_blocks = []
    for i in range(len(all_rows)):
        r = all_rows[i]
        if r[0] and isinstance(r[0], str) and r[1] is None:
            cell = r[0].strip()
            code = _resolve_state_code(cell)
            if code is not None:
                state_blocks.append((i, cell, code))

    # Process each state block
    for block_idx, (start_row, state_name, state_code) in enumerate(state_blocks):
        if state_code == "US":
            continue

        # Find data start: skip headers until we hit actual data rows
        data_start = start_row + 1
        while data_start < len(all_rows):
            r = all_rows[data_start]
            if r[0] and isinstance(r[0], str):
                cell = r[0].strip()
                # Skip header-like rows
                if cell in ("", " ", "Medical Assistance Program", "Service Category",
                            "Administration") or cell.startswith("FY ") or cell.startswith("Net "):
                    data_start += 1
                    continue
                if r[1] is not None:
                    # This is a data row (has a number in col 1)
                    break
                else:
                    data_start += 1
                    continue
            elif r[0] is None:
                data_start += 1
                continue
            else:
                break

        # Find end of this state's block
        if block_idx + 1 < len(state_blocks):
            end_row = state_blocks[block_idx + 1][0]
        else:
            end_row = len(all_rows)

        # Parse data rows
        for i in range(data_start, end_row):
            r = all_rows[i]

            # MAP side
            map_svc_idx = map_cols["svc"]
            map_svc = r[map_svc_idx] if len(r) > map_svc_idx and r[map_svc_idx] else None
            if not _should_skip_row(map_svc):
                map_svc = str(map_svc).strip()
                rec = {
                    "state_code": state_code,
                    "program": "Medical Assistance Program",
                    "service_category": map_svc,
                    "total_computable": _safe_float(r[map_cols["total"]]),
                    "federal_share": _safe_float(r[map_cols["fed"]]),
                    "federal_share_medicaid": _safe_float(r[map_cols["fed_medicaid"]]) if "fed_medicaid" in map_cols else None,
                    "federal_share_arra": _safe_float(r[map_cols["fed_arra"]]) if "fed_arra" in map_cols else None,
                    "federal_share_covid": None,
                    "state_share": _safe_float(r[map_cols["state"]]),
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }
                rows.append(rec)

            # ADM side
            adm_svc_idx = adm_cols["svc"]
            adm_svc = r[adm_svc_idx] if len(r) > adm_svc_idx and r[adm_svc_idx] else None
            if not _should_skip_row(adm_svc):
                adm_svc = str(adm_svc).strip()
                adm_state_idx = adm_cols["state"]
                rec = {
                    "state_code": state_code,
                    "program": "Administration",
                    "service_category": adm_svc,
                    "total_computable": _safe_float(r[adm_cols["total"]]),
                    "federal_share": _safe_float(r[adm_cols["fed"]]),
                    "federal_share_medicaid": None,
                    "federal_share_arra": None,
                    "federal_share_covid": None,
                    "state_share": _safe_float(r[adm_state_idx]) if len(r) > adm_state_idx else None,
                    "fiscal_year": year,
                    "source": "medicaid.gov/fmr",
                    "snapshot_date": SNAPSHOT_DATE,
                    "snapshot": SNAPSHOT_DATE,
                }
                rows.append(rec)

    wb.close()
    zf.close()
    return rows


def parse_year(year: int) -> list[dict]:
    """Route to the correct parser based on file format era."""
    zip_key = YEAR_TO_ZIP_KEY[year]
    zip_path = RAW_DIR / ZIP_FILENAMES[zip_key]

    if not zip_path.exists():
        print(f"  ERROR: {zip_path.name} not found")
        return []

    if year >= 2013:
        return parse_format_a(year, zip_path)
    elif year == 2012:
        return parse_format_b(year, zip_path)
    else:
        return parse_format_c(year, zip_path)


def main():
    parser = argparse.ArgumentParser(
        description="CMS-64 FMR historical ETL: FY1997-2017 Excel -> Parquet"
    )
    parser.add_argument(
        "--years",
        nargs="+",
        type=int,
        default=ALL_YEARS,
        help=f"Fiscal years to process (default: {ALL_YEARS[0]}-{ALL_YEARS[-1]})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate but do not write files",
    )
    args = parser.parse_args()

    print(f"=== CMS-64 FMR Historical ETL ===")
    print(f"Years: {args.years}")
    print(f"Snapshot: {SNAPSHOT_DATE}")
    print(f"Run ID: {RUN_ID}")
    if args.dry_run:
        print("[DRY RUN]")
    print()

    # Step 0: Ensure all required ZIPs are downloaded
    print("--- Checking downloads ---")
    needed_zips = set()
    for year in sorted(args.years):
        needed_zips.add(YEAR_TO_ZIP_KEY[year])

    for zip_key in sorted(needed_zips, key=str):
        zip_path = ensure_zip(zip_key)
        if zip_path is None:
            print(f"  FATAL: Could not download {ZIP_FILENAMES[zip_key]}")
            sys.exit(1)
    print()

    all_rows = []
    year_stats = {}

    for year in sorted(args.years):
        print(f"--- FY{year} ---")

        year_rows = parse_year(year)
        print(f"  Parsed {len(year_rows):,} rows from FY{year}")

        if year_rows:
            states = set(r["state_code"] for r in year_rows)
            programs = set(r["program"] for r in year_rows)
            total_comp = sum(r["total_computable"] or 0 for r in year_rows)
            year_stats[year] = {
                "rows": len(year_rows),
                "states": len(states),
                "programs": sorted(programs),
                "total_computable": total_comp,
            }
            print(f"  {len(states)} states/territories, {sorted(programs)}")
            print(f"  Total computable: ${total_comp:,.0f}")

        all_rows.extend(year_rows)
        print()

    if not all_rows:
        print("No data parsed. Exiting.")
        sys.exit(1)

    # Step 3: Load into DuckDB and write Parquet
    print(f"=== Combined: {len(all_rows):,} rows across {len(year_stats)} years ===")

    con = duckdb.connect()

    con.execute("""
        CREATE TABLE _cms64_historical (
            state_code VARCHAR,
            program VARCHAR,
            service_category VARCHAR,
            total_computable DOUBLE,
            federal_share DOUBLE,
            federal_share_medicaid DOUBLE,
            federal_share_arra DOUBLE,
            federal_share_covid DOUBLE,
            state_share DOUBLE,
            fiscal_year INTEGER,
            source VARCHAR,
            snapshot_date DATE,
            snapshot DATE
        )
    """)

    # Insert in batches
    BATCH_SIZE = 5000
    for i in range(0, len(all_rows), BATCH_SIZE):
        batch = all_rows[i : i + BATCH_SIZE]
        values_list = []
        for r in batch:
            values_list.append((
                r["state_code"],
                r["program"],
                r["service_category"],
                r["total_computable"],
                r["federal_share"],
                r["federal_share_medicaid"],
                r["federal_share_arra"],
                r["federal_share_covid"],
                r["state_share"],
                r["fiscal_year"],
                r["source"],
                r["snapshot_date"],
                r["snapshot"],
            ))
        con.executemany(
            """INSERT INTO _cms64_historical VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?::DATE, ?::DATE
            )""",
            values_list,
        )

    # Validation stats
    print()
    print("=== Validation ===")

    stats = con.execute("""
        SELECT
            COUNT(*) AS total_rows,
            COUNT(DISTINCT state_code) AS state_count,
            COUNT(DISTINCT fiscal_year) AS year_count,
            MIN(fiscal_year) AS min_year,
            MAX(fiscal_year) AS max_year,
            COUNT(DISTINCT service_category) AS category_count,
            SUM(total_computable) AS total_spending,
            SUM(federal_share) AS total_federal,
            SUM(state_share) AS total_state
        FROM _cms64_historical
    """).fetchone()

    print(f"  Total rows:        {stats[0]:,}")
    print(f"  States/territories: {stats[1]}")
    print(f"  Years:              {stats[2]} ({stats[3]}-{stats[4]})")
    print(f"  Service categories: {stats[5]}")
    print(f"  Total computable:   ${stats[6]:,.0f}")
    print(f"  Federal share:      ${stats[7]:,.0f}")
    print(f"  State share:        ${stats[8]:,.0f}")

    # Per-year spending
    print()
    print("  Per-year total computable:")
    year_totals = con.execute("""
        SELECT fiscal_year, COUNT(*) AS rows,
               COUNT(DISTINCT state_code) AS states,
               SUM(total_computable) AS total
        FROM _cms64_historical
        GROUP BY fiscal_year ORDER BY fiscal_year
    """).fetchall()
    for yr, rows, states, total in year_totals:
        print(f"    FY{yr}: {rows:,} rows, {states} states, ${total:,.0f}")

    # Per-program breakdown
    print()
    print("  Per-program breakdown:")
    prog_totals = con.execute("""
        SELECT program, COUNT(*) AS rows, SUM(total_computable) AS total
        FROM _cms64_historical
        GROUP BY program ORDER BY total DESC
    """).fetchall()
    for prog, rows, total in prog_totals:
        print(f"    {prog}: {rows:,} rows, ${total:,.0f}")

    # Top 5 states by total computable (most recent year in range)
    print()
    max_yr = stats[4]
    print(f"  Top 5 states by total computable (FY{max_yr}):")
    top_states = con.execute(f"""
        SELECT state_code, SUM(total_computable) AS total
        FROM _cms64_historical
        WHERE fiscal_year = {max_yr}
        GROUP BY state_code ORDER BY total DESC LIMIT 5
    """).fetchall()
    for sc, total in top_states:
        print(f"    {sc}: ${total:,.0f}")

    # Write Parquet
    out_path = _snapshot_path("cms64_historical")
    total_written = write_parquet(con, "_cms64_historical", out_path, args.dry_run)

    # Write manifest
    manifest = {
        "snapshot_date": SNAPSHOT_DATE,
        "pipeline_run_id": RUN_ID,
        "created_at": datetime.utcnow().isoformat() + "Z",
        "source_files": sorted(set(
            str(RAW_DIR / ZIP_FILENAMES[YEAR_TO_ZIP_KEY[y]])
            for y in sorted(year_stats.keys())
        )),
        "source_url": "https://www.medicaid.gov/medicaid/financial-management/state-budget-expenditure-reporting-for-medicaid-and-chip/expenditure-reports-mbes/cbes",
        "tables": {
            "fact_cms64_historical": {
                "rows": total_written,
                "years": sorted(year_stats.keys()),
                "states": stats[1],
                "per_year": {
                    str(yr): {"rows": ys["rows"], "states": ys["states"]}
                    for yr, ys in year_stats.items()
                },
            }
        },
        "total_rows": total_written,
        "notes": (
            "CMS-64 Financial Management Report (FMR) net expenditures by state, "
            "service category, and program (MAP/ADM). Historical FY1997-2017. "
            "FY2012-2017 have federal_share breakdown (Medicaid/ARRA/BIPP). "
            "FY1997-2011 have aggregate federal_share only (no Medicaid/ARRA/BIPP split). "
            "CHIP excluded. National Totals excluded. Source: medicaid.gov/fmr."
        ),
    }

    if not args.dry_run:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest_path = META_DIR / f"manifest_cms64_historical_{SNAPSHOT_DATE}.json"
        manifest_path.write_text(json.dumps(manifest, indent=2))
        print(f"\n  Manifest: {manifest_path.relative_to(LAKE_DIR)}")

    con.execute("DROP TABLE IF EXISTS _cms64_historical")
    con.close()

    print()
    print(f"Done. {total_written:,} rows written to fact_cms64_historical.")


if __name__ == "__main__":
    main()
