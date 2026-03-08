#!/usr/bin/env python3
"""
build_lake_round2.py — Ingest remaining Ring 0 datasets into the lake.

Sources:
  1. NSDUH 2024 state prevalence (replaces 2023 data — newer survey years)
  2. Hospice quality measures (facility-level, 331K rows from CMS Care Compare)
  3. CHIP eligibility levels by state (52 rows)
  4. Continuous eligibility policies (35 rows)
  5. HCBS spending by authority type (Medicaid Scorecard)
  6. Managed Care quality features (81 rows)

Tables built:
  fact_nsduh_prevalence       — UPDATED to 2023-2024 survey years
  fact_hospice_quality        — Hospice facility quality measures
  fact_chip_eligibility       — CHIP/Medicaid eligibility income thresholds by state
  fact_continuous_eligibility — Continuous eligibility policy by state
  fact_hcbs_authority         — HCBS spending by authority type (scorecard)
  fact_mc_quality_features    — Managed care quality features

Usage:
  python3 scripts/build_lake_round2.py
  python3 scripts/build_lake_round2.py --dry-run
  python3 scripts/build_lake_round2.py --table nsduh_2024
"""

import argparse
import json
import uuid
from datetime import date, datetime
from pathlib import Path

import duckdb

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
}


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


# NSDUH table numbers → measure IDs (same mapping as 2023)
NSDUH_TABLES = {
    1: ("illicit_drug_use_past_month", "Illicit Drug Use in Past Month"),
    3: ("marijuana_use_past_month", "Marijuana Use in Past Month"),
    6: ("illicit_drug_other_than_mj_past_month", "Illicit Drug Use Other Than Marijuana in Past Month"),
    9: ("heroin_use_past_year", "Heroin Use in Past Year"),
    12: ("meth_use_past_year", "Methamphetamine Use in Past Year"),
    13: ("rx_pain_reliever_misuse_past_year", "Prescription Pain Reliever Misuse in Past Year"),
    14: ("opioid_misuse_past_year", "Opioid Misuse in Past Year"),
    15: ("alcohol_use_past_month", "Alcohol Use in Past Month"),
    16: ("binge_alcohol_past_month", "Binge Alcohol Use in Past Month"),
    19: ("tobacco_use_past_month", "Tobacco Product Use in Past Month"),
    24: ("sud_past_year", "Substance Use Disorder in Past Year"),
    25: ("aud_past_year", "Alcohol Use Disorder in Past Year"),
    27: ("dud_past_year", "Drug Use Disorder in Past Year"),
    29: ("oud_past_year", "Opioid Use Disorder in Past Year"),
    30: ("su_treatment_past_year", "Received Substance Use Treatment in Past Year"),
    31: ("needing_su_treatment", "Classified as Needing SU Treatment in Past Year"),
    32: ("su_treatment_gap", "Did Not Receive SU Treatment Among Those Needing It"),
    33: ("any_mental_illness", "Any Mental Illness in Past Year (18+)"),
    34: ("serious_mental_illness", "Serious Mental Illness in Past Year (18+)"),
    35: ("co_occurring_sud_ami", "Co-occurring SUD and Any Mental Illness (18+)"),
    36: ("co_occurring_sud_smi", "Co-occurring SUD and Serious Mental Illness (18+)"),
    37: ("mh_treatment_past_year", "Received Mental Health Treatment in Past Year"),
    38: ("major_depressive_episode", "Major Depressive Episode in Past Year"),
    39: ("suicidal_thoughts", "Serious Thoughts of Suicide in Past Year"),
    40: ("suicide_plans", "Made Any Suicide Plans in Past Year"),
    41: ("suicide_attempt", "Attempted Suicide in Past Year"),
}


def build_nsduh_2024(con, dry_run: bool) -> int:
    """Parse NSDUH 2024 Excel into fact_nsduh_prevalence (replaces 2023 data)."""
    print("Building fact_nsduh_prevalence (2023-2024 survey)...")
    xlsx_path = RAW_DIR / "nsduh_2024_state_tables.xlsx"
    if not xlsx_path.exists():
        print("  SKIPPED — nsduh_2024_state_tables.xlsx not found")
        return 0

    try:
        import openpyxl
    except ImportError:
        print("  SKIPPED — openpyxl not installed")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    all_rows = []

    for tab_num, (measure_id, measure_name) in NSDUH_TABLES.items():
        sheet_name = f"Table {tab_num}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))

        # Find the header row (contains "State" and "Estimate")
        header_idx = None
        for i, row in enumerate(rows_data):
            if row and len(row) >= 2 and row[1] and "State" in str(row[1]):
                header_idx = i
                break
        if header_idx is None:
            continue

        headers = rows_data[header_idx]
        # Parse age groups from header strings
        # Typical: Order, State, 12+ Estimate, 12+ CI Lower, 12+ CI Upper, 12-17 Estimate, ...
        # Or for 18+ only tables: Order, State, 18+ Estimate, 18+ CI Lower, 18+ CI Upper, 18-25 Estimate, ...
        age_groups = []
        col = 2
        while col + 2 < len(headers) and headers[col] is not None:
            header_text = str(headers[col])
            # Extract age group from header like "18+\nEstimate"
            age_label = header_text.split("\n")[0].strip() if "\n" in header_text else header_text.split(" ")[0].strip()
            if age_label and any(c.isdigit() for c in age_label):
                age_groups.append((age_label, col, col + 1, col + 2))
            col += 3

        # Parse state rows (skip Total U.S., regions)
        for row in rows_data[header_idx + 1:]:
            if not row or not row[1]:
                break
            state_name = str(row[1]).strip()
            if state_name not in STATE_NAME_TO_CODE:
                continue
            state_code = STATE_NAME_TO_CODE[state_name]

            for age_label, est_idx, ci_lo_idx, ci_hi_idx in age_groups:
                try:
                    estimate = float(row[est_idx]) if row[est_idx] is not None else None
                except (ValueError, TypeError):
                    estimate = None
                try:
                    ci_low = float(row[ci_lo_idx]) if row[ci_lo_idx] is not None else None
                except (ValueError, TypeError):
                    ci_low = None
                try:
                    ci_high = float(row[ci_hi_idx]) if row[ci_hi_idx] is not None else None
                except (ValueError, TypeError):
                    ci_high = None

                if estimate is not None:
                    # Convert from decimal to percentage
                    all_rows.append({
                        "state_code": state_code,
                        "measure_id": measure_id,
                        "measure_name": measure_name,
                        "age_group": age_label,
                        "estimate_pct": round(estimate * 100, 2),
                        "ci_lower_pct": round(ci_low * 100, 2) if ci_low else None,
                        "ci_upper_pct": round(ci_high * 100, 2) if ci_high else None,
                        "survey_years": "2023-2024",
                        "source": "samhsa_nsduh_2024",
                    })

    wb.close()

    if not all_rows:
        print("  No NSDUH 2024 data parsed")
        return 0

    con.execute("""
        CREATE OR REPLACE TABLE _fact_nsduh (
            state_code VARCHAR, measure_id VARCHAR, measure_name VARCHAR,
            age_group VARCHAR, estimate_pct DOUBLE, ci_lower_pct DOUBLE,
            ci_upper_pct DOUBLE, survey_years VARCHAR, source VARCHAR,
            snapshot_date DATE
        )
    """)
    for row in all_rows:
        con.execute("INSERT INTO _fact_nsduh VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", [
            row["state_code"], row["measure_id"], row["measure_name"],
            row["age_group"], row["estimate_pct"], row["ci_lower_pct"],
            row["ci_upper_pct"], row["survey_years"], row["source"],
            SNAPSHOT_DATE,
        ])

    count = write_parquet(con, "_fact_nsduh", _snapshot_path("nsduh_prevalence"), dry_run)
    measures = con.execute("SELECT COUNT(DISTINCT measure_id) FROM _fact_nsduh").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_nsduh").fetchone()[0]

    # AMI ranking
    top_ami = con.execute("""
        SELECT state_code, estimate_pct FROM _fact_nsduh
        WHERE measure_id = 'any_mental_illness' AND age_group = '18+'
        ORDER BY estimate_pct DESC LIMIT 5
    """).fetchall()
    print(f"  {count:,} rows, {measures} measures, {states} states, survey years 2023-2024")
    if top_ami:
        print(f"  Top AMI (18+): {', '.join(f'{s[0]} {s[1]:.1f}%' for s in top_ami)}")

    con.execute("DROP TABLE IF EXISTS _fact_nsduh")
    return count


def build_hospice_quality(con, dry_run: bool) -> int:
    """Build fact_hospice_quality from CMS Care Compare hospice data."""
    print("Building fact_hospice_quality...")
    json_path = RAW_DIR / "hospice_national.json"
    if not json_path.exists():
        print("  SKIPPED — hospice_national.json not found")
        return 0

    # DuckDB can read JSON directly for this size
    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_hospice AS
        SELECT
            cms_certification_number_ccn AS ccn,
            facility_name,
            state,
            citytown AS city,
            zip_code,
            countyparish AS county,
            cms_region,
            measure_code,
            measure_name,
            TRY_CAST(score AS DOUBLE) AS score,
            footnote,
            measure_date_range,
            'cms_care_compare_hospice' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_json_auto('{json_path}')
        WHERE state IS NOT NULL AND LENGTH(state) <= 2
    """)

    count = write_parquet(con, "_fact_hospice", _snapshot_path("hospice_quality"), dry_run)
    facilities = con.execute("SELECT COUNT(DISTINCT ccn) FROM _fact_hospice").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state) FROM _fact_hospice").fetchone()[0]
    measures = con.execute("SELECT COUNT(DISTINCT measure_code) FROM _fact_hospice").fetchone()[0]

    print(f"  {count:,} rows, {facilities:,} facilities, {states} states, {measures} measures")

    # Top measures
    top = con.execute("""
        SELECT measure_code, measure_name, COUNT(*) AS cnt
        FROM _fact_hospice WHERE measure_name != ''
        GROUP BY measure_code, measure_name ORDER BY cnt DESC LIMIT 5
    """).fetchall()
    for m in top:
        print(f"    {m[0]}: {m[1][:50]} ({m[2]:,} rows)")

    con.execute("DROP TABLE IF EXISTS _fact_hospice")
    return count


def build_chip_eligibility(con, dry_run: bool) -> int:
    """Build fact_chip_eligibility from CHIP eligibility income levels."""
    print("Building fact_chip_eligibility...")
    csv_path = RAW_DIR / "chip_eligibility_levels.csv"
    if not csv_path.exists():
        print("  SKIPPED — chip_eligibility_levels.csv not found")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_chip_elig AS
        SELECT
            State AS state_name,
            Notes AS notes,
            "Medicaid Ages 0-1" AS medicaid_ages_0_1,
            "Medicaid Ages 1-5" AS medicaid_ages_1_5,
            "Medicaid Ages 6-18" AS medicaid_ages_6_18,
            "Separate CHIP" AS separate_chip,
            "Pregnant Women Medicaid" AS pregnant_women_medicaid,
            "Pregnant Women CHIP" AS pregnant_women_chip,
            "Parent/Caretaker" AS parent_caretaker,
            "Expansion to Adults" AS expansion_adults,
            "Separate CHIP Ages" AS separate_chip_ages,
            "Parent/Caretaker Income Standard" AS parent_income_standard,
            'kff_chip_eligibility' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', header=true)
        WHERE State IS NOT NULL AND LENGTH(State) > 1
    """)

    count = write_parquet(con, "_fact_chip_elig", _snapshot_path("chip_eligibility"), dry_run)
    print(f"  {count} state eligibility records")

    # Expansion status
    expansion = con.execute("""
        SELECT expansion_adults, COUNT(*) FROM _fact_chip_elig
        GROUP BY expansion_adults ORDER BY COUNT(*) DESC
    """).fetchall()
    for e in expansion:
        print(f"    Expansion '{e[0]}': {e[1]} states")

    con.execute("DROP TABLE IF EXISTS _fact_chip_elig")
    return count


def build_continuous_eligibility(con, dry_run: bool) -> int:
    """Build fact_continuous_eligibility from state policy data."""
    print("Building fact_continuous_eligibility...")
    csv_path = RAW_DIR / "continuous_eligibility.csv"
    if not csv_path.exists():
        print("  SKIPPED — continuous_eligibility.csv not found")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_cont_elig AS
        SELECT
            State AS state_name,
            CHIP AS chip_continuous,
            Medicaid AS medicaid_continuous,
            Exceptions AS exceptions,
            'medicaid_gov_continuous_eligibility' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', header=true)
        WHERE State IS NOT NULL AND LENGTH(State) > 1
    """)

    count = write_parquet(con, "_fact_cont_elig", _snapshot_path("continuous_eligibility"), dry_run)
    chip_yes = con.execute("SELECT COUNT(*) FROM _fact_cont_elig WHERE chip_continuous = 'true'").fetchone()[0]
    med_yes = con.execute("SELECT COUNT(*) FROM _fact_cont_elig WHERE medicaid_continuous = 'true'").fetchone()[0]
    print(f"  {count} states — CHIP continuous: {chip_yes}, Medicaid continuous: {med_yes}")

    con.execute("DROP TABLE IF EXISTS _fact_cont_elig")
    return count


def build_hcbs_authority(con, dry_run: bool) -> int:
    """Build fact_hcbs_authority from Medicaid Scorecard HCBS data."""
    print("Building fact_hcbs_authority...")
    csv_path = RAW_DIR / "hcbs_by_authority.csv"
    if not csv_path.exists():
        print("  SKIPPED — hcbs_by_authority.csv not found")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_hcbs_auth AS
        SELECT
            measureAbbreviation AS measure_id,
            measureName AS measure_name,
            description,
            measureType AS measure_type,
            dataPeriodType AS period_type,
            dataRange AS data_range,
            TRY_CAST(mdlNumberStates AS INTEGER) AS num_states,
            medianLabel AS median_label,
            TRY_CAST(medianType AS DOUBLE) AS median_value,
            meanLabel AS mean_label,
            TRY_CAST(meanType AS DOUBLE) AS mean_value,
            reportingProgram AS reporting_program,
            pillarId AS pillar_id,
            'medicaid_scorecard_hcbs' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', header=true)
        WHERE measureAbbreviation IS NOT NULL
    """)

    count = write_parquet(con, "_fact_hcbs_auth", _snapshot_path("hcbs_authority"), dry_run)
    print(f"  {count} HCBS/waiver measures from Medicaid Scorecard")

    con.execute("DROP TABLE IF EXISTS _fact_hcbs_auth")
    return count


def build_mc_quality_features(con, dry_run: bool) -> int:
    """Build fact_mc_quality_features from managed care quality features data."""
    print("Building fact_mc_quality_features...")
    csv_path = RAW_DIR / "mc_features_qa_performance.csv"
    if not csv_path.exists():
        print("  SKIPPED — mc_features_qa_performance.csv not found")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_mc_qa AS
        SELECT
            Features AS feature,
            "Comprehensive MCO with or without MLTSS" AS comprehensive_mco,
            PCCM AS pccm,
            "PCCM Entity" AS pccm_entity,
            MLTSS AS mltss,
            "BHO (PIHP and/or PAHP)" AS bho,
            Dental AS dental,
            Transportation AS transportation,
            "Other PHP" AS other_php,
            PACE AS pace,
            TRY_CAST(Year AS INTEGER) AS year,
            'medicaid_gov_mc_features' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', header=true)
        WHERE Features IS NOT NULL
    """)

    count = write_parquet(con, "_fact_mc_qa", _snapshot_path("mc_quality_features"), dry_run)
    year = con.execute("SELECT DISTINCT year FROM _fact_mc_qa").fetchone()
    print(f"  {count} managed care quality features (year: {year[0] if year else 'N/A'})")

    con.execute("DROP TABLE IF EXISTS _fact_mc_qa")
    return count


ALL_TABLES = {
    "nsduh_2024": ("fact_nsduh_prevalence", build_nsduh_2024),
    "hospice": ("fact_hospice_quality", build_hospice_quality),
    "chip_eligibility": ("fact_chip_eligibility", build_chip_eligibility),
    "continuous_eligibility": ("fact_continuous_eligibility", build_continuous_eligibility),
    "hcbs_authority": ("fact_hcbs_authority", build_hcbs_authority),
    "mc_quality_features": ("fact_mc_quality_features", build_mc_quality_features),
}


def main():
    parser = argparse.ArgumentParser(description="Ingest remaining Ring 0 datasets into Aradune lake")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--table", choices=list(ALL_TABLES.keys()) + ["all"], default="all")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"Round 2 Data Ingestion — {SNAPSHOT_DATE}")
    print(f"{'='*60}")
    print(f"Run ID: {RUN_ID}\n")

    con = duckdb.connect()
    totals = {}

    tables_to_build = ALL_TABLES if args.table == "all" else {args.table: ALL_TABLES[args.table]}
    for key, (fact_name, builder) in tables_to_build.items():
        totals[fact_name] = builder(con, args.dry_run)
        print()

    con.close()

    print("=" * 60)
    print("ROUND 2 LAKE INGESTION COMPLETE")
    print("=" * 60)
    total_rows = sum(totals.values())
    for name, count in totals.items():
        status = "written" if not args.dry_run else "dry-run"
        print(f"  {name:40s} {count:>12,} rows  [{status}]")
    print(f"  {'TOTAL':40s} {total_rows:>12,} rows")

    if not args.dry_run and total_rows > 0:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "snapshot_date": SNAPSHOT_DATE,
            "pipeline_run_id": RUN_ID,
            "created_at": datetime.now().isoformat() + "Z",
            "tables": {name: {"rows": count} for name, count in totals.items()},
            "total_rows": total_rows,
        }
        manifest_file = META_DIR / f"manifest_round2_{SNAPSHOT_DATE}.json"
        with open(manifest_file, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"\n  Manifest: {manifest_file}")

    print("\nDone.")


if __name__ == "__main__":
    main()
