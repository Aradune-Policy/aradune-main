#!/usr/bin/env python3
"""
build_lake_macpac_exhibit14_v2.py — Re-ingest MACPAC Exhibit 14 with all dual-status
age breakdowns into the Aradune data lake.

Source: data/raw/EXHIBIT-14.-Medicaid-Enrollment-by-State-Eligibility-Group-and-Dually-Eligible-Status-FY-2023.xlsx

The existing macpac_enrollment table is missing:
  - state_code (only has state_name)
  - dual age 65+ breakdowns (dual_total_age65, full_dual_age65, partial_dual_age65)

Table built:
  macpac_enrollment_v2 — Full Exhibit 14 with state codes and age 65+ dual breakdowns

Usage:
  python3 scripts/build_lake_macpac_exhibit14_v2.py
"""

import json
import re
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_PATH = PROJECT_ROOT / "data" / "raw" / "EXHIBIT-14.-Medicaid-Enrollment-by-State-Eligibility-Group-and-Dually-Eligible-Status-FY-2023.xlsx"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

STATE_MAP = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN",
    "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA",
    "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "District of Columbia": "DC",
    "Total": "US", "Totals": "US", "U.S. Total": "US",
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
}


def _parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if s in ("", "-", "--", "N/A", "n/a", "*"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def build():
    print(f"MACPAC Exhibit 14 v2 ETL -- snapshot {SNAPSHOT_DATE}")
    print(f"  Source: {RAW_PATH.name}")

    if not RAW_PATH.exists():
        print("  ERROR: Source file not found")
        return

    wb = openpyxl.load_workbook(RAW_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Structure from inspection:
    # Row 1: Title
    # Row 3-5: Multi-row header
    # Row 5: Column sub-headers
    # Row 6+: Data
    #
    # Columns (values in thousands):
    # A: State
    # B: Total
    # C: Child
    # D: New adult group
    # E: Other adult
    # F: Disabled
    # G: Aged
    # H: All dual total
    # I: All dual age 65+
    # J: Full dual total
    # K: Full dual age 65+
    # L: Partial dual total
    # M: Partial dual age 65+

    rows = []
    for row_num in range(6, ws.max_row + 1):
        raw_state = ws.cell(row=row_num, column=1).value
        if not raw_state or not isinstance(raw_state, str):
            continue
        clean_state = re.sub(r"\d+$", "", raw_state.strip()).strip()
        state_code = STATE_MAP.get(clean_state)
        if not state_code:
            continue

        # Values are in thousands
        total = _parse_numeric(ws.cell(row=row_num, column=2).value)
        if total is None:
            continue

        record = {
            "state_code": state_code,
            "state_name": clean_state if state_code != "US" else "United States",
            "fiscal_year": 2023,
            "total_enrollment_k": total,
            "child_k": _parse_numeric(ws.cell(row=row_num, column=3).value),
            "new_adult_group_k": _parse_numeric(ws.cell(row=row_num, column=4).value),
            "other_adult_k": _parse_numeric(ws.cell(row=row_num, column=5).value),
            "disabled_k": _parse_numeric(ws.cell(row=row_num, column=6).value),
            "aged_k": _parse_numeric(ws.cell(row=row_num, column=7).value),
            "dual_total_k": _parse_numeric(ws.cell(row=row_num, column=8).value),
            "dual_total_age65_k": _parse_numeric(ws.cell(row=row_num, column=9).value),
            "full_dual_total_k": _parse_numeric(ws.cell(row=row_num, column=10).value),
            "full_dual_age65_k": _parse_numeric(ws.cell(row=row_num, column=11).value),
            "partial_dual_total_k": _parse_numeric(ws.cell(row=row_num, column=12).value),
            "partial_dual_age65_k": _parse_numeric(ws.cell(row=row_num, column=13).value),
            "source": "macpac",
            "snapshot_date": SNAPSHOT_DATE,
        }
        rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return

    print(f"  Parsed {len(rows)} state records")

    con = duckdb.connect()
    columns = list(rows[0].keys())
    num_cols = {c for c in columns if c.endswith("_k")}
    int_cols = {"fiscal_year"}
    col_defs = ", ".join(
        f"{c} {'DOUBLE' if c in num_cols else 'INTEGER' if c in int_cols else 'VARCHAR'}"
        for c in columns
    )
    con.execute(f"CREATE TABLE _ex14 ({col_defs})")
    placeholders = ", ".join(["?"] * len(columns))
    con.executemany(
        f"INSERT INTO _ex14 VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    out_path = FACT_DIR / "macpac_enrollment_v2" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    count = con.execute("SELECT COUNT(*) FROM _ex14").fetchone()[0]
    con.execute(f"COPY _ex14 TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_mb = out_path.stat().st_size / (1024 * 1024)
    print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count} rows, {size_mb:.2f} MB)")

    # Validation
    states = [r for r in rows if r["state_code"] not in ("US", "PR", "GU", "VI", "AS", "MP")]
    us = [r for r in rows if r["state_code"] == "US"]
    print(f"  {count} rows, {len(states)} states + territories + US total")
    if us:
        u = us[0]
        print(f"  US total enrollment: {u['total_enrollment_k']:,.1f}K ({u['total_enrollment_k'] * 1000:,.0f})")
        print(f"  US duals total: {u['dual_total_k']:,.1f}K")
        print(f"  US duals age 65+: {u['dual_total_age65_k']:,.1f}K")
        if u["dual_total_k"] and u["total_enrollment_k"]:
            pct = u["dual_total_k"] / u["total_enrollment_k"] * 100
            print(f"  Dual share of total: {pct:.1f}%")
        if u["dual_total_age65_k"] and u["dual_total_k"]:
            pct = u["dual_total_age65_k"] / u["dual_total_k"] * 100
            print(f"  Age 65+ share of duals: {pct:.1f}%")

    # Check non-expansion states (new_adult should be 0 or None)
    non_exp = [r for r in rows if r["new_adult_group_k"] == 0 and r["state_code"] != "US"]
    print(f"  Non-expansion states (new adult = 0): {len(non_exp)}")

    con.close()

    # Manifest
    manifest = {
        "pipeline_run": "macpac_exhibit14_v2",
        "run_id": RUN_ID,
        "snapshot_date": SNAPSHOT_DATE,
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "total_rows": count,
        "tables": ["macpac_enrollment_v2"],
        "source_files": [RAW_PATH.name],
        "source": "MACPAC MACStats Exhibit 14 (macpac.gov)",
        "notes": "Values in thousands. Includes dual age 65+ breakdowns not in v1.",
    }
    manifest_path = META_DIR / f"manifest_macpac_enrollment_v2_{SNAPSHOT_DATE}.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"  Manifest: {manifest_path.relative_to(LAKE_DIR)}")

    print(f"\n== Summary ==")
    print(f"  Table: macpac_enrollment_v2 ({count} rows)")
    print(f"  Adds: state_code, dual age 65+ breakdowns over v1")


if __name__ == "__main__":
    build()
