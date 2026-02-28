#!/usr/bin/env python3
"""
Build reference data JSON files for the Aradune T-MSIS Explorer.

Generates three files in public/data/:
  - medicare_rates.json   (Medicare PFS rates by HCPCS)
  - fee_schedules.json    (State Medicaid fee schedule rates by HCPCS)
  - risk_adj.json         (Eligibility-mix risk adjustment factors by state)

Usage:
  python scripts/build_reference_data.py

Expects source files in data/ directory:
  - PPRRVU2026_Jan_nonQPP.csv              (Medicare RVU file from CMS)
  - Florida_2026_Practitioner_Fee_Schedule.xlsx   (FL AHCA practitioner fee schedule)
  - Florida_2026_Practitioner_Laboratory_Fee_Schedule.xlsx  (FL AHCA lab fee schedule)
  - EXHIBIT-14_*.xlsx                       (MACPAC enrollment by eligibility)
  - EXHIBIT-22_*.xlsx                       (MACPAC spending per enrollee by eligibility)
"""

import openpyxl
import csv
import json
import os
import re
import glob
import sys

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "public", "data")

def safe_float(v, default=0.0):
    if v is None:
        return default
    try:
        s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
        if not s or s in ("—", "-", ".", "N/A", "n.a.", "..", "DSH"):
            return default
        return float(s)
    except:
        return default

def clean_state_name(s):
    """Strip footnote markers from MACPAC state names (e.g., 'Illinois7' -> 'Illinois')."""
    if not s:
        return ""
    return re.sub(r'\d+$', '', str(s).strip())

def normalize_code(code):
    """Normalize HCPCS code to 5-char string."""
    code = str(code).strip()
    if code.replace(".", "").isdigit():
        code = str(int(float(code)))
    if code.isdigit() and len(code) < 5:
        code = code.zfill(5)
    return code

ST_MAP = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR", "California": "CA",
    "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE", "District of Columbia": "DC",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL",
    "Indiana": "IN", "Iowa": "IA", "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA",
    "Maine": "ME", "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT",
    "Virginia": "VA", "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY"
}


def build_medicare_rates():
    """Build medicare_rates.json from CMS PFS RVU file."""
    rvu_file = glob.glob(os.path.join(DATA_DIR, "PPRRVU*nonQPP*.csv"))
    if not rvu_file:
        print("  SKIP: No RVU file found (PPRRVU*nonQPP*.csv)")
        return None

    print(f"  Source: {os.path.basename(rvu_file[0])}")
    rates = {}
    cf = 33.4009  # CY2026 conversion factor — update when CMS releases new CF

    with open(rvu_file[0], "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header_found = False
        for row in reader:
            if not header_found:
                if len(row) > 0 and row[0].strip() == "HCPCS":
                    header_found = True
                continue
            if len(row) < 13:
                continue

            hcpcs = row[0].strip()
            mod = row[1].strip()
            status = row[3].strip()
            work = safe_float(row[5])
            nf_total = safe_float(row[11])
            fac_total = safe_float(row[12])

            # Skip bundled, tracking, carrier-priced, non-payable
            if status in ("B", "T", "C", "N"):
                continue
            # Prefer base code (no modifier); skip modifier duplicates
            if mod and hcpcs in rates:
                continue

            nf_rate = round(nf_total * cf, 2) if nf_total > 0 else 0
            fac_rate = round(fac_total * cf, 2) if fac_total > 0 else 0

            if nf_rate > 0 or fac_rate > 0:
                entry = {"r": nf_rate, "fr": fac_rate, "rvu": round(nf_total, 4), "w": round(work, 2)}
                desc = row[2].strip()
                if desc:
                    entry["d"] = desc[:60]
                rates[hcpcs] = entry

    result = {"rates": rates, "cf": cf, "year": 2026}
    print(f"  Medicare codes: {len(rates)}")
    return result


def build_fee_schedules():
    """Build fee_schedules.json from state fee schedule Excel files."""
    states_data = {}

    # --- Florida Practitioner ---
    fl_prac = glob.glob(os.path.join(DATA_DIR, "*Practitioner_Fee_Schedule*"))
    fl_prac = [f for f in fl_prac if "Laboratory" not in f]
    if fl_prac:
        print(f"  FL Practitioner: {os.path.basename(fl_prac[0])}")
        fl_rates = {}
        wb = openpyxl.load_workbook(fl_prac[0], read_only=True)
        # Find the fee schedule sheet (not the notes sheet)
        ws_name = [s for s in wb.sheetnames if "Fee Schedule" in s and "Note" not in s]
        ws = wb[ws_name[0]] if ws_name else wb[wb.sheetnames[-1]]

        for row in ws.iter_rows(min_row=4, values_only=True):
            code = str(row[1]).strip() if row[1] else ""
            mod = str(row[2]).strip() if row[2] else ""
            fsi = safe_float(row[3])
            facility = safe_float(row[4])
            pci = safe_float(row[5])
            tci = safe_float(row[6])

            if not code:
                continue
            code = normalize_code(code)

            if mod in ("", "*"):
                entry = {"r": fsi}
                if facility > 0:
                    entry["fr"] = facility
                if pci > 0:
                    entry["pc"] = pci
                if tci > 0:
                    entry["tc"] = tci
                fl_rates[code] = entry
            elif code not in fl_rates:
                entry = {"r": fsi}
                if facility > 0:
                    entry["fr"] = facility
                fl_rates[code] = entry
        wb.close()

        # --- Florida Lab ---
        fl_lab = glob.glob(os.path.join(DATA_DIR, "*Laboratory_Fee_Schedule*"))
        if fl_lab:
            print(f"  FL Lab: {os.path.basename(fl_lab[0])}")
            wb = openpyxl.load_workbook(fl_lab[0], read_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=5, values_only=True):
                code = str(row[1]).strip() if row[1] else ""
                fs_rate = safe_float(row[2])
                pc = safe_float(row[3])
                tc = safe_float(row[4])
                if not code:
                    continue
                code = normalize_code(code)
                if fs_rate > 0 and code not in fl_rates:
                    entry = {"r": fs_rate}
                    if pc > 0:
                        entry["pc"] = pc
                    if tc > 0:
                        entry["tc"] = tc
                    fl_rates[code] = entry
            wb.close()

        states_data["FL"] = {
            "rates": fl_rates,
            "name": "Florida",
            "year": 2026,
            "source": "AHCA Practitioner Fee Schedule, Jan 2026",
            "n": len(fl_rates)
        }
        print(f"  FL total codes: {len(fl_rates)}")

    # Add more states here as fee schedules are collected
    # Example:
    # states_data["TX"] = { "rates": tx_rates, "name": "Texas", ... }

    if not states_data:
        print("  SKIP: No fee schedule files found")
        return None

    return {"states": states_data}


def build_risk_adjustment():
    """Build risk_adj.json from MACPAC Exhibits 14 and 22."""
    ex14 = glob.glob(os.path.join(DATA_DIR, "EXHIBIT-14*"))
    ex22 = glob.glob(os.path.join(DATA_DIR, "EXHIBIT-22*"))

    if not ex14 or not ex22:
        print(f"  SKIP: Missing MACPAC files (found Ex14={bool(ex14)}, Ex22={bool(ex22)})")
        return None

    print(f"  Enrollment: {os.path.basename(ex14[0])}")
    print(f"  Spending:   {os.path.basename(ex22[0])}")

    # Parse Exhibit 14: Enrollment by eligibility group
    enrollment = {}
    wb = openpyxl.load_workbook(ex14[0], read_only=True)
    for row in wb[wb.sheetnames[0]].iter_rows(min_row=6, values_only=True):
        state = clean_state_name(row[0])
        if not state or state.startswith("Source") or state.startswith("Note"):
            continue
        total = safe_float(row[1]) * 1000  # Values in thousands
        if total > 0:
            enrollment[state] = {
                "total": total,
                "child": safe_float(row[2]) * 1000,
                "new_adult": safe_float(row[3]) * 1000,
                "other_adult": safe_float(row[4]) * 1000,
                "disabled": safe_float(row[5]) * 1000,
                "aged": safe_float(row[6]) * 1000,
                "dual": safe_float(row[7]) * 1000
            }
    wb.close()

    # Parse Exhibit 22: Per-enrollee spending by eligibility group
    spending_pe = {}
    wb = openpyxl.load_workbook(ex22[0], read_only=True)
    for row in wb[wb.sheetnames[0]].iter_rows(min_row=5, values_only=True):
        state = clean_state_name(row[0])
        if not state or state.startswith("Source") or state.startswith("Note") or state.startswith("—"):
            continue
        total_pe = safe_float(row[1])
        if total_pe > 0:
            spending_pe[state] = {
                "total": total_pe,
                "child": safe_float(row[3]),       # All enrollees column
                "new_adult": safe_float(row[5]),
                "other_adult": safe_float(row[7]),
                "disabled": safe_float(row[9]),
                "aged": safe_float(row[11])
            }
    wb.close()

    # Compute adjustment factors
    natl_pe = spending_pe.get("Total", {})
    natl_total_pe = natl_pe.get("total", 9254)

    risk_adj_states = {}
    for state_name, enr in enrollment.items():
        if state_name == "Total":
            continue
        ab = ST_MAP.get(state_name)
        if not ab:
            continue
        total = enr["total"]
        if total <= 0:
            continue

        # State enrollment shares by group
        groups = ["child", "new_adult", "other_adult", "disabled", "aged"]
        shares = {k: enr[k] / total for k in groups}

        # Expected PE = national per-capita by group × state enrollment shares
        expected = sum(natl_pe.get(k, 0) * shares[k] for k in groups)

        # Factor: >1 means sicker/costlier mix than national average
        factor = round(expected / natl_total_pe, 4) if natl_total_pe > 0 else 1.0

        state_pe = spending_pe.get(state_name, {})
        actual_pe = state_pe.get("total", 0)
        adjusted = round(actual_pe / factor, 2) if factor > 0 else actual_pe
        dual_share = round(enr.get("dual", 0) / total * 100, 1) if total > 0 else 0

        risk_adj_states[ab] = {
            "factor": factor,
            "adjusted_pe": adjusted,
            "actual_pe": round(actual_pe, 2),
            "expected_pe": round(expected, 2),
            "dual_pct": dual_share,
            "mix": {k: round(v * 100, 1) for k, v in shares.items()}
        }

    print(f"  States with adjustment: {len(risk_adj_states)}")
    return {
        "states": risk_adj_states,
        "method": "eligibility_mix_adjustment",
        "source": "MACPAC Exhibits 14 & 22, FY2023",
        "national_pe": {k: round(v, 2) for k, v in natl_pe.items()}
    }


def main():
    print(f"Data directory: {DATA_DIR}")
    print(f"Output directory: {OUT_DIR}")
    os.makedirs(OUT_DIR, exist_ok=True)

    builders = [
        ("medicare_rates.json", build_medicare_rates),
        ("fee_schedules.json", build_fee_schedules),
        ("risk_adj.json", build_risk_adjustment),
    ]

    for filename, builder in builders:
        print(f"\n{'='*50}")
        print(f"Building {filename}")
        print(f"{'='*50}")
        data = builder()
        if data:
            path = os.path.join(OUT_DIR, filename)
            with open(path, "w") as f:
                json.dump(data, f, separators=(",", ":"))
            size_kb = os.path.getsize(path) / 1024
            print(f"  Written: {path} ({size_kb:.0f} KB)")
        else:
            print(f"  SKIPPED: {filename}")

    print(f"\n{'='*50}")
    print("Done! Files are ready in public/data/")
    print("Deploy with: npm run build && npm run deploy")


if __name__ == "__main__":
    main()
