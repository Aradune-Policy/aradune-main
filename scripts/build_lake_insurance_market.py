#!/usr/bin/env python3
"""
build_lake_insurance_market.py — Ingest health insurance market & coverage data.

Tables built:
  fact_mlr_market            — Commercial MLR (individual/small/large group) by issuer/state, 2022-2024
  fact_risk_adjustment       — ACA risk adjustment state-level transfers by market (2024 benefit year)
  fact_ma_star_ratings       — Medicare Advantage contract-level star ratings (2025-2026)
  fact_census_health_insurance — Health insurance coverage by type and state, 2008-2024 (ACS)
  fact_meps_employer_insurance — MEPS-IC state-level employer health insurance premiums (2020)

Data sources:
  - CMS CCIIO MLR Public Use Files: cms.gov/cciio/resources/data-resources/mlr
  - CMS Risk Adjustment Summary Reports: cms.gov/cciio/programs-and-initiatives/premium-stabilization-programs
  - CMS MA Star Ratings: cms.gov/medicare/health-drug-plans/part-c-d-performance-data
  - Census ACS HIC tables: census.gov
  - AHRQ MEPS-IC: meps.ahrq.gov

Usage:
  python3 scripts/build_lake_insurance_market.py
  python3 scripts/build_lake_insurance_market.py --dry-run
  python3 scripts/build_lake_insurance_market.py --only fact_mlr_market,fact_ma_star_ratings
"""

import argparse
import csv
import io
import json
import os
import re
import subprocess
import uuid
import zipfile
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path

import duckdb

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "insurance_market"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.2f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _fact_path(name: str) -> Path:
    return FACT_DIR / name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def _download(url: str, dest: Path) -> bool:
    """Download a file using curl. Returns True on success."""
    if dest.exists() and dest.stat().st_size > 1000:
        print(f"  Using cached: {dest.name}")
        return True
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"  Downloading: {url}")
    result = subprocess.run(
        ["curl", "-L", "-s", "-o", str(dest), url],
        capture_output=True, timeout=120
    )
    if result.returncode != 0 or not dest.exists() or dest.stat().st_size < 100:
        print(f"  FAILED to download {url}")
        return False
    return True


def _unzip(zip_path: Path, dest_dir: Path) -> list[Path]:
    """Unzip a file. Returns list of extracted files."""
    dest_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(dest_dir)
        return [dest_dir / name for name in z.namelist()]


# ── 1. MLR Market Data ──────────────────────────────────────────────────

MLR_URLS = {
    2024: "https://www.cms.gov/files/zip/mlr-public-use-file-2024.zip",
    2023: "https://www.cms.gov/files/zip/mlr-public-use-file-2023.zip",
    2022: "https://www.cms.gov/files/zip/mlr-public-use-file-2022.zip",
}

# Key row codes for MLR analysis
MLR_KEY_ROWS = {
    "MLR_NUMERATOR": "mlr_numerator",
    "MLR_DENOMINATOR": "mlr_denominator",
    "PRELIMINARY_MLR": "preliminary_mlr",
    "CREDIBILITY_ADJUSTED_MLR": "credibility_adjusted_mlr",
    "ADJ_EARNED_PREMIUM_LIC_REG_FEE": "adjusted_premium",
    "ADJ_INCURRED_CLAIMS": "adjusted_incurred_claims",
    "ADJ_INCURRED_CLAIMS_RESTATED_Q1": "adjusted_claims_restated",
    "QUALITY_IMPROVEMENT_EXPENSES": "quality_improvement_expenses",
    "REBATE_AMT_CREDIBILITY_ADJ_MLR": "rebate_amount",
    "REBATE_LIABILITY_TOTAL": "rebate_liability_total",
    "FED_STATE_TAXES_LIC_OR_REG_FEE": "fed_state_taxes",
    "FED_RISK_ADJ_NET_PAYMENTS_HHS": "risk_adjustment_net",
    "PREMIUM_EARNED_INCLUDING_FSHRP": "premium_earned",
    "TP_COVERED_LIVES": "covered_lives",
    "TP_LIFE_YEARS": "life_years",
}


def build_mlr_market(con, dry_run: bool) -> int:
    """Build fact_mlr_market: Commercial MLR data by issuer, state, market, year."""
    print("Building fact_mlr_market...")

    all_records = []

    for year, url in sorted(MLR_URLS.items()):
        zip_path = RAW_DIR / f"mlr_{year}.zip"
        extract_dir = RAW_DIR / f"mlr_{year}"

        if not _download(url, zip_path):
            continue

        if not extract_dir.exists():
            _unzip(zip_path, extract_dir)

        # Read template header (issuer info)
        header_file = extract_dir / "MR_Submission_Template_Header.csv"
        part3_file = extract_dir / "Part3_MLR_Rebate_Calculation.csv"

        if not header_file.exists() or not part3_file.exists():
            print(f"  SKIPPED {year} — missing required files")
            continue

        # Build issuer lookup from header (handle case-insensitive column names)
        issuers = {}
        with open(header_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Normalize keys to lowercase for cross-year compatibility
                lr = {k.lower(): v for k, v in row.items()}
                tid = lr.get("mr_submission_template_id", "").strip()
                state = lr.get("business_state", "").strip()
                if tid and state and len(state) == 2 and state != "GT" and state.upper() != "GRAND TOTAL":
                    issuers[tid] = {
                        "state_code": state,
                        "hios_issuer_id": lr.get("hios_issuer_id", "").strip(),
                        "company_name": lr.get("company_name", "").strip(),
                        "domiciliary_state": lr.get("domiciliary_state", "").strip(),
                        "not_for_profit": lr.get("not_for_profit", "").strip(),
                    }

        # Read Part3 (MLR calculations) — pivot from long to wide per issuer
        issuer_data = {}  # tid -> {market -> {metric -> value}}
        with open(part3_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                lr = {k.lower(): v for k, v in row.items()}
                tid = lr.get("mr_submission_template_id", "").strip()
                code = lr.get("row_lookup_code", "").strip()

                if code not in MLR_KEY_ROWS or tid not in issuers:
                    continue

                metric = MLR_KEY_ROWS[code]

                if tid not in issuer_data:
                    issuer_data[tid] = {}

                # Extract per-market values (case-insensitive)
                for market_prefix, market_name in [
                    ("cmm_individual_total", "individual"),
                    ("cmm_small_group_total", "small_group"),
                    ("cmm_large_group_total", "large_group"),
                ]:
                    val = lr.get(market_prefix, "").strip()
                    if val and val != "":
                        if market_name not in issuer_data[tid]:
                            issuer_data[tid][market_name] = {}
                        try:
                            issuer_data[tid][market_name][metric] = float(val)
                        except ValueError:
                            pass

        # Flatten to records
        for tid, markets in issuer_data.items():
            issuer = issuers.get(tid, {})
            for market_name, metrics in markets.items():
                if not metrics:
                    continue
                record = {
                    "report_year": year,
                    "state_code": issuer.get("state_code", ""),
                    "market_type": market_name,
                    "hios_issuer_id": issuer.get("hios_issuer_id", ""),
                    "company_name": issuer.get("company_name", ""),
                    "domiciliary_state": issuer.get("domiciliary_state", ""),
                    "not_for_profit": issuer.get("not_for_profit", "") == "1",
                }
                for metric_name in MLR_KEY_ROWS.values():
                    record[metric_name] = metrics.get(metric_name)
                all_records.append(record)

        print(f"  {year}: {len(issuer_data)} issuers, {sum(len(m) for m in issuer_data.values())} market-entries")

    if not all_records:
        print("  SKIPPED — no MLR data loaded")
        return 0

    # Load into DuckDB
    import pandas as pd
    df = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _mlr_market AS SELECT * FROM df")

    # Clean up and add metadata
    con.execute(f"""
        CREATE OR REPLACE TABLE _mlr_final AS
        SELECT
            CAST(report_year AS INTEGER) AS report_year,
            state_code,
            market_type,
            hios_issuer_id,
            company_name,
            domiciliary_state,
            not_for_profit,
            CAST(mlr_numerator AS DOUBLE) AS mlr_numerator,
            CAST(mlr_denominator AS DOUBLE) AS mlr_denominator,
            CAST(preliminary_mlr AS DOUBLE) AS preliminary_mlr,
            CAST(credibility_adjusted_mlr AS DOUBLE) AS credibility_adjusted_mlr,
            CAST(adjusted_premium AS DOUBLE) AS adjusted_premium,
            CAST(adjusted_incurred_claims AS DOUBLE) AS adjusted_incurred_claims,
            CAST(adjusted_claims_restated AS DOUBLE) AS adjusted_claims_restated,
            CAST(quality_improvement_expenses AS DOUBLE) AS quality_improvement_expenses,
            CAST(rebate_amount AS DOUBLE) AS rebate_amount,
            CAST(rebate_liability_total AS DOUBLE) AS rebate_liability_total,
            CAST(fed_state_taxes AS DOUBLE) AS fed_state_taxes,
            CAST(risk_adjustment_net AS DOUBLE) AS risk_adjustment_net,
            CAST(premium_earned AS DOUBLE) AS premium_earned,
            CAST(covered_lives AS DOUBLE) AS covered_lives,
            CAST(life_years AS DOUBLE) AS life_years,
            'cms.gov/cciio/mlr' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _mlr_market
        WHERE state_code IS NOT NULL AND LENGTH(state_code) = 2
        ORDER BY report_year, state_code, market_type, company_name
    """)

    count = write_parquet(con, "_mlr_final", _fact_path("mlr_market"), dry_run)
    if count > 0:
        stats = con.execute("""
            SELECT
                COUNT(DISTINCT state_code) AS states,
                COUNT(DISTINCT report_year) AS years,
                COUNT(DISTINCT hios_issuer_id) AS issuers,
                COUNT(DISTINCT market_type) AS markets,
                ROUND(AVG(CASE WHEN preliminary_mlr > 0 AND preliminary_mlr < 2
                    THEN preliminary_mlr END), 4) AS avg_mlr
            FROM _mlr_final
        """).fetchone()
        print(f"  {stats[0]} states, {stats[1]} years, {stats[2]} issuers, {stats[3]} markets, avg MLR: {stats[4]}")

    con.execute("DROP TABLE IF EXISTS _mlr_market")
    con.execute("DROP TABLE IF EXISTS _mlr_final")
    return count


# ── 2. Risk Adjustment Transfers ────────────────────────────────────────

RA_URLS = {
    "appendixA": "https://www.cms.gov/files/document/final-by2024-appendixa-revised.xlsx",
    "appendixC": "https://www.cms.gov/files/document/final-by2024-appendixc-revised.xlsx",
    "appendixD": "https://www.cms.gov/files/document/final-by2024-appendixd-revised.xlsx",
}


def build_risk_adjustment(con, dry_run: bool) -> int:
    """Build fact_risk_adjustment: ACA risk adjustment state-level + issuer-level transfers."""
    print("Building fact_risk_adjustment...")

    con.execute("INSTALL spatial; LOAD spatial;")

    all_records = []

    # 1. State-level data from Appendix A
    app_a_path = RAW_DIR / "risk_adj_2024_appendixA.xlsx"
    if not _download(RA_URLS["appendixA"], app_a_path):
        print("  SKIPPED — could not download Appendix A")
        return 0

    import openpyxl
    wb = openpyxl.load_workbook(app_a_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Header row has column names
        header = [str(c).replace("\n", " ").strip() if c else "" for c in rows[0]]
        market_type = sheet_name.strip().lower().replace(" ", "_")

        for row in rows[1:]:
            if not row[0] or row[0] is None:
                continue
            state = str(row[0]).strip()
            if len(state) != 2:
                continue

            record = {
                "benefit_year": 2024,
                "state_code": state,
                "market_type": market_type,
                "level": "state",
                "avg_premium": row[1] if len(row) > 1 else None,
                "avg_premium_before_adj": row[2] if len(row) > 2 else None,
                "avg_plan_liability_risk_score": row[3] if len(row) > 3 else None,
                "allowable_rating_factor": row[4] if len(row) > 4 else None,
                "avg_actuarial_value": row[5] if len(row) > 5 else None,
                "induced_demand_factor": row[6] if len(row) > 6 else None,
                "billable_member_months": row[7] if len(row) > 7 else None,
            }
            all_records.append(record)

    wb.close()
    print(f"  Appendix A: {len(all_records)} state-market entries")

    # 2. Issuer-level transfers from Appendix C (non-merged)
    app_c_path = RAW_DIR / "risk_adj_2024_appendixC.xlsx"
    issuer_records = []

    if _download(RA_URLS["appendixC"], app_c_path):
        wb = openpyxl.load_workbook(app_c_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            for row in rows[1:]:
                if not row or not row[2]:
                    continue
                state = str(row[2]).strip()
                if len(state) != 2:
                    continue

                def _safe_float(v):
                    if v is None or v == "-" or str(v).strip() == "-":
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None

                issuer_records.append({
                    "benefit_year": 2024,
                    "state_code": state,
                    "market_type": "non_merged",
                    "level": "issuer",
                    "hios_id": str(row[0]).strip() if row[0] else None,
                    "issuer_name": str(row[1]).strip() if row[1] else None,
                    "high_cost_pool_individual": _safe_float(row[3]),
                    "high_cost_pool_small_group": _safe_float(row[4]),
                    "ra_transfer_individual": _safe_float(row[5]),
                    "ra_transfer_catastrophic": _safe_float(row[6]),
                    "ra_transfer_small_group": _safe_float(row[7]),
                })
        wb.close()
        print(f"  Appendix C: {len(issuer_records)} issuer non-merged entries")

    # 3. Issuer-level transfers from Appendix D (merged market)
    app_d_path = RAW_DIR / "risk_adj_2024_appendixD.xlsx"
    merged_records = []

    if _download(RA_URLS["appendixD"], app_d_path):
        wb = openpyxl.load_workbook(app_d_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            for row in rows[1:]:
                if not row or not row[2]:
                    continue
                state = str(row[2]).strip()
                if len(state) != 2:
                    continue

                def _safe_float(v):
                    if v is None or v == "-" or str(v).strip() == "-":
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None

                merged_records.append({
                    "benefit_year": 2024,
                    "state_code": state,
                    "market_type": "merged",
                    "level": "issuer",
                    "hios_id": str(row[0]).strip() if row[0] else None,
                    "issuer_name": str(row[1]).strip() if row[1] else None,
                    "high_cost_pool_merged": _safe_float(row[3]),
                    "ra_transfer_merged": _safe_float(row[4]),
                })
        wb.close()
        print(f"  Appendix D: {len(merged_records)} issuer merged entries")

    if not all_records and not issuer_records and not merged_records:
        print("  SKIPPED — no risk adjustment data")
        return 0

    # Build state-level table
    import pandas as pd

    if all_records:
        df_state = pd.DataFrame(all_records)
        con.execute("CREATE OR REPLACE TABLE _ra_state AS SELECT * FROM df_state")

    # Build issuer-level table (combine non-merged + merged)
    all_issuer = issuer_records + merged_records
    if all_issuer:
        df_issuer = pd.DataFrame(all_issuer)
        con.execute("CREATE OR REPLACE TABLE _ra_issuer AS SELECT * FROM df_issuer")

    # Create unified table with explicit types (NULL columns need explicit casts)
    con.execute(f"""
        CREATE OR REPLACE TABLE _ra_final AS
        -- State-level aggregates
        SELECT
            CAST(benefit_year AS INTEGER) AS benefit_year,
            state_code,
            market_type,
            'state' AS level,
            NULL::VARCHAR AS hios_id,
            NULL::VARCHAR AS issuer_name,
            CAST(avg_premium AS DOUBLE) AS avg_premium,
            CAST(avg_premium_before_adj AS DOUBLE) AS avg_premium_before_adj,
            CAST(avg_plan_liability_risk_score AS DOUBLE) AS avg_risk_score,
            CAST(allowable_rating_factor AS DOUBLE) AS allowable_rating_factor,
            CAST(avg_actuarial_value AS DOUBLE) AS avg_actuarial_value,
            CAST(induced_demand_factor AS DOUBLE) AS induced_demand_factor,
            CAST(billable_member_months AS DOUBLE) AS billable_member_months,
            NULL::DOUBLE AS ra_transfer_amount,
            NULL::DOUBLE AS high_cost_pool_amount,
            'cms.gov/cciio/risk-adjustment' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _ra_state
        WHERE state_code IS NOT NULL AND LENGTH(state_code) = 2
    """)

    if all_issuer:
        con.execute(f"""
            INSERT INTO _ra_final (
                benefit_year, state_code, market_type, level,
                hios_id, issuer_name,
                avg_premium, avg_premium_before_adj, avg_risk_score,
                allowable_rating_factor, avg_actuarial_value,
                induced_demand_factor, billable_member_months,
                ra_transfer_amount, high_cost_pool_amount,
                source, snapshot_date
            )
            SELECT
                CAST(benefit_year AS INTEGER),
                state_code,
                market_type,
                'issuer',
                CAST(hios_id AS VARCHAR),
                CAST(issuer_name AS VARCHAR),
                NULL::DOUBLE, NULL::DOUBLE, NULL::DOUBLE,
                NULL::DOUBLE, NULL::DOUBLE,
                NULL::DOUBLE, NULL::DOUBLE,
                COALESCE(
                    CAST(ra_transfer_individual AS DOUBLE),
                    CAST(ra_transfer_small_group AS DOUBLE),
                    CAST(ra_transfer_merged AS DOUBLE)
                ),
                COALESCE(
                    CAST(high_cost_pool_individual AS DOUBLE),
                    CAST(high_cost_pool_small_group AS DOUBLE),
                    CAST(high_cost_pool_merged AS DOUBLE)
                ),
                'cms.gov/cciio/risk-adjustment',
                DATE '{SNAPSHOT_DATE}'
            FROM _ra_issuer
            WHERE state_code IS NOT NULL AND LENGTH(state_code) = 2
        """)

    count = write_parquet(con, "_ra_final", _fact_path("risk_adjustment"), dry_run)
    if count > 0:
        stats = con.execute("""
            SELECT COUNT(DISTINCT state_code), COUNT(DISTINCT market_type),
                   SUM(CASE WHEN level='state' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN level='issuer' THEN 1 ELSE 0 END)
            FROM _ra_final
        """).fetchone()
        print(f"  {stats[0]} states, {stats[1]} markets, {stats[2]} state-level, {stats[3]} issuer-level")

    for t in ["_ra_state", "_ra_issuer", "_ra_final"]:
        con.execute(f"DROP TABLE IF EXISTS {t}")
    return count


# ── 3. Medicare Advantage Star Ratings ──────────────────────────────────

STAR_URLS = {
    2026: "https://www.cms.gov/files/zip/2026-star-ratings-data-tables.zip",
    2025: "https://www.cms.gov/files/zip/2025-star-ratings-data-tables.zip",
}


def build_ma_star_ratings(con, dry_run: bool) -> int:
    """Build fact_ma_star_ratings: MA plan-level star ratings."""
    print("Building fact_ma_star_ratings...")

    all_records = []

    for year, url in sorted(STAR_URLS.items()):
        zip_path = RAW_DIR / f"star_ratings_{year}.zip"
        extract_dir = RAW_DIR / f"star_{year}"

        if not _download(url, zip_path):
            continue

        if not extract_dir.exists():
            _unzip(zip_path, extract_dir)

        # Find the Summary Ratings file
        summary_files = list(extract_dir.glob("*Summary Ratings*"))
        domain_files = list(extract_dir.glob("*Domain Stars*"))

        if not summary_files:
            print(f"  SKIPPED {year} — no Summary Ratings file")
            continue

        # Parse Summary Ratings (row 0 = title, row 1 = actual header)
        # Use cp1252 encoding (Windows smart quotes in contract names)
        with open(summary_files[0], newline="", encoding="cp1252") as f:
            reader = csv.reader(f)
            title_row = next(reader)  # Skip title row
            header = next(reader)     # Actual header

            for row in reader:
                if not row or not row[0].strip():
                    continue

                contract_id = row[0].strip()
                if len(contract_id) < 3:
                    continue

                def _parse_star(val):
                    v = val.strip() if val else ""
                    if v in ("", "Not Applicable", "Not enough data available",
                             "Plan too new to be measured", "QBP    Not enough data"):
                        return None
                    try:
                        return float(v)
                    except ValueError:
                        return None

                record = {
                    "star_year": year,
                    "contract_id": contract_id,
                    "organization_type": row[1].strip() if len(row) > 1 else None,
                    "contract_name": row[2].strip() if len(row) > 2 else None,
                    "marketing_name": row[3].strip() if len(row) > 3 else None,
                    "parent_organization": row[4].strip() if len(row) > 4 else None,
                    "snp": row[5].strip() if len(row) > 5 else None,
                }

                # Column positions vary by year; try to extract Part C/D/Overall
                for i, col in enumerate(header):
                    col_clean = col.strip().lower()
                    if "part c summary" in col_clean:
                        record["part_c_summary"] = _parse_star(row[i]) if len(row) > i else None
                    elif "part d summary" in col_clean:
                        record["part_d_summary"] = _parse_star(row[i]) if len(row) > i else None
                    elif col_clean.endswith("overall"):
                        record["overall_rating"] = _parse_star(row[i]) if len(row) > i else None
                    elif "disaster" in col_clean and "2024" in col_clean:
                        record["disaster_pct_2024"] = _parse_star(row[i]) if len(row) > i else None
                    elif "disaster" in col_clean and "2023" in col_clean:
                        record["disaster_pct_2023"] = _parse_star(row[i]) if len(row) > i else None

                all_records.append(record)

        # Parse Domain Stars for domain-level detail
        if domain_files:
            domain_records = []
            with open(domain_files[0], newline="", encoding="cp1252") as f:
                reader = csv.reader(f)
                next(reader)  # Skip title
                header = next(reader)

                for row in reader:
                    if not row or not row[0].strip() or len(row[0].strip()) < 3:
                        continue

                    contract_id = row[0].strip()

                    for i in range(5, len(header)):
                        col = header[i].strip() if i < len(header) else ""
                        if not col:
                            continue
                        val = row[i].strip() if i < len(row) else ""
                        if val in ("", "Plan not required to report measure",
                                   "Not enough data available", "Plan too new to be measured"):
                            continue
                        try:
                            star_val = float(val)
                        except ValueError:
                            continue

                        domain_records.append({
                            "star_year": year,
                            "contract_id": contract_id,
                            "domain": col,
                            "domain_star": star_val,
                        })

        print(f"  {year}: {len(all_records)} contracts")

    if not all_records:
        print("  SKIPPED — no star ratings data")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _stars AS SELECT * FROM df")

    con.execute(f"""
        CREATE OR REPLACE TABLE _stars_final AS
        SELECT
            CAST(star_year AS INTEGER) AS star_year,
            TRIM(contract_id) AS contract_id,
            TRIM(organization_type) AS organization_type,
            TRIM(contract_name) AS contract_name,
            TRIM(marketing_name) AS marketing_name,
            TRIM(parent_organization) AS parent_organization,
            CASE WHEN TRIM(snp) = 'Yes' THEN TRUE WHEN TRIM(snp) = 'No' THEN FALSE ELSE NULL END AS is_snp,
            CAST(part_c_summary AS DOUBLE) AS part_c_summary,
            CAST(part_d_summary AS DOUBLE) AS part_d_summary,
            CAST(overall_rating AS DOUBLE) AS overall_rating,
            'cms.gov/ma-star-ratings' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _stars
        ORDER BY star_year, contract_id
    """)

    count = write_parquet(con, "_stars_final", _fact_path("ma_star_ratings"), dry_run)
    if count > 0:
        stats = con.execute("""
            SELECT COUNT(DISTINCT star_year), COUNT(DISTINCT contract_id),
                   ROUND(AVG(overall_rating), 2),
                   SUM(CASE WHEN overall_rating >= 4 THEN 1 ELSE 0 END),
                   COUNT(overall_rating)
            FROM _stars_final
        """).fetchone()
        print(f"  {stats[0]} years, {stats[1]} contracts, avg overall: {stats[2]}, "
              f"{stats[3]}/{stats[4]} rated 4+")

    con.execute("DROP TABLE IF EXISTS _stars")
    con.execute("DROP TABLE IF EXISTS _stars_final")
    return count


# ── 4. Census Health Insurance Coverage ─────────────────────────────────

CENSUS_HIC_URLS = {
    "hic04_acs": "https://www2.census.gov/programs-surveys/demo/tables/health-insurance/time-series/acs/hic04_acs.xlsx",
    "hic05_acs": "https://www2.census.gov/programs-surveys/demo/tables/health-insurance/time-series/acs/hic05_acs.xlsx",
    "hic06_acs": "https://www2.census.gov/programs-surveys/demo/tables/health-insurance/time-series/acs/hic06_acs.xlsx",
}

STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "United States": "US",
}


def build_census_health_insurance(con, dry_run: bool) -> int:
    """Build fact_census_health_insurance: Coverage by type and state, 2008-2024."""
    print("Building fact_census_health_insurance...")

    con.execute("INSTALL spatial; LOAD spatial;")

    # Download HIC-04 (main coverage by type and state)
    hic04_path = RAW_DIR / "hic04_acs.xlsx"
    if not _download(CENSUS_HIC_URLS["hic04_acs"], hic04_path):
        print("  SKIPPED — could not download HIC-04")
        return 0

    # Read via DuckDB spatial (handles messy xlsx)
    df = con.execute(f"SELECT * FROM st_read('{hic04_path}')").fetchdf()

    # Find header row (has "Nation/State" in first column)
    header_idx = None
    for i in range(min(10, len(df))):
        val = str(df.iloc[i, 0]).strip()
        if "Nation/State" in val:
            header_idx = i
            break

    if header_idx is None:
        print("  SKIPPED — could not find header row (looking for 'Nation/State')")
        return 0

    print(f"  Header found at row {header_idx}")

    # Data starts 2 rows after Nation/State header row (skip sub-header)
    data_start = header_idx + 2

    # Build year list from header row
    # Each year occupies 4 columns: Estimate, MOE, Percent, MOE
    # Years appear at col 2, 6, 10, 14, 18, ... (stride of 4)
    year_row = df.iloc[header_idx]
    years = []  # list of (col_start_of_estimate, year_int)

    for col_idx in range(2, len(year_row)):
        val = str(year_row.iloc[col_idx]).strip()
        if val in ("nan", "None", ""):
            continue
        # Handle "2020 2" (footnote numbers appended to year)
        year_match = re.search(r"(\d{4})", val)
        if year_match:
            year_int = int(year_match.group(1))
            if 2000 <= year_int <= 2030:
                years.append((col_idx, year_int))

    print(f"  Found {len(years)} years: {[y for _, y in years]}")

    def _clean_num(v):
        if v is None:
            return None
        s = str(v).strip().replace(",", "")
        if s in ("", "nan", "None", "N", "(X)", "***", "...", "-"):
            return None
        try:
            return float(s)
        except ValueError:
            return None

    # Parse data rows
    all_records = []
    current_state = None

    for i in range(data_start, len(df)):
        state_val = str(df.iloc[i, 0]).strip()
        coverage_val = str(df.iloc[i, 1]).strip() if df.iloc[i, 1] is not None else ""

        if state_val and state_val not in ("nan", "None", ""):
            current_state = state_val.strip()

        if not current_state or not coverage_val or coverage_val in ("nan", "None", ""):
            continue

        # Skip footnote rows
        if coverage_val.startswith("Source:") or coverage_val.startswith("Note:"):
            break

        # Clean coverage type (remove leading dots used for indentation)
        coverage_clean = coverage_val.lstrip(".")
        indent = len(coverage_val) - len(coverage_clean)

        state_code = STATE_NAME_TO_CODE.get(current_state)
        if not state_code:
            continue

        for col_start, year in years:
            # Estimate is at col_start, Percent is at col_start + 2
            est_val = df.iloc[i, col_start] if col_start < len(df.columns) else None
            pct_val = df.iloc[i, col_start + 2] if col_start + 2 < len(df.columns) else None

            est = _clean_num(est_val)
            pct = _clean_num(pct_val)

            if est is None and pct is None:
                continue

            all_records.append({
                "year": year,
                "state_code": state_code,
                "coverage_type": coverage_clean,
                "coverage_level": indent,  # 0=top, 1=sub, 2=sub-sub
                "estimate_thousands": est,
                "pct_of_population": pct,
            })

    if not all_records:
        print("  SKIPPED — no Census data parsed")
        return 0

    import pandas as pd
    df_out = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _census_hi AS SELECT * FROM df_out")

    con.execute(f"""
        CREATE OR REPLACE TABLE _census_hi_final AS
        SELECT
            CAST(year AS INTEGER) AS year,
            state_code,
            coverage_type,
            CAST(coverage_level AS INTEGER) AS coverage_level,
            CAST(estimate_thousands AS DOUBLE) AS estimate_thousands,
            CAST(pct_of_population AS DOUBLE) AS pct_of_population,
            'census.gov/acs/hic-04' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _census_hi
        WHERE state_code IS NOT NULL
        ORDER BY year DESC, state_code, coverage_level, coverage_type
    """)

    count = write_parquet(con, "_census_hi_final", _fact_path("census_health_insurance"), dry_run)
    if count > 0:
        stats = con.execute("""
            SELECT COUNT(DISTINCT year), COUNT(DISTINCT state_code),
                   COUNT(DISTINCT coverage_type)
            FROM _census_hi_final
        """).fetchone()
        print(f"  {stats[0]} years, {stats[1]} states, {stats[2]} coverage types")

    con.execute("DROP TABLE IF EXISTS _census_hi")
    con.execute("DROP TABLE IF EXISTS _census_hi_final")
    return count


# ── 5. MEPS-IC Employer Insurance ───────────────────────────────────────

MEPS_TABLES = {
    "single_premium": {
        "url": "https://meps.ahrq.gov/data_stats/summ_tables/insr/state/series_2/2020/tiic1.htm",
        "metric": "avg_single_premium",
    },
    "family_premium": {
        "url": "https://meps.ahrq.gov/data_stats/summ_tables/insr/state/series_2/2020/tiid1.htm",
        "metric": "avg_family_premium",
    },
    "single_contribution_pct": {
        "url": "https://meps.ahrq.gov/data_stats/summ_tables/insr/state/series_2/2020/tiic3.htm",
        "metric": "employee_single_share_pct",
        "is_pct": True,
    },
    "family_contribution_pct": {
        "url": "https://meps.ahrq.gov/data_stats/summ_tables/insr/state/series_2/2020/tiid3.htm",
        "metric": "employee_family_share_pct",
        "is_pct": True,
    },
}


class MepsTableParser(HTMLParser):
    """Parse MEPS-IC HTML summary tables. Only reads the first table
    (second table is typically standard errors / RSEs)."""

    def __init__(self):
        super().__init__()
        self.in_table = False
        self.in_row = False
        self.in_cell = False
        self.table_count = 0
        self.current_row = []
        self.rows = []
        self.cell_text = ""

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self.table_count += 1
            if self.table_count == 1:
                self.in_table = True
        elif tag == "tr" and self.in_table:
            self.in_row = True
            self.current_row = []
        elif tag in ("td", "th") and self.in_row:
            self.in_cell = True
            self.cell_text = ""

    def handle_endtag(self, tag):
        if tag == "table" and self.in_table:
            self.in_table = False
        elif tag == "tr" and self.in_row:
            self.in_row = False
            if self.current_row:
                self.rows.append(self.current_row)
        elif tag in ("td", "th") and self.in_cell:
            self.in_cell = False
            self.current_row.append(self.cell_text.strip())

    def handle_data(self, data):
        if self.in_cell:
            self.cell_text += data


def _parse_meps_html(html_path: Path, is_pct: bool = False) -> list[dict]:
    """Parse a MEPS-IC HTML table into state-level records."""
    with open(html_path, encoding="utf-8") as f:
        content = f.read()

    parser = MepsTableParser()
    parser.feed(content)

    results = []

    # Map header names to our firm_size keys
    header_map = {
        "Total": "total",
        "Less than 10 employees": "lt_10",
        "10-24 employees": "10_24",
        "25-99 employees": "25_99",
        "100-999 employees": "100_999",
        "1000 or more employees": "1000_plus",
        "Less than 50 employees": "lt_50",
        "50 or more employees": "50_plus",
    }

    # First row is the header
    if not parser.rows:
        return results

    header_row = parser.rows[0]
    col_mapping = []  # list of (col_index, firm_size)
    for i, col_name in enumerate(header_row):
        key = header_map.get(col_name.strip())
        if key:
            col_mapping.append((i, key))

    for row in parser.rows[1:]:  # skip header
        if not row or len(row) < 2:
            continue

        state_name = row[0].strip().rstrip(":")
        state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            continue

        for col_idx, size in col_mapping:
            if col_idx >= len(row):
                break
            val_str = row[col_idx].strip()
            # Remove non-breaking spaces, asterisks, dollar signs, commas, percent signs
            val_str = val_str.replace("\xa0", " ").replace("*", "").replace("$", "")
            val_str = val_str.replace(",", "").replace("%", "").strip()
            if val_str in ("", "--", "NSD", "N/A"):
                continue
            try:
                val = float(val_str)
            except ValueError:
                continue

            results.append({
                "state_code": state_code,
                "firm_size": size,
                "value": val,
            })

    return results


def build_meps_employer_insurance(con, dry_run: bool) -> int:
    """Build fact_meps_employer_insurance: MEPS-IC state-level employer premiums."""
    print("Building fact_meps_employer_insurance...")

    all_records = {}  # (state, firm_size) -> {metric: value}

    for table_name, config in MEPS_TABLES.items():
        htm_path = RAW_DIR / f"meps_{table_name}.htm"
        if not _download(config["url"], htm_path):
            continue

        is_pct = config.get("is_pct", False)
        parsed = _parse_meps_html(htm_path, is_pct=is_pct)
        metric = config["metric"]

        for entry in parsed:
            key = (entry["state_code"], entry["firm_size"])
            if key not in all_records:
                all_records[key] = {"state_code": entry["state_code"],
                                    "firm_size": entry["firm_size"]}
            all_records[key][metric] = entry["value"]

        print(f"  {table_name}: {len(parsed)} entries")

    if not all_records:
        print("  SKIPPED — no MEPS data")
        return 0

    records = list(all_records.values())
    import pandas as pd
    df = pd.DataFrame(records)
    con.execute("CREATE OR REPLACE TABLE _meps AS SELECT * FROM df")

    # Build column list dynamically based on what metrics we have
    metrics_available = set()
    for rec in records:
        for k in rec.keys():
            if k not in ("state_code", "firm_size"):
                metrics_available.add(k)

    metric_cols = []
    for m in ["avg_single_premium", "avg_family_premium",
              "employee_single_share_pct", "employee_family_share_pct"]:
        if m in metrics_available:
            metric_cols.append(f"TRY_CAST({m} AS DOUBLE) AS {m}")
        else:
            metric_cols.append(f"NULL::DOUBLE AS {m}")

    metric_sql = ",\n            ".join(metric_cols)

    con.execute(f"""
        CREATE OR REPLACE TABLE _meps_final AS
        SELECT
            2020 AS data_year,
            state_code,
            firm_size,
            {metric_sql},
            'meps.ahrq.gov/ic' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _meps
        WHERE state_code IS NOT NULL
        ORDER BY state_code, firm_size
    """)

    count = write_parquet(con, "_meps_final", _fact_path("meps_employer_insurance"), dry_run)
    if count > 0:
        stats = con.execute("""
            SELECT COUNT(DISTINCT state_code), COUNT(DISTINCT firm_size),
                   ROUND(AVG(avg_single_premium), 0),
                   ROUND(AVG(avg_family_premium), 0)
            FROM _meps_final WHERE firm_size = 'total'
        """).fetchone()
        single_avg = f"${stats[2]:,.0f}" if stats[2] else "N/A"
        family_avg = f"${stats[3]:,.0f}" if stats[3] else "N/A"
        print(f"  {stats[0]} states, {stats[1]} firm sizes, "
              f"avg single: {single_avg}, avg family: {family_avg}")

    con.execute("DROP TABLE IF EXISTS _meps")
    con.execute("DROP TABLE IF EXISTS _meps_final")
    return count


# ── Main ────────────────────────────────────────────────────────────────

BUILDERS = {
    "fact_mlr_market": build_mlr_market,
    "fact_risk_adjustment": build_risk_adjustment,
    "fact_ma_star_ratings": build_ma_star_ratings,
    "fact_census_health_insurance": build_census_health_insurance,
    "fact_meps_employer_insurance": build_meps_employer_insurance,
}


def main():
    parser = argparse.ArgumentParser(description="Build insurance market lake tables")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only", type=str, default="",
                        help="Comma-separated list of table names to build")
    args = parser.parse_args()

    targets = set(args.only.split(",")) if args.only else set(BUILDERS.keys())

    RAW_DIR.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect()
    con.execute("SET memory_limit='4GB'")

    results = {}
    total = 0

    print(f"=== Insurance Market Lake Build ({SNAPSHOT_DATE}) ===")
    print(f"Run ID: {RUN_ID}")
    print()

    for name, builder in BUILDERS.items():
        if name not in targets:
            continue
        try:
            count = builder(con, args.dry_run)
            results[name] = count
            total += count
            print()
        except Exception as e:
            print(f"  ERROR building {name}: {e}")
            import traceback
            traceback.print_exc()
            results[name] = 0
            print()

    # Write manifest
    if not args.dry_run and total > 0:
        manifest = {
            "run_id": RUN_ID,
            "snapshot_date": SNAPSHOT_DATE,
            "script": "build_lake_insurance_market.py",
            "tables": results,
            "total_rows": total,
            "created_at": datetime.now().isoformat(),
        }
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest_path = META_DIR / f"insurance_market_{SNAPSHOT_DATE}.json"
        with open(manifest_path, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"Manifest: {manifest_path.relative_to(LAKE_DIR)}")

    print(f"\n=== Summary ===")
    for name, count in results.items():
        status = f"{count:,} rows" if count > 0 else "SKIPPED"
        print(f"  {name}: {status}")
    print(f"  Total: {total:,} rows")


if __name__ == "__main__":
    main()
