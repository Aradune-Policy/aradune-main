#!/usr/bin/env python3
"""
build_lake_federal_data.py — Ingest federal regulatory and policy data into the lake.

Downloads and processes:
  1. Federal Register CMS documents (5,900+ docs, all types)
  2. MFCU Statistics (Medicaid Fraud Control Units, per-state)
  3. National Health Expenditure (NHE) historical data
  4. NHE Projections
  5. PERM Improper Payment Rates

Tables built:
  fact_federal_register_cms  — CMS Federal Register documents (Rules, Notices, Proposed Rules)
  fact_mfcu_stats            — MFCU per-state fraud/recovery statistics FY2024
  fact_nhe                   — National Health Expenditure historical (1960-2024)
  fact_nhe_projections       — NHE projections (2024-2033)
  fact_perm_rates            — PERM improper payment rates (2020-2025)

Usage:
  python3 scripts/build_lake_federal_data.py
  python3 scripts/build_lake_federal_data.py --dry-run
  python3 scripts/build_lake_federal_data.py --only fact_federal_register_cms
"""

import argparse
import json
import time
import uuid
from datetime import date, datetime
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


# ---------------------------------------------------------------------------
# 1. Federal Register CMS Documents
# ---------------------------------------------------------------------------

def _fetch_federal_register_cms() -> list[dict]:
    """Fetch all CMS documents from the Federal Register API."""
    import subprocess

    all_docs = []
    page = 1
    total_pages = None
    fields = [
        "document_number", "title", "type", "publication_date", "abstract",
        "html_url", "pdf_url", "cfr_references",
        "docket_ids", "citation", "action", "agencies",
    ]
    field_params = "&".join(f"fields%5B%5D={f}" for f in fields)

    while True:
        url = (
            f"https://www.federalregister.gov/api/v1/documents.json"
            f"?conditions%5Bagencies%5D%5B%5D=centers-for-medicare-medicaid-services"
            f"&per_page=1000&page={page}&{field_params}"
        )
        result = subprocess.run(
            ["curl", "-s", url],
            capture_output=True, text=True, timeout=60
        )
        try:
            data = json.loads(result.stdout)
        except json.JSONDecodeError:
            print(f"  ERROR: Failed to parse response on page {page}")
            break

        # Check for API error
        if "status" in data and data.get("status") != 200:
            print(f"  API Error: {data.get('message', 'unknown')}")
            break

        if total_pages is None:
            total_pages = data.get("total_pages", 1)
            total_count = data.get("count", "?")
            print(f"  Federal Register: {total_count} total documents, {total_pages} pages")

        results = data.get("results", [])
        if not results:
            break

        all_docs.extend(results)
        print(f"  Page {page}/{total_pages}: fetched {len(results)} docs (total: {len(all_docs)})")

        if page >= total_pages:
            break
        page += 1
        time.sleep(0.5)  # Be polite to the API

    return all_docs


def build_fact_federal_register_cms(con, dry_run: bool) -> int:
    print("Building fact_federal_register_cms...")

    # Save raw JSON
    json_path = RAW_DIR / "federal_register_cms.json"
    if not json_path.exists() or json_path.stat().st_size < 10000:
        docs = _fetch_federal_register_cms()
        with open(json_path, "w") as f:
            json.dump(docs, f)
        print(f"  Saved {len(docs)} docs to {json_path.name}")
    else:
        with open(json_path) as f:
            docs = json.load(f)
        print(f"  Using cached {json_path.name} ({len(docs)} docs)")

    if not docs:
        print("  SKIPPED: no documents fetched")
        return 0

    # Process into rows
    rows = []
    for doc in docs:
        cfr_refs = doc.get("cfr_references", []) or []
        cfr_str = "; ".join(
            f"{r.get('title','')} CFR {r.get('part','')}"
            for r in cfr_refs if r
        ) if cfr_refs else None

        docket_ids = doc.get("docket_ids", []) or []
        docket_str = "; ".join(docket_ids) if docket_ids else None

        agencies = doc.get("agencies", []) or []
        agency_names = "; ".join(
            a.get("name", "") for a in agencies if a and a.get("name")
        ) if agencies else "CMS"

        rows.append({
            "document_number": doc.get("document_number"),
            "title": doc.get("title"),
            "doc_type": doc.get("type"),
            "publication_date": doc.get("publication_date"),
            "abstract": doc.get("abstract"),
            "action": doc.get("action"),
            "citation": doc.get("citation"),
            "html_url": doc.get("html_url"),
            "pdf_url": doc.get("pdf_url"),
            "cfr_references": cfr_str,
            "docket_ids": docket_str,
            "agencies": agency_names,
        })

    # Load into DuckDB
    import pandas as pd
    df = pd.DataFrame(rows)
    con.register("_fr_df", df)

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_fr_cms AS
        SELECT
            document_number,
            title,
            doc_type,
            CAST(publication_date AS DATE) AS publication_date,
            EXTRACT(YEAR FROM CAST(publication_date AS DATE)) AS year,
            abstract,
            action,
            citation,
            html_url,
            pdf_url,
            cfr_references,
            docket_ids,
            agencies,
            'federalregister.gov' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _fr_df
        WHERE document_number IS NOT NULL
        ORDER BY publication_date DESC
    """)

    count = write_parquet(con, "_fact_fr_cms", _snapshot_path("federal_register_cms"), dry_run)

    # Stats
    types = con.execute("""
        SELECT doc_type, COUNT(*) as cnt
        FROM _fact_fr_cms GROUP BY doc_type ORDER BY cnt DESC
    """).fetchall()
    date_range = con.execute("""
        SELECT MIN(publication_date), MAX(publication_date) FROM _fact_fr_cms
    """).fetchone()
    print(f"  {count:,} rows, types: {dict(types)}")
    print(f"  Date range: {date_range[0]} to {date_range[1]}")

    con.execute("DROP TABLE IF EXISTS _fact_fr_cms")
    return count


# ---------------------------------------------------------------------------
# 2. MFCU Statistics
# ---------------------------------------------------------------------------

def build_fact_mfcu_stats(con, dry_run: bool) -> int:
    print("Building fact_mfcu_stats...")

    xlsx_path = RAW_DIR / "mfcu_fy2024_statistical_chart.xlsx"
    if not xlsx_path.exists():
        print(f"  SKIPPED: {xlsx_path.name} not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Row 2 is the header
    headers = [cell.value for cell in ws[2]]
    # Clean header names
    header_map = {
        0: "state_name",
        1: "total_investigations",
        2: "fraud_investigations",
        3: "abuse_neglect_investigations",
        4: "total_indictments",
        5: "fraud_indictments",
        6: "abuse_neglect_indictments",
        7: "total_convictions",
        8: "fraud_convictions",
        9: "abuse_neglect_convictions",
        10: "civil_settlements_judgments",
        11: "total_recoveries",
        12: "total_criminal_recoveries",
        13: "civil_recoveries_global",
        14: "civil_recoveries_other",
        15: "mfcu_grant_expenditures",
        16: "total_medicaid_expenditures",
        17: "staff_on_board",
    }

    STATE_NAME_TO_CODE = {
        "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR",
        "CALIFORNIA": "CA", "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE",
        "DISTRICT OF COLUMBIA": "DC", "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI",
        "IDAHO": "ID", "ILLINOIS": "IL", "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS",
        "KENTUCKY": "KY", "LOUISIANA": "LA", "MAINE": "ME", "MARYLAND": "MD",
        "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN", "MISSISSIPPI": "MS",
        "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
        "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
        "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK",
        "OREGON": "OR", "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI",
        "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD", "TENNESSEE": "TN", "TEXAS": "TX",
        "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA", "WASHINGTON": "WA",
        "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
        "PUERTO RICO": "PR", "GUAM": "GU", "VIRGIN ISLANDS": "VI",
        "AMERICAN SAMOA": "AS", "NORTHERN MARIANA ISLANDS": "MP",
        "U.S. VIRGIN ISLANDS": "VI",
    }

    rows = []
    for row_idx in range(3, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=c + 1).value for c in range(18)]
        state_name = row_data[0]
        if not state_name or not isinstance(state_name, str):
            continue
        state_name = state_name.strip().upper()
        # Skip totals row
        if "TOTAL" in state_name or "AVERAGE" in state_name or state_name == "":
            continue
        state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            print(f"  WARNING: unknown state '{state_name}'")
            continue

        row_dict = {"state_code": state_code, "state_name": state_name, "fiscal_year": 2024}
        for i in range(1, 18):
            col_name = header_map.get(i)
            if col_name:
                val = row_data[i]
                if val is not None and isinstance(val, (int, float)):
                    row_dict[col_name] = float(val)
                else:
                    row_dict[col_name] = 0.0
        rows.append(row_dict)

    if not rows:
        print("  SKIPPED: no state rows found")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.register("_mfcu_df", df)

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_mfcu AS
        SELECT
            state_code,
            state_name,
            CAST(fiscal_year AS INTEGER) AS fiscal_year,
            CAST(total_investigations AS INTEGER) AS total_investigations,
            CAST(fraud_investigations AS INTEGER) AS fraud_investigations,
            CAST(abuse_neglect_investigations AS INTEGER) AS abuse_neglect_investigations,
            CAST(total_indictments AS INTEGER) AS total_indictments,
            CAST(fraud_indictments AS INTEGER) AS fraud_indictments,
            CAST(abuse_neglect_indictments AS INTEGER) AS abuse_neglect_indictments,
            CAST(total_convictions AS INTEGER) AS total_convictions,
            CAST(fraud_convictions AS INTEGER) AS fraud_convictions,
            CAST(abuse_neglect_convictions AS INTEGER) AS abuse_neglect_convictions,
            CAST(civil_settlements_judgments AS INTEGER) AS civil_settlements_judgments,
            total_recoveries,
            total_criminal_recoveries,
            civil_recoveries_global,
            civil_recoveries_other,
            mfcu_grant_expenditures,
            total_medicaid_expenditures,
            CAST(staff_on_board AS INTEGER) AS staff_on_board,
            CASE WHEN total_medicaid_expenditures > 0
                THEN total_recoveries / total_medicaid_expenditures * 100
                ELSE NULL
            END AS recovery_pct_of_medicaid,
            CASE WHEN mfcu_grant_expenditures > 0
                THEN total_recoveries / mfcu_grant_expenditures
                ELSE NULL
            END AS roi_ratio,
            'oig.hhs.gov' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _mfcu_df
        ORDER BY state_code
    """)

    count = write_parquet(con, "_fact_mfcu", _snapshot_path("mfcu_stats"), dry_run)

    # Stats
    total_rec = con.execute("SELECT SUM(total_recoveries) FROM _fact_mfcu").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_mfcu").fetchone()[0]
    print(f"  {count:,} rows, {states} states/territories")
    print(f"  Total recoveries: ${total_rec:,.2f}")

    con.execute("DROP TABLE IF EXISTS _fact_mfcu")
    return count


# ---------------------------------------------------------------------------
# 3. NHE Historical Data
# ---------------------------------------------------------------------------

def _parse_nhe_table03_historical() -> list[dict]:
    """Parse Table 03 (NHE by Source of Funds) from historical NHE data."""
    xlsx_path = RAW_DIR / "nhe_historical" / "Table 03 National Health Expenditures, by Source of Funds.xlsx"
    if not xlsx_path.exists():
        return []

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Row 2: headers (years)
    years = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and isinstance(val, (int, float)):
            years.append(int(val))
        elif val and str(val).strip().isdigit():
            years.append(int(str(val).strip()))

    if not years:
        return []

    rows = []
    # Rows 4 onward: data
    for row_idx in range(4, ws.max_row + 1):
        item = ws.cell(row=row_idx, column=1).value
        if not item or not isinstance(item, str):
            continue
        item = item.strip()
        if not item or item.startswith("NOTE") or item.startswith("SOURCE"):
            continue

        for i, year in enumerate(years):
            val = ws.cell(row=row_idx, column=i + 2).value
            if val is None or val == "" or val == "---":
                continue
            try:
                amount = float(val)
            except (ValueError, TypeError):
                continue

            rows.append({
                "category": item,
                "year": year,
                "amount_billions": amount,
                "table_name": "Table 03",
                "data_type": "historical",
            })

    return rows


def _parse_nhe_table01_historical() -> list[dict]:
    """Parse Table 01 (NHE Aggregate and Per Capita) from historical NHE data."""
    xlsx_path = RAW_DIR / "nhe_historical" / "Table 01 National Health Expenditures; Aggregate and Per Capita Amounts.xlsx"
    if not xlsx_path.exists():
        return []

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Row 2: years
    years = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and isinstance(val, (int, float)):
            years.append(int(val))
        elif val and str(val).strip().isdigit():
            years.append(int(str(val).strip()))

    if not years:
        return []

    rows = []
    for row_idx in range(4, ws.max_row + 1):
        item = ws.cell(row=row_idx, column=1).value
        if not item or not isinstance(item, str):
            continue
        item = item.strip()
        if not item or item.startswith("NOTE") or item.startswith("SOURCE") or item.startswith("1 "):
            continue

        for i, year in enumerate(years):
            val = ws.cell(row=row_idx, column=i + 2).value
            if val is None or val == "" or val == "---":
                continue
            try:
                amount = float(val)
            except (ValueError, TypeError):
                continue

            rows.append({
                "category": item,
                "year": year,
                "amount_billions": amount,
                "table_name": "Table 01",
                "data_type": "historical",
            })

    return rows


def _parse_nhe_table21_historical() -> list[dict]:
    """Parse Table 21 (Insurance Enrollment and Per Enrollee) from historical NHE data."""
    xlsx_path = RAW_DIR / "nhe_historical" / "Table 21 Expenditures, Enrollment and Per Enrollee Estimates of Health Insurance.xlsx"
    if not xlsx_path.exists():
        return []

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    years = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and isinstance(val, (int, float)):
            years.append(int(val))
        elif val and str(val).strip().isdigit():
            years.append(int(str(val).strip()))

    if not years:
        return []

    rows = []
    for row_idx in range(4, ws.max_row + 1):
        item = ws.cell(row=row_idx, column=1).value
        if not item or not isinstance(item, str):
            continue
        item = item.strip()
        if not item or item.startswith("NOTE") or item.startswith("SOURCE") or item.startswith("1 "):
            continue

        for i, year in enumerate(years):
            val = ws.cell(row=row_idx, column=i + 2).value
            if val is None or val == "" or val == "---":
                continue
            try:
                amount = float(val)
            except (ValueError, TypeError):
                continue

            rows.append({
                "category": item,
                "year": year,
                "amount_billions": amount,
                "table_name": "Table 21",
                "data_type": "historical",
            })

    return rows


def build_fact_nhe(con, dry_run: bool) -> int:
    print("Building fact_nhe...")

    nhe_dir = RAW_DIR / "nhe_historical"
    if not nhe_dir.exists():
        print(f"  SKIPPED: {nhe_dir} not found. Run download first.")
        return 0

    all_rows = []
    # Parse key tables
    rows01 = _parse_nhe_table01_historical()
    print(f"  Table 01 (Aggregate): {len(rows01)} data points")
    all_rows.extend(rows01)

    rows03 = _parse_nhe_table03_historical()
    print(f"  Table 03 (Source of Funds): {len(rows03)} data points")
    all_rows.extend(rows03)

    rows21 = _parse_nhe_table21_historical()
    print(f"  Table 21 (Enrollment/Per Enrollee): {len(rows21)} data points")
    all_rows.extend(rows21)

    if not all_rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_rows)
    con.register("_nhe_df", df)

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_nhe AS
        SELECT
            category,
            CAST(year AS INTEGER) AS year,
            CAST(amount_billions AS DOUBLE) AS amount_billions,
            table_name,
            data_type,
            'cms.gov/nhe' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _nhe_df
        ORDER BY table_name, category, year
    """)

    count = write_parquet(con, "_fact_nhe", _snapshot_path("nhe"), dry_run)
    years_range = con.execute("SELECT MIN(year), MAX(year) FROM _fact_nhe").fetchone()
    cats = con.execute("SELECT COUNT(DISTINCT category) FROM _fact_nhe").fetchone()[0]
    print(f"  {count:,} rows, {cats} categories, years {years_range[0]}-{years_range[1]}")

    # Medicaid-specific highlight (amount > 100 to skip pct distribution rows)
    med = con.execute("""
        SELECT year, amount_billions FROM _fact_nhe
        WHERE category LIKE '%Medicaid%' AND table_name = 'Table 03'
          AND amount_billions > 100
        ORDER BY year DESC LIMIT 1
    """).fetchone()
    if med:
        print(f"  Latest Medicaid spending (Table 03): ${med[1]:,.1f}B ({med[0]})")

    con.execute("DROP TABLE IF EXISTS _fact_nhe")
    return count


# ---------------------------------------------------------------------------
# 4. NHE Projections
# ---------------------------------------------------------------------------

def _parse_nhe_projections_table03() -> list[dict]:
    """Parse projections Table 03 (NHE by Sources of Funds)."""
    xlsx_path = RAW_DIR / "nhe_projections" / "Table03 National Health Expenditures by Sources of Funds.xlsx"
    if not xlsx_path.exists():
        return []

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Row 3 has column headers
    col_names = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val:
            col_names.append(str(val).strip())
        else:
            col_names.append(None)

    rows = []
    current_section = "historical"

    for row_idx in range(4, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell is None:
            continue

        # Handle string labels
        if isinstance(first_cell, str):
            first_cell_str = first_cell.strip()
            if first_cell_str == "Projected":
                current_section = "projected"
                continue
            if first_cell_str in ("Historical Estimates", "Amount in Billions",
                                   "Average Annual Percent Change from Previous Year Shown"):
                if "Percent Change" in first_cell_str:
                    current_section = "pct_change"
                continue
            if "Per Capita" in first_cell_str:
                current_section = "per_capita"
                continue
            if first_cell_str.startswith("NOTE") or first_cell_str.startswith("SOURCE"):
                continue
            # Try parsing string as year
            try:
                year = int(first_cell_str)
            except (ValueError, TypeError):
                continue
        elif isinstance(first_cell, (int, float)):
            year = int(first_cell)
        else:
            continue

        # This is a data row with a year
        for col_idx in range(1, min(len(col_names), ws.max_column)):
            col_name = col_names[col_idx]
            if not col_name:
                continue
            val = ws.cell(row=row_idx, column=col_idx + 1).value
            if val is None or val == "":
                continue
            # Skip em-dash or non-numeric strings
            if isinstance(val, str):
                val_str = val.strip()
                if val_str in ("---", "—", "-"):
                    continue
                try:
                    amount = float(val_str.replace(",", ""))
                except (ValueError, TypeError):
                    continue
            elif isinstance(val, (int, float)):
                amount = float(val)
            else:
                continue

            rows.append({
                "category": col_name,
                "year": year,
                "amount_billions": amount,
                "data_type": current_section,
            })

    return rows


def _parse_nhe_projections_table01() -> list[dict]:
    """Parse projections Table 01 (NHE and Economic Indicators)."""
    xlsx_path = RAW_DIR / "nhe_projections" / "Table01 National Health Expenditures and Selected Economic Indicators.xlsx"
    if not xlsx_path.exists():
        return []

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Row 2 has "Projected" label at the column where projections start
    projected_start_col = None
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and isinstance(val, str) and "Projected" in val:
            projected_start_col = col
            break

    # Row 3 has years
    years = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val:
            try:
                years.append((int(float(str(val).strip())), col))
            except (ValueError, TypeError):
                years.append(None)
        else:
            years.append(None)

    rows = []
    current_section = "amount"
    for row_idx in range(4, ws.max_row + 1):
        item = ws.cell(row=row_idx, column=1).value
        if not item or not isinstance(item, str):
            continue
        item = item.strip()
        if not item or item.startswith("NOTE") or item.startswith("SOURCE") or item.startswith("1 "):
            continue
        if item == "Amount in Billions":
            current_section = "amount"
            continue
        if item == "Level":
            current_section = "index"
            continue
        if "Percent" in item and "Change" in item:
            current_section = "pct_change"
            continue

        for i, year_info in enumerate(years):
            if year_info is None:
                continue
            year, col_idx = year_info
            val = ws.cell(row=row_idx, column=i + 2).value
            if val is None or val == "" or val == "---":
                continue
            try:
                amount = float(val)
            except (ValueError, TypeError):
                continue

            # Determine if historical or projected based on column position
            if projected_start_col and col_idx >= projected_start_col:
                data_type = "projected"
            else:
                data_type = "historical"

            rows.append({
                "category": item,
                "year": year,
                "amount_billions": amount,
                "data_type": data_type,
                "metric_type": current_section,
            })

    return rows


def build_fact_nhe_projections(con, dry_run: bool) -> int:
    print("Building fact_nhe_projections...")

    proj_dir = RAW_DIR / "nhe_projections"
    if not proj_dir.exists():
        print(f"  SKIPPED: {proj_dir} not found.")
        return 0

    all_rows = []

    rows03 = _parse_nhe_projections_table03()
    print(f"  Table 03 (Source of Funds): {len(rows03)} data points")
    for r in rows03:
        r["table_name"] = "Table 03"
        r.setdefault("metric_type", "amount")
    all_rows.extend(rows03)

    rows01 = _parse_nhe_projections_table01()
    print(f"  Table 01 (NHE + Economic Indicators): {len(rows01)} data points")
    for r in rows01:
        r["table_name"] = "Table 01"
    all_rows.extend(rows01)

    if not all_rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_rows)
    con.register("_nhep_df", df)

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_nhe_proj AS
        SELECT
            category,
            CAST(year AS INTEGER) AS year,
            CAST(amount_billions AS DOUBLE) AS value,
            data_type,
            table_name,
            metric_type,
            'cms.gov/nhe-projections' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _nhep_df
        ORDER BY table_name, category, year
    """)

    count = write_parquet(con, "_fact_nhe_proj", _snapshot_path("nhe_projections"), dry_run)
    years_range = con.execute("SELECT MIN(year), MAX(year) FROM _fact_nhe_proj").fetchone()
    cats = con.execute("SELECT COUNT(DISTINCT category) FROM _fact_nhe_proj").fetchone()[0]
    print(f"  {count:,} rows, {cats} categories, years {years_range[0]}-{years_range[1]}")

    # Medicaid projection
    med = con.execute("""
        SELECT year, value FROM _fact_nhe_proj
        WHERE category LIKE '%Medicaid%' AND data_type = 'projected'
        ORDER BY year DESC LIMIT 1
    """).fetchone()
    if med:
        print(f"  Latest Medicaid projection: ${med[1]:,.1f}B ({med[0]})")

    con.execute("DROP TABLE IF EXISTS _fact_nhe_proj")
    return count


# ---------------------------------------------------------------------------
# 5. PERM Improper Payment Rates
# ---------------------------------------------------------------------------

def build_fact_perm_rates(con, dry_run: bool) -> int:
    print("Building fact_perm_rates...")

    # Hard-coded from CMS PERM findings page (2020-2025)
    # These are national rolling 3-year rates
    perm_data = [
        # Medicaid
        {"program": "Medicaid", "year": 2025, "overall_rate": 6.12, "ffs_rate": 4.60, "mc_rate": 0.00, "eligibility_rate": 4.42, "estimated_improper_payments_billions": 37.39},
        {"program": "Medicaid", "year": 2024, "overall_rate": 5.09, "ffs_rate": 4.83, "mc_rate": 0.00, "eligibility_rate": 3.31, "estimated_improper_payments_billions": 31.10},
        {"program": "Medicaid", "year": 2023, "overall_rate": 8.58, "ffs_rate": 6.90, "mc_rate": 0.00, "eligibility_rate": 5.95, "estimated_improper_payments_billions": 50.33},
        {"program": "Medicaid", "year": 2022, "overall_rate": 15.62, "ffs_rate": 10.42, "mc_rate": 0.03, "eligibility_rate": 11.89, "estimated_improper_payments_billions": 80.57},
        {"program": "Medicaid", "year": 2021, "overall_rate": 21.69, "ffs_rate": 13.90, "mc_rate": 0.04, "eligibility_rate": 16.62, "estimated_improper_payments_billions": 98.72},
        {"program": "Medicaid", "year": 2020, "overall_rate": 21.36, "ffs_rate": 16.84, "mc_rate": 0.06, "eligibility_rate": 14.94, "estimated_improper_payments_billions": 86.49},
        # CHIP
        {"program": "CHIP", "year": 2025, "overall_rate": 7.05, "ffs_rate": 4.65, "mc_rate": 0.94, "eligibility_rate": 5.23, "estimated_improper_payments_billions": 1.37},
        {"program": "CHIP", "year": 2024, "overall_rate": 6.11, "ffs_rate": 4.72, "mc_rate": 0.72, "eligibility_rate": 4.44, "estimated_improper_payments_billions": 1.07},
        {"program": "CHIP", "year": 2023, "overall_rate": 12.81, "ffs_rate": 7.09, "mc_rate": 0.59, "eligibility_rate": 10.86, "estimated_improper_payments_billions": 2.14},
        {"program": "CHIP", "year": 2022, "overall_rate": 26.75, "ffs_rate": 11.23, "mc_rate": 0.62, "eligibility_rate": 24.01, "estimated_improper_payments_billions": 4.30},
        {"program": "CHIP", "year": 2021, "overall_rate": 31.84, "ffs_rate": 13.67, "mc_rate": 0.48, "eligibility_rate": 28.71, "estimated_improper_payments_billions": 5.37},
        {"program": "CHIP", "year": 2020, "overall_rate": 27.00, "ffs_rate": 14.15, "mc_rate": 0.49, "eligibility_rate": 23.53, "estimated_improper_payments_billions": 4.78},
    ]

    import pandas as pd
    df = pd.DataFrame(perm_data)
    con.register("_perm_df", df)

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_perm AS
        SELECT
            program,
            CAST(year AS INTEGER) AS year,
            CAST(overall_rate AS DOUBLE) AS overall_rate_pct,
            CAST(ffs_rate AS DOUBLE) AS ffs_rate_pct,
            CAST(mc_rate AS DOUBLE) AS mc_rate_pct,
            CAST(eligibility_rate AS DOUBLE) AS eligibility_rate_pct,
            CAST(estimated_improper_payments_billions AS DOUBLE) AS estimated_improper_payments_billions,
            'cms.gov/perm' AS source,
            'National rolling 3-year rate' AS rate_methodology,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _perm_df
        ORDER BY program, year DESC
    """)

    count = write_parquet(con, "_fact_perm", _snapshot_path("perm_rates"), dry_run)

    latest_medicaid = con.execute("""
        SELECT year, overall_rate_pct, estimated_improper_payments_billions
        FROM _fact_perm WHERE program = 'Medicaid' ORDER BY year DESC LIMIT 1
    """).fetchone()
    print(f"  {count:,} rows (Medicaid + CHIP, 2020-2025)")
    if latest_medicaid:
        print(f"  Latest Medicaid: {latest_medicaid[1]}% = ${latest_medicaid[2]:.1f}B ({latest_medicaid[0]})")

    con.execute("DROP TABLE IF EXISTS _fact_perm")
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

ALL_TABLES = {
    "fact_federal_register_cms": build_fact_federal_register_cms,
    "fact_mfcu_stats": build_fact_mfcu_stats,
    "fact_nhe": build_fact_nhe,
    "fact_nhe_projections": build_fact_nhe_projections,
    "fact_perm_rates": build_fact_perm_rates,
}


def main():
    parser = argparse.ArgumentParser(description="Ingest federal regulatory/policy data")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only", type=str, default=None)
    args = parser.parse_args()

    tables = list(ALL_TABLES.keys())
    if args.only:
        tables = [t.strip() for t in args.only.split(",")]

    print(f"Snapshot: {SNAPSHOT_DATE}")
    print(f"Run ID:   {RUN_ID}")
    print(f"Building: {', '.join(tables)}")
    print()

    con = duckdb.connect()
    totals = {}
    for name in tables:
        if name not in ALL_TABLES:
            print(f"Unknown table: {name}")
            continue
        totals[name] = ALL_TABLES[name](con, args.dry_run)
        print()

    con.close()

    print("=" * 60)
    print("FEDERAL DATA LAKE INGESTION COMPLETE")
    print("=" * 60)
    total_rows = sum(totals.values())
    for name, count in totals.items():
        status = "written" if not args.dry_run else "dry-run"
        print(f"  {name:35s} {count:>12,} rows  [{status}]")
    print(f"  {'TOTAL':35s} {total_rows:>12,} rows")

    if not args.dry_run and total_rows > 0:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "snapshot_date": SNAPSHOT_DATE,
            "pipeline_run_id": RUN_ID,
            "created_at": datetime.now().isoformat() + "Z",
            "tables": {name: {"rows": count} for name, count in totals.items()},
            "total_rows": total_rows,
        }
        manifest_file = META_DIR / f"manifest_federal_data_{SNAPSHOT_DATE}.json"
        with open(manifest_file, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"\n  Manifest: {manifest_file}")

    print("\nDone.")


if __name__ == "__main__":
    main()
