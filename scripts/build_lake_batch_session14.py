#!/usr/bin/env python3
"""
build_lake_batch_session14.py — Batch ingest multiple datasets.

Datasets:
  1. Medicaid Opioid Prescribing Rates (geographic, 539K rows)
  2. Medicaid Spending by Drug (CMS, 2019-2023 wide→long)
  3. MH/SUD Services 2020-2022 (Medicaid behavioral health utilization)
  4. Managed Care Annual (updated Oct 2025)
  5. Managed Care Monthly (updated Oct 2025)
  6. MA/PACE Directory (921 plans)
  7. Presumptive Eligibility (state program features)
  8. IPPS Impact File FY2025 (hospital payment parameters)
  9. VIII Group Expenditures Q3 FY2025 (ACA expansion spending)

Usage:
  python3 scripts/build_lake_batch_session14.py
"""

import json
import uuid
from datetime import date, datetime
from pathlib import Path

import duckdb

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

results = {}


def write_parquet(con: duckdb.DuckDBPyConnection, table: str, path: Path) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    con.execute(
        f"COPY {table} TO '{path}' (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 500000)"
    )
    count = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    size_mb = path.stat().st_size / 1_048_576
    print(f"  -> {path.relative_to(PROJECT_ROOT)} ({count:,} rows, {size_mb:.1f} MB)")
    return count


# ── 1. Medicaid Opioid Prescribing Rates ────────────────────────────
def build_opioid_prescribing_rates():
    csv_path = RAW_DIR / "medicaid_opioid_prescribing_rates.csv"
    if not csv_path.exists():
        print("  Opioid prescribing rates CSV not found, skipping")
        return 0

    print("\n--- Medicaid Opioid Prescribing Rates (Geographic) ---")
    con = duckdb.connect()

    con.execute(f"""
        CREATE TABLE fact_opioid_prescribing_geo AS
        SELECT
            CAST(Year AS INTEGER) AS year,
            TRIM(Geo_Lvl) AS geo_level,
            TRIM(Geo_Cd) AS geo_code,
            TRIM(Geo_Desc) AS geo_description,
            TRIM(Plan_Type) AS plan_type,
            TRY_CAST(Tot_Opioid_Clms AS BIGINT) AS total_opioid_claims,
            TRY_CAST(Tot_Clms AS BIGINT) AS total_claims,
            TRY_CAST(Opioid_Prscrbng_Rate AS DOUBLE) AS opioid_prescribing_rate,
            TRY_CAST(Opioid_Prscrbng_Rate_5Y_Chg AS DOUBLE) AS rate_5yr_change,
            TRY_CAST(Opioid_Prscrbng_Rate_1Y_Chg AS DOUBLE) AS rate_1yr_change,
            TRY_CAST(LA_Tot_Opioid_Clms AS BIGINT) AS la_total_opioid_claims,
            TRY_CAST(LA_Opioid_Prscrbng_Rate AS DOUBLE) AS la_opioid_prescribing_rate,
            TRY_CAST(LA_Opioid_Prscrbng_Rate_5Y_Chg AS DOUBLE) AS la_rate_5yr_change,
            TRY_CAST(LA_Opioid_Prscrbng_Rate_1Y_Chg AS DOUBLE) AS la_rate_1yr_change,
            'https://data.cms.gov/summary-statistics-on-use-and-payments/medicare-medicaid-opioid-prescribing-rates/medicaid-opioid-prescribing-rates-by-geography' AS source,
            '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', ignore_errors=true)
        WHERE Year IS NOT NULL
    """)

    count = con.execute("SELECT COUNT(*) FROM fact_opioid_prescribing_geo").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT geo_code) FROM fact_opioid_prescribing_geo WHERE geo_level = 'State'").fetchone()[0]
    years = con.execute("SELECT DISTINCT year FROM fact_opioid_prescribing_geo ORDER BY year").fetchall()
    print(f"  {count:,} rows, {states} states, years: {[y[0] for y in years]}")

    out_path = FACT_DIR / "opioid_prescribing_geo" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    row_count = write_parquet(con, "fact_opioid_prescribing_geo", out_path)
    con.close()
    return row_count


# ── 2. Medicaid Spending by Drug (wide → long) ──────────────────────
def build_spending_by_drug():
    csv_path = RAW_DIR / "medicaid_spending_by_drug.csv"
    if not csv_path.exists():
        print("  Spending by drug CSV not found, skipping")
        return 0

    print("\n--- Medicaid Spending by Drug (2019-2023) ---")
    con = duckdb.connect()

    # Wide format: columns have year suffixes. Unpivot to long.
    years = [2019, 2020, 2021, 2022, 2023]
    union_parts = []
    for yr in years:
        union_parts.append(f"""
            SELECT
                TRIM(Brnd_Name) AS brand_name,
                TRIM(Gnrc_Name) AS generic_name,
                TRY_CAST(Tot_Mftr AS INTEGER) AS total_manufacturers,
                TRIM(Mftr_Name) AS manufacturer_name,
                {yr} AS year,
                TRY_CAST(Tot_Spndng_{yr} AS DOUBLE) AS total_spending,
                TRY_CAST(Tot_Dsg_Unts_{yr} AS DOUBLE) AS total_dosage_units,
                TRY_CAST(Tot_Clms_{yr} AS BIGINT) AS total_claims,
                TRY_CAST(Avg_Spnd_Per_Dsg_Unt_Wghtd_{yr} AS DOUBLE) AS avg_spend_per_dosage_unit,
                TRY_CAST(Avg_Spnd_Per_Clm_{yr} AS DOUBLE) AS avg_spend_per_claim,
                TRY_CAST(Outlier_Flag_{yr} AS INTEGER) AS outlier_flag
            FROM read_csv_auto('{csv_path}', all_varchar=true, ignore_errors=true)
            WHERE Brnd_Name IS NOT NULL AND TRIM(Brnd_Name) != ''
              AND Tot_Spndng_{yr} IS NOT NULL AND TRIM(Tot_Spndng_{yr}) != ''
        """)

    union_sql = "\nUNION ALL\n".join(union_parts)
    con.execute(f"""
        CREATE TABLE fact_drug_spending_trend AS
        SELECT *,
            'https://data.cms.gov/summary-statistics-on-use-and-payments/medicare-medicaid-spending-by-drug' AS source,
            '{SNAPSHOT_DATE}' AS snapshot_date
        FROM ({union_sql})
        WHERE total_spending IS NOT NULL AND total_spending > 0
    """)

    count = con.execute("SELECT COUNT(*) FROM fact_drug_spending_trend").fetchone()[0]
    drugs = con.execute("SELECT COUNT(DISTINCT generic_name) FROM fact_drug_spending_trend").fetchone()[0]
    total = con.execute("SELECT ROUND(SUM(total_spending)/1e9, 2) FROM fact_drug_spending_trend").fetchone()[0]
    print(f"  {count:,} rows, {drugs:,} unique drugs, ${total}B total spending")

    # Top drugs by spending
    top = con.execute("""
        SELECT generic_name, ROUND(SUM(total_spending)/1e9, 2) as spend_B
        FROM fact_drug_spending_trend WHERE manufacturer_name = 'Overall'
        GROUP BY generic_name ORDER BY spend_B DESC LIMIT 5
    """).fetchall()
    print("  Top drugs by spending:")
    for row in top:
        print(f"    {row[0]}: ${row[1]}B")

    out_path = FACT_DIR / "drug_spending_trend" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    row_count = write_parquet(con, "fact_drug_spending_trend", out_path)
    con.close()
    return row_count


# ── 3. MH/SUD Services 2020-2022 ────────────────────────────────────
def build_mh_sud_services():
    csv_path = RAW_DIR / "medicaid_mh_sud_services_2020_2022.csv"
    if not csv_path.exists():
        print("  MH/SUD services CSV not found, skipping")
        return 0

    print("\n--- MH/SUD Services 2020-2022 ---")
    con = duckdb.connect()

    con.execute(f"""
        CREATE TABLE fact_mh_sud_services AS
        SELECT
            CAST(Year AS INTEGER) AS year,
            TRIM(Geography) AS geography,
            TRIM("Subpopulation topic") AS subpopulation_topic,
            TRIM(Subpopulation) AS subpopulation,
            TRIM(Category) AS category,
            TRY_CAST(REPLACE("Count of enrollees", ',', '') AS BIGINT) AS enrollee_count,
            TRY_CAST(REPLACE("Denominator count of enrollees", ',', '') AS BIGINT) AS denominator_count,
            TRY_CAST("Percentage of enrollees" AS DOUBLE) AS pct_enrollees,
            TRIM("Data version") AS data_version,
            'https://www.medicaid.gov/medicaid/benefits/behavioral-health-services' AS source,
            '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', ignore_errors=true)
        WHERE Year IS NOT NULL
    """)

    count = con.execute("SELECT COUNT(*) FROM fact_mh_sud_services").fetchone()[0]
    cats = con.execute("SELECT DISTINCT category FROM fact_mh_sud_services ORDER BY category").fetchall()
    print(f"  {count:,} rows")
    print(f"  Categories: {[c[0] for c in cats]}")

    out_path = FACT_DIR / "mh_sud_services" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    row_count = write_parquet(con, "fact_mh_sud_services", out_path)
    con.close()
    return row_count


# ── 4 & 5. Managed Care Annual + Monthly (Oct 2025 update) ──────────
def build_mc_enrollment_updated():
    annual_path = RAW_DIR / "managed_care_annual_oct2025.csv"
    monthly_path = RAW_DIR / "managed_care_monthly_oct2025.csv"

    total_rows = 0

    if annual_path.exists():
        print("\n--- Managed Care Annual (Oct 2025 update) ---")
        con = duckdb.connect()
        con.execute(f"""
            CREATE TABLE fact_mc_enrollment_annual AS
            SELECT
                TRIM(State) AS state_name,
                CAST(Year AS INTEGER) AS year,
                TRIM("managedcare participation") AS mc_type,
                TRY_CAST(REPLACE(CountEverEnrolled, ',', '') AS BIGINT) AS ever_enrolled,
                TRY_CAST(REPLACE(CountLastMonthEnrollment, ',', '') AS BIGINT) AS last_month_enrollment,
                TRY_CAST(REPLACE(AverageEnrollmentPerMonth, ',', '') AS BIGINT) AS avg_monthly_enrollment,
                TRIM(dunusable) AS dq_status,
                'https://www.medicaid.gov/medicaid/managed-care/enrollment-report' AS source,
                '{SNAPSHOT_DATE}' AS snapshot_date
            FROM read_csv_auto('{annual_path}', all_varchar=true, ignore_errors=true)
            WHERE State IS NOT NULL AND TRIM(State) != ''
        """)

        count = con.execute("SELECT COUNT(*) FROM fact_mc_enrollment_annual").fetchone()[0]
        states = con.execute("SELECT COUNT(DISTINCT state_name) FROM fact_mc_enrollment_annual").fetchone()[0]
        years = con.execute("SELECT MIN(year), MAX(year) FROM fact_mc_enrollment_annual").fetchone()
        print(f"  {count:,} rows, {states} states, {years[0]}-{years[1]}")

        out_path = FACT_DIR / "mc_enrollment_annual" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
        annual_rows = write_parquet(con, "fact_mc_enrollment_annual", out_path)
        con.close()
        total_rows += annual_rows

    if monthly_path.exists():
        print("\n--- Managed Care Monthly (Oct 2025 update) ---")
        con = duckdb.connect()
        con.execute(f"""
            CREATE TABLE fact_mc_enrollment_monthly AS
            SELECT
                TRIM(State) AS state_name,
                TRIM(Month) AS month_str,
                TRY_CAST(LEFT(TRIM(Month), 4) AS INTEGER) AS year,
                TRY_CAST(RIGHT(TRIM(Month), 2) AS INTEGER) AS month,
                TRIM("managedcare participation") AS mc_type,
                TRY_CAST(REPLACE(CountEnrolled, ',', '') AS BIGINT) AS enrolled,
                TRIM(dunusable) AS dq_status,
                'https://www.medicaid.gov/medicaid/managed-care/enrollment-report' AS source,
                '{SNAPSHOT_DATE}' AS snapshot_date
            FROM read_csv_auto('{monthly_path}', all_varchar=true, ignore_errors=true)
            WHERE State IS NOT NULL AND TRIM(State) != ''
        """)

        count = con.execute("SELECT COUNT(*) FROM fact_mc_enrollment_monthly").fetchone()[0]
        states = con.execute("SELECT COUNT(DISTINCT state_name) FROM fact_mc_enrollment_monthly").fetchone()[0]
        months = con.execute("SELECT MIN(month_str), MAX(month_str) FROM fact_mc_enrollment_monthly").fetchone()
        print(f"  {count:,} rows, {states} states, {months[0]}-{months[1]}")

        out_path = FACT_DIR / "mc_enrollment_monthly" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
        monthly_rows = write_parquet(con, "fact_mc_enrollment_monthly", out_path)
        con.close()
        total_rows += monthly_rows

    return total_rows


# ── 6. MA/PACE Directory ────────────────────────────────────────────
def build_pace_directory():
    csv_path = RAW_DIR / "ma_pace_directory" / "MA_Plan_Directory_2026_02" / "MA_Contract_directory_2026_02.csv"
    if not csv_path.exists():
        print("  PACE directory CSV not found, skipping")
        return 0

    print("\n--- MA/PACE Directory (Feb 2026) ---")
    con = duckdb.connect()

    con.execute(f"""
        CREATE TABLE fact_pace_directory AS
        SELECT
            TRIM("Legal Entity Name") AS legal_entity_name,
            TRIM("Organization Marketing Name") AS organization_name,
            TRIM("Contract Number") AS contract_number,
            TRIM("Organization Type") AS organization_type,
            TRIM("Plan Type") AS plan_type,
            CAST("Contract Effective Date" AS VARCHAR) AS contract_effective_date,
            TRIM("Tax Status") AS tax_status,
            TRIM("Parent Organization") AS parent_organization,
            TRIM("CMS Region Responsible") AS cms_region,
            TRY_CAST(Enrollment AS INTEGER) AS enrollment,
            TRIM("Legal Entity City") AS city,
            TRIM("Legal Entity State Code") AS state_code,
            TRIM("Legal Entity Zip Code") AS zip_code,
            'https://www.cms.gov/research-statistics-data-and-systems/statistics-trends-and-reports/mcradvpartdenroldata/ma-contract-directory' AS source,
            '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', ignore_errors=true)
        WHERE "Contract Number" IS NOT NULL
    """)

    count = con.execute("SELECT COUNT(*) FROM fact_pace_directory").fetchone()[0]
    pace = con.execute("SELECT COUNT(*) FROM fact_pace_directory WHERE plan_type = 'National PACE'").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM fact_pace_directory").fetchone()[0]
    total_enroll = con.execute("SELECT SUM(enrollment) FROM fact_pace_directory WHERE enrollment IS NOT NULL").fetchone()[0]
    print(f"  {count:,} plans ({pace} PACE), {states} states, {total_enroll:,} total enrollment")

    out_path = FACT_DIR / "pace_directory" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    row_count = write_parquet(con, "fact_pace_directory", out_path)
    con.close()
    return row_count


# ── 7. Presumptive Eligibility ──────────────────────────────────────
def build_presumptive_eligibility():
    csv_path = RAW_DIR / "medicaid_presumptive_eligibility.csv"
    if not csv_path.exists():
        print("  Presumptive eligibility CSV not found, skipping")
        return 0

    print("\n--- Presumptive Eligibility (State Features) ---")
    con = duckdb.connect()

    con.execute(f"""
        CREATE TABLE fact_presumptive_eligibility AS
        SELECT
            TRIM(State) AS state_name,
            CASE WHEN CAST(COALESCE(CHIP, '') AS VARCHAR) IN ('true', 'TRUE', 'True') THEN true ELSE false END AS chip_presumptive,
            CASE WHEN CAST(COALESCE(Medicaid, '') AS VARCHAR) IN ('true', 'TRUE', 'True') THEN true ELSE false END AS medicaid_presumptive,
            'https://www.medicaid.gov/medicaid/eligibility' AS source,
            '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', ignore_errors=true)
        WHERE State IS NOT NULL AND TRIM(State) != ''
    """)

    count = con.execute("SELECT COUNT(*) FROM fact_presumptive_eligibility").fetchone()[0]
    chip = con.execute("SELECT COUNT(*) FROM fact_presumptive_eligibility WHERE chip_presumptive").fetchone()[0]
    mcaid = con.execute("SELECT COUNT(*) FROM fact_presumptive_eligibility WHERE medicaid_presumptive").fetchone()[0]
    print(f"  {count:,} states, {mcaid} with Medicaid PE, {chip} with CHIP PE")

    out_path = FACT_DIR / "presumptive_eligibility" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    row_count = write_parquet(con, "fact_presumptive_eligibility", out_path)
    con.close()
    return row_count


# ── 8. IPPS Impact File FY2025 ──────────────────────────────────────
def build_ipps_impact():
    txt_path = RAW_DIR / "ipps_impact" / "FY 2025 IPPS Impact File - Final Rule.txt"
    if not txt_path.exists():
        print("  IPPS impact file not found, skipping")
        return 0

    print("\n--- IPPS Impact File FY2025 (Final Rule) ---")
    con = duckdb.connect()

    # Tab-delimited, first row is title, second row is headers
    con.execute(f"""
        CREATE TABLE fact_ipps_impact AS
        SELECT
            TRIM("Provider Number") AS provider_number,
            TRIM(Name) AS hospital_name,
            TRIM("Geographic Labor Market Area") AS geo_labor_market_area,
            TRIM("Payment Labor Market Area") AS payment_labor_market_area,
            TRIM("FIPS County Code") AS fips_county_code,
            TRY_CAST(Region AS INTEGER) AS region,
            TRIM(URGEO) AS urban_rural_geo,
            TRIM(URSPA) AS urban_rural_spa,
            TRIM(RECLASS) AS reclassification,
            TRY_CAST("FY 2025 Wage Index" AS DOUBLE) AS wage_index_fy2025,
            TRY_CAST(LUGAR AS VARCHAR) AS lugar,
            2025 AS fiscal_year,
            'https://www.cms.gov/medicare/payment/prospective-payment-systems/acute-inpatient-pps/fy-2025-ipps-final-rule-home-page' AS source,
            '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{txt_path}', delim='\t', skip=1, header=true,
             all_varchar=true, ignore_errors=true, sample_size=5000)
        WHERE "Provider Number" IS NOT NULL AND TRIM("Provider Number") != ''
    """)

    count = con.execute("SELECT COUNT(*) FROM fact_ipps_impact").fetchone()[0]
    avg_wi = con.execute("SELECT ROUND(AVG(wage_index_fy2025), 4) FROM fact_ipps_impact WHERE wage_index_fy2025 IS NOT NULL").fetchone()[0]
    print(f"  {count:,} hospitals, avg wage index: {avg_wi}")

    out_path = FACT_DIR / "ipps_impact" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    row_count = write_parquet(con, "fact_ipps_impact", out_path)
    con.close()
    return row_count


# ── 9. VIII Group Expenditures (ACA Expansion) ──────────────────────
def build_viii_group_expenditures():
    xlsx_path = RAW_DIR / "viii_group_expenditures_q3_2025.xlsx"
    if not xlsx_path.exists():
        print("  VIII group expenditures file not found, skipping")
        return 0

    print("\n--- VIII Group Expenditures Q3 FY2025 ---")

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        if len(data) < 2:
            continue

        # Find header row
        header_row = None
        for i, row in enumerate(data):
            row_str = str(row).lower()
            if 'state' in row_str and ('expenditure' in row_str or 'enrollment' in row_str or 'total' in row_str):
                header_row = i
                break
        if header_row is None:
            # Try first row with content
            for i, row in enumerate(data):
                if row[0] and str(row[0]).strip():
                    header_row = i
                    break
        if header_row is None:
            continue

        headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(data[header_row])]
        print(f"  Sheet '{sheet_name}': {len(headers)} cols, headers at row {header_row}")

        for row in data[header_row + 1:]:
            if not row[0] or not str(row[0]).strip():
                continue
            state = str(row[0]).strip()
            if state.lower() in ('total', 'totals', 'national', ''):
                continue
            row_dict = {"state_name": state, "sheet": sheet_name}
            for j, val in enumerate(row[1:], 1):
                if j < len(headers):
                    col = headers[j]
                    if val is not None:
                        try:
                            row_dict[col] = float(val)
                        except (ValueError, TypeError):
                            row_dict[col] = str(val)
            rows.append(row_dict)

    if not rows:
        print("  No data found in workbook")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    df["source"] = "https://www.medicaid.gov/medicaid/financial-management"
    df["snapshot_date"] = SNAPSHOT_DATE

    con = duckdb.connect()
    con.execute("CREATE TABLE fact_viii_group_expenditure AS SELECT * FROM df")
    count = con.execute("SELECT COUNT(*) FROM fact_viii_group_expenditure").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state_name) FROM fact_viii_group_expenditure").fetchone()[0]
    print(f"  {count:,} rows, {states} states")

    out_path = FACT_DIR / "viii_group_expenditure" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    row_count = write_parquet(con, "fact_viii_group_expenditure", out_path)
    con.close()
    return row_count


def write_manifest():
    META_DIR.mkdir(parents=True, exist_ok=True)
    manifest = {
        "run_id": RUN_ID,
        "snapshot_date": SNAPSHOT_DATE,
        "script": "build_lake_batch_session14.py",
        "tables": results,
        "completed_at": datetime.now().isoformat() + "Z",
    }
    manifest_path = META_DIR / f"manifest_session14_{SNAPSHOT_DATE}.json"
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"\n  Manifest: {manifest_path}")


def main():
    print("=" * 60)
    print("Session 14 Batch Ingestion")
    print(f"  Snapshot: {SNAPSHOT_DATE}")
    print(f"  Run ID:   {RUN_ID}")

    builders = [
        ("opioid_prescribing_geo", build_opioid_prescribing_rates),
        ("drug_spending_trend", build_spending_by_drug),
        ("mh_sud_services", build_mh_sud_services),
        ("mc_enrollment_annual+monthly", build_mc_enrollment_updated),
        ("pace_directory", build_pace_directory),
        ("presumptive_eligibility", build_presumptive_eligibility),
        ("ipps_impact", build_ipps_impact),
        ("viii_group_expenditure", build_viii_group_expenditures),
    ]

    total = 0
    for name, builder in builders:
        try:
            rows = builder()
            results[name] = {"rows": rows}
            total += rows
        except Exception as e:
            print(f"\n  ERROR in {name}: {e}")
            import traceback
            traceback.print_exc()
            results[name] = {"rows": 0, "error": str(e)}

    write_manifest()

    print("\n" + "=" * 60)
    print("SESSION 14 BATCH INGESTION COMPLETE")
    for name, info in results.items():
        rows = info.get("rows", 0)
        err = info.get("error", "")
        status = f"ERROR: {err}" if err else f"{rows:,} rows"
        print(f"  {name}: {status}")
    print(f"  TOTAL: {total:,} rows")
    print("=" * 60)


if __name__ == "__main__":
    main()
