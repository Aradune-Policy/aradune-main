"""
Round 11b — Ingest CMS Financial Management Report (FMR) FY 2024 and
VIII Group (New Adult) Expenditures Q3 2025.

Tables produced:
  fact_fmr_fy2024          — FY 2024 Medicaid net expenditures by state and service category
  fact_new_adult_spending  — New Adult Group (VIII Group) expenditures by state and quarter
"""

import duckdb
import json
import openpyxl
from pathlib import Path
from datetime import date

LAKE = Path(__file__).resolve().parent.parent / "data" / "lake"
RAW  = Path(__file__).resolve().parent.parent / "data" / "raw"
SNAP = str(date.today())

STATE_NAME_TO_CODE = {
    'Alabama': 'AL', 'Alaska': 'AK', 'Arizona': 'AZ', 'Arkansas': 'AR',
    'California': 'CA', 'Colorado': 'CO', 'Connecticut': 'CT', 'Delaware': 'DE',
    'Dist. Of Col.': 'DC', 'District of Columbia': 'DC',
    'Florida': 'FL', 'Georgia': 'GA', 'Hawaii': 'HI', 'Idaho': 'ID',
    'Illinois': 'IL', 'Indiana': 'IN', 'Iowa': 'IA', 'Kansas': 'KS',
    'Kentucky': 'KY', 'Louisiana': 'LA', 'Maine': 'ME', 'Maryland': 'MD',
    'Massachusetts': 'MA', 'Michigan': 'MI', 'Minnesota': 'MN', 'Mississippi': 'MS',
    'Missouri': 'MO', 'Montana': 'MT', 'Nebraska': 'NE', 'Nevada': 'NV',
    'New Hampshire': 'NH', 'New Jersey': 'NJ', 'New Mexico': 'NM', 'New York': 'NY',
    'North Carolina': 'NC', 'North Dakota': 'ND', 'Ohio': 'OH', 'Oklahoma': 'OK',
    'Oregon': 'OR', 'Pennsylvania': 'PA', 'Rhode Island': 'RI',
    'South Carolina': 'SC', 'South Dakota': 'SD', 'Tennessee': 'TN', 'Texas': 'TX',
    'Utah': 'UT', 'Vermont': 'VT', 'Virginia': 'VA', 'Washington': 'WA',
    'West Virginia': 'WV', 'Wisconsin': 'WI', 'Wyoming': 'WY',
    'Puerto Rico': 'PR', 'Amer. Samoa': 'AS', 'Guam': 'GU',
    'Virgin Islands': 'VI', 'N. Mariana Islands': 'MP',
}


def _write_parquet(con, table_name, fact_name):
    """Write table to Parquet in the lake."""
    out_dir = LAKE / "fact" / fact_name / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "data.parquet"
    con.execute(f"""
        COPY {table_name} TO '{out_path}'
        (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 100000)
    """)
    cnt = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    print(f"  ✓ {fact_name}: {cnt:,} rows → {out_path}")
    return cnt


def build_fmr_fy2024(con):
    """Ingest FY 2024 Financial Management Report — Medicaid net expenditures by service category."""
    print("\n── FMR FY 2024 ──")
    path = RAW / "fin_mgmt_fy2024" / "FY 2024 FMR NET EXPENDITURES.xlsx"
    if not path.exists():
        print("  File not found")
        return 0

    wb = openpyxl.load_workbook(path, read_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("MAP - "):
            continue
        state_name = sheet_name.replace("MAP - ", "").strip()
        if state_name == "National Totals":
            state_code = "US"
        else:
            state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            print(f"  Skipping: {state_name}")
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=7, values_only=True):
            category = row[0]
            if not category or not isinstance(category, str):
                continue
            category = category.strip()
            if category in ("Service Category", "") or category.startswith("Total"):
                # Capture total rows too
                if category.startswith("Total"):
                    pass  # include totals
                else:
                    continue

            def _num(v):
                if v is None:
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            total_computable = _num(row[1])
            federal_share = _num(row[2])
            federal_medicaid = _num(row[3])
            federal_arra = _num(row[4])
            federal_covid = _num(row[5])
            state_share = _num(row[6]) if len(row) > 6 else None

            # Skip rows where everything is None/0
            if total_computable is None and federal_share is None:
                continue

            rows.append({
                'state_code': state_code,
                'fiscal_year': 2024,
                'report_type': 'MAP',
                'service_category': category,
                'total_computable': total_computable,
                'federal_share': federal_share,
                'federal_medicaid': federal_medicaid,
                'federal_arra': federal_arra,
                'federal_covid': federal_covid,
                'state_share': state_share,
            })

    # Also read ADM (Administration) sheets
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("ADM - "):
            continue
        state_name = sheet_name.replace("ADM - ", "").strip()
        if state_name == "National Totals":
            state_code = "US"
        else:
            state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=7, values_only=True):
            category = row[0]
            if not category or not isinstance(category, str):
                continue
            category = category.strip()
            if category in ("Administration Category", ""):
                continue

            def _num(v):
                if v is None:
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            total_computable = _num(row[1])
            federal_share = _num(row[2])
            if total_computable is None and federal_share is None:
                continue

            rows.append({
                'state_code': state_code,
                'fiscal_year': 2024,
                'report_type': 'ADM',
                'service_category': category,
                'total_computable': total_computable,
                'federal_share': federal_share,
                'federal_medicaid': _num(row[3]) if len(row) > 3 else None,
                'federal_arra': _num(row[4]) if len(row) > 4 else None,
                'federal_covid': _num(row[5]) if len(row) > 5 else None,
                'state_share': _num(row[6]) if len(row) > 6 else None,
            })

    wb.close()

    if not rows:
        print("  No data found")
        return 0

    print(f"  Parsed {len(rows)} rows from {len(set(r['state_code'] for r in rows))} states")

    con.execute("DROP TABLE IF EXISTS _fmr")
    con.execute("""
        CREATE TABLE _fmr (
            state_code VARCHAR, fiscal_year INTEGER, report_type VARCHAR,
            service_category VARCHAR, total_computable DOUBLE,
            federal_share DOUBLE, federal_medicaid DOUBLE,
            federal_arra DOUBLE, federal_covid DOUBLE,
            state_share DOUBLE
        )
    """)
    con.executemany(
        "INSERT INTO _fmr VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [(r['state_code'], r['fiscal_year'], r['report_type'],
          r['service_category'], r['total_computable'],
          r['federal_share'], r['federal_medicaid'],
          r['federal_arra'], r['federal_covid'],
          r['state_share']) for r in rows]
    )
    return _write_parquet(con, '_fmr', 'fmr_fy2024')


def build_new_adult_spending(con):
    """Ingest New Adult (VIII Group) quarterly expenditures."""
    print("\n── VIII Group (New Adult) Expenditures ──")
    path = RAW / "viii_group_expenditures_q3_2025.xlsx"
    if not path.exists():
        print("  File not found")
        return 0

    wb = openpyxl.load_workbook(path, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]

    # Read header to understand columns
    rows_data = list(ws.iter_rows(values_only=True))
    # Find header row
    header_idx = 0
    for i, row in enumerate(rows_data):
        if row[0] and 'State' in str(row[0]):
            header_idx = i
            break

    headers = [str(h).strip() if h else f'col_{j}' for j, h in enumerate(rows_data[header_idx])]
    print(f"  Headers: {headers[:10]}")

    rows = []
    for row in rows_data[header_idx + 1:]:
        state = str(row[0]).strip() if row[0] else None
        if not state or state == 'Total':
            continue
        state_code = STATE_NAME_TO_CODE.get(state)
        if not state_code:
            continue

        def _num(v):
            if v is None:
                return None
            try:
                return float(str(v).replace(',', '').replace('$', ''))
            except (ValueError, TypeError):
                return None

        row_data = {'state_code': state_code}
        for j, h in enumerate(headers[1:], 1):
            if j < len(row):
                row_data[h] = _num(row[j])
        rows.append(row_data)

    wb.close()

    if not rows:
        print("  No data found")
        return 0

    print(f"  Parsed {len(rows)} states")

    # Create a simple table with key columns
    con.execute("DROP TABLE IF EXISTS _new_adult")
    con.execute("""
        CREATE TABLE _new_adult (
            state_code VARCHAR,
            data_json VARCHAR
        )
    """)
    for r in rows:
        sc = r.pop('state_code')
        con.execute("INSERT INTO _new_adult VALUES (?, ?)", [sc, json.dumps(r)])

    return _write_parquet(con, '_new_adult', 'new_adult_spending')


def main():
    con = duckdb.connect()
    total = 0

    total += build_fmr_fy2024(con)
    total += build_new_adult_spending(con)

    con.close()

    print(f"\n{'='*60}")
    print(f"Round 11b complete: {total:,} total rows")

    # Write manifest
    manifest = {
        "pipeline_run": "round11b",
        "snapshot_date": SNAP,
        "total_rows": total,
        "tables": ["fact_fmr_fy2024", "fact_new_adult_spending"],
    }
    manifest_path = LAKE / "metadata" / f"manifest_round11b_{SNAP}.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"Manifest: {manifest_path}")


if __name__ == "__main__":
    main()
