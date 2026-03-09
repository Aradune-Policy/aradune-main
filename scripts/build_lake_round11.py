"""
Round 11 — Ingest LTSS expenditure/users, CDC vital stats, CDC maternal mortality, MACPAC HCBS payment scan.

Tables produced:
  fact_ltss_expenditure    — LTSS expenditure by state, delivery system, year (CY 2022-2023)
  fact_ltss_users          — LTSS users by state, delivery system, year (CY 2022-2023)
  fact_ltss_rebalancing    — HCBS rebalancing measures by state, demographic group (CY 2022-2023)
  fact_vital_stats_monthly — CDC VSRR monthly births, deaths, infant deaths by state
  fact_maternal_mortality_monthly — CDC provisional maternal mortality rolling 12-month
  ref_hcbs_payment_method  — MACPAC 1915(c) waiver payment methodology by state
"""

import duckdb
import json
import openpyxl
from pathlib import Path
from datetime import date

LAKE = Path(__file__).resolve().parent.parent / "data" / "lake"
RAW  = Path(__file__).resolve().parent.parent / "data" / "raw"
SNAP = str(date.today())

STATE_NAME_TO_CODE = {
    'Alabama': 'AL', 'Alaska': 'AK', 'Arizona': 'AZ', 'Arkansas': 'AR',
    'California': 'CA', 'Colorado': 'CO', 'Connecticut': 'CT', 'Delaware': 'DE',
    'District of Columbia': 'DC', 'Florida': 'FL', 'Georgia': 'GA', 'Hawaii': 'HI',
    'Idaho': 'ID', 'Illinois': 'IL', 'Indiana': 'IN', 'Iowa': 'IA',
    'Kansas': 'KS', 'Kentucky': 'KY', 'Louisiana': 'LA', 'Maine': 'ME',
    'Maryland': 'MD', 'Massachusetts': 'MA', 'Michigan': 'MI', 'Minnesota': 'MN',
    'Mississippi': 'MS', 'Missouri': 'MO', 'Montana': 'MT', 'Nebraska': 'NE',
    'Nevada': 'NV', 'New Hampshire': 'NH', 'New Jersey': 'NJ', 'New Mexico': 'NM',
    'New York': 'NY', 'North Carolina': 'NC', 'North Dakota': 'ND', 'Ohio': 'OH',
    'Oklahoma': 'OK', 'Oregon': 'OR', 'Pennsylvania': 'PA', 'Rhode Island': 'RI',
    'South Carolina': 'SC', 'South Dakota': 'SD', 'Tennessee': 'TN', 'Texas': 'TX',
    'Utah': 'UT', 'Vermont': 'VT', 'Virginia': 'VA', 'Washington': 'WA',
    'West Virginia': 'WV', 'Wisconsin': 'WI', 'Wyoming': 'WY',
    'Puerto Rico': 'PR', 'US Virgin Islands': 'VI', 'Guam': 'GU',
    'American Samoa': 'AS', 'Northern Mariana Islands': 'MP',
}

# Also map state names used in CDC VSRR (all caps)
STATE_UPPER_TO_CODE = {k.upper(): v for k, v in STATE_NAME_TO_CODE.items()}
STATE_UPPER_TO_CODE['UNITED STATES'] = 'US'


def _write_parquet(con, table_name, fact_name, category="fact"):
    """Write table to Parquet in the lake."""
    out_dir = LAKE / category / fact_name / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "data.parquet"
    con.execute(f"""
        COPY {table_name} TO '{out_path}'
        (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 100000)
    """)
    cnt = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    print(f"  ✓ {fact_name}: {cnt:,} rows → {out_path}")
    return cnt


def _num(v):
    """Parse a numeric value, returning None for suppressed/missing."""
    if v is None or v == 'DS' or v == 'blank cell' or v == '–' or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _find_ltss_file(year, prefix):
    """Find the LTSS Excel file for a given year, checking both directory structures."""
    for subdir in [f"ltss_{year}", "ltss_2019_2021"]:
        path = RAW / subdir / f"{prefix}_{year}.xlsx"
        if path.exists():
            return path
    return None


def _read_ltss_exp_sheet(wb, year):
    """Read expenditure data — handles both 2019-2021 and 2022+ formats."""
    rows = []
    # 2022+ has combined sheet 'A.2.1 All-Total' with all columns
    if 'A.2.1 All-Total' in wb.sheetnames:
        ws = wb['A.2.1 All-Total']
        header_row = 1
        # Check if row 0 is a title or the header
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if first_row[0] and 'State' not in str(first_row[0]):
            header_row = 2  # skip title row
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            state_name = str(row[0]).strip() if row[0] else None
            if not state_name:
                continue
            state_code = 'US' if state_name == 'National' else STATE_NAME_TO_CODE.get(state_name)
            if not state_code:
                continue
            rows.append({
                'state_code': state_code, 'year': year,
                'ltss_total': _num(row[1]),
                'institutional_total': _num(row[2]), 'institutional_pct': _num(row[3]),
                'hcbs_total': _num(row[4]), 'hcbs_pct': _num(row[5]),
            })
    else:
        # 2019-2021: separate inst and HCBS sheets, need to merge
        inst_data = {}
        hcbs_data = {}
        # Read institutional totals
        for sn in wb.sheetnames:
            if 'All-Inst' in sn or 'Inst' in sn:
                ws = wb[sn]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    state_name = str(row[0]).strip() if row[0] else None
                    if not state_name:
                        continue
                    state_code = 'US' if state_name == 'National' else STATE_NAME_TO_CODE.get(state_name)
                    if state_code:
                        inst_data[state_code] = _num(row[1])
                break
        # Read HCBS totals
        for sn in wb.sheetnames:
            if 'All-HCBS' in sn or 'HCBS' in sn:
                ws = wb[sn]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    state_name = str(row[0]).strip() if row[0] else None
                    if not state_name:
                        continue
                    state_code = 'US' if state_name == 'National' else STATE_NAME_TO_CODE.get(state_name)
                    if state_code:
                        hcbs_data[state_code] = _num(row[1])
                break
        # Combine
        all_states = set(inst_data.keys()) | set(hcbs_data.keys())
        for sc in sorted(all_states):
            inst = inst_data.get(sc)
            hcbs = hcbs_data.get(sc)
            total = None
            if inst is not None and hcbs is not None:
                total = inst + hcbs
            inst_pct = round(inst / total * 100, 2) if total and inst else None
            hcbs_pct = round(hcbs / total * 100, 2) if total and hcbs else None
            rows.append({
                'state_code': sc, 'year': year,
                'ltss_total': total,
                'institutional_total': inst, 'institutional_pct': inst_pct,
                'hcbs_total': hcbs, 'hcbs_pct': hcbs_pct,
            })
    return rows


def build_ltss_expenditure(con):
    """Ingest LTSS expenditure by state and delivery system for 2019-2023."""
    print("\n── LTSS Expenditure ──")
    rows = []
    for year in range(2019, 2024):
        path = _find_ltss_file(year, "A2_LTSSExpDlvrySystm")
        if not path:
            print(f"  Skipping {year} — file not found")
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        yr_rows = _read_ltss_exp_sheet(wb, year)
        rows.extend(yr_rows)
        print(f"  {year}: {len(yr_rows)} states")
        wb.close()

    if not rows:
        print("  No data found")
        return 0

    con.execute("DROP TABLE IF EXISTS _ltss_exp")
    con.execute("""
        CREATE TABLE _ltss_exp (
            state_code VARCHAR, year INTEGER,
            ltss_total DOUBLE, institutional_total DOUBLE, institutional_pct DOUBLE,
            hcbs_total DOUBLE, hcbs_pct DOUBLE
        )
    """)
    con.executemany(
        "INSERT INTO _ltss_exp VALUES (?, ?, ?, ?, ?, ?, ?)",
        [(r['state_code'], r['year'], r['ltss_total'], r['institutional_total'],
          r['institutional_pct'], r['hcbs_total'], r['hcbs_pct']) for r in rows]
    )
    return _write_parquet(con, '_ltss_exp', 'ltss_expenditure')


def build_ltss_users(con):
    """Ingest LTSS users by state and delivery system for 2019-2023."""
    print("\n── LTSS Users ──")
    rows = []
    for year in range(2019, 2024):
        path = _find_ltss_file(year, "A1_LTSSUsrDlvrySystm")
        if not path:
            print(f"  Skipping {year} — file not found")
            continue
        wb = openpyxl.load_workbook(path, read_only=True)

        # Try the combined sheet first (2022+)
        if 'A.1.1 All-Total' in wb.sheetnames:
            ws = wb['A.1.1 All-Total']
            for row in ws.iter_rows(min_row=3, values_only=True):
                state_name = str(row[0]).strip() if row[0] else None
                if not state_name:
                    continue
                state_code = 'US' if state_name == 'National' else STATE_NAME_TO_CODE.get(state_name)
                if not state_code:
                    continue
                rows.append({
                    'state_code': state_code, 'year': year,
                    'ltss_total': _num(row[1]),
                    'institutional_total': _num(row[2]), 'institutional_pct': _num(row[3]),
                    'hcbs_total': _num(row[4]), 'hcbs_pct': _num(row[5]),
                    'both_total': _num(row[6]) if len(row) > 6 else None,
                    'both_pct': _num(row[7]) if len(row) > 7 else None,
                })
        else:
            # 2019-2021: separate sheets — read inst and HCBS totals
            inst_data = {}
            hcbs_data = {}
            for sn in wb.sheetnames:
                if 'All-Inst' in sn:
                    ws = wb[sn]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        sn2 = str(row[0]).strip() if row[0] else None
                        if not sn2:
                            continue
                        sc = 'US' if sn2 == 'National' else STATE_NAME_TO_CODE.get(sn2)
                        if sc:
                            inst_data[sc] = _num(row[1])
                    break
            for sn in wb.sheetnames:
                if 'All-HCBS' in sn:
                    ws = wb[sn]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        sn2 = str(row[0]).strip() if row[0] else None
                        if not sn2:
                            continue
                        sc = 'US' if sn2 == 'National' else STATE_NAME_TO_CODE.get(sn2)
                        if sc:
                            hcbs_data[sc] = _num(row[1])
                    break
            for sc in sorted(set(inst_data.keys()) | set(hcbs_data.keys())):
                inst = inst_data.get(sc)
                hcbs = hcbs_data.get(sc)
                total = (inst or 0) + (hcbs or 0) if inst is not None or hcbs is not None else None
                rows.append({
                    'state_code': sc, 'year': year,
                    'ltss_total': total,
                    'institutional_total': inst,
                    'institutional_pct': round(inst / total * 100, 2) if total and inst else None,
                    'hcbs_total': hcbs,
                    'hcbs_pct': round(hcbs / total * 100, 2) if total and hcbs else None,
                    'both_total': None, 'both_pct': None,
                })

        print(f"  {year}: {sum(1 for r in rows if r['year'] == year)} states")
        wb.close()

    if not rows:
        print("  No data found")
        return 0

    con.execute("DROP TABLE IF EXISTS _ltss_users")
    con.execute("""
        CREATE TABLE _ltss_users (
            state_code VARCHAR, year INTEGER,
            ltss_total DOUBLE, institutional_total DOUBLE, institutional_pct DOUBLE,
            hcbs_total DOUBLE, hcbs_pct DOUBLE,
            both_total DOUBLE, both_pct DOUBLE
        )
    """)
    con.executemany(
        "INSERT INTO _ltss_users VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [(r['state_code'], r['year'], r['ltss_total'], r['institutional_total'],
          r['institutional_pct'], r['hcbs_total'], r['hcbs_pct'],
          r.get('both_total'), r.get('both_pct')) for r in rows]
    )
    return _write_parquet(con, '_ltss_users', 'ltss_users')


def build_ltss_rebalancing(con):
    """Ingest LTSS HCBS rebalancing measures by demographics for 2019-2023."""
    print("\n── LTSS Rebalancing ──")
    rows = []
    for year in range(2019, 2024):
        path = _find_ltss_file(year, "B2_LTSSExpRebalMeasr")
        if not path:
            print(f"  Skipping {year} — file not found")
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        # Find the BalanceByAge sheet
        age_sheet = None
        for sn in wb.sheetnames:
            if 'BalanceByAge' in sn or 'Balance' in sn:
                age_sheet = sn
                break
        if not age_sheet:
            wb.close()
            continue
        ws = wb[age_sheet]
        # Find header row
        start_row = 1
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            if row[0] and 'State' in str(row[0]):
                start_row = i + 1
                break
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            state_name = str(row[0]).strip() if row[0] else None
            if not state_name:
                continue
            state_code = 'US' if state_name == 'National' else STATE_NAME_TO_CODE.get(state_name)
            if not state_code:
                continue

            # Overall HCBS %
            rows.append({'state_code': state_code, 'year': year,
                         'group': 'overall', 'subgroup': 'all',
                         'hcbs_pct': _num(row[1])})
            # By age
            age_groups = ['0-20', '21-44', '45-64', '65+']
            for j, ag in enumerate(age_groups):
                if len(row) > 2 + j:
                    rows.append({'state_code': state_code, 'year': year,
                                 'group': 'age', 'subgroup': ag,
                                 'hcbs_pct': _num(row[2 + j])})
        print(f"  {year}: {sum(1 for r in rows if r['year'] == year)} entries")
        wb.close()

    if not rows:
        print("  No data found")
        return 0

    con.execute("DROP TABLE IF EXISTS _ltss_rebal")
    con.execute("""
        CREATE TABLE _ltss_rebal (
            state_code VARCHAR, year INTEGER,
            demographic_group VARCHAR, subgroup VARCHAR,
            hcbs_pct DOUBLE
        )
    """)
    con.executemany(
        "INSERT INTO _ltss_rebal VALUES (?, ?, ?, ?, ?)",
        [(r['state_code'], r['year'], r['group'], r['subgroup'], r['hcbs_pct']) for r in rows]
    )
    return _write_parquet(con, '_ltss_rebal', 'ltss_rebalancing')


def build_vital_stats_monthly(con):
    """Ingest CDC VSRR monthly births, deaths, infant deaths by state."""
    print("\n── CDC VSRR Vital Stats Monthly ──")
    path = RAW / "cdc_vsrr_births_deaths_infant.csv"
    if not path.exists():
        print("  File not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _vsrr")
    con.execute(f"""
        CREATE TABLE _vsrr AS
        SELECT
            CASE
                WHEN UPPER(TRIM(state)) = 'UNITED STATES' THEN 'US'
                ELSE NULL
            END AS state_code_raw,
            TRIM(state) AS state_name,
            TRY_CAST(year AS INTEGER) AS year,
            month AS month_name,
            period,
            indicator,
            TRY_CAST(REPLACE(CAST(data_value AS VARCHAR), ',', '') AS BIGINT) AS value
        FROM read_csv_auto('{path}', header=true)
        WHERE state IS NOT NULL
    """)

    # Map state names to codes
    # Build mapping from state names
    mapping_values = ", ".join(
        f"('{name}', '{code}')" for name, code in STATE_UPPER_TO_CODE.items()
    )
    con.execute(f"""
        CREATE TABLE _state_map AS
        SELECT * FROM (VALUES {mapping_values}) AS t(state_upper, state_code)
    """)

    con.execute("""
        CREATE TABLE _vsrr2 AS
        SELECT
            COALESCE(m.state_code, v.state_code_raw) AS state_code,
            v.year, v.month_name, v.period, v.indicator, v.value
        FROM _vsrr v
        LEFT JOIN _state_map m ON UPPER(TRIM(v.state_name)) = m.state_upper
        WHERE COALESCE(m.state_code, v.state_code_raw) IS NOT NULL
    """)

    con.execute("DROP TABLE IF EXISTS _vsrr")
    con.execute("DROP TABLE IF EXISTS _state_map")
    con.execute("ALTER TABLE _vsrr2 RENAME TO _vsrr")
    return _write_parquet(con, '_vsrr', 'vital_stats_monthly')


def build_maternal_mortality_monthly(con):
    """Ingest CDC provisional maternal mortality (12-month rolling rates)."""
    print("\n── CDC Maternal Mortality Monthly ──")
    path = RAW / "cdc_maternal_mortality_provisional.csv"
    if not path.exists():
        print("  File not found")
        return 0

    con.execute("DROP TABLE IF EXISTS _matmort")
    con.execute(f"""
        CREATE TABLE _matmort AS
        SELECT
            TRIM(jurisdiction) AS jurisdiction,
            TRIM("group") AS demographic_group,
            TRIM(subgroup) AS subgroup,
            TRY_CAST(year_of_death AS INTEGER) AS year,
            TRY_CAST(month_of_death AS INTEGER) AS month,
            time_period,
            TRY_CAST(maternal_deaths AS INTEGER) AS maternal_deaths,
            TRY_CAST(live_births AS BIGINT) AS live_births,
            TRY_CAST(maternal_mortality_rate AS DOUBLE) AS maternal_mortality_rate
        FROM read_csv_auto('{path}', header=true)
        WHERE jurisdiction IS NOT NULL
    """)
    return _write_parquet(con, '_matmort', 'maternal_mortality_monthly')


def build_hcbs_payment_method(con):
    """Ingest MACPAC 1915(c) HCBS payment methodology summary by state."""
    print("\n── MACPAC HCBS Payment Methodology ──")
    path = RAW / "macpac_hcbs_1915c_payment_scan.xlsx"
    if not path.exists():
        print("  File not found")
        return 0

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Summary']

    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        state_name = str(row[0]).strip() if row[0] else None
        if not state_name:
            continue
        state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            continue

        def _str(v):
            if v is None or v == '–' or v == 'blank cell':
                return None
            return str(v).strip()[:500]

        rows.append({
            'state_code': state_code,
            'total_1915c_waivers': _str(row[1]),
            'delivery_system': _str(row[2]),
            'rate_study_provider_survey': _str(row[3]),
            'rate_study_stakeholder': _str(row[4]),
            'rate_study_adopted': _str(row[5]),
            'rate_review_periodicity': _str(row[6]),
            'rate_review_method': _str(row[7]),
            'rate_adjustment_sources': _str(row[8]),
            'home_rate_approach': _str(row[9]),
            'home_wage_sources': _str(row[10]),
            'home_self_directed': _str(row[11]),
            'day_rate_approach': _str(row[12]),
            'day_wage_sources': _str(row[13]),
            'day_self_directed': _str(row[14]),
            'rtc_rate_approach': _str(row[15]),
            'rtc_wage_sources': _str(row[16]),
            'rtc_self_directed': _str(row[17]),
            'appendix_k_adjustments': _str(row[18]),
            'pass_through_requirement': _str(row[19]),
        })
    wb.close()

    if not rows:
        print("  No data found")
        return 0

    con.execute("DROP TABLE IF EXISTS _hcbs_pay")
    con.execute("""
        CREATE TABLE _hcbs_pay (
            state_code VARCHAR,
            total_1915c_waivers VARCHAR,
            delivery_system VARCHAR,
            rate_study_provider_survey VARCHAR,
            rate_study_stakeholder VARCHAR,
            rate_study_adopted VARCHAR,
            rate_review_periodicity VARCHAR,
            rate_review_method VARCHAR,
            rate_adjustment_sources VARCHAR,
            home_rate_approach VARCHAR,
            home_wage_sources VARCHAR,
            home_self_directed VARCHAR,
            day_rate_approach VARCHAR,
            day_wage_sources VARCHAR,
            day_self_directed VARCHAR,
            rtc_rate_approach VARCHAR,
            rtc_wage_sources VARCHAR,
            rtc_self_directed VARCHAR,
            appendix_k_adjustments VARCHAR,
            pass_through_requirement VARCHAR
        )
    """)
    con.executemany(
        "INSERT INTO _hcbs_pay VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [tuple(r.values()) for r in rows]
    )
    return _write_parquet(con, '_hcbs_pay', 'hcbs_payment_method', category="reference")


def main():
    con = duckdb.connect()
    total = 0

    total += build_ltss_expenditure(con)
    total += build_ltss_users(con)
    total += build_ltss_rebalancing(con)
    total += build_vital_stats_monthly(con)
    total += build_maternal_mortality_monthly(con)
    total += build_hcbs_payment_method(con)

    con.close()

    print(f"\n{'='*60}")
    print(f"Round 11 complete: {total:,} total rows")

    # Write manifest
    manifest = {
        "pipeline_run": "round11",
        "snapshot_date": SNAP,
        "total_rows": total,
        "tables": [
            "fact_ltss_expenditure", "fact_ltss_users", "fact_ltss_rebalancing",
            "fact_vital_stats_monthly", "fact_maternal_mortality_monthly",
            "ref_hcbs_payment_method",
        ],
    }
    manifest_path = LAKE / "metadata" / f"manifest_round11_{SNAP}.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"Manifest: {manifest_path}")


if __name__ == "__main__":
    main()
