#!/usr/bin/env python3
"""
build_lake_macpac_hcbs_payment_scan.py — Ingest MACPAC HCBS 1915(c) Payment Scan Summary
into the Aradune data lake.

Source: data/raw/macpac_hcbs_1915c_payment_scan.xlsx (Summary tab)

Table built:
  macpac_hcbs_payment_scan — 1915(c) waiver payment rate-setting approaches by state

The Summary tab contains per-state information about:
  - Total 1915(c) waivers, delivery system, rate study practices
  - FFS rate approach, wage sources, self-direction for home, day, and round-the-clock services
  - Appendix K COVID adjustments

Usage:
  python3 scripts/build_lake_macpac_hcbs_payment_scan.py
"""

import json
import re
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_PATH = PROJECT_ROOT / "data" / "raw" / "macpac_hcbs_1915c_payment_scan.xlsx"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

STATE_MAP = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN",
    "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA",
    "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "District of Columbia": "DC",
}


def _clean_cell(val):
    """Clean cell value to string, handling None and whitespace."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "-", "--", "n/a", "N/A"):
        return None
    return s


def _parse_int(val):
    """Parse integer from cell."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(",", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def build():
    print(f"MACPAC HCBS 1915(c) Payment Scan ETL -- snapshot {SNAPSHOT_DATE}")
    print(f"  Source: {RAW_PATH}")

    if not RAW_PATH.exists():
        print("  ERROR: Source file not found")
        return

    wb = openpyxl.load_workbook(RAW_PATH, read_only=True, data_only=True)
    ws = wb["Summary"]

    # Column mapping from inspection:
    # A: State
    # B: Total 1915(c) waivers
    # C: Delivery system
    # D-F: Rate study (provider survey, stakeholder engagement, results adopted)
    # G-I: FFS rate reviews (periodicity, use of rate study/indexing, adjustment sources)
    # J: Home-based FFS rate approach
    # K: Home-based wage sources
    # L: Home-based self-directed
    # M: Day services FFS rate approach
    # N: Day services wage sources
    # O: Day services self-directed
    # P: Round-the-clock FFS rate approach
    # Q: Round-the-clock wage sources
    # R: Round-the-clock self-directed
    # S: Appendix K adjustments

    rows = []
    for row_num in range(6, ws.max_row + 1):
        raw_state = _clean_cell(ws.cell(row=row_num, column=1).value)
        if not raw_state:
            continue
        # Clean footnote numbers from state names
        clean_state = re.sub(r"\s*\d+$", "", raw_state).strip()
        state_code = STATE_MAP.get(clean_state)
        if not state_code:
            continue

        # Check if this is a "no 1915(c) waivers" state (like Arizona)
        total_waivers = _parse_int(ws.cell(row=row_num, column=2).value)

        record = {
            "state_code": state_code,
            "total_1915c_waivers": total_waivers,
            "delivery_system": _clean_cell(ws.cell(row=row_num, column=3).value),
            "rate_study_provider_survey": _clean_cell(ws.cell(row=row_num, column=4).value),
            "rate_study_stakeholder_engagement": _clean_cell(ws.cell(row=row_num, column=5).value),
            "rate_study_results_adopted": _clean_cell(ws.cell(row=row_num, column=6).value),
            "rate_review_periodicity": _clean_cell(ws.cell(row=row_num, column=7).value),
            "rate_review_method": _clean_cell(ws.cell(row=row_num, column=8).value),
            "rate_adjustment_sources": _clean_cell(ws.cell(row=row_num, column=9).value),
            "home_ffs_rate_approach": _clean_cell(ws.cell(row=row_num, column=10).value),
            "home_wage_sources": _clean_cell(ws.cell(row=row_num, column=11).value),
            "home_self_directed": _clean_cell(ws.cell(row=row_num, column=12).value),
            "day_ffs_rate_approach": _clean_cell(ws.cell(row=row_num, column=13).value),
            "day_wage_sources": _clean_cell(ws.cell(row=row_num, column=14).value),
            "day_self_directed": _clean_cell(ws.cell(row=row_num, column=15).value),
            "roundtheclock_ffs_rate_approach": _clean_cell(ws.cell(row=row_num, column=16).value),
            "roundtheclock_wage_sources": _clean_cell(ws.cell(row=row_num, column=17).value),
            "roundtheclock_self_directed": _clean_cell(ws.cell(row=row_num, column=18).value),
            "appendix_k_adjustments": _clean_cell(ws.cell(row=row_num, column=19).value),
            "source": "macpac",
            "reference_year": 2023,
            "snapshot_date": SNAPSHOT_DATE,
        }
        rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return

    print(f"  Parsed {len(rows)} state records")

    # Load into DuckDB and write
    con = duckdb.connect()

    columns = list(rows[0].keys())
    col_defs = ", ".join(
        f"{c} {'INTEGER' if c in ('total_1915c_waivers', 'reference_year') else 'VARCHAR'}"
        for c in columns
    )
    con.execute(f"CREATE TABLE _scan ({col_defs})")
    placeholders = ", ".join(["?"] * len(columns))
    con.executemany(
        f"INSERT INTO _scan VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    out_path = FACT_DIR / "macpac_hcbs_payment_scan" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    count = con.execute("SELECT COUNT(*) FROM _scan").fetchone()[0]
    con.execute(f"COPY _scan TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_mb = out_path.stat().st_size / (1024 * 1024)
    print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count} rows, {size_mb:.2f} MB)")

    # Validation
    has_waivers = con.execute("SELECT COUNT(*) FROM _scan WHERE total_1915c_waivers IS NOT NULL AND total_1915c_waivers > 0").fetchone()[0]
    ffs_count = con.execute("SELECT COUNT(*) FROM _scan WHERE delivery_system ILIKE '%fee%' OR delivery_system = 'Both'").fetchone()[0]
    mc_count = con.execute("SELECT COUNT(*) FROM _scan WHERE delivery_system ILIKE '%managed%'").fetchone()[0]
    rate_study = con.execute("SELECT COUNT(*) FROM _scan WHERE rate_study_results_adopted ILIKE '%yes%'").fetchone()[0]
    print(f"  States with 1915(c) waivers: {has_waivers}")
    print(f"  Delivery: FFS/Both={ffs_count}, Managed care={mc_count}")
    print(f"  States with rate study results adopted: {rate_study}")

    # Sample
    sample = con.execute("SELECT state_code, total_1915c_waivers, delivery_system, rate_review_periodicity FROM _scan LIMIT 5").fetchall()
    for s in sample:
        print(f"    {s}")

    con.close()

    # Manifest
    manifest = {
        "pipeline_run": "macpac_hcbs_payment_scan",
        "run_id": RUN_ID,
        "snapshot_date": SNAPSHOT_DATE,
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "total_rows": count,
        "tables": ["macpac_hcbs_payment_scan"],
        "source_files": ["data/raw/macpac_hcbs_1915c_payment_scan.xlsx"],
        "source": "MACPAC HCBS 1915(c) Payment Scan Compendium",
    }
    manifest_path = META_DIR / f"manifest_macpac_hcbs_payment_{SNAPSHOT_DATE}.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"  Manifest: {manifest_path.relative_to(LAKE_DIR)}")

    print(f"\n== Summary ==")
    print(f"  Table: macpac_hcbs_payment_scan ({count} rows)")
    print(f"  HCBS rate-setting approaches for 1915(c) waivers across {len(rows)} states/DC")


if __name__ == "__main__":
    build()
