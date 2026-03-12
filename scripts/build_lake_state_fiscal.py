#!/usr/bin/env python3
"""
build_lake_state_fiscal.py — Ingest state budget, fiscal, and tax data into the lake.

Downloads and parses:
  1. Census Annual Survey of State Government Finances (2018-2023)
  2. Tax Foundation Facts & Figures 2025 (43 tables of state tax data)
  3. MACPAC FMAP Historical (FY 2023-2026 with emergency adjustments)

Tables built:
  fact_census_state_finances    — State government revenue, expenditure, debt (2017-2023)
  fact_tax_burden               — State-local tax burden as % of income + per capita
  fact_state_tax_collections    — State tax collections per capita + revenue per capita
  fact_federal_aid_share        — Federal aid as % of state general revenue
  fact_state_debt               — State debt per capita
  fact_pension_funded_ratio     — Public pension plan funded ratios by state
  fact_state_tax_rates          — Key state tax rates (sales, gasoline, cigarette, competitiveness)
  fact_tax_revenue_sources      — Tax revenue composition by source (property, sales, income, etc.)
  fact_income_per_capita        — Per capita income by state
  fact_property_tax_rate        — Effective property tax rate as % of housing value
  fact_fmap_historical          — FMAP and E-FMAP by state, FY 2023-2026 (quarterly)

Sources:
  - Census: https://www2.census.gov/programs-surveys/state/tables/
  - Tax Foundation: https://taxfoundation.org/data/all/state/2025-state-tax-data/
  - MACPAC: https://www.macpac.gov/subtopic/matching-rates/

Usage:
  python3 scripts/build_lake_state_fiscal.py
  python3 scripts/build_lake_state_fiscal.py --dry-run
  python3 scripts/build_lake_state_fiscal.py --only fact_census_state_finances,fact_tax_burden
"""

import argparse
import os
import re
import subprocess
import uuid
from datetime import date
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "state_fiscal"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

# Abbreviated state names used by Tax Foundation
TF_STATE_TO_CODE = {
    'Ala.': 'AL', 'Alaska': 'AK', 'Ariz.': 'AZ', 'Ark.': 'AR',
    'Calif.': 'CA', 'Colo.': 'CO', 'Conn.': 'CT', 'Del.': 'DE',
    'Fla.': 'FL', 'Ga.': 'GA', 'Hawaii': 'HI', 'Idaho': 'ID',
    'Ill.': 'IL', 'Ind.': 'IN', 'Iowa': 'IA', 'Kans.': 'KS',
    'Ky.': 'KY', 'La.': 'LA', 'Maine': 'ME', 'Md.': 'MD',
    'Mass.': 'MA', 'Mich.': 'MI', 'Minn.': 'MN', 'Miss.': 'MS',
    'Mo.': 'MO', 'Mont.': 'MT', 'Nebr.': 'NE', 'Nev.': 'NV',
    'N.H.': 'NH', 'N.J.': 'NJ', 'N.Mex.': 'NM', 'N.Y.': 'NY',
    'N.C.': 'NC', 'N.D.': 'ND', 'Ohio': 'OH', 'Okla.': 'OK',
    'Ore.': 'OR', 'Pa.': 'PA', 'R.I.': 'RI', 'S.C.': 'SC',
    'S.D.': 'SD', 'Tenn.': 'TN', 'Texas': 'TX', 'Utah': 'UT',
    'Vt.': 'VT', 'Va.': 'VA', 'Wash.': 'WA', 'W.Va.': 'WV',
    'Wis.': 'WI', 'Wyo.': 'WY', 'D.C.': 'DC',
    'U.S.': 'US', 'U.S. Average': 'US',
}

STATE_NAME_TO_CODE = {
    'Alabama': 'AL', 'Alaska': 'AK', 'Arizona': 'AZ', 'Arkansas': 'AR',
    'California': 'CA', 'Colorado': 'CO', 'Connecticut': 'CT', 'Delaware': 'DE',
    'District of Columbia': 'DC', 'Florida': 'FL', 'Georgia': 'GA', 'Hawaii': 'HI',
    'Idaho': 'ID', 'Illinois': 'IL', 'Indiana': 'IN', 'Iowa': 'IA',
    'Kansas': 'KS', 'Kentucky': 'KY', 'Louisiana': 'LA', 'Maine': 'ME',
    'Maryland': 'MD', 'Massachusetts': 'MA', 'Michigan': 'MI', 'Minnesota': 'MN',
    'Mississippi': 'MS', 'Missouri': 'MO', 'Montana': 'MT', 'Nebraska': 'NE',
    'Nevada': 'NV', 'New Hampshire': 'NH', 'New Jersey': 'NJ', 'New Mexico': 'NM',
    'New York': 'NY', 'North Carolina': 'NC', 'North Dakota': 'ND', 'Ohio': 'OH',
    'Oklahoma': 'OK', 'Oregon': 'OR', 'Pennsylvania': 'PA', 'Puerto Rico': 'PR',
    'Rhode Island': 'RI', 'South Carolina': 'SC', 'South Dakota': 'SD',
    'Tennessee': 'TN', 'Texas': 'TX', 'Utah': 'UT', 'Vermont': 'VT',
    'Virginia': 'VA', 'Washington': 'WA', 'West Virginia': 'WV',
    'Wisconsin': 'WI', 'Wyoming': 'WY', 'United States': 'US',
}


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.2f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def _download_file(url: str, dest: Path) -> bool:
    """Download a file using curl. Returns True on success."""
    if dest.exists() and dest.stat().st_size > 1000:
        print(f"  [cached] {dest.name}")
        return True
    print(f"  Downloading {dest.name}...")
    result = subprocess.run(
        ["curl", "-sL", "-o", str(dest), url],
        capture_output=True, text=True, timeout=120
    )
    if result.returncode != 0 or not dest.exists() or dest.stat().st_size < 100:
        print(f"  FAILED to download {url}")
        return False
    return True


def _safe_float(v) -> float | None:
    """Convert a value to float, handling X, N/A, None, etc."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    s = s.replace("\xa0", "")
    if s in ("", "X", "N/A", "n.a.", "-", "—", "(X)", "N.A.", "..."):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _safe_int(v) -> int | None:
    f = _safe_float(v)
    return int(f) if f is not None else None


def _clean_label(s: str) -> str:
    """Normalize a Census finance category label."""
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


# ---------------------------------------------------------------------------
# 1. Census Annual Survey of State Government Finances
# ---------------------------------------------------------------------------

def _download_census_files() -> list[Path]:
    """Download Census ASFIN files for 2018-2023."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    files = []

    # 2022-2023: new format (single year, states as columns)
    for year in [2023, 2022]:
        dest = RAW_DIR / f"census_state_finances_{year}.xlsx"
        url = f"https://www2.census.gov/programs-surveys/state/tables/{year}/{year}%20ASFIN%20State%20Totals.xlsx"
        if _download_file(url, dest):
            files.append(dest)

    # 2018-2021: old format (two fiscal years per file, states in groups of 3 cols)
    old_files = {
        2021: "ASFIN%20FY2021_2020.xlsx",
        2020: "ASFIN%20FY2020_2019.xlsx",
        2019: "ASFIN%20FY2019_2018.xlsx",
        2018: "ASFIN%20FY2018_2017.xlsx",
    }
    for year, fname in old_files.items():
        dest = RAW_DIR / f"census_state_finances_{year}.xlsx"
        url = f"https://www2.census.gov/programs-surveys/state/tables/{year}/{fname}"
        if _download_file(url, dest):
            files.append(dest)

    return files


def _parse_census_new_format(filepath: Path, year: int) -> list[dict]:
    """Parse 2022-2023 format: single year, states as columns B-AZ."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = []

    # Find header row with state names (row with "United States" in col B)
    header_row = None
    data_start = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        for cell in row:
            if cell.value and "United States" in str(cell.value):
                header_row = r_idx
                break
        if header_row:
            data_start = header_row + 1
            break

    if not header_row:
        print(f"  WARNING: Could not find header row in {filepath.name}")
        return []

    # Extract state names from header row
    state_names = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        state_names = [_clean_label(str(v)) if v else None for v in row]
        break

    # Map columns to state codes
    col_to_state = {}
    for i, name in enumerate(state_names):
        if name and name != "(Thousands of Dollars)":
            code = STATE_NAME_TO_CODE.get(name.strip())
            if code:
                col_to_state[i] = code

    # Parse data rows
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        category = _clean_label(str(row[0])) if row[0] else None
        if not category:
            continue
        # Skip footnotes, headers, etc.
        if category.startswith("Footnote") or category.startswith("X ") or category.startswith("Last Revised"):
            break
        if category.startswith("Source:") or category.startswith("Table "):
            continue

        # Determine hierarchy level from leading spaces
        raw_label = str(row[0]).replace("\xa0", " ") if row[0] else ""
        indent = len(raw_label) - len(raw_label.lstrip())
        level = indent // 5  # Approx 5 spaces per indent level

        for col_idx, state_code in col_to_state.items():
            if col_idx < len(row):
                amount = _safe_int(row[col_idx])
                if amount is not None:
                    rows.append({
                        "state_code": state_code,
                        "fiscal_year": year,
                        "category": category,
                        "level": level,
                        "amount_thousands": amount,
                        "source": "census.gov/state-finances",
                        "snapshot_date": SNAPSHOT_DATE,
                    })

    return rows


def _parse_census_old_format(filepath: Path) -> list[dict]:
    """Parse 2018-2021 format: two FYs per file, states in groups of 3 cols."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = []

    # Row 1: state names in every 3rd column starting at C
    # Row 2: FY labels (e.g., FY2019, FY2018, Percent change)
    # Data starts at row 3

    # Get state names from row 1
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    # Build mapping: (col_idx, fiscal_year) -> state_code
    col_map = {}  # col_idx -> (state_code, fiscal_year)
    current_state = None
    for i, v in enumerate(row1):
        if v:
            name = str(v).strip()
            if name in STATE_NAME_TO_CODE:
                current_state = STATE_NAME_TO_CODE[name]
            elif name == "(Thousands of Dollars)":
                continue
        if current_state and i < len(row2) and row2[i]:
            fy_label = str(row2[i]).strip()
            m = re.match(r"FY(\d{4})", fy_label)
            if m:
                col_map[i] = (current_state, int(m.group(1)))

    # Parse data rows (start at row 3)
    for row in ws.iter_rows(min_row=3, values_only=True):
        category = _clean_label(str(row[0])) if row[0] else None
        if not category:
            continue
        if category.startswith("Source:") or category.startswith("Note:"):
            break

        raw_label = str(row[0]) if row[0] else ""
        indent = len(raw_label) - len(raw_label.lstrip())
        level = indent // 2

        for col_idx, (state_code, fiscal_year) in col_map.items():
            if col_idx < len(row):
                amount = _safe_int(row[col_idx])
                if amount is not None:
                    rows.append({
                        "state_code": state_code,
                        "fiscal_year": fiscal_year,
                        "category": category,
                        "level": level,
                        "amount_thousands": amount,
                        "source": "census.gov/state-finances",
                        "snapshot_date": SNAPSHOT_DATE,
                    })

    return rows


def build_census_state_finances(con, dry_run: bool) -> int:
    """Build fact_census_state_finances from Census ASFIN Excel files."""
    print("Building fact_census_state_finances...")
    files = _download_census_files()

    all_rows = []

    for f in files:
        year_match = re.search(r"(\d{4})", f.name)
        if not year_match:
            continue
        year = int(year_match.group(1))

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            print(f"  ERROR loading {f.name}: {e}")
            continue

        # Detect format: new (2022-2023) has "Table 1." in A1
        cell_a1 = str(ws.cell(1, 1).value or "")
        if "Table" in cell_a1:
            parsed = _parse_census_new_format(f, year)
        else:
            parsed = _parse_census_old_format(f)

        print(f"  {f.name}: {len(parsed):,} records")
        all_rows.extend(parsed)

    if not all_rows:
        print("  SKIPPED: no data parsed")
        return 0

    # Deduplicate (old format files contain overlapping years)
    seen = set()
    deduped = []
    for r in all_rows:
        key = (r["state_code"], r["fiscal_year"], r["category"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    print(f"  Total: {len(deduped):,} unique records (from {len(all_rows):,} raw)")

    import pandas as pd
    df = pd.DataFrame(deduped)

    con.execute("DROP TABLE IF EXISTS _fact_census_state_finances")
    con.execute("""
        CREATE TABLE _fact_census_state_finances AS
        SELECT
            state_code::VARCHAR AS state_code,
            fiscal_year::INTEGER AS fiscal_year,
            category::VARCHAR AS category,
            level::INTEGER AS level,
            amount_thousands::BIGINT AS amount_thousands,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY fiscal_year, state_code, category
    """)

    return write_parquet(con, "_fact_census_state_finances",
                         _snapshot_path("fact_census_state_finances"), dry_run)


# ---------------------------------------------------------------------------
# 2. Tax Foundation Facts & Figures
# ---------------------------------------------------------------------------

def _download_tax_foundation() -> Path | None:
    """Download Tax Foundation Facts & Figures 2025 Excel."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    dest = RAW_DIR / "tax_foundation_facts_figures_2025.xlsx"
    url = "https://taxfoundation.org/wp-content/uploads/2025/05/Facts_and_Figures_2025-v2.xlsx"
    if _download_file(url, dest):
        return dest
    return None


def _parse_tf_simple_table(ws, value_col: int = 1, rank_col: int = 2,
                           data_start: int = 5, state_col: int = 0) -> list[dict]:
    """Parse a simple Tax Foundation table: State | Value | Rank."""
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        state_name = str(row[state_col]).strip() if row[state_col] else None
        if not state_name:
            continue
        code = TF_STATE_TO_CODE.get(state_name)
        if not code:
            continue
        value = _safe_float(row[value_col]) if value_col < len(row) else None
        rank = _safe_int(row[rank_col]) if rank_col < len(row) else None
        if value is not None:
            rows.append({"state_code": code, "value": value, "rank": rank})
    return rows


def build_tax_burden(con, dry_run: bool) -> int:
    """Build fact_tax_burden from Tax Foundation Table 2."""
    print("Building fact_tax_burden...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    ws = wb["2"]

    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        state_name = str(row[0]).strip() if row[0] else None
        if not state_name:
            continue
        code = TF_STATE_TO_CODE.get(state_name)
        if not code:
            continue
        burden_pct = _safe_float(row[1])
        rank = _safe_int(row[2])
        burden_per_capita = _safe_float(row[3])
        if burden_pct is not None:
            rows.append({
                "state_code": code,
                "year": 2022,
                "tax_burden_pct": burden_pct,
                "rank": rank,
                "tax_burden_per_capita": burden_per_capita,
                "source": "taxfoundation.org",
                "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_tax_burden")
    con.execute("""
        CREATE TABLE _fact_tax_burden AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            tax_burden_pct::DOUBLE AS tax_burden_pct,
            rank::INTEGER AS rank,
            tax_burden_per_capita::DOUBLE AS tax_burden_per_capita,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code
    """)
    return write_parquet(con, "_fact_tax_burden",
                         _snapshot_path("fact_tax_burden"), dry_run)


def build_state_tax_collections(con, dry_run: bool) -> int:
    """Build fact_state_tax_collections from Tax Foundation Tables 1, 4, 5."""
    print("Building fact_state_tax_collections...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    rows = []

    # Table 1: State Tax Collections per Capita (FY 2023)
    ws1 = wb["1"]
    for row in ws1.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2023,
                "metric": "state_tax_collections_per_capita",
                "value": val, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 4: State Revenue per Capita (FY 2022)
    ws4 = wb["4"]
    for row in ws4.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2022,
                "metric": "state_revenue_per_capita",
                "value": val, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 5: State & Local Tax Collections per Capita (FY 2022)
    ws5 = wb["5"]
    for row in ws5.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2022,
                "metric": "state_local_tax_collections_per_capita",
                "value": val, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 12: State Individual Income Tax Collections per Capita (FY 2023)
    ws12 = wb["12"]
    for row in ws12.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2023,
                "metric": "individual_income_tax_per_capita",
                "value": val, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 15: State Corporate Income Tax Collections per Capita (FY 2023)
    ws15 = wb["15"]
    for row in ws15.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2023,
                "metric": "corporate_income_tax_per_capita",
                "value": val, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 19: State General Sales Tax Collections per Capita (FY 2023)
    ws19 = wb["19"]
    for row in ws19.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2023,
                "metric": "general_sales_tax_per_capita",
                "value": val, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_state_tax_collections")
    con.execute("""
        CREATE TABLE _fact_state_tax_collections AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            metric::VARCHAR AS metric,
            value::DOUBLE AS value,
            rank::INTEGER AS rank,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY metric, state_code
    """)
    return write_parquet(con, "_fact_state_tax_collections",
                         _snapshot_path("fact_state_tax_collections"), dry_run)


def build_federal_aid_share(con, dry_run: bool) -> int:
    """Build fact_federal_aid_share from Tax Foundation Table 8."""
    print("Building fact_federal_aid_share...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    ws = wb["8"]
    rows = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        pct = _safe_float(row[1])
        rank = _safe_int(row[2])
        if pct is not None:
            rows.append({
                "state_code": code, "year": 2022,
                "federal_aid_pct_of_general_revenue": pct,
                "rank": rank,
                "source": "taxfoundation.org",
                "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_federal_aid_share")
    con.execute("""
        CREATE TABLE _fact_federal_aid_share AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            federal_aid_pct_of_general_revenue::DOUBLE AS federal_aid_pct_of_general_revenue,
            rank::INTEGER AS rank,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code
    """)
    return write_parquet(con, "_fact_federal_aid_share",
                         _snapshot_path("fact_federal_aid_share"), dry_run)


def build_state_debt(con, dry_run: bool) -> int:
    """Build fact_state_debt from Tax Foundation Table 38."""
    print("Building fact_state_debt...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    ws = wb["38"]
    rows = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2022,
                "debt_per_capita": val,
                "rank": rank,
                "source": "taxfoundation.org",
                "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_state_debt")
    con.execute("""
        CREATE TABLE _fact_state_debt AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            debt_per_capita::DOUBLE AS debt_per_capita,
            rank::INTEGER AS rank,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code
    """)
    return write_parquet(con, "_fact_state_debt",
                         _snapshot_path("fact_state_debt"), dry_run)


def build_pension_funded_ratio(con, dry_run: bool) -> int:
    """Build fact_pension_funded_ratio from Tax Foundation Table 40."""
    print("Building fact_pension_funded_ratio...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    ws = wb["40"]
    rows = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        ratio = _safe_float(row[1])
        rank = _safe_int(row[2])
        if ratio is not None:
            rows.append({
                "state_code": code, "year": 2023,
                "funded_ratio": ratio,
                "rank": rank,
                "source": "taxfoundation.org",
                "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_pension_funded_ratio")
    con.execute("""
        CREATE TABLE _fact_pension_funded_ratio AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            funded_ratio::DOUBLE AS funded_ratio,
            rank::INTEGER AS rank,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code
    """)
    return write_parquet(con, "_fact_pension_funded_ratio",
                         _snapshot_path("fact_pension_funded_ratio"), dry_run)


def build_state_tax_rates(con, dry_run: bool) -> int:
    """Build fact_state_tax_rates from Tax Foundation Tables 3, 11, 14, 18, 22, 24."""
    print("Building fact_state_tax_rates...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    rows = []

    # Table 3: State Tax Competitiveness Index (2025) - overall + component ranks
    ws3 = wb["3"]
    for row in ws3.iter_rows(min_row=6, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        overall_rank = _safe_int(row[1]) if len(row) > 1 else None
        if overall_rank is not None:
            rows.append({
                "state_code": code, "year": 2025,
                "metric": "tax_competitiveness_rank",
                "value": float(overall_rank), "rate_text": None,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 18: State & Local Sales Tax Rates (2025)
    ws18 = wb["18"]
    for row in ws18.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        state_rate = _safe_float(row[1])
        if state_rate is not None:
            rows.append({
                "state_code": code, "year": 2025,
                "metric": "state_sales_tax_rate",
                "value": state_rate, "rate_text": None,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })
        combined = _safe_float(row[3]) if len(row) > 3 else None
        if combined is not None:
            rows.append({
                "state_code": code, "year": 2025,
                "metric": "combined_sales_tax_rate",
                "value": combined, "rate_text": None,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 22: Gasoline Tax Rates (cents per gallon)
    ws22 = wb["22"]
    for row in ws22.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        rate = _safe_float(row[1])
        if rate is not None:
            rows.append({
                "state_code": code, "year": 2025,
                "metric": "gasoline_tax_cents_per_gallon",
                "value": rate, "rate_text": None,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    # Table 24: Cigarette Excise Tax Rates ($ per 20-pack)
    ws24 = wb["24"]
    for row in ws24.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        rate = _safe_float(row[1])
        if rate is not None:
            rows.append({
                "state_code": code, "year": 2025,
                "metric": "cigarette_tax_per_pack",
                "value": rate, "rate_text": None,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_state_tax_rates")
    con.execute("""
        CREATE TABLE _fact_state_tax_rates AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            metric::VARCHAR AS metric,
            value::DOUBLE AS value,
            rate_text::VARCHAR AS rate_text,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY metric, state_code
    """)
    return write_parquet(con, "_fact_state_tax_rates",
                         _snapshot_path("fact_state_tax_rates"), dry_run)


def build_tax_revenue_sources(con, dry_run: bool) -> int:
    """Build fact_tax_revenue_sources from Tax Foundation Table 7."""
    print("Building fact_tax_revenue_sources...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    ws = wb["7"]
    rows = []

    # Columns: State | Property | General Sales | Individual Income | Corporate Income | Other
    for row in ws.iter_rows(min_row=6, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        metrics = [
            ("property_tax_pct", 1),
            ("general_sales_tax_pct", 2),
            ("individual_income_tax_pct", 3),
            ("corporate_income_tax_pct", 4),
            ("other_taxes_pct", 5),
        ]
        for metric_name, col in metrics:
            val = _safe_float(row[col]) if col < len(row) else None
            if val is not None:
                rows.append({
                    "state_code": code, "year": 2022,
                    "metric": metric_name, "value": val,
                    "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
                })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_tax_revenue_sources")
    con.execute("""
        CREATE TABLE _fact_tax_revenue_sources AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            metric::VARCHAR AS metric,
            value::DOUBLE AS value,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code, metric
    """)
    return write_parquet(con, "_fact_tax_revenue_sources",
                         _snapshot_path("fact_tax_revenue_sources"), dry_run)


def build_income_per_capita(con, dry_run: bool) -> int:
    """Build fact_income_per_capita from Tax Foundation Table 41."""
    print("Building fact_income_per_capita...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    ws = wb["41"]
    rows = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        val = _safe_float(row[1])
        rank = _safe_int(row[2])
        if val is not None:
            rows.append({
                "state_code": code, "year": 2023,
                "income_per_capita": val, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_income_per_capita")
    con.execute("""
        CREATE TABLE _fact_income_per_capita AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            income_per_capita::DOUBLE AS income_per_capita,
            rank::INTEGER AS rank,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code
    """)
    return write_parquet(con, "_fact_income_per_capita",
                         _snapshot_path("fact_income_per_capita"), dry_run)


def build_property_tax_rate(con, dry_run: bool) -> int:
    """Build fact_property_tax_rate from Tax Foundation Table 33."""
    print("Building fact_property_tax_rate...")
    tf_path = _download_tax_foundation()
    if not tf_path:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(tf_path, data_only=True)
    ws = wb["33"]
    rows = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        state = str(row[0]).strip() if row[0] else None
        code = TF_STATE_TO_CODE.get(state) if state else None
        if not code:
            continue
        rate = _safe_float(row[1])
        rank = _safe_int(row[2])
        if rate is not None:
            rows.append({
                "state_code": code, "year": 2023,
                "effective_property_tax_rate": rate, "rank": rank,
                "source": "taxfoundation.org", "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_property_tax_rate")
    con.execute("""
        CREATE TABLE _fact_property_tax_rate AS
        SELECT
            state_code::VARCHAR AS state_code,
            year::INTEGER AS year,
            effective_property_tax_rate::DOUBLE AS effective_property_tax_rate,
            rank::INTEGER AS rank,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code
    """)
    return write_parquet(con, "_fact_property_tax_rate",
                         _snapshot_path("fact_property_tax_rate"), dry_run)


# ---------------------------------------------------------------------------
# 3. MACPAC FMAP Historical
# ---------------------------------------------------------------------------

def _download_macpac_fmap() -> Path | None:
    """Download MACPAC FMAP Excel file."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    dest = RAW_DIR / "macpac_fmap_2023_2026.xlsx"
    # URL has em-dash encoded
    url = "https://www.macpac.gov/wp-content/uploads/2026/01/EXHIBIT-6.-Federal-Medical-Assistance-Percentages-and-Enhanced-Federal-Medical-Assistance-Percentages-by-State-FYs-2023%E2%80%932026.xlsx"
    if _download_file(url, dest):
        return dest
    return None


def build_fmap_historical(con, dry_run: bool) -> int:
    """Build fact_fmap_historical from MACPAC Exhibit 6."""
    print("Building fact_fmap_historical...")
    fpath = _download_macpac_fmap()
    if not fpath:
        print("  SKIPPED: file not available")
        return 0

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    rows = []

    # Row 3 has period labels, Row 4 has sub-labels
    # Columns: A=State, B-G=FMAP periods, H-M=E-FMAP periods
    # Periods: FY2023 Q1-2 (Emergency), FY2023 Q3 (Emergency), FY2023 Q4 (Emergency),
    #          FY2024, FY2025, FY2026

    periods = [
        {"col": 1, "fiscal_year": 2023, "quarter": "Q1-Q2", "emergency": True, "type": "fmap"},
        {"col": 2, "fiscal_year": 2023, "quarter": "Q3", "emergency": True, "type": "fmap"},
        {"col": 3, "fiscal_year": 2023, "quarter": "Q4", "emergency": True, "type": "fmap"},
        {"col": 4, "fiscal_year": 2024, "quarter": "annual", "emergency": False, "type": "fmap"},
        {"col": 5, "fiscal_year": 2025, "quarter": "annual", "emergency": False, "type": "fmap"},
        {"col": 6, "fiscal_year": 2026, "quarter": "annual", "emergency": False, "type": "fmap"},
        {"col": 7, "fiscal_year": 2023, "quarter": "Q1-Q2", "emergency": True, "type": "efmap"},
        {"col": 8, "fiscal_year": 2023, "quarter": "Q3", "emergency": True, "type": "efmap"},
        {"col": 9, "fiscal_year": 2023, "quarter": "Q4", "emergency": True, "type": "efmap"},
        {"col": 10, "fiscal_year": 2024, "quarter": "annual", "emergency": False, "type": "efmap"},
        {"col": 11, "fiscal_year": 2025, "quarter": "annual", "emergency": False, "type": "efmap"},
        {"col": 12, "fiscal_year": 2026, "quarter": "annual", "emergency": False, "type": "efmap"},
    ]

    for row in ws.iter_rows(min_row=5, values_only=True):
        state_name = str(row[0]).strip() if row[0] else None
        if not state_name:
            continue
        code = STATE_NAME_TO_CODE.get(state_name)
        if not code:
            continue

        for p in periods:
            val = _safe_float(row[p["col"]]) if p["col"] < len(row) else None
            if val is not None:
                rows.append({
                    "state_code": code,
                    "fiscal_year": p["fiscal_year"],
                    "quarter": p["quarter"],
                    "emergency_period": p["emergency"],
                    "rate_type": p["type"],
                    "rate": val,
                    "source": "macpac.gov",
                    "snapshot_date": SNAPSHOT_DATE,
                })

    if not rows:
        print("  SKIPPED: no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS _fact_fmap_historical")
    con.execute("""
        CREATE TABLE _fact_fmap_historical AS
        SELECT
            state_code::VARCHAR AS state_code,
            fiscal_year::INTEGER AS fiscal_year,
            quarter::VARCHAR AS quarter,
            emergency_period::BOOLEAN AS emergency_period,
            rate_type::VARCHAR AS rate_type,
            rate::DOUBLE AS rate,
            source::VARCHAR AS source,
            snapshot_date::DATE AS snapshot_date
        FROM df
        ORDER BY state_code, fiscal_year, quarter, rate_type
    """)
    return write_parquet(con, "_fact_fmap_historical",
                         _snapshot_path("fact_fmap_historical"), dry_run)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

BUILDERS = {
    "fact_census_state_finances": build_census_state_finances,
    "fact_tax_burden": build_tax_burden,
    "fact_state_tax_collections": build_state_tax_collections,
    "fact_federal_aid_share": build_federal_aid_share,
    "fact_state_debt": build_state_debt,
    "fact_pension_funded_ratio": build_pension_funded_ratio,
    "fact_state_tax_rates": build_state_tax_rates,
    "fact_tax_revenue_sources": build_tax_revenue_sources,
    "fact_income_per_capita": build_income_per_capita,
    "fact_property_tax_rate": build_property_tax_rate,
    "fact_fmap_historical": build_fmap_historical,
}


def main():
    parser = argparse.ArgumentParser(description="Build state fiscal data lake tables")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only", type=str, help="Comma-separated list of tables to build")
    args = parser.parse_args()

    con = duckdb.connect()
    total_rows = 0

    targets = BUILDERS
    if args.only:
        names = [n.strip() for n in args.only.split(",")]
        targets = {k: v for k, v in BUILDERS.items() if k in names}

    print(f"=== build_lake_state_fiscal.py === (run_id={RUN_ID[:8]})")
    print(f"  Snapshot date: {SNAPSHOT_DATE}")
    print(f"  Tables to build: {', '.join(targets.keys())}")
    print()

    for name, builder in targets.items():
        try:
            count = builder(con, args.dry_run)
            total_rows += count
            print(f"  {name}: {count:,} rows")
        except Exception as e:
            print(f"  ERROR building {name}: {e}")
            import traceback
            traceback.print_exc()
        print()

    print(f"=== DONE: {total_rows:,} total rows across {len(targets)} tables ===")
    con.close()


if __name__ == "__main__":
    main()
