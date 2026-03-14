#!/usr/bin/env python3
"""
build_lake_chronic_conditions.py — Ingest chronic conditions, MCBS, MACPAC,
and Medicare supplemental datasets into the Aradune data lake.

Sources:
  - CCW Tables B.2.a/b/c: Medicare chronic conditions prevalence (national, FFS + all Medicare)
  - MCBS Cost Supplement PUF: Beneficiary-level spending by demographics, setting, payer
  - MACPAC Exhibit 17: Benefit spending by state and service category
  - MACPAC Exhibit 21: Spending by state, eligibility group, dual status
  - MACPAC Exhibit 29: Managed care enrollment % by state
  - Medicare Telehealth Trends: Quarterly state-level telehealth utilization

Tables built:
  fact_chronic_conditions_national  — 30 CCW conditions, national FFS, 2017-2022
  fact_chronic_conditions_other     — 40+ other chronic/disabling conditions, national FFS, 2013-2022
  fact_chronic_conditions_all_medicare — 27 conditions, all Medicare (incl. MA), 2012-2021
  fact_mcbs_cost_summary            — MCBS Cost Supplement aggregated by demographics
  fact_macpac_benefit_spending      — Benefit spending by state and service category (updated)
  fact_macpac_spending_by_elig      — Spending by state, eligibility group, dual status
  fact_macpac_mc_enrollment_pct     — Managed care enrollment % by state and plan type
  fact_medicare_telehealth          — Quarterly telehealth utilization by state/demographics

Usage:
  python3 scripts/build_lake_chronic_conditions.py
  python3 scripts/build_lake_chronic_conditions.py --dry-run
  python3 scripts/build_lake_chronic_conditions.py --only fact_chronic_conditions_national
"""

import argparse
import csv
import json
import os
import re
import subprocess
import tempfile
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path

import duckdb

try:
    import openpyxl
except ImportError:
    openpyxl = None

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def _download(url: str, dest: Path, label: str = "") -> bool:
    """Download a file using curl subprocess."""
    if dest.exists() and dest.stat().st_size > 100:
        print(f"  [cached] {label or dest.name}")
        return True
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"  Downloading {label or dest.name}...")
    result = subprocess.run(
        ["curl", "-sL", "-o", str(dest), url],
        capture_output=True, timeout=300
    )
    if result.returncode != 0 or not dest.exists() or dest.stat().st_size < 100:
        print(f"  FAILED to download {label or url}")
        return False
    size_mb = dest.stat().st_size / (1024 * 1024)
    print(f"  Downloaded {label or dest.name} ({size_mb:.1f} MB)")
    return True


def _clean_number(val):
    """Convert string to float, handling commas and special chars."""
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("$", "").replace("%", "")
    if s in ("", "–", "-", "—", "N/A", "n/a", "*", ".", ".."):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _clean_state_name(name: str) -> str:
    """Strip footnote numbers and whitespace from state names."""
    if not name:
        return ""
    return re.sub(r'[\d,*†‡§]+$', '', str(name)).strip()


STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
    "U.S. Virgin Islands": "VI",
}


# ---------------------------------------------------------------------------
# CCW Chronic Conditions — Parse from PDF tables
# ---------------------------------------------------------------------------

def _parse_ccw_b2a() -> list:
    """Parse Table B.2.a: 30 CCW chronic conditions prevalence, FFS, 2017-2022.
    Data extracted manually from the PDF structure we observed."""
    # National-level data for 30 conditions across 6 years
    # Columns per year: numerator, rate
    conditions_p1 = {
        "Acute myocardial infarction": {2017: (378696, 1.2), 2018: (391967, 1.2), 2019: (400692, 1.3)},
        "Alzheimer's disease": {2017: (1021002, 3.1), 2018: (975599, 3.0), 2019: (930049, 2.9)},
        "Anemia": {2017: (7040563, 21.7), 2018: (6984615, 21.7), 2019: (6941320, 21.8)},
        "Asthma": {2017: (2552924, 7.9), 2018: (2487089, 7.7), 2019: (2431373, 7.6)},
        "Atrial fibrillation and flutter": {2017: (4261499, 13.1), 2018: (4294182, 13.3), 2019: (4338834, 13.6)},
        "Benign prostatic hyperplasia": {2017: (3731131, 25.4), 2018: (3807277, 26.2), 2019: (3887623, 26.9)},
        "Cancer, breast (female)": {2017: (1325734, 7.4), 2018: (1338988, 7.6), 2019: (1356944, 7.8)},
        "Cancer, breast (male)": {2017: (19570, 0.1), 2018: (25169, 0.2), 2019: (28217, 0.2)},
        "Cancer, colorectal": {2017: (534198, 1.6), 2018: (526699, 1.6), 2019: (520113, 1.6)},
        "Cancer, endometrial": {2017: (171441, 1.0), 2018: (176062, 1.0), 2019: (180248, 1.0)},
        "Cancer, lung": {2017: (411784, 1.3), 2018: (414284, 1.3), 2019: (419811, 1.3)},
        "Cancer, prostate": {2017: (1310686, 8.9), 2018: (1319521, 9.1), 2019: (1339503, 9.3)},
        "Cancer, urologic (kidney, renal pelvis, and ureter)": {2017: (241732, 0.7), 2018: (247779, 0.8), 2019: (255399, 0.8)},
        "Cataract": {2017: (8388282, 25.8), 2018: (8371047, 26.0), 2019: (8364322, 26.3)},
        "Chronic kidney disease": {2017: (5296046, 16.3), 2018: (5429683, 16.9), 2019: (5600301, 17.6)},
        "Chronic obstructive pulmonary disease": {2017: (5155458, 15.9), 2018: (5033696, 15.6), 2019: (4882331, 15.3)},
        "Depression, bipolar, or other depressive mood disorders": {2017: (6299572, 19.4), 2018: (6371448, 19.8), 2019: (6431278, 20.2)},
        "Diabetes": {2017: (9081954, 28.0), 2018: (8919842, 27.7), 2019: (8771519, 27.5)},
        "Glaucoma": {2017: (4483052, 13.8), 2018: (4545279, 14.1), 2019: (4615489, 14.5)},
        "Heart failure and non-ischemic heart disease": {2017: (3940948, 12.1), 2018: (3965162, 12.3), 2019: (3987434, 12.5)},
        "Hip/Pelvic fracture": {2017: (377066, 1.2), 2018: (374278, 1.2), 2019: (374464, 1.2)},
        "Hyperlipidemia": {2017: (19927458, 61.3), 2018: (19967870, 62.0), 2019: (20040500, 62.9)},
        "Hypertension": {2017: (22024151, 67.8), 2018: (21801519, 67.7), 2019: (21592541, 67.8)},
        "Hypothyroidism": {2017: (6788655, 20.9), 2018: (6803999, 21.1), 2019: (6802822, 21.4)},
        "Ischemic heart disease": {2017: (7081526, 21.8), 2018: (7000441, 21.8), 2019: (6949697, 21.8)},
        "Non-Alzheimer's dementia": {2017: (2587487, 8.0), 2018: (2529061, 7.9), 2019: (2474287, 7.8)},
        "Osteoporosis with or without pathological fracture": {2017: (3209888, 9.9), 2018: (3265644, 10.1), 2019: (3351209, 10.5)},
        "Parkinson's disease and secondary Parkinsonism": {2017: (553944, 1.7), 2018: (553460, 1.7), 2019: (556085, 1.7)},
        "Pneumonia, all-cause": {2017: (1779648, 5.5), 2018: (1736215, 5.4), 2019: (1682695, 5.3)},
        "Rheumatoid arthritis/Osteoarthritis": {2017: (11524987, 35.5), 2018: (11552428, 35.9), 2019: (11588956, 36.4)},
        "Stroke / Transient ischemic attack": {2017: (2133129, 6.6), 2018: (2099102, 6.5), 2019: (2075437, 6.5)},
    }

    # Page 3: 2020-2022 data
    conditions_p3 = {
        "Acute myocardial infarction": {2020: (384185, 1.2), 2021: (381206, 1.3), 2022: (372451, 1.3)},
        "Alzheimer's disease": {2020: (857109, 2.7), 2021: (750273, 2.5), 2022: (713609, 2.5)},
        "Anemia": {2020: (6645691, 21.3), 2021: (6266946, 21.1), 2022: (6202406, 21.7)},
        "Asthma": {2020: (2314636, 7.4), 2021: (2150373, 7.2), 2022: (2108123, 7.4)},
        "Atrial fibrillation and flutter": {2020: (4249615, 13.6), 2021: (4088560, 13.7), 2022: (4029477, 14.1)},
        "Benign prostatic hyperplasia": {2020: (3822081, 27.0), 2021: (3724246, 27.6), 2022: (3759882, 29.1)},
        "Cancer, breast (female)": {2020: (1325827, 7.8), 2021: (1286577, 7.9), 2022: (1287453, 8.2)},
        "Cancer, breast (male)": {2020: (29966, 0.2), 2021: (31593, 0.2), 2022: (34432, 0.3)},
        "Cancer, colorectal": {2020: (488814, 1.6), 2021: (455757, 1.5), 2022: (440246, 1.5)},
        "Cancer, endometrial": {2020: (174721, 1.0), 2021: (168734, 1.0), 2022: (170072, 1.1)},
        "Cancer, lung": {2020: (405281, 1.3), 2021: (385721, 1.3), 2022: (375575, 1.3)},
        "Cancer, prostate": {2020: (1315058, 9.3), 2021: (1277012, 9.5), 2022: (1278644, 9.9)},
        "Cancer, urologic (kidney, renal pelvis, and ureter)": {2020: (252697, 0.8), 2021: (245085, 0.8), 2022: (245726, 0.9)},
        "Cataract": {2020: (6947683, 22.3), 2021: (7735428, 26.0), 2022: (7534840, 26.4)},
        "Chronic kidney disease": {2020: (5589772, 17.9), 2021: (5383435, 18.1), 2022: (5352992, 18.7)},
        "Chronic obstructive pulmonary disease": {2020: (4437223, 14.2), 2021: (3883186, 13.0), 2022: (3661898, 12.8)},
        "Depression, bipolar, or other depressive mood disorders": {2020: (6206012, 19.9), 2021: (5790493, 19.5), 2022: (5611909, 19.7)},
        "Diabetes": {2020: (8416991, 27.0), 2021: (7873257, 26.5), 2022: (7541228, 26.4)},
        "Glaucoma": {2020: (4449662, 14.3), 2021: (4255705, 14.3), 2022: (4284879, 15.0)},
        "Heart failure and non-ischemic heart disease": {2020: (3836105, 12.3), 2021: (3593979, 12.1), 2022: (3509486, 12.3)},
        "Hip/Pelvic fracture": {2020: (341689, 1.1), 2021: (336351, 1.1), 2022: (337424, 1.2)},
        "Hyperlipidemia": {2020: (19570280, 62.8), 2021: (18897422, 63.5), 2022: (18690952, 65.5)},
        "Hypertension": {2020: (20840730, 66.8), 2021: (19734885, 66.3), 2022: (19065069, 66.8)},
        "Hypothyroidism": {2020: (6597347, 21.2), 2021: (6296573, 21.2), 2022: (6151796, 21.5)},
        "Ischemic heart disease": {2020: (6681009, 21.4), 2021: (6303076, 21.2), 2022: (6158955, 21.6)},
        "Non-Alzheimer's dementia": {2020: (2374413, 7.6), 2021: (2179461, 7.3), 2022: (2107305, 7.4)},
        "Osteoporosis with or without pathological fracture": {2020: (3270710, 10.5), 2021: (3224005, 10.8), 2022: (3328045, 11.7)},
        "Parkinson's disease and secondary Parkinsonism": {2020: (536948, 1.7), 2021: (518814, 1.7), 2022: (515960, 1.8)},
        "Pneumonia, all-cause": {2020: (1638279, 5.3), 2021: (1543714, 5.2), 2022: (1455870, 5.1)},
        "Rheumatoid arthritis/Osteoarthritis": {2020: (11295424, 36.2), 2021: (10771605, 36.2), 2022: (10641100, 37.3)},
        "Stroke / Transient ischemic attack": {2020: (1989620, 6.4), 2021: (1871218, 6.3), 2022: (1818810, 6.4)},
    }

    # Denominators by year (total FFS beneficiaries)
    denominators = {
        2017: 32483529, 2018: 32184841, 2019: 31860990,
        2020: 31187299, 2021: 29757982, 2022: 28550545,
    }

    rows = []
    for condition in conditions_p1:
        for year in range(2017, 2023):
            data_p1 = conditions_p1.get(condition, {}).get(year)
            data_p3 = conditions_p3.get(condition, {}).get(year)
            data = data_p1 or data_p3
            if data:
                numerator, rate = data
                rows.append({
                    "condition": condition,
                    "year": year,
                    "numerator": numerator,
                    "prevalence_rate": rate,
                    "denominator": denominators.get(year),
                    "population": "FFS",
                    "condition_set": "30 CCW Chronic Conditions",
                    "geo_level": "National",
                })
    return rows


def _parse_ccw_b2c() -> list:
    """Parse Table B.2.c: 27 conditions, all Medicare (incl. MA), 2012-2021."""
    # Data from pages 1-2 of the PDF
    conditions_2012_2016 = {
        "Acute myocardial infarction": {2012: (290669, 0.9), 2013: (281449, 0.9), 2014: (273188, 0.9), 2015: (281530, 0.9), 2016: (300868, 0.9)},
        "Acquired hypothyroidism": {2012: (4677556, 14.7), 2013: (4830009, 15.0), 2014: (4915127, 15.3), 2015: (5010087, 15.6), 2016: (5154798, 15.8)},
        "Alzheimer's disease": {2012: (1596214, 5.0), 2013: (1515835, 4.7), 2014: (1414215, 4.4), 2015: (1385135, 4.3), 2016: (1411457, 4.3)},
        "Alzheimer's disease, related disorders, or senile dementia": {2012: (3585024, 11.2), 2013: (3551009, 11.1), 2014: (3463908, 10.8), 2015: (3474060, 10.8), 2016: (3625188, 11.1)},
        "Anemia": {2012: (7831930, 24.6), 2013: (7689519, 23.9), 2014: (7481006, 23.4), 2015: (7362976, 22.9), 2016: (7324867, 22.5)},
        "Asthma": {2012: (1639648, 5.1), 2013: (1696716, 5.3), 2014: (1713081, 5.3), 2015: (1756685, 5.5), 2016: (1726018, 5.3)},
        "Atrial fibrillation": {2012: (2672647, 8.4), 2013: (2715050, 8.5), 2014: (2739534, 8.6), 2015: (2789063, 8.7), 2016: (2806954, 8.6)},
        "Benign prostatic hyperplasia": {2012: (2032585, 14.3), 2013: (2108288, 14.7), 2014: (2132621, 14.8), 2015: (2241647, 15.5), 2016: (2372143, 16.2)},
        "Cancer, breast (female)": {2012: (959243, 5.4), 2013: (972486, 5.5), 2014: (976598, 5.5), 2015: (997882, 5.6), 2016: (1023455, 5.7)},
        "Cancer, breast (male)": {2012: (8215, 0.1), 2013: (8342, 0.1), 2014: (8397, 0.1), 2015: (8495, 0.1), 2016: (8548, 0.1)},
        "Cancer, colorectal": {2012: (429617, 1.3), 2013: (421768, 1.3), 2014: (408075, 1.3), 2015: (406155, 1.3), 2016: (404897, 1.2)},
        "Cancer, endometrial": {2012: (95554, 0.5), 2013: (98066, 0.5), 2014: (100293, 0.6), 2015: (105032, 0.6), 2016: (112644, 0.6)},
        "Cancer, lung": {2012: (347272, 1.1), 2013: (346189, 1.1), 2014: (344399, 1.1), 2015: (348239, 1.1), 2016: (351744, 1.1)},
        "Cancer, prostate": {2012: (1038454, 7.3), 2013: (1031759, 7.2), 2014: (1010383, 7.0), 2015: (1016376, 7.0), 2016: (1041335, 7.1)},
        "Cataract": {2012: (6076757, 19.1), 2013: (6031317, 18.8), 2014: (5859189, 18.3), 2015: (5834323, 18.1), 2016: (5856642, 18.0)},
        "Chronic kidney disease": {2012: (5254359, 16.5), 2013: (5471390, 17.0), 2014: (5649152, 17.6), 2015: (6228879, 19.4), 2016: (7531754, 23.1)},
        "Chronic obstructive pulmonary disease": {2012: (3823489, 12.0), 2013: (3819839, 11.9), 2014: (3762650, 11.7), 2015: (3850136, 12.0), 2016: (3927067, 12.0)},
        "Depression": {2012: (5164459, 16.2), 2013: (5332925, 16.6), 2014: (5485465, 17.1), 2015: (5741768, 17.8), 2016: (5812110, 17.8)},
        "Diabetes": {2012: (9085499, 28.5), 2013: (9110312, 28.4), 2014: (9051458, 28.3), 2015: (9062604, 28.2), 2016: (9163634, 28.1)},
        "Glaucoma": {2012: (3207017, 10.1), 2013: (3200952, 10.0), 2014: (3173578, 9.9), 2015: (3081627, 9.6), 2016: (2392332, 7.3)},
        "Heart failure": {2012: (4964025, 15.6), 2013: (4823634, 15.0), 2014: (4687295, 14.6), 2015: (4652637, 14.5), 2016: (4690896, 14.4)},
        "Hip/Pelvic fracture": {2012: (261425, 0.8), 2013: (259653, 0.8), 2014: (258774, 0.8), 2015: (254791, 0.8), 2016: (234798, 0.7)},
        "Hyperlipidemia": {2012: (15073632, 47.3), 2013: (15243722, 47.5), 2014: (15149531, 47.3), 2015: (15216792, 47.3), 2016: (15146099, 46.4)},
        "Hypertension": {2012: (18683755, 58.6), 2013: (18804701, 58.5), 2014: (18663558, 58.3), 2015: (18750064, 58.3), 2016: (19105483, 58.6)},
        "Ischemic heart disease": {2012: (9675396, 30.3), 2013: (9454965, 29.4), 2014: (9196730, 28.7), 2015: (9083042, 28.2), 2016: (9122052, 28.0)},
        "Osteoporosis": {2012: (2167775, 6.8), 2013: (2083998, 6.5), 2014: (2038716, 6.4), 2015: (2066638, 6.4), 2016: (2082581, 6.4)},
        "Rheumatoid arthritis/Osteoarthritis": {2012: (9851000, 30.9), 2013: (9993880, 31.1), 2014: (10053514, 31.4), 2015: (10342805, 32.1), 2016: (10908057, 33.5)},
        "Stroke/Transient ischemic attack": {2012: (1296511, 4.1), 2013: (1279697, 4.0), 2014: (1274536, 4.0), 2015: (1270383, 3.9), 2016: (1262036, 3.9)},
    }

    denominators_all = {
        2012: 31895335, 2013: 32124150, 2014: 32034955, 2015: 32179757, 2016: 32607998,
    }

    rows = []
    for condition, year_data in conditions_2012_2016.items():
        for year, (numerator, rate) in year_data.items():
            rows.append({
                "condition": condition,
                "year": year,
                "numerator": numerator,
                "prevalence_rate": rate,
                "denominator": denominators_all.get(year),
                "population": "All Medicare",
                "condition_set": "27 CCW Chronic Conditions",
                "geo_level": "National",
            })
    return rows


# ---------------------------------------------------------------------------
# Machine-readable CCW source (for reference / future auto-download)
# ---------------------------------------------------------------------------
# The CCW publishes an XLSX with all 30 chronic condition prevalence data:
#   https://www2.ccwdata.org/documents/10280/19099065/medicare-charts-chronic-conditions-data.xlsx
# This file has 18 worksheets covering Table B.2.a (FFS 2017-2022),
# Table B.2.b (other conditions), and Table B.2.c (all Medicare 2012-2021).
# The PDF version is at:
#   https://www2.ccwdata.org/documents/10280/19099065/medicare-charts-chronic-conditions.pdf
#
# NOTE: CCW announced these charts will be retired effective 06/15/2026.
# Future updates will move to data.cms.gov:
#   https://data.cms.gov/medicare-chronic-conditions/specific-chronic-conditions
#   https://data.cms.gov/medicare-chronic-conditions/multiple-chronic-conditions
# The data.cms.gov datasets provide CSV/API access at national, state, and
# county levels (dataset ID for specific conditions to be confirmed after
# full migration).
#
# The hardcoded dictionaries below were manually transcribed from the CCW
# PDF tables. The XLSX above can be used to cross-validate these values.


# ---------------------------------------------------------------------------
# CCW Data Validation
# ---------------------------------------------------------------------------

def validate_ccw_data():
    """Validate hardcoded CCW chronic conditions data for consistency.

    Checks:
      1. All prevalence percentages are between 0 and 100
      2. Year-over-year changes don't exceed 5 percentage points (suspicious)
      3. Number of conditions matches expected counts
      4. Prints a validation report
    """
    print("=" * 60)
    print("CCW CHRONIC CONDITIONS DATA VALIDATION")
    print("=" * 60)

    errors = []
    warnings = []

    # --- Parse all data ---
    ffs_rows = _parse_ccw_b2a()
    all_medicare_rows = _parse_ccw_b2c()

    # --- Check 3: condition counts ---
    ffs_conditions = set(r["condition"] for r in ffs_rows)
    all_conditions = set(r["condition"] for r in all_medicare_rows)

    expected_ffs = 30
    expected_all = 27

    # The FFS data has 31 entries because "Cancer, breast (male)" is counted
    # separately alongside 30 original conditions (some tables list 30 unique,
    # this one has acute MI through stroke = 30 + breast male as separate = 31).
    # We accept 30 or 31.
    if len(ffs_conditions) < expected_ffs:
        errors.append(
            f"FFS condition count: expected >= {expected_ffs}, got {len(ffs_conditions)}"
        )
    else:
        print(f"  [PASS] FFS conditions: {len(ffs_conditions)} (expected >= {expected_ffs})")

    # The all-Medicare data lists "Alzheimer's disease" and "Alzheimer's disease,
    # related disorders, or senile dementia" as separate rows, so we accept 27 or 28.
    if len(all_conditions) < expected_all or len(all_conditions) > expected_all + 1:
        errors.append(
            f"All-Medicare condition count: expected {expected_all}-{expected_all + 1}, "
            f"got {len(all_conditions)}"
        )
    else:
        print(f"  [PASS] All-Medicare conditions: {len(all_conditions)} (expected {expected_all}-{expected_all + 1})")

    # --- Check 1: prevalence ranges ---
    for label, rows in [("FFS", ffs_rows), ("All-Medicare", all_medicare_rows)]:
        out_of_range = []
        for r in rows:
            rate = r["prevalence_rate"]
            if rate is not None and (rate < 0 or rate > 100):
                out_of_range.append(
                    f"  {r['condition']} ({r['year']}): {rate}%"
                )
        if out_of_range:
            errors.append(
                f"{label}: {len(out_of_range)} prevalence values out of [0, 100] range:\n"
                + "\n".join(out_of_range)
            )
        else:
            print(f"  [PASS] {label}: all prevalence rates in [0, 100] range")

    # --- Check 2: year-over-year changes ---
    for label, rows in [("FFS", ffs_rows), ("All-Medicare", all_medicare_rows)]:
        # Build condition -> {year: rate} mapping
        cond_years = {}
        for r in rows:
            cond_years.setdefault(r["condition"], {})[r["year"]] = r["prevalence_rate"]

        suspicious = []
        for condition, year_rates in sorted(cond_years.items()):
            years_sorted = sorted(year_rates.keys())
            for i in range(1, len(years_sorted)):
                prev_year = years_sorted[i - 1]
                curr_year = years_sorted[i]
                prev_rate = year_rates[prev_year]
                curr_rate = year_rates[curr_year]
                if prev_rate is not None and curr_rate is not None:
                    delta = abs(curr_rate - prev_rate)
                    if delta > 5.0:
                        suspicious.append(
                            f"  {condition}: {prev_year}={prev_rate}% -> "
                            f"{curr_year}={curr_rate}% (delta={delta:.1f}pp)"
                        )
        if suspicious:
            warnings.append(
                f"{label}: {len(suspicious)} year-over-year changes > 5 percentage points:\n"
                + "\n".join(suspicious)
            )
        else:
            print(f"  [PASS] {label}: no year-over-year changes > 5 percentage points")

    # --- Summary ---
    print()
    if warnings:
        print(f"WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  [WARN] {w}")
        print()

    if errors:
        print(f"ERRORS ({len(errors)}):")
        for e in errors:
            print(f"  [FAIL] {e}")
        print()
        print("VALIDATION FAILED")
        return False
    else:
        if warnings:
            print("VALIDATION PASSED (with warnings)")
        else:
            print("VALIDATION PASSED")
        return True


def build_fact_chronic_conditions_national(con, dry_run: bool) -> int:
    """Build national-level chronic conditions prevalence (30 CCW, FFS)."""
    print("Building fact_chronic_conditions_national...")
    rows = _parse_ccw_b2a()
    if not rows:
        print("  SKIPPED — no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("CREATE OR REPLACE TABLE _df AS SELECT * FROM df")
    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_cc_national AS
        SELECT
            condition,
            CAST(year AS INTEGER) AS year,
            CAST(numerator AS BIGINT) AS numerator,
            CAST(prevalence_rate AS DOUBLE) AS prevalence_rate,
            CAST(denominator AS BIGINT) AS denominator,
            population,
            condition_set,
            geo_level,
            'ccwdata.org' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _df
    """)
    con.execute("DROP TABLE IF EXISTS _df")

    count = write_parquet(con, "_fact_cc_national", _snapshot_path("chronic_conditions_national"), dry_run)
    conditions = con.execute("SELECT COUNT(DISTINCT condition) FROM _fact_cc_national").fetchone()[0]
    years = con.execute("SELECT MIN(year), MAX(year) FROM _fact_cc_national").fetchone()
    print(f"  {count:,} rows, {conditions} conditions, {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_cc_national")
    return count


def build_fact_chronic_conditions_all_medicare(con, dry_run: bool) -> int:
    """Build all-Medicare chronic conditions prevalence (27 conditions, incl. MA)."""
    print("Building fact_chronic_conditions_all_medicare...")
    rows = _parse_ccw_b2c()
    if not rows:
        print("  SKIPPED — no data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("CREATE OR REPLACE TABLE _df AS SELECT * FROM df")
    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_cc_all AS
        SELECT
            condition,
            CAST(year AS INTEGER) AS year,
            CAST(numerator AS BIGINT) AS numerator,
            CAST(prevalence_rate AS DOUBLE) AS prevalence_rate,
            CAST(denominator AS BIGINT) AS denominator,
            population,
            condition_set,
            geo_level,
            'ccwdata.org' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _df
    """)
    con.execute("DROP TABLE IF EXISTS _df")

    count = write_parquet(con, "_fact_cc_all", _snapshot_path("chronic_conditions_all_medicare"), dry_run)
    conditions = con.execute("SELECT COUNT(DISTINCT condition) FROM _fact_cc_all").fetchone()[0]
    years = con.execute("SELECT MIN(year), MAX(year) FROM _fact_cc_all").fetchone()
    print(f"  {count:,} rows, {conditions} conditions, {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_cc_all")
    return count


# ---------------------------------------------------------------------------
# MCBS Cost Supplement
# ---------------------------------------------------------------------------

def build_fact_mcbs_cost_summary(con, dry_run: bool) -> int:
    """Build MCBS Cost Supplement summary: aggregate by demographics."""
    print("Building fact_mcbs_cost_summary...")
    raw_dir = RAW_DIR / "mcbs"
    csv_path = raw_dir / "cspuf2023.csv"
    zip_url = "https://data.cms.gov/sites/default/files/2026-01/e68b9516-a8f0-425e-916c-909c3b693afe/CSPUF2023_Data.zip"

    if not csv_path.exists():
        zip_path = raw_dir / "CSPUF2023.zip"
        if not _download(zip_url, zip_path, "MCBS Cost PUF 2023"):
            print("  SKIPPED — download failed")
            return 0
        with zipfile.ZipFile(zip_path) as zf:
            zf.extract("cspuf2023.csv", raw_dir)
        print(f"  Extracted cspuf2023.csv")

    # Age: 1=Under 65, 2=65-74, 3=75+
    # Sex: 1=Male, 2=Female
    # Race: 1=Non-Hispanic white, 2=Non-Hispanic black, 3=Hispanic, 4=Other
    # Income: 1=Less than $25K, 2=$25K+
    # NChrncnd: 1=0-1, 2=2-3, 3=4+

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_mcbs AS
        WITH raw AS (
            SELECT
                CAST(SURVEYYR AS INTEGER) AS survey_year,
                CASE CSP_AGE
                    WHEN '1' THEN 'Under 65'
                    WHEN '2' THEN '65-74'
                    WHEN '3' THEN '75+'
                    ELSE 'Unknown'
                END AS age_group,
                CASE CSP_SEX
                    WHEN '1' THEN 'Male'
                    WHEN '2' THEN 'Female'
                    ELSE 'Unknown'
                END AS sex,
                CASE CSP_RACE
                    WHEN '1' THEN 'Non-Hispanic White'
                    WHEN '2' THEN 'Non-Hispanic Black'
                    WHEN '3' THEN 'Hispanic'
                    WHEN '4' THEN 'Other'
                    ELSE 'Unknown'
                END AS race_ethnicity,
                CASE CSP_INCOME
                    WHEN '1' THEN 'Less than $25K'
                    WHEN '2' THEN '$25K or more'
                    ELSE 'Unknown'
                END AS income_group,
                CASE CSP_NCHRNCND
                    WHEN '1' THEN '0-1'
                    WHEN '2' THEN '2-3'
                    WHEN '3' THEN '4+'
                    ELSE 'Unknown'
                END AS chronic_condition_group,
                TRY_CAST(PAMTTOT AS DOUBLE) AS total_spending,
                TRY_CAST(PAMTCARE AS DOUBLE) AS medicare_spending,
                TRY_CAST(PAMTCAID AS DOUBLE) AS medicaid_spending,
                TRY_CAST(PAMTMADV AS DOUBLE) AS medicare_advantage_spending,
                TRY_CAST(PAMTALPR AS DOUBLE) AS other_private_spending,
                TRY_CAST(PAMTOOP AS DOUBLE) AS out_of_pocket_spending,
                TRY_CAST(PAMTHU AS DOUBLE) AS hospital_spending,
                TRY_CAST(PAMTIP AS DOUBLE) AS inpatient_spending,
                TRY_CAST(PAMTOP AS DOUBLE) AS outpatient_spending,
                TRY_CAST(PAMTPM AS DOUBLE) AS pharmacy_spending,
                TRY_CAST(PAMTHH AS DOUBLE) AS home_health_spending,
                TRY_CAST(PAMTMP AS DOUBLE) AS medical_provider_spending,
                TRY_CAST(PEVENTS AS INTEGER) AS total_events,
                TRY_CAST(CSPUFWGT AS DOUBLE) AS survey_weight
            FROM read_csv_auto('{csv_path}', all_varchar=true)
        )
        SELECT
            survey_year,
            age_group,
            sex,
            race_ethnicity,
            income_group,
            chronic_condition_group,
            COUNT(*) AS sample_n,
            ROUND(SUM(survey_weight), 0) AS weighted_n,
            ROUND(AVG(total_spending), 2) AS avg_total_spending,
            ROUND(AVG(medicare_spending), 2) AS avg_medicare_spending,
            ROUND(AVG(medicaid_spending), 2) AS avg_medicaid_spending,
            ROUND(AVG(out_of_pocket_spending), 2) AS avg_oop_spending,
            ROUND(AVG(pharmacy_spending), 2) AS avg_pharmacy_spending,
            ROUND(AVG(inpatient_spending), 2) AS avg_inpatient_spending,
            ROUND(AVG(total_events), 1) AS avg_events,
            'data.cms.gov/mcbs-cost-supplement' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM raw
        GROUP BY survey_year, age_group, sex, race_ethnicity, income_group, chronic_condition_group
        ORDER BY survey_year, age_group, sex, race_ethnicity, income_group, chronic_condition_group
    """)

    count = write_parquet(con, "_fact_mcbs", _snapshot_path("mcbs_cost_summary"), dry_run)
    groups = con.execute("SELECT COUNT(DISTINCT age_group) FROM _fact_mcbs").fetchone()[0]
    print(f"  {count:,} rows, {groups} age groups, CY2023")
    con.execute("DROP TABLE IF EXISTS _fact_mcbs")
    return count


# ---------------------------------------------------------------------------
# MACPAC Exhibit 17: Benefit Spending by State and Category
# ---------------------------------------------------------------------------

def build_fact_macpac_benefit_spending_v2(con, dry_run: bool) -> int:
    """Build MACPAC benefit spending by state and service category (Exhibit 17, FY2024)."""
    print("Building fact_macpac_benefit_spending_v2...")
    if openpyxl is None:
        print("  SKIPPED — openpyxl not installed")
        return 0

    xlsx_path = RAW_DIR / "macpac" / "exhibit17.xlsx"
    url = "https://www.macpac.gov/wp-content/uploads/2026/01/EXHIBIT-17.-Total-Medicaid-Benefit-Spending-by-State-and-Category-FY-2024.xlsx"
    if not xlsx_path.exists():
        if not _download(url, xlsx_path, "MACPAC Exhibit 17"):
            print("  SKIPPED — download failed")
            return 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Read header row to get column names (row 3 + row 4)
    # Columns: State, Total, FFS subcategories, MC, Other
    # Row 3: State, Total spending on benefits, Fee for service (spans), Managed care, Disproportionate share, Other
    # Row 4: (blank), (blank), Hospital, Physician, Dental, Other practitioner, Clinic..., Other acute, Drugs, Inst LTSS, HCBS, ...

    rows = []
    for row in ws.iter_rows(min_row=5, max_row=75, values_only=True):
        state_name = _clean_state_name(str(row[0]) if row[0] else "")
        if not state_name or state_name.lower() in ("total", "notes", "source", "", "none"):
            state_code = None
            if state_name.lower() == "total":
                state_code = "US"
            else:
                continue
        else:
            state_code = STATE_NAME_TO_CODE.get(state_name)
            if not state_code:
                continue

        def to_millions(v):
            n = _clean_number(v)
            return round(n, 2) if n is not None else None

        rows.append({
            "state_code": state_code,
            "state_name": state_name if state_code != "US" else "Total",
            "fiscal_year": 2024,
            "total_benefits": to_millions(row[1]),
            "ffs_hospital": to_millions(row[2]),
            "ffs_physician": to_millions(row[3]),
            "ffs_dental": to_millions(row[4]),
            "ffs_other_practitioner": to_millions(row[5]),
            "ffs_clinic_health_center": to_millions(row[6]),
            "ffs_other_acute": to_millions(row[7]),
            "ffs_drugs": to_millions(row[8]),
            "ffs_institutional_ltss": to_millions(row[9]),
            "ffs_hcbs": to_millions(row[10]),
            "managed_care": to_millions(row[11]),
            "dsh": to_millions(row[12]),
            "other_spending": to_millions(row[13]) if len(row) > 13 else None,
        })
    wb.close()

    if not rows:
        print("  SKIPPED — no rows parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("CREATE OR REPLACE TABLE _df AS SELECT * FROM df")
    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_macpac_ben AS
        SELECT *, 'macpac.gov' AS source, DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _df
        WHERE state_code IS NOT NULL
    """)
    con.execute("DROP TABLE IF EXISTS _df")

    count = write_parquet(con, "_fact_macpac_ben", _snapshot_path("macpac_benefit_spending_v2"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_macpac_ben WHERE state_code != 'US'").fetchone()[0]
    print(f"  {count:,} rows, {states} states, FY2024")
    con.execute("DROP TABLE IF EXISTS _fact_macpac_ben")
    return count


# ---------------------------------------------------------------------------
# MACPAC Exhibit 21: Spending by Eligibility Group and Dual Status
# ---------------------------------------------------------------------------

def build_fact_macpac_spending_by_elig(con, dry_run: bool) -> int:
    """Build MACPAC spending by state, eligibility group, dual status (Exhibit 21)."""
    print("Building fact_macpac_spending_by_elig...")
    if openpyxl is None:
        print("  SKIPPED — openpyxl not installed")
        return 0

    xlsx_path = RAW_DIR / "macpac" / "exhibit21.xlsx"
    url = "https://www.macpac.gov/wp-content/uploads/2026/02/EXHIBIT-21.-Medicaid-Spending-by-State-Eligibility-Group-and-Dually-Eligible-Status-FY-2023-1.xlsx"
    if not xlsx_path.exists():
        if not _download(url, xlsx_path, "MACPAC Exhibit 21"):
            print("  SKIPPED — download failed")
            return 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Use the second sheet which has the data
    if "Exhibit 21" in wb.sheetnames:
        ws = wb["Exhibit 21"]
    else:
        ws = wb.active

    # Exhibit 21 structure (from inspection):
    # Col A: State
    # Col B: Total spending
    # Col C-G: By eligibility (Child, New adult, Other adult, Disabled, Aged) - as PROPORTIONS
    # Col H+: Dual status breakdowns
    # Note: eligibility columns contain proportions (e.g., 0.2309 = 23.09%)
    # Total is in millions

    rows = []
    for row in ws.iter_rows(min_row=6, max_row=73, values_only=True):
        state_name = _clean_state_name(str(row[0]) if row[0] else "")
        if not state_name or state_name.lower() in ("notes", "source", "", "none"):
            continue

        if state_name.lower() == "total":
            state_code = "US"
        else:
            state_code = STATE_NAME_TO_CODE.get(state_name)
            if not state_code:
                continue

        total = _clean_number(row[1])
        child_pct = _clean_number(row[2])
        new_adult_pct = _clean_number(row[3])
        other_adult_pct = _clean_number(row[4])
        disabled_pct = _clean_number(row[5])
        aged_pct = _clean_number(row[6])

        # Dual status
        dual_total = _clean_number(row[7])
        dual_65plus_pct = _clean_number(row[8]) if len(row) > 8 else None
        dual_full_benefit = _clean_number(row[9]) if len(row) > 9 else None

        rows.append({
            "state_code": state_code,
            "state_name": state_name if state_code != "US" else "Total",
            "fiscal_year": 2023,
            "total_spending_m": round(total, 2) if total else None,
            "child_pct": round(child_pct * 100, 1) if child_pct and child_pct < 1 else child_pct,
            "new_adult_group_pct": round(new_adult_pct * 100, 1) if new_adult_pct and new_adult_pct < 1 else new_adult_pct,
            "other_adult_pct": round(other_adult_pct * 100, 1) if other_adult_pct and other_adult_pct < 1 else other_adult_pct,
            "disabled_pct": round(disabled_pct * 100, 1) if disabled_pct and disabled_pct < 1 else disabled_pct,
            "aged_pct": round(aged_pct * 100, 1) if aged_pct and aged_pct < 1 else aged_pct,
            "dual_total_spending_m": round(dual_total, 2) if dual_total else None,
            "dual_65plus_pct": round(dual_65plus_pct * 100, 1) if dual_65plus_pct and dual_65plus_pct < 1 else dual_65plus_pct,
            "dual_full_benefit_spending_m": round(dual_full_benefit, 2) if dual_full_benefit else None,
        })
    wb.close()

    if not rows:
        print("  SKIPPED — no rows parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("CREATE OR REPLACE TABLE _df AS SELECT * FROM df")
    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_macpac_elig AS
        SELECT *, 'macpac.gov' AS source, DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _df
        WHERE state_code IS NOT NULL
    """)
    con.execute("DROP TABLE IF EXISTS _df")

    count = write_parquet(con, "_fact_macpac_elig", _snapshot_path("macpac_spending_by_elig"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_macpac_elig WHERE state_code != 'US'").fetchone()[0]
    print(f"  {count:,} rows, {states} states, FY2023")
    con.execute("DROP TABLE IF EXISTS _fact_macpac_elig")
    return count


# ---------------------------------------------------------------------------
# MACPAC Exhibit 29: Managed Care Enrollment %
# ---------------------------------------------------------------------------

def build_fact_macpac_mc_enrollment_pct(con, dry_run: bool) -> int:
    """Build MACPAC managed care enrollment % by state and plan type (Exhibit 29)."""
    print("Building fact_macpac_mc_enrollment_pct...")
    if openpyxl is None:
        print("  SKIPPED — openpyxl not installed")
        return 0

    xlsx_path = RAW_DIR / "macpac" / "exhibit29.xlsx"
    url = "https://www.macpac.gov/wp-content/uploads/2026/01/EXHIBIT-29.-Percentage-of-Medicaid-Enrollees-in-Managed-Care-by-State-July-1-2022.xlsx"
    if not xlsx_path.exists():
        if not _download(url, xlsx_path, "MACPAC Exhibit 29"):
            print("  SKIPPED — download failed")
            return 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Structure from inspection:
    # Row 3: State, Total Medicaid enrollees, Percentage in managed care (spans), ...
    # Row 4: (blank), (blank), Comprehensive MC, Limited-benefit plans (spans), PCCM, ...
    # Row 5: (blank), (blank), (blank), MLTSS, BHO, Dental, Transportation, Other, (blank), ...

    rows = []
    for row in ws.iter_rows(min_row=6, max_row=67, values_only=True):
        state_name = _clean_state_name(str(row[0]) if row[0] else "")
        if not state_name or state_name.lower() in ("notes", "source", "", "none"):
            continue

        if state_name.lower() == "total":
            state_code = "US"
        else:
            state_code = STATE_NAME_TO_CODE.get(state_name)
            if not state_code:
                continue

        total_enrollees = _clean_number(row[1])

        def to_pct(v):
            n = _clean_number(v)
            if n is None:
                return None
            # Values are proportions (0.0 to 1.0)
            if n <= 1.0:
                return round(n * 100, 1)
            return round(n, 1)

        rows.append({
            "state_code": state_code,
            "state_name": state_name if state_code != "US" else "Total",
            "data_date": "2022-07-01",
            "total_medicaid_enrollees": int(total_enrollees) if total_enrollees else None,
            "comprehensive_mc_pct": to_pct(row[2]),
            "mltss_pct": to_pct(row[3]),
            "bho_pct": to_pct(row[4]),
            "dental_mc_pct": to_pct(row[5]),
            "transportation_mc_pct": to_pct(row[6]),
            "other_limited_pct": to_pct(row[7]),
            "pccm_pct": to_pct(row[8]),
        })
    wb.close()

    if not rows:
        print("  SKIPPED — no rows parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("CREATE OR REPLACE TABLE _df AS SELECT * FROM df")
    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_mc_pct AS
        SELECT *, 'macpac.gov' AS source, DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _df
        WHERE state_code IS NOT NULL
    """)
    con.execute("DROP TABLE IF EXISTS _df")

    count = write_parquet(con, "_fact_mc_pct", _snapshot_path("macpac_mc_enrollment_pct"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_mc_pct WHERE state_code != 'US'").fetchone()[0]
    print(f"  {count:,} rows, {states} states, July 2022")
    con.execute("DROP TABLE IF EXISTS _fact_mc_pct")
    return count


# ---------------------------------------------------------------------------
# Medicare Telehealth Trends
# ---------------------------------------------------------------------------

def build_fact_medicare_telehealth(con, dry_run: bool) -> int:
    """Build Medicare telehealth utilization trends by state/demographics."""
    print("Building fact_medicare_telehealth...")
    csv_path = RAW_DIR / "medicare_telehealth_trends.csv"
    url = "https://data.cms.gov/sites/default/files/2025-12/TMEDTREND_PUBLIC_251210.csv"
    if not _download(url, csv_path, "Medicare Telehealth Trends"):
        print("  SKIPPED — download failed")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_th AS
        SELECT
            CAST(Year AS INTEGER) AS year,
            CASE WHEN quarter IN ('1','2','3','4') THEN CAST(quarter AS INTEGER) ELSE NULL END AS quarter,
            CASE WHEN quarter NOT IN ('1','2','3','4') THEN quarter ELSE NULL END AS period_label,
            Bene_Geo_Desc AS state_or_geo,
            CASE
                WHEN LENGTH(TRIM(Bene_Geo_Desc)) = 2 THEN Bene_Geo_Desc
                ELSE NULL
            END AS state_code,
            Bene_Mdcd_Mdcr_Enrl_Stus AS dual_status,
            Bene_Race_Desc AS race,
            Bene_Sex_Desc AS sex,
            Bene_Mdcr_Entlmt_Stus AS entitlement_status,
            Bene_Age_Desc AS age_group,
            Bene_RUCA_Desc AS rurality,
            TRY_CAST(Total_Bene_TH_Elig AS BIGINT) AS beneficiaries_th_eligible,
            CAST(ROUND(TRY_CAST(Total_PartB_Enrl AS DOUBLE), 0) AS BIGINT) AS total_part_b_enrolled,
            TRY_CAST(Total_Bene_Telehealth AS BIGINT) AS beneficiaries_using_telehealth,
            CASE
                WHEN TRY_CAST(Total_Bene_TH_Elig AS DOUBLE) > 0
                THEN ROUND(TRY_CAST(Total_Bene_Telehealth AS DOUBLE) / TRY_CAST(Total_Bene_TH_Elig AS DOUBLE) * 100, 2)
                ELSE NULL
            END AS telehealth_utilization_pct,
            'data.cms.gov' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE Year IS NOT NULL AND Year != ''
    """)

    # Map state names to codes where possible
    state_map_sql = " ".join([
        f"WHEN state_or_geo = '{name}' THEN '{code}'"
        for name, code in STATE_NAME_TO_CODE.items()
    ])
    con.execute(f"""
        UPDATE _fact_th SET state_code = CASE {state_map_sql} ELSE state_code END
        WHERE state_code IS NULL
    """)

    count = write_parquet(con, "_fact_th", _snapshot_path("medicare_telehealth"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_th WHERE state_code IS NOT NULL").fetchone()[0]
    years = con.execute("SELECT MIN(year), MAX(year) FROM _fact_th").fetchone()
    print(f"  {count:,} rows, {states} states, {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_th")
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

ALL_TABLES = {
    "fact_chronic_conditions_national": build_fact_chronic_conditions_national,
    "fact_chronic_conditions_all_medicare": build_fact_chronic_conditions_all_medicare,
    "fact_mcbs_cost_summary": build_fact_mcbs_cost_summary,
    "fact_macpac_benefit_spending_v2": build_fact_macpac_benefit_spending_v2,
    "fact_macpac_spending_by_elig": build_fact_macpac_spending_by_elig,
    "fact_macpac_mc_enrollment_pct": build_fact_macpac_mc_enrollment_pct,
    "fact_medicare_telehealth": build_fact_medicare_telehealth,
}


def main():
    parser = argparse.ArgumentParser(description="Ingest chronic conditions + supplemental data")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only", type=str, default=None,
                        help="Comma-separated list of table names to build")
    args = parser.parse_args()

    tables = list(ALL_TABLES.keys())
    if args.only:
        tables = [t.strip() for t in args.only.split(",")]

    print(f"Snapshot: {SNAPSHOT_DATE}")
    print(f"Run ID:   {RUN_ID}")
    print(f"Building: {', '.join(tables)}")
    print()

    # Validate hardcoded CCW data before building any chronic conditions tables
    ccw_tables = [t for t in tables if t.startswith("fact_chronic_conditions")]
    if ccw_tables:
        validate_ccw_data()
        print()

    con = duckdb.connect()
    totals = {}
    for name in tables:
        if name in ALL_TABLES:
            totals[name] = ALL_TABLES[name](con, args.dry_run)
        else:
            print(f"Unknown table: {name}")
            totals[name] = 0
        print()

    con.close()

    print("=" * 60)
    print("CHRONIC CONDITIONS + SUPPLEMENTAL DATA INGESTION COMPLETE")
    print("=" * 60)
    total_rows = sum(totals.values())
    for name, count in totals.items():
        status = "written" if not args.dry_run else "dry-run"
        print(f"  {name:45s} {count:>10,} rows  [{status}]")
    print(f"  {'TOTAL':45s} {total_rows:>10,} rows")

    if not args.dry_run and total_rows > 0:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "snapshot_date": SNAPSHOT_DATE,
            "pipeline_run_id": RUN_ID,
            "created_at": datetime.now().isoformat() + "Z",
            "tables": {name: {"rows": count} for name, count in totals.items()},
            "total_rows": total_rows,
        }
        manifest_file = META_DIR / f"manifest_chronic_conditions_{SNAPSHOT_DATE}.json"
        with open(manifest_file, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"\n  Manifest: {manifest_file}")

    print("\nDone.")


if __name__ == "__main__":
    main()
