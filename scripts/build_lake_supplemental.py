#!/usr/bin/env python3
"""
Build supplemental payment lake tables from:
  1. CMS-64 Financial Management Reports (FMR) — DSH, supplemental, GME line items per state
  2. MACPAC Exhibit 24 — State-level hospital supplemental payment summary

Produces:
  fact/fmr_supplemental/  — Per-state, per-category DSH & supplemental payment amounts (FY 2019-2024)
  fact/macpac_supplemental/ — MACPAC Exhibit 24 summary (DSH + non-DSH + 1115 waiver by state)

Data sources:
  - https://www.medicaid.gov/medicaid/financial-management/state-expenditure-reporting-for-medicaid-chip/expenditure-reports-mbescbes
  - https://www.macpac.gov/publication/medicaid-supplemental-payments-to-hospital-providers-by-state/
"""
import argparse
import os
import re
import sys
import zipfile
import tempfile
from datetime import date
from pathlib import Path

import duckdb
import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
DOWNLOAD_DIR = Path(tempfile.gettempdir()) / "aradune_supplemental"
SNAPSHOT_DATE = date.today().isoformat()

# State name → abbreviation mapping
STATE_ABBREVS = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Dist. Of Col.": "DC",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ",
    "New Mexico": "NM", "New York": "NY", "North Carolina": "NC",
    "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR",
    "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "Amer. Samoa": "AS", "N. Mariana Islands": "MP",
}

# Categories of interest from FMR — anything with DSH, Sup., or GME
SUPPLEMENTAL_KEYWORDS = ["DSH", "Sup. Payments", "Sup. Pay", "GME", "Supplemental"]

FMR_FISCAL_YEARS = [2024, 2023, 2022, 2021, 2020, 2019]

FMR_BASE_URL = "https://www.medicaid.gov/medicaid/financial-management/downloads/financial-management-report-fy{fy}.zip"


def download_file(url: str, dest: Path) -> bool:
    """Download a file if not already cached."""
    if dest.exists() and dest.stat().st_size > 1000:
        print(f"  Using cached: {dest.name}")
        return True
    import urllib.request
    print(f"  Downloading: {url}")
    try:
        urllib.request.urlretrieve(url, dest)
        return dest.exists() and dest.stat().st_size > 1000
    except Exception as e:
        print(f"  Download failed: {e}")
        return False


def parse_fmr_excel(xlsx_path: str, fiscal_year: int) -> list[dict]:
    """Parse a CMS-64 FMR Excel file and extract supplemental payment line items."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("MAP - "):
            continue
        state_name = sheet_name.replace("MAP - ", "").strip()
        if state_name == "National Totals":
            continue

        state_code = STATE_ABBREVS.get(state_name)
        if not state_code:
            # Try cleaning footnote numbers
            clean_name = "".join(c for c in state_name if not c.isdigit()).strip().rstrip(",")
            state_code = STATE_ABBREVS.get(clean_name)
            if not state_code:
                print(f"  Warning: Unknown state '{state_name}', skipping")
                continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=7, max_row=300, values_only=True):
            category = row[0]
            if not category or not isinstance(category, str):
                continue
            category = category.strip()

            # Check if this is a supplemental/DSH/GME line
            is_supplemental = any(kw in category for kw in SUPPLEMENTAL_KEYWORDS)
            if not is_supplemental:
                continue

            total_computable = row[1] if len(row) > 1 else None
            federal_share = row[2] if len(row) > 2 else None
            state_share = row[6] if len(row) > 6 else None

            # Classify the payment type
            if "DSH" in category:
                payment_type = "dsh"
            elif "GME" in category:
                payment_type = "gme"
            else:
                payment_type = "supplemental"

            # Classify the service category
            if "Inpatient Hospital" in category or "Critical Access Hospitals Inpatient" in category:
                service = "inpatient_hospital"
            elif "Outpatient Hospital" in category or "Critical Access Hospitals Outpatient" in category:
                service = "outpatient_hospital"
            elif "Mental Health" in category:
                service = "mental_health"
            elif "Nursing Facility" in category:
                service = "nursing_facility"
            elif "Intermediate Care" in category or "ICF" in category:
                service = "icf_iid"
            elif "Physician" in category or "Surgical" in category:
                service = "physician"
            elif "Clinic" in category:
                service = "clinic"
            elif "Other Practitioner" in category:
                service = "other_practitioner"
            elif "Transportation" in category:
                service = "transportation"
            else:
                service = "other"

            # Determine if this is Title XIX (T-), CHIP (C-), or regular
            if category.startswith("T-"):
                program = "title_xix_expansion"
            elif category.startswith("C-"):
                program = "chip"
            else:
                program = "medicaid"

            def safe_float(v):
                if v is None:
                    return 0.0
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return 0.0

            tc = safe_float(total_computable)
            fs = safe_float(federal_share)
            ss = safe_float(state_share)

            # Only include if there's any actual spending
            if tc != 0 or fs != 0 or ss != 0:
                rows.append({
                    "state_code": state_code,
                    "fiscal_year": fiscal_year,
                    "category": re.sub(r'^[TCM]-\s*', '', category).strip(),
                    "payment_type": payment_type,
                    "service": service,
                    "program": program,
                    "total_computable": tc,
                    "federal_share": fs,
                    "state_share": ss,
                    "source": "cms_fmr",
                })

    wb.close()
    return rows


def parse_macpac_exhibit24(xlsx_path: str, fiscal_year: int) -> list[dict]:
    """Parse MACPAC Exhibit 24 Excel file."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []

    for row in ws.iter_rows(min_row=4, max_row=60, values_only=True):
        state_raw = row[0]
        if not state_raw or not isinstance(state_raw, str):
            continue
        # Clean footnote markers (e.g., "California4" → "California")
        state_clean = "".join(c for c in state_raw if not c.isdigit()).strip().rstrip(",")
        if state_clean in ("Total", "Notes", "Source"):
            continue

        state_code = STATE_ABBREVS.get(state_clean)
        if not state_code:
            continue

        def safe_float(v):
            if v is None:
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        total_payments = safe_float(row[1])  # millions
        dsh = safe_float(row[2])
        non_dsh_supplemental = safe_float(row[3])
        sec_1115_waiver = safe_float(row[4])
        pct_supplemental = safe_float(row[5])

        rows.append({
            "state_code": state_code,
            "fiscal_year": fiscal_year,
            "total_hospital_payments_m": total_payments,
            "dsh_payments_m": dsh,
            "non_dsh_supplemental_m": non_dsh_supplemental,
            "sec_1115_waiver_m": sec_1115_waiver,
            "supplemental_pct": round(pct_supplemental * 100, 2) if pct_supplemental else None,
            "source": "macpac_exhibit24",
        })

    wb.close()
    return rows


def write_parquet(con: duckdb.DuckDBPyConnection, table_name: str, out_dir: Path, dry_run: bool) -> int:
    """Write a DuckDB table to partitioned Parquet."""
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if dry_run:
        print(f"  [DRY RUN] Would write {count:,} rows to {out_dir}")
        return count
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    print(f"  Wrote {count:,} rows to {out_path}")
    return count


def main():
    parser = argparse.ArgumentParser(description="Build supplemental payment lake tables")
    parser.add_argument("--dry-run", action="store_true", help="Parse but don't write Parquet")
    parser.add_argument("--tables", nargs="*", default=["fmr", "macpac"],
                        help="Which tables to build: fmr, macpac")
    args = parser.parse_args()

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect()
    total_rows = 0

    # ── FMR Supplemental Payments ──────────────────────────────────────
    if "fmr" in args.tables:
        print("\n=== CMS-64 FMR Supplemental Payments ===")
        all_fmr_rows = []

        for fy in FMR_FISCAL_YEARS:
            zip_path = DOWNLOAD_DIR / f"fmr_fy{fy}.zip"
            # Check /tmp first (may already be downloaded)
            tmp_path = Path(f"/tmp/fmr_fy{fy}.zip")
            if tmp_path.exists() and tmp_path.stat().st_size > 10000:
                zip_path = tmp_path
            else:
                url = FMR_BASE_URL.format(fy=fy)
                if not download_file(url, zip_path):
                    continue

            # Extract the NET EXPENDITURES xlsx
            with zipfile.ZipFile(zip_path) as zf:
                xlsx_names = [n for n in zf.namelist() if "NET EXPENDITURES" in n and n.endswith(".xlsx") and "CHIP" not in n]
                if not xlsx_names:
                    print(f"  FY {fy}: No NET EXPENDITURES xlsx found in zip")
                    continue
                xlsx_name = xlsx_names[0]
                extract_path = DOWNLOAD_DIR / f"fmr_fy{fy}.xlsx"
                with zf.open(xlsx_name) as src, open(extract_path, "wb") as dst:
                    dst.write(src.read())

            print(f"  Parsing FY {fy}...")
            rows = parse_fmr_excel(str(extract_path), fy)
            print(f"  FY {fy}: {len(rows):,} supplemental payment line items")
            all_fmr_rows.extend(rows)

        if all_fmr_rows:
            con.execute("CREATE TABLE _fmr AS SELECT * FROM (VALUES (NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL)) LIMIT 0")
            con.execute("DROP TABLE _fmr")

            # Create table from records
            con.execute("""
                CREATE TABLE _fmr_supplemental (
                    state_code VARCHAR,
                    fiscal_year INTEGER,
                    category VARCHAR,
                    payment_type VARCHAR,
                    service VARCHAR,
                    program VARCHAR,
                    total_computable DOUBLE,
                    federal_share DOUBLE,
                    state_share DOUBLE,
                    source VARCHAR,
                    snapshot_date DATE
                )
            """)
            for r in all_fmr_rows:
                con.execute("""
                    INSERT INTO _fmr_supplemental VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, [r["state_code"], r["fiscal_year"], r["category"],
                      r["payment_type"], r["service"], r["program"],
                      r["total_computable"], r["federal_share"], r["state_share"],
                      r["source"], SNAPSHOT_DATE])

            count = write_parquet(con, "_fmr_supplemental",
                                 LAKE_DIR / "fact" / "fmr_supplemental", args.dry_run)
            total_rows += count
            con.execute("DROP TABLE _fmr_supplemental")
        else:
            print("  No FMR data parsed")

    # ── MACPAC Exhibit 24 ──────────────────────────────────────────────
    if "macpac" in args.tables:
        print("\n=== MACPAC Exhibit 24 — Hospital Supplemental Payments ===")
        all_macpac_rows = []

        macpac_files = {
            2024: "https://www.macpac.gov/wp-content/uploads/2026/01/EXHIBIT-24.-Medicaid-Supplemental-Payments-to-Hospital-Providers-by-State-FY-2024.xlsx",
            2023: "https://www.macpac.gov/wp-content/uploads/2024/12/EXHIBIT-24.-Medicaid-Supplemental-Payments-to-Hospital-Providers-by-State-FY-2023.xlsx",
        }

        for fy, url in macpac_files.items():
            xlsx_path = DOWNLOAD_DIR / f"macpac_exhibit24_fy{fy}.xlsx"
            # Check /tmp first
            tmp_path = Path(f"/tmp/macpac_exhibit24_fy{fy}.xlsx")
            if tmp_path.exists() and tmp_path.stat().st_size > 5000:
                xlsx_path = tmp_path
            else:
                if not download_file(url, xlsx_path):
                    continue

            print(f"  Parsing FY {fy}...")
            rows = parse_macpac_exhibit24(str(xlsx_path), fy)
            print(f"  FY {fy}: {len(rows)} states")
            all_macpac_rows.extend(rows)

        if all_macpac_rows:
            con.execute("""
                CREATE TABLE _macpac_supplemental (
                    state_code VARCHAR,
                    fiscal_year INTEGER,
                    total_hospital_payments_m DOUBLE,
                    dsh_payments_m DOUBLE,
                    non_dsh_supplemental_m DOUBLE,
                    sec_1115_waiver_m DOUBLE,
                    supplemental_pct DOUBLE,
                    source VARCHAR,
                    snapshot_date DATE
                )
            """)
            for r in all_macpac_rows:
                con.execute("""
                    INSERT INTO _macpac_supplemental VALUES (?,?,?,?,?,?,?,?,?)
                """, [r["state_code"], r["fiscal_year"], r["total_hospital_payments_m"],
                      r["dsh_payments_m"], r["non_dsh_supplemental_m"],
                      r["sec_1115_waiver_m"], r["supplemental_pct"],
                      r["source"], SNAPSHOT_DATE])

            count = write_parquet(con, "_macpac_supplemental",
                                 LAKE_DIR / "fact" / "macpac_supplemental", args.dry_run)
            total_rows += count
            con.execute("DROP TABLE _macpac_supplemental")
        else:
            print("  No MACPAC data parsed")

    print(f"\n=== Total: {total_rows:,} rows written ===")
    con.close()


if __name__ == "__main__":
    main()
