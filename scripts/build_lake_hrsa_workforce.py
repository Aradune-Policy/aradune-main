#!/usr/bin/env python3
"""
build_lake_hrsa_workforce.py — Ingest HRSA workforce and health center data into the lake.

Reads from: data/raw/hrsa_awarded_grants.csv
             data/raw/hrsa_nhsc_scholar_pipeline_2025.xlsx
             data/lake/fact/workforce_projections/ (existing)
             data/lake/fact/nursing_workforce/ (existing)
Writes to:  data/lake/

Tables built:
  fact_health_center_awards       — HRSA health center program awarded grants by state
  fact_bh_workforce_projections   — Behavioral health workforce supply/demand projections
  fact_np_pa_supply               — Nurse practitioner and physician assistant supply
  fact_nhsc_scholar_pipeline      — NHSC Scholar Pipeline by state and discipline

Skipped (already in lake):
  fact_nhsc_field_strength        — 222 rows (snapshot 2026-03-09)
  fact_mua_designation            — 19,645 rows (snapshot 2026-03-09)
  fact_health_center_sites        — 8,121 rows (snapshot 2026-03-09)
  fact_workforce_projections      — 102,528 rows (snapshot 2026-03-09)
  fact_nursing_workforce          — 17,640 rows (snapshot 2026-03-09)

UDS aggregate data (fact_uds_fqhc) is not available for bulk download from HRSA.
The BPHC Electronic Reading Room is behind authentication (403).
Available UDS subsets (hypertension, CHQR badges) are already ingested.

Usage:
  python3 scripts/build_lake_hrsa_workforce.py
  python3 scripts/build_lake_hrsa_workforce.py --dry-run
  python3 scripts/build_lake_hrsa_workforce.py --only fact_health_center_awards
"""

import argparse
import csv
import json
import uuid
from datetime import date, datetime
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "Virgin Islands": "VI", "Guam": "GU",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
    "Federated States of Micronesia": "FM", "Marshall Islands": "MH",
    "Palau": "PW",
}


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


# ---------------------------------------------------------------------------
# Health Center Program Awarded Grants
# ---------------------------------------------------------------------------

def build_fact_health_center_awards(con, dry_run: bool) -> int:
    """
    Build fact_health_center_awards from HRSA awarded grants CSV.

    The CSV has a trailing comma in the header which creates an extra empty column.
    We use Python csv module to parse, then load into DuckDB via DataFrame.
    """
    print("Building fact_health_center_awards...")
    csv_path = RAW_DIR / "hrsa_awarded_grants.csv"
    if not csv_path.exists():
        print(f"  SKIPPED - {csv_path.name} not found")
        return 0

    # Parse with Python csv (DuckDB struggles with trailing comma format)
    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            award_year = row.get("Award Year", "").strip()
            financial_assistance = row.get("Financial Assistance", "").strip()
            state_abbr = row.get("Grantee State Abbreviation", "").strip()
            program_area = row.get("HRSA Program Area Name", "").strip()
            program_name = row.get("Grant Program Name", "").strip()
            grantee_name = row.get("Grantee Name", "").strip()
            grantee_city = row.get("Grantee City", "").strip()
            county = row.get("Complete County Name", "").strip()
            fips = row.get("State and County Federal Information Processing Standard Code", "").strip()
            grant_number = row.get("Grant Number", "").strip()
            activity_code = row.get("Grant Activity Code", "").strip()
            uds_desc = row.get("Uniform Data System Grant Program Description", "").strip()
            grantee_type = row.get("Grantee Type Description", "").strip()
            uei = row.get("Unique Entity Identifier", "").strip()
            start_date = row.get("Project Period Start Date Text String", "").strip()
            end_date = row.get("Grant Project Period End Date Text", "").strip()
            hhs_region = row.get("HHS Region Number", "").strip()
            lon = row.get("Geocoding Artifact Address Primary X Coordinate", "").strip()
            lat = row.get("Geocoding Artifact Address Primary Y Coordinate", "").strip()

            if not state_abbr or not award_year:
                continue

            try:
                fa = float(financial_assistance) if financial_assistance else None
            except ValueError:
                fa = None

            try:
                yr = int(award_year)
            except ValueError:
                continue

            rows.append({
                "award_year": yr,
                "financial_assistance": fa,
                "state_code": state_abbr if len(state_abbr) == 2 else None,
                "program_area": program_area,
                "program_name": program_name,
                "grantee_name": grantee_name,
                "grantee_city": grantee_city,
                "county": county,
                "fips_code": fips if fips else None,
                "grant_number": grant_number,
                "activity_code": activity_code,
                "uds_description": uds_desc if uds_desc else None,
                "grantee_type": grantee_type if grantee_type else None,
                "unique_entity_id": uei if uei else None,
                "project_start_date": start_date if start_date else None,
                "project_end_date": end_date if end_date else None,
                "hhs_region": hhs_region if hhs_region else None,
                "longitude": float(lon) if lon else None,
                "latitude": float(lat) if lat else None,
                "source": "data.hrsa.gov",
                "snapshot_date": SNAPSHOT_DATE,
            })

    if not rows:
        print("  SKIPPED - no rows parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.register("_df_awards", df)

    con.execute("""
        CREATE OR REPLACE TABLE _fact_hc_awards AS
        SELECT
            CAST(award_year AS INTEGER) AS award_year,
            CAST(financial_assistance AS DOUBLE) AS financial_assistance,
            state_code,
            program_area,
            program_name,
            grantee_name,
            grantee_city,
            county,
            fips_code,
            grant_number,
            activity_code,
            uds_description,
            grantee_type,
            unique_entity_id,
            project_start_date,
            project_end_date,
            hhs_region,
            CAST(longitude AS DOUBLE) AS longitude,
            CAST(latitude AS DOUBLE) AS latitude,
            source,
            CAST(snapshot_date AS DATE) AS snapshot_date
        FROM _df_awards
    """)

    count = write_parquet(con, "_fact_hc_awards", _snapshot_path("health_center_awards"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_hc_awards").fetchone()[0]
    programs = con.execute("SELECT COUNT(DISTINCT program_area) FROM _fact_hc_awards").fetchone()[0]
    total_fa = con.execute("SELECT SUM(financial_assistance) FROM _fact_hc_awards").fetchone()[0]
    print(f"  {count:,} rows, {states} states, {programs} program areas, ${total_fa:,.0f} total")
    con.execute("DROP TABLE IF EXISTS _fact_hc_awards")
    return count


# ---------------------------------------------------------------------------
# Behavioral Health Workforce Projections (subset of existing)
# ---------------------------------------------------------------------------

def build_fact_bh_workforce_projections(con, dry_run: bool) -> int:
    """
    Extract behavioral health workforce subset from existing workforce_projections table.
    Includes both the Behavioral Health profession group and BH-relevant professions
    from other groups (e.g., Psychiatric NPs from Primary Care).
    """
    print("Building fact_bh_workforce_projections...")

    # Find existing workforce projections parquet
    wp_dir = FACT_DIR / "workforce_projections"
    if not wp_dir.exists():
        print(f"  SKIPPED - workforce_projections not found in lake")
        return 0

    snapshots = sorted(wp_dir.iterdir())
    if not snapshots:
        print(f"  SKIPPED - no snapshots found")
        return 0

    latest_snap = snapshots[-1] / "data.parquet"
    if not latest_snap.exists():
        print(f"  SKIPPED - {latest_snap} not found")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_bh_wp AS
        SELECT
            year,
            profession_group,
            profession,
            state,
            rurality,
            supply_fte,
            demand_fte,
            pct_adequacy,
            region,
            'data.hrsa.gov (BHW projections)' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_parquet('{latest_snap}')
        WHERE profession_group = 'Behavioral Health'
           OR profession IN (
               'Psychiatric Nurse Practitioners',
               'Psychiatric Physician Assistants',
               'Psychiatric Aides',
               'Psychiatric Technicians'
           )
    """)

    count = write_parquet(con, "_fact_bh_wp", _snapshot_path("bh_workforce_projections"), dry_run)
    professions = con.execute("SELECT COUNT(DISTINCT profession) FROM _fact_bh_wp").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state) FROM _fact_bh_wp").fetchone()[0]
    year_range = con.execute("SELECT MIN(year), MAX(year) FROM _fact_bh_wp").fetchone()
    print(f"  {count:,} rows, {professions} professions, {states} states, {year_range[0]}-{year_range[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_bh_wp")
    return count


# ---------------------------------------------------------------------------
# NP/PA Supply (from nursing workforce + workforce projections)
# ---------------------------------------------------------------------------

def build_fact_np_pa_supply(con, dry_run: bool) -> int:
    """
    Build NP/PA supply table combining:
    1. NSSRN 2022 survey data (nursing_workforce) - current headcounts by state
    2. HRSA projections (workforce_projections) - supply/demand forecasts
    """
    print("Building fact_np_pa_supply...")

    # Find existing parquets
    nw_dir = FACT_DIR / "nursing_workforce"
    wp_dir = FACT_DIR / "workforce_projections"

    nw_snap = None
    wp_snap = None

    if nw_dir.exists():
        snaps = sorted(nw_dir.iterdir())
        if snaps:
            nw_snap = snaps[-1] / "data.parquet"

    if wp_dir.exists():
        snaps = sorted(wp_dir.iterdir())
        if snaps:
            wp_snap = snaps[-1] / "data.parquet"

    if not nw_snap or not nw_snap.exists():
        print("  SKIPPED - nursing_workforce not found")
        return 0
    if not wp_snap or not wp_snap.exists():
        print("  SKIPPED - workforce_projections not found")
        return 0

    # Part 1: NSSRN survey counts (state-level NP totals, 2022)
    con.execute(f"""
        CREATE OR REPLACE TABLE _nssrn_np AS
        SELECT
            state_name,
            license_type AS provider_type,
            weighted_count AS headcount_2022,
            '2022 NSSRN' AS data_source,
            'survey' AS data_type
        FROM read_parquet('{nw_snap}')
        WHERE license_type IN ('Nurse Practitioners', 'Nurse practitioners')
          AND status LIKE '%All%'
          AND age = 'All'
          AND sex = 'All'
          AND race_ethnicity = 'All'
          AND veteran_status = 'All'
          AND languages = 'All'
    """)

    # Part 2: HRSA projections for NP and PA
    con.execute(f"""
        CREATE OR REPLACE TABLE _proj_np_pa AS
        SELECT
            profession AS provider_type,
            state AS state_name,
            year,
            supply_fte,
            demand_fte,
            pct_adequacy,
            rurality,
            'HRSA BHW Projections' AS data_source,
            'projection' AS data_type
        FROM read_parquet('{wp_snap}')
        WHERE profession IN (
            'Nurse Practitioners', 'Nurse Practitioners (PC)', 'Nurse Practitioners (WH)',
            'Physician Assistants', 'Physician Assistants (PC)', 'Physician Assistants (WH)',
            'Nurse Midwives', 'Nurse Anesthetists'
        )
    """)

    # Combine into a unified table
    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_np_pa AS
        SELECT
            provider_type,
            state_name,
            2022 AS year,
            CAST(headcount_2022 AS DOUBLE) AS supply_fte,
            NULL::DOUBLE AS demand_fte,
            NULL::DOUBLE AS pct_adequacy,
            'All' AS rurality,
            data_source,
            data_type,
            'data.hrsa.gov' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _nssrn_np
        UNION ALL
        SELECT
            provider_type,
            state_name,
            year,
            supply_fte,
            demand_fte,
            pct_adequacy,
            rurality,
            data_source,
            data_type,
            'data.hrsa.gov' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM _proj_np_pa
    """)

    count = write_parquet(con, "_fact_np_pa", _snapshot_path("np_pa_supply"), dry_run)
    types = con.execute("SELECT COUNT(DISTINCT provider_type) FROM _fact_np_pa").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state_name) FROM _fact_np_pa").fetchone()[0]
    print(f"  {count:,} rows, {types} provider types, {states} states/territories")

    con.execute("DROP TABLE IF EXISTS _nssrn_np")
    con.execute("DROP TABLE IF EXISTS _proj_np_pa")
    con.execute("DROP TABLE IF EXISTS _fact_np_pa")
    return count


# ---------------------------------------------------------------------------
# NHSC Scholar Pipeline
# ---------------------------------------------------------------------------

def build_fact_nhsc_scholar_pipeline(con, dry_run: bool) -> int:
    """
    Build NHSC Scholar Pipeline table from Excel.
    Shows scholars in training by state and discipline.
    """
    print("Building fact_nhsc_scholar_pipeline...")
    xlsx_path = RAW_DIR / "hrsa_nhsc_scholar_pipeline_2025.xlsx"
    if not xlsx_path.exists():
        print(f"  SKIPPED - {xlsx_path.name} not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row (contains 'State')
    header = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip() if v else "" for v in row]
        if "State" in vals:
            header = vals
            header_row_idx = i
            break

    if not header:
        print("  SKIPPED - could not find header row")
        wb.close()
        return 0

    # Parse data rows
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 2, values_only=True)):
        vals = list(row)
        state_code = str(vals[0]).strip() if vals[0] else ""
        if not state_code or len(state_code) != 2:
            continue

        total = vals[1] if len(vals) > 1 else 0
        nhsc_sp = vals[2] if len(vals) > 2 else 0
        s2s_lrp = vals[3] if len(vals) > 3 else 0
        physician = vals[4] if len(vals) > 4 else 0
        dentist = vals[5] if len(vals) > 5 else 0
        pa = vals[6] if len(vals) > 6 else 0
        np = vals[7] if len(vals) > 7 else 0
        cnm = vals[8] if len(vals) > 8 else 0

        def safe_int(v):
            try:
                return int(v) if v is not None else 0
            except (ValueError, TypeError):
                return 0

        rows.append({
            "state_code": state_code,
            "fiscal_year": 2025,
            "total_scholars": safe_int(total),
            "nhsc_scholarship": safe_int(nhsc_sp),
            "s2s_lrp": safe_int(s2s_lrp),
            "physicians": safe_int(physician),
            "dentists": safe_int(dentist),
            "physician_assistants": safe_int(pa),
            "nurse_practitioners": safe_int(np),
            "certified_nurse_midwives": safe_int(cnm),
            "source": "data.hrsa.gov",
            "snapshot_date": SNAPSHOT_DATE,
        })

    wb.close()

    if not rows:
        print("  SKIPPED - no rows parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    con.register("_df_pipeline", df)

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_scholar AS
        SELECT
            state_code,
            CAST(fiscal_year AS INTEGER) AS fiscal_year,
            CAST(total_scholars AS INTEGER) AS total_scholars,
            CAST(nhsc_scholarship AS INTEGER) AS nhsc_scholarship,
            CAST(s2s_lrp AS INTEGER) AS s2s_lrp,
            CAST(physicians AS INTEGER) AS physicians,
            CAST(dentists AS INTEGER) AS dentists,
            CAST(physician_assistants AS INTEGER) AS physician_assistants,
            CAST(nurse_practitioners AS INTEGER) AS nurse_practitioners,
            CAST(certified_nurse_midwives AS INTEGER) AS certified_nurse_midwives,
            source,
            CAST(snapshot_date AS DATE) AS snapshot_date
        FROM _df_pipeline
    """)

    count = write_parquet(con, "_fact_scholar", _snapshot_path("nhsc_scholar_pipeline"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_scholar").fetchone()[0]
    total_scholars = con.execute("SELECT SUM(total_scholars) FROM _fact_scholar").fetchone()[0]
    print(f"  {count:,} rows, {states} states, {total_scholars:,} total scholars")
    con.execute("DROP TABLE IF EXISTS _fact_scholar")
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

ALL_TABLES = {
    "fact_health_center_awards": build_fact_health_center_awards,
    "fact_bh_workforce_projections": build_fact_bh_workforce_projections,
    "fact_np_pa_supply": build_fact_np_pa_supply,
    "fact_nhsc_scholar_pipeline": build_fact_nhsc_scholar_pipeline,
}


def main():
    parser = argparse.ArgumentParser(description="Ingest HRSA workforce & health center data")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only", type=str, default=None,
                        help="Comma-separated list of tables to build")
    args = parser.parse_args()

    tables = list(ALL_TABLES.keys())
    if args.only:
        tables = [t.strip() for t in args.only.split(",")]

    print(f"Snapshot: {SNAPSHOT_DATE}")
    print(f"Run ID:   {RUN_ID}")
    print(f"Building: {', '.join(tables)}")
    print()

    print("=" * 60)
    print("EXISTING TABLES (skipped):")
    print("  nhsc_field_strength       — 222 rows")
    print("  mua_designation           — 19,645 rows")
    print("  health_center_sites       — 8,121 rows")
    print("  workforce_projections     — 102,528 rows")
    print("  nursing_workforce         — 17,640 rows")
    print()
    print("UDS AGGREGATE DATA (fact_uds_fqhc):")
    print("  Not available for bulk download from HRSA.")
    print("  BPHC Electronic Reading Room returns 403.")
    print("  UDS subsets already in lake: fqhc_hypertension, fqhc_quality_badges.")
    print("=" * 60)
    print()

    con = duckdb.connect()
    totals = {}
    for name in tables:
        if name not in ALL_TABLES:
            print(f"  UNKNOWN table: {name}")
            continue
        totals[name] = ALL_TABLES[name](con, args.dry_run)
        print()

    con.close()

    print("=" * 60)
    print("HRSA WORKFORCE DATA LAKE INGESTION COMPLETE")
    print("=" * 60)
    total_rows = sum(totals.values())
    for name, count in totals.items():
        status = "written" if not args.dry_run else "dry-run"
        print(f"  {name:40s} {count:>12,} rows  [{status}]")
    print(f"  {'TOTAL':40s} {total_rows:>12,} rows")

    if not args.dry_run and total_rows > 0:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "snapshot_date": SNAPSHOT_DATE,
            "pipeline_run_id": RUN_ID,
            "created_at": datetime.now().isoformat() + "Z",
            "tables": {name: {"rows": count} for name, count in totals.items()},
            "total_rows": total_rows,
        }
        manifest_file = META_DIR / f"manifest_hrsa_workforce_{SNAPSHOT_DATE}.json"
        with open(manifest_file, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"\n  Manifest: {manifest_file}")

    print("\nDone.")


if __name__ == "__main__":
    main()
