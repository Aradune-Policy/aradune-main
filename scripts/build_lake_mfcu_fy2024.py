#!/usr/bin/env python3
"""
build_lake_mfcu_fy2024.py — Ingest MFCU FY2024 detailed reports into the Aradune data lake.

Sources:
  1. data/raw/mfcu_fy2024_statistical_chart.xlsx — State-level MFCU activity (investigations,
     indictments, convictions, recoveries, expenditures, staffing)
  2. data/raw/mfcu_fy2024_open_cases.xlsx — Open cases by provider type
  3. data/raw/mfcu_fy2024_case_outcomes.xlsx — Case outcomes by provider type

Tables built:
  mfcu_statistical_chart — State-level MFCU activity for FY 2024
  mfcu_open_cases        — Open investigations by provider type for FY 2024
  mfcu_case_outcomes     — Convictions, settlements, recoveries by provider type for FY 2024

Usage:
  python3 scripts/build_lake_mfcu_fy2024.py
"""

import json
import re
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

STATE_MAP = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR",
    "CALIFORNIA": "CA", "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE",
    "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID",
    "ILLINOIS": "IL", "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS",
    "KENTUCKY": "KY", "LOUISIANA": "LA", "MAINE": "ME", "MARYLAND": "MD",
    "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN", "MISSISSIPPI": "MS",
    "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK",
    "OREGON": "OR", "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI",
    "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD", "TENNESSEE": "TN",
    "TEXAS": "TX", "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA",
    "WASHINGTON": "WA", "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
    "TOTAL": "US", "TOTALS": "US",
}


def _parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "")
    if s in ("", "-", "--", "N/A", "n/a", "*"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(val):
    n = _parse_numeric(val)
    return int(n) if n is not None else None


def _snapshot_path(fact_name):
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def _write_parquet(con, table_name, fact_name):
    out_path = _snapshot_path(fact_name)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if count > 0:
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.2f} MB)")
    return count


# ──────────────────────────────────────────────────
# 1. Statistical Chart (state-level)
# ──────────────────────────────────────────────────
def build_statistical_chart(con):
    print("\n== MFCU FY2024 Statistical Chart (state-level) ==")
    xlsx_path = RAW_DIR / "mfcu_fy2024_statistical_chart.xlsx"
    if not xlsx_path.exists():
        print("  SKIPPED - file not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Row 2 = header, data starts row 3
    # Cols: State, Total Investigations, Fraud Investigations, Abuse/Neglect Investigations,
    # Total Indictments, Fraud Indictments, Abuse/Neglect Indictments,
    # Total Convictions, Fraud Convictions, Abuse/Neglect Convictions,
    # Civil Settlements, Total Recoveries, Criminal Recoveries,
    # Civil Recoveries Global, Civil Recoveries Other,
    # MFCU Grant Expenditures, Total Medicaid Expenditures, Staff On Board

    rows = []
    for row_num in range(3, ws.max_row + 1):
        raw_state = ws.cell(row=row_num, column=1).value
        if not raw_state or not isinstance(raw_state, str):
            continue
        clean = re.sub(r"\d+$", "", raw_state.strip()).strip().upper()
        state_code = STATE_MAP.get(clean)
        if not state_code:
            continue

        record = {
            "state_code": state_code,
            "fiscal_year": 2024,
            "total_investigations": _parse_int(ws.cell(row=row_num, column=2).value),
            "fraud_investigations": _parse_int(ws.cell(row=row_num, column=3).value),
            "abuse_neglect_investigations": _parse_int(ws.cell(row=row_num, column=4).value),
            "total_indictments": _parse_int(ws.cell(row=row_num, column=5).value),
            "fraud_indictments": _parse_int(ws.cell(row=row_num, column=6).value),
            "abuse_neglect_indictments": _parse_int(ws.cell(row=row_num, column=7).value),
            "total_convictions": _parse_int(ws.cell(row=row_num, column=8).value),
            "fraud_convictions": _parse_int(ws.cell(row=row_num, column=9).value),
            "abuse_neglect_convictions": _parse_int(ws.cell(row=row_num, column=10).value),
            "civil_settlements_judgments": _parse_int(ws.cell(row=row_num, column=11).value),
            "total_recoveries": _parse_numeric(ws.cell(row=row_num, column=12).value),
            "criminal_recoveries": _parse_numeric(ws.cell(row=row_num, column=13).value),
            "civil_recoveries_global": _parse_numeric(ws.cell(row=row_num, column=14).value),
            "civil_recoveries_other": _parse_numeric(ws.cell(row=row_num, column=15).value),
            "mfcu_grant_expenditures": _parse_numeric(ws.cell(row=row_num, column=16).value),
            "total_medicaid_expenditures": _parse_numeric(ws.cell(row=row_num, column=17).value),
            "staff_on_board": _parse_numeric(ws.cell(row=row_num, column=18).value),
            "source": "oig_mfcu",
            "snapshot_date": SNAPSHOT_DATE,
        }
        rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return 0

    columns = list(rows[0].keys())
    int_cols = {"total_investigations", "fraud_investigations", "abuse_neglect_investigations",
                "total_indictments", "fraud_indictments", "abuse_neglect_indictments",
                "total_convictions", "fraud_convictions", "abuse_neglect_convictions",
                "civil_settlements_judgments", "fiscal_year"}
    float_cols = {"total_recoveries", "criminal_recoveries", "civil_recoveries_global",
                  "civil_recoveries_other", "mfcu_grant_expenditures",
                  "total_medicaid_expenditures", "staff_on_board"}
    col_defs = ", ".join(
        f"{c} {'INTEGER' if c in int_cols else 'DOUBLE' if c in float_cols else 'VARCHAR'}"
        for c in columns
    )
    con.execute(f"CREATE OR REPLACE TABLE _stat ({col_defs})")
    placeholders = ", ".join(["?"] * len(columns))
    con.executemany(
        f"INSERT INTO _stat VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    count = _write_parquet(con, "_stat", "mfcu_statistical_chart")

    # Validation
    state_count = len([r for r in rows if r["state_code"] != "US"])
    us = [r for r in rows if r["state_code"] == "US"]
    print(f"  {count} rows, {state_count} states + US total")
    if us:
        print(f"  US total investigations: {us[0]['total_investigations']:,}")
        print(f"  US total convictions: {us[0]['total_convictions']:,}")
        if us[0]["total_recoveries"]:
            print(f"  US total recoveries: ${us[0]['total_recoveries']:,.0f}")

    return count


# ──────────────────────────────────────────────────
# 2. Open Cases (by provider type)
# ──────────────────────────────────────────────────
def build_open_cases(con):
    print("\n== MFCU FY2024 Open Cases (by provider type) ==")
    xlsx_path = RAW_DIR / "mfcu_fy2024_open_cases.xlsx"
    if not xlsx_path.exists():
        print("  SKIPPED - file not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Row 2 = header, data starts row 3
    # Provider Type, Criminal Fraud, Criminal Abuse/Neglect, Civil Fraud, Civil Abuse/Neglect

    rows = []
    for row_num in range(3, ws.max_row + 1):
        ptype = ws.cell(row=row_num, column=1).value
        if not ptype or not isinstance(ptype, str) or not ptype.strip():
            continue
        ptype = ptype.strip()
        if ptype.lower().startswith("total") or ptype.lower().startswith("source"):
            # Still capture totals
            if not ptype.lower().startswith("total"):
                continue

        record = {
            "provider_type": ptype,
            "fiscal_year": 2024,
            "criminal_open_fraud": _parse_int(ws.cell(row=row_num, column=2).value),
            "criminal_open_abuse_neglect": _parse_int(ws.cell(row=row_num, column=3).value),
            "civil_open_fraud": _parse_int(ws.cell(row=row_num, column=4).value),
            "civil_open_abuse_neglect": _parse_int(ws.cell(row=row_num, column=5).value),
            "source": "oig_mfcu",
            "snapshot_date": SNAPSHOT_DATE,
        }
        rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return 0

    columns = list(rows[0].keys())
    int_cols = {"criminal_open_fraud", "criminal_open_abuse_neglect",
                "civil_open_fraud", "civil_open_abuse_neglect", "fiscal_year"}
    col_defs = ", ".join(
        f"{c} {'INTEGER' if c in int_cols else 'VARCHAR'}" for c in columns
    )
    con.execute(f"CREATE OR REPLACE TABLE _cases ({col_defs})")
    placeholders = ", ".join(["?"] * len(columns))
    con.executemany(
        f"INSERT INTO _cases VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    count = _write_parquet(con, "_cases", "mfcu_open_cases")

    # Validation
    total_row = [r for r in rows if r["provider_type"].lower().startswith("total")]
    provider_types = len([r for r in rows if not r["provider_type"].lower().startswith("total")])
    print(f"  {count} rows, {provider_types} provider types")
    if total_row:
        t = total_row[0]
        print(f"  Total criminal fraud open: {t['criminal_open_fraud']:,}")
        print(f"  Total civil fraud open: {t['civil_open_fraud']:,}")

    return count


# ──────────────────────────────────────────────────
# 3. Case Outcomes (by provider type)
# ──────────────────────────────────────────────────
def build_case_outcomes(con):
    print("\n== MFCU FY2024 Case Outcomes (by provider type) ==")
    xlsx_path = RAW_DIR / "mfcu_fy2024_case_outcomes.xlsx"
    if not xlsx_path.exists():
        print("  SKIPPED - file not found")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Row 2 = header, data starts row 3
    # Provider Type, Criminal Convictions Fraud, Criminal Convictions Abuse/Neglect,
    # Criminal Recoveries Fraud, Criminal Recoveries Abuse/Neglect,
    # Civil Settlements Fraud, Civil Settlements Abuse/Neglect,
    # Civil Recoveries Fraud, Civil Recoveries Abuse/Neglect

    rows = []
    for row_num in range(3, ws.max_row + 1):
        ptype = ws.cell(row=row_num, column=1).value
        if not ptype or not isinstance(ptype, str) or not ptype.strip():
            continue
        ptype = ptype.strip()
        if ptype.lower().startswith("source"):
            continue

        record = {
            "provider_type": ptype,
            "fiscal_year": 2024,
            "criminal_convictions_fraud": _parse_int(ws.cell(row=row_num, column=2).value),
            "criminal_convictions_abuse_neglect": _parse_int(ws.cell(row=row_num, column=3).value),
            "criminal_recoveries_fraud": _parse_numeric(ws.cell(row=row_num, column=4).value),
            "criminal_recoveries_abuse_neglect": _parse_numeric(ws.cell(row=row_num, column=5).value),
            "civil_settlements_fraud": _parse_int(ws.cell(row=row_num, column=6).value),
            "civil_settlements_abuse_neglect": _parse_int(ws.cell(row=row_num, column=7).value),
            "civil_recoveries_fraud": _parse_numeric(ws.cell(row=row_num, column=8).value),
            "civil_recoveries_abuse_neglect": _parse_numeric(ws.cell(row=row_num, column=9).value),
            "source": "oig_mfcu",
            "snapshot_date": SNAPSHOT_DATE,
        }
        rows.append(record)

    wb.close()

    if not rows:
        print("  No data parsed")
        return 0

    columns = list(rows[0].keys())
    int_cols = {"criminal_convictions_fraud", "criminal_convictions_abuse_neglect",
                "civil_settlements_fraud", "civil_settlements_abuse_neglect", "fiscal_year"}
    float_cols = {"criminal_recoveries_fraud", "criminal_recoveries_abuse_neglect",
                  "civil_recoveries_fraud", "civil_recoveries_abuse_neglect"}
    col_defs = ", ".join(
        f"{c} {'INTEGER' if c in int_cols else 'DOUBLE' if c in float_cols else 'VARCHAR'}"
        for c in columns
    )
    con.execute(f"CREATE OR REPLACE TABLE _outcomes ({col_defs})")
    placeholders = ", ".join(["?"] * len(columns))
    con.executemany(
        f"INSERT INTO _outcomes VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    count = _write_parquet(con, "_outcomes", "mfcu_case_outcomes")

    # Validation
    total_row = [r for r in rows if r["provider_type"].lower().startswith("total")]
    provider_types = len([r for r in rows if not r["provider_type"].lower().startswith("total")])
    print(f"  {count} rows, {provider_types} provider types")
    if total_row:
        t = total_row[0]
        total_crim_rec = (t["criminal_recoveries_fraud"] or 0) + (t["criminal_recoveries_abuse_neglect"] or 0)
        total_civil_rec = (t["civil_recoveries_fraud"] or 0) + (t["civil_recoveries_abuse_neglect"] or 0)
        print(f"  Total criminal convictions: {(t['criminal_convictions_fraud'] or 0) + (t['criminal_convictions_abuse_neglect'] or 0):,}")
        print(f"  Total criminal recoveries: ${total_crim_rec:,.0f}")
        print(f"  Total civil recoveries: ${total_civil_rec:,.0f}")

    return count


# ──────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────
def main():
    print(f"MFCU FY2024 ETL -- snapshot {SNAPSHOT_DATE}")
    con = duckdb.connect()
    total_rows = 0
    tables_built = []

    n = build_statistical_chart(con)
    if n > 0:
        total_rows += n
        tables_built.append("mfcu_statistical_chart")

    n = build_open_cases(con)
    if n > 0:
        total_rows += n
        tables_built.append("mfcu_open_cases")

    n = build_case_outcomes(con)
    if n > 0:
        total_rows += n
        tables_built.append("mfcu_case_outcomes")

    con.close()

    print(f"\n== Summary ==")
    print(f"  Tables built: {len(tables_built)}")
    print(f"  Total rows: {total_rows:,}")
    for t in tables_built:
        print(f"    - {t}")

    if tables_built:
        manifest = {
            "pipeline_run": "mfcu_fy2024",
            "run_id": RUN_ID,
            "snapshot_date": SNAPSHOT_DATE,
            "run_timestamp": datetime.now(timezone.utc).isoformat(),
            "total_rows": total_rows,
            "tables": tables_built,
            "source_files": [
                "data/raw/mfcu_fy2024_statistical_chart.xlsx",
                "data/raw/mfcu_fy2024_open_cases.xlsx",
                "data/raw/mfcu_fy2024_case_outcomes.xlsx",
            ],
            "source": "OIG MFCU Statistical Data FY 2024",
        }
        manifest_path = META_DIR / f"manifest_mfcu_fy2024_{SNAPSHOT_DATE}.json"
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(json.dumps(manifest, indent=2))
        print(f"  Manifest: {manifest_path.relative_to(LAKE_DIR)}")


if __name__ == "__main__":
    main()
