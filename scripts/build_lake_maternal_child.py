#!/usr/bin/env python3
"""
build_lake_maternal_child.py -- Ingest maternal & child health data into the lake.

Downloads and processes:
  1. fact_cdc_natality         -- VSRR state births, deaths, infant deaths (CDC hmz2-vwda)
  2. fact_infant_mortality_state -- Infant mortality by state, race (CDC pjb2-jvdr)
  3. fact_infant_mortality_quarterly -- VSRR quarterly infant mortality (CDC jqwm-z2g9)
  4. fact_child_vaccination    -- NIS vaccination coverage, ages 0-35 months (CDC fhky-rtsk)
  5. fact_adolescent_vaccination -- NIS vaccination coverage, ages 13-17 (CDC ee48-w5t6)
  6. fact_teen_birth_rate      -- Teen birth rates by county (CDC 3h58-x6cd)
  7. fact_wic_nutrition        -- WIC child obesity/weight data (CDC 735e-byxc)
  8. fact_wic_participation    -- WIC participation by state (USDA Excel)
  9. fact_foster_care          -- Children in foster care by state (Annie E. Casey / AFCARS)
 10. fact_title_v_mch          -- Title V MCH block grant funding by state (HRSA TVIS API)

Usage:
  python3 scripts/build_lake_maternal_child.py
  python3 scripts/build_lake_maternal_child.py --dry-run
  python3 scripts/build_lake_maternal_child.py --table fact_wic_participation
"""

import argparse
import json
import subprocess
import tempfile
import uuid
from datetime import date
from pathlib import Path

import duckdb

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
}


def _curl_download(url: str, out_path: Path) -> bool:
    """Download a URL via curl (avoids Python urllib issues with some federal sites)."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        ["curl", "-sL", "-o", str(out_path), url],
        capture_output=True, text=True, timeout=120,
    )
    if result.returncode != 0 or not out_path.exists() or out_path.stat().st_size < 100:
        print(f"  DOWNLOAD FAILED: {url}")
        return False
    return True


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.2f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


# ---------------------------------------------------------------------------
# 1. VSRR State Births / Deaths / Infant Deaths
# ---------------------------------------------------------------------------
def build_cdc_natality(con, dry_run: bool) -> int:
    """CDC VSRR provisional counts for live births, deaths, and infant deaths by state."""
    print("\n[1/10] Building fact_cdc_natality...")
    url = "https://data.cdc.gov/api/views/hmz2-vwda/rows.csv?accessType=DOWNLOAD"
    csv_path = RAW_DIR / "cdc_vsrr_state_births_deaths.csv"

    if not _curl_download(url, csv_path):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_natality AS
        SELECT
            TRIM(State) AS state_name,
            TRY_CAST(Year AS INTEGER) AS year,
            TRIM(Month) AS month,
            TRIM(Period) AS period,
            TRIM(Indicator) AS indicator,
            TRY_CAST("Data Value" AS BIGINT) AS data_value,
            'data.cdc.gov/hmz2-vwda' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE "Data Value" IS NOT NULL
          AND TRIM("Data Value") != ''
    """)

    count = write_parquet(con, "_fact_natality", _snapshot_path("cdc_natality"), dry_run)
    if count > 0:
        states = con.execute("SELECT COUNT(DISTINCT state_name) FROM _fact_natality").fetchone()[0]
        indicators = con.execute("SELECT DISTINCT indicator FROM _fact_natality").fetchall()
        print(f"  {count:,} rows, {states} states/territories, indicators: {[r[0] for r in indicators]}")
    con.execute("DROP TABLE IF EXISTS _fact_natality")
    return count


# ---------------------------------------------------------------------------
# 2. Infant Mortality by State and Race
# ---------------------------------------------------------------------------
def build_infant_mortality_state(con, dry_run: bool) -> int:
    """CDC DQS infant mortality rates by race, Hispanic origin, state."""
    print("\n[2/10] Building fact_infant_mortality_state...")
    url = "https://data.cdc.gov/api/views/pjb2-jvdr/rows.csv?accessType=DOWNLOAD"
    csv_path = RAW_DIR / "cdc_infant_mortality_state_race.csv"

    if not _curl_download(url, csv_path):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_imr AS
        SELECT
            TRIM(TOPIC) AS topic,
            TRIM(SUBTOPIC) AS subtopic,
            TRIM(CLASSIFICATION) AS classification,
            TRIM("GROUP") AS group_name,
            TRIM(SUBGROUP) AS subgroup,
            TRIM(ESTIMATE_TYPE) AS estimate_type,
            TRIM(TIME_PERIOD) AS time_period,
            TRY_CAST(ESTIMATE AS DOUBLE) AS estimate,
            TRY_CAST(STANDARD_ERROR AS DOUBLE) AS standard_error,
            TRY_CAST(ESTIMATE_LCI AS DOUBLE) AS estimate_lci,
            TRY_CAST(ESTIMATE_UCI AS DOUBLE) AS estimate_uci,
            TRIM(FLAG) AS flag,
            TRIM(CAST(STATE_FIPS AS VARCHAR)) AS state_fips,
            'data.cdc.gov/pjb2-jvdr' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE ESTIMATE IS NOT NULL
          AND TRIM(ESTIMATE) != ''
    """)

    count = write_parquet(con, "_fact_imr", _snapshot_path("infant_mortality_state"), dry_run)
    if count > 0:
        states = con.execute("SELECT COUNT(DISTINCT state_fips) FROM _fact_imr").fetchone()[0]
        periods = con.execute("SELECT COUNT(DISTINCT time_period) FROM _fact_imr").fetchone()[0]
        print(f"  {count:,} rows, {states} state FIPS codes, {periods} time periods")
    con.execute("DROP TABLE IF EXISTS _fact_imr")
    return count


# ---------------------------------------------------------------------------
# 3. VSRR Quarterly Infant Mortality
# ---------------------------------------------------------------------------
def build_infant_mortality_quarterly(con, dry_run: bool) -> int:
    """CDC VSRR quarterly provisional estimates for infant mortality (national)."""
    print("\n[3/10] Building fact_infant_mortality_quarterly...")
    url = "https://data.cdc.gov/api/views/jqwm-z2g9/rows.csv?accessType=DOWNLOAD"
    csv_path = RAW_DIR / "cdc_vsrr_infant_mortality_quarterly.csv"

    if not _curl_download(url, csv_path):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_imq AS
        SELECT
            TRIM("Year and Quarter") AS year_quarter,
            TRIM(Topic) AS topic,
            TRIM(Indicator) AS indicator,
            TRIM("Time Period") AS time_period,
            TRY_CAST(Rate AS DOUBLE) AS rate,
            TRIM(Unit) AS unit,
            TRIM(Significant) AS significant,
            TRY_CAST("Standard Error" AS DOUBLE) AS standard_error,
            TRIM("Footnote Symbol") AS footnote_symbol,
            TRIM(Footnote) AS footnote,
            'data.cdc.gov/jqwm-z2g9' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE Rate IS NOT NULL
          AND TRIM(Rate) != ''
    """)

    count = write_parquet(con, "_fact_imq", _snapshot_path("infant_mortality_quarterly"), dry_run)
    if count > 0:
        indicators = con.execute("SELECT DISTINCT indicator FROM _fact_imq").fetchall()
        print(f"  {count:,} rows, indicators: {[r[0] for r in indicators]}")
    con.execute("DROP TABLE IF EXISTS _fact_imq")
    return count


# ---------------------------------------------------------------------------
# 4. Child Vaccination Coverage (NIS, 0-35 months)
# ---------------------------------------------------------------------------
def build_child_vaccination(con, dry_run: bool) -> int:
    """CDC NIS vaccination coverage among young children (0-35 months) by state."""
    print("\n[4/10] Building fact_child_vaccination...")
    url = "https://data.cdc.gov/api/views/fhky-rtsk/rows.csv?accessType=DOWNLOAD"
    csv_path = RAW_DIR / "cdc_child_vaccination_coverage.csv"

    if not _curl_download(url, csv_path):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_cvax AS
        SELECT
            TRIM(Vaccine) AS vaccine,
            TRIM(Dose) AS dose,
            TRIM("Geography Type") AS geography_type,
            TRIM(Geography) AS geography,
            TRIM("Birth Year/Birth Cohort") AS birth_year_cohort,
            TRIM("Dimension Type") AS dimension_type,
            TRIM(Dimension) AS dimension,
            TRY_CAST(REPLACE("Estimate (%)", '%', '') AS DOUBLE) AS estimate_pct,
            TRIM("95% CI (%)") AS confidence_interval,
            TRY_CAST("Sample Size" AS INTEGER) AS sample_size,
            'data.cdc.gov/fhky-rtsk' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE "Estimate (%)" IS NOT NULL
          AND TRIM("Estimate (%)") != ''
          AND TRIM("Estimate (%)") != 'NR'
    """)

    count = write_parquet(con, "_fact_cvax", _snapshot_path("child_vaccination"), dry_run)
    if count > 0:
        vaccines = con.execute("SELECT COUNT(DISTINCT vaccine) FROM _fact_cvax").fetchone()[0]
        geos = con.execute("SELECT COUNT(DISTINCT geography) FROM _fact_cvax WHERE geography_type = 'States/Local Areas'").fetchone()[0]
        print(f"  {count:,} rows, {vaccines} vaccines, {geos} state/local areas")
    con.execute("DROP TABLE IF EXISTS _fact_cvax")
    return count


# ---------------------------------------------------------------------------
# 5. Adolescent Vaccination Coverage (NIS-Teen, 13-17)
# ---------------------------------------------------------------------------
def build_adolescent_vaccination(con, dry_run: bool) -> int:
    """CDC NIS-Teen vaccination coverage among adolescents (13-17) by state."""
    print("\n[5/10] Building fact_adolescent_vaccination...")
    url = "https://data.cdc.gov/api/views/ee48-w5t6/rows.csv?accessType=DOWNLOAD"
    csv_path = RAW_DIR / "cdc_adolescent_vaccination_coverage.csv"

    if not _curl_download(url, csv_path):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_avax AS
        SELECT
            TRIM("Vaccine/Sample") AS vaccine,
            TRIM(Dose) AS dose,
            TRIM("Geography Type") AS geography_type,
            TRIM(Geography) AS geography,
            TRY_CAST("Survey Year" AS INTEGER) AS survey_year,
            TRIM("Dimension Type") AS dimension_type,
            TRIM(Dimension) AS dimension,
            TRY_CAST(REPLACE("Estimate (%)", '%', '') AS DOUBLE) AS estimate_pct,
            TRIM("95% CI (%)") AS confidence_interval,
            TRY_CAST("Sample Size" AS INTEGER) AS sample_size,
            'data.cdc.gov/ee48-w5t6' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE "Estimate (%)" IS NOT NULL
          AND TRIM("Estimate (%)") != ''
          AND TRIM("Estimate (%)") != 'NR'
    """)

    count = write_parquet(con, "_fact_avax", _snapshot_path("adolescent_vaccination"), dry_run)
    if count > 0:
        vaccines = con.execute("SELECT COUNT(DISTINCT vaccine) FROM _fact_avax").fetchone()[0]
        years = con.execute("SELECT MIN(survey_year), MAX(survey_year) FROM _fact_avax").fetchone()
        print(f"  {count:,} rows, {vaccines} vaccines, years {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_avax")
    return count


# ---------------------------------------------------------------------------
# 6. Teen Birth Rates by County
# ---------------------------------------------------------------------------
def build_teen_birth_rate(con, dry_run: bool) -> int:
    """CDC NCHS teen birth rates (15-19) by county."""
    print("\n[6/10] Building fact_teen_birth_rate...")
    url = "https://data.cdc.gov/api/views/3h58-x6cd/rows.csv?accessType=DOWNLOAD"
    csv_path = RAW_DIR / "cdc_teen_birth_rates_county.csv"

    if not _curl_download(url, csv_path):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_tbr AS
        SELECT
            TRY_CAST(Year AS INTEGER) AS year,
            TRIM(State) AS state_name,
            TRIM(County) AS county_name,
            LPAD(TRIM(CAST("State FIPS Code" AS VARCHAR)), 2, '0') AS state_fips,
            LPAD(TRIM(CAST("County FIPS Code" AS VARCHAR)), 3, '0') AS county_fips,
            LPAD(TRIM(CAST("Combined FIPS Code" AS VARCHAR)), 5, '0') AS fips_code,
            TRY_CAST("Birth Rate" AS DOUBLE) AS birth_rate,
            TRY_CAST("Lower Confidence Limit" AS DOUBLE) AS lower_ci,
            TRY_CAST("Upper Confidence Limit" AS DOUBLE) AS upper_ci,
            'data.cdc.gov/3h58-x6cd' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE "Birth Rate" IS NOT NULL
          AND TRIM("Birth Rate") != ''
    """)

    count = write_parquet(con, "_fact_tbr", _snapshot_path("teen_birth_rate"), dry_run)
    if count > 0:
        states = con.execute("SELECT COUNT(DISTINCT state_name) FROM _fact_tbr").fetchone()[0]
        counties = con.execute("SELECT COUNT(DISTINCT fips_code) FROM _fact_tbr").fetchone()[0]
        years = con.execute("SELECT MIN(year), MAX(year) FROM _fact_tbr").fetchone()
        print(f"  {count:,} rows, {states} states, {counties} counties, years {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_tbr")
    return count


# ---------------------------------------------------------------------------
# 7. WIC Nutrition (CDC Socrata)
# ---------------------------------------------------------------------------
def build_wic_nutrition(con, dry_run: bool) -> int:
    """CDC WIC child obesity and weight status data by state."""
    print("\n[7/10] Building fact_wic_nutrition...")
    url = "https://data.cdc.gov/api/views/735e-byxc/rows.csv?accessType=DOWNLOAD"
    csv_path = RAW_DIR / "cdc_wic_nutrition.csv"

    if not _curl_download(url, csv_path):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_wicn AS
        SELECT
            TRY_CAST(YearStart AS INTEGER) AS year_start,
            TRY_CAST(YearEnd AS INTEGER) AS year_end,
            TRIM(LocationAbbr) AS state_code,
            TRIM(LocationDesc) AS state_name,
            TRIM(Class) AS class,
            TRIM(Topic) AS topic,
            TRIM(Question) AS question,
            TRY_CAST(Data_Value AS DOUBLE) AS data_value,
            TRY_CAST(Low_Confidence_Limit AS DOUBLE) AS lower_ci,
            TRY_CAST(High_Confidence_Limit AS DOUBLE) AS upper_ci,
            TRY_CAST(Sample_Size AS INTEGER) AS sample_size,
            TRIM("Age(months)") AS age_months,
            TRIM(Sex) AS sex,
            TRIM("Race/Ethnicity") AS race_ethnicity,
            TRIM(StratificationCategory1) AS stratification_category,
            TRIM(Stratification1) AS stratification,
            'data.cdc.gov/735e-byxc' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', all_varchar=true)
        WHERE Data_Value IS NOT NULL
          AND TRIM(Data_Value) != ''
    """)

    count = write_parquet(con, "_fact_wicn", _snapshot_path("wic_nutrition"), dry_run)
    if count > 0:
        states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_wicn").fetchone()[0]
        years = con.execute("SELECT MIN(year_start), MAX(year_end) FROM _fact_wicn").fetchone()
        print(f"  {count:,} rows, {states} states, years {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_wicn")
    return count


# ---------------------------------------------------------------------------
# 8. WIC Participation (USDA Excel)
# ---------------------------------------------------------------------------
def build_wic_participation(con, dry_run: bool) -> int:
    """USDA WIC total participation by state, FY2021-2025."""
    print("\n[8/10] Building fact_wic_participation...")

    url = "https://www.fns.usda.gov/sites/default/files/resource-files/26wifypart-2.xlsx"
    xlsx_path = RAW_DIR / "wic_participation_annual.xlsx"

    if not _curl_download(url, xlsx_path):
        return 0

    # Parse with openpyxl since the Excel has simple layout
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Part"]

    # Find header row and year columns
    # Row 3 has: State/Indian Tribe, FY2021, FY2022, FY2023, FY2024, FY2025
    headers = []
    for cell in ws[3]:
        v = str(cell.value).strip() if cell.value else ""
        headers.append(v)

    fy_cols = []
    for i, h in enumerate(headers):
        if h.startswith("FY"):
            fy_cols.append((i, int(h.replace("FY ", "").replace("FY", ""))))

    rows = []
    for row_idx in range(5, ws.max_row + 1):  # Data starts row 5
        state_val = ws.cell(row=row_idx, column=1).value
        if not state_val:
            continue
        state_name = str(state_val).strip()
        # Skip sub-agencies (indented), totals, and non-state rows
        if state_name.startswith(" ") or state_name.startswith("  "):
            continue
        if state_name.lower() in ("total", "grand total", "preliminary", ""):
            continue
        if "total" in state_name.lower() and "territory" not in state_name.lower():
            continue

        state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            continue

        for col_idx, fy in fy_cols:
            val = ws.cell(row=row_idx, column=col_idx + 1).value
            if val is not None:
                try:
                    participation = float(val)
                    rows.append({
                        "state_code": state_code,
                        "state_name": state_name,
                        "fiscal_year": fy,
                        "avg_monthly_participation": round(participation, 1),
                    })
                except (ValueError, TypeError):
                    pass

    if not rows:
        print("  SKIPPED -- no rows parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    df["source"] = "fns.usda.gov/wic"
    df["snapshot_date"] = SNAPSHOT_DATE

    con.execute("CREATE OR REPLACE TABLE _fact_wicp AS SELECT * FROM df")

    count = write_parquet(con, "_fact_wicp", _snapshot_path("wic_participation"), dry_run)
    if count > 0:
        states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_wicp").fetchone()[0]
        years = con.execute("SELECT MIN(fiscal_year), MAX(fiscal_year) FROM _fact_wicp").fetchone()
        print(f"  {count:,} rows, {states} states, FY{years[0]}-FY{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_wicp")
    return count


# ---------------------------------------------------------------------------
# 9. Foster Care (Annie E. Casey Kids Count / AFCARS)
# ---------------------------------------------------------------------------
def build_foster_care(con, dry_run: bool) -> int:
    """Children in foster care by state, 2000-2023 (AFCARS via Annie E. Casey Kids Count)."""
    print("\n[9/10] Building fact_foster_care...")

    url = "https://datacenter.aecf.org/rawdata.axd?ind=6243&dtm=14413"
    xlsx_path = RAW_DIR / "foster_care_kidscount.xlsx"

    if not _curl_download(url, xlsx_path):
        return 0

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        loc_type, location, timeframe, data_format, data_val = row[:5]
        if loc_type != "State":
            continue
        if not data_val:
            continue

        state_name = str(location).strip()
        state_code = STATE_NAME_TO_CODE.get(state_name)
        if not state_code:
            continue

        try:
            count_val = int(float(str(data_val)))
            year = int(str(timeframe))
        except (ValueError, TypeError):
            continue

        rows.append({
            "state_code": state_code,
            "state_name": state_name,
            "year": year,
            "children_in_foster_care": count_val,
        })

    if not rows:
        print("  SKIPPED -- no rows parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(rows)
    df["source"] = "datacenter.aecf.org/AFCARS"
    df["snapshot_date"] = SNAPSHOT_DATE

    con.execute("CREATE OR REPLACE TABLE _fact_fc AS SELECT * FROM df")

    count = write_parquet(con, "_fact_fc", _snapshot_path("foster_care"), dry_run)
    if count > 0:
        states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_fc").fetchone()[0]
        years = con.execute("SELECT MIN(year), MAX(year) FROM _fact_fc").fetchone()
        print(f"  {count:,} rows, {states} states, {years[0]}-{years[1]}")
    con.execute("DROP TABLE IF EXISTS _fact_fc")
    return count


# ---------------------------------------------------------------------------
# 10. Title V MCH Block Grant Funding (HRSA TVIS API)
# ---------------------------------------------------------------------------
def build_title_v_mch(con, dry_run: bool) -> int:
    """Title V MCH block grant funding by source and individuals served, by state."""
    print("\n[10/10] Building fact_title_v_mch...")

    all_rows = []
    state_codes = list(STATE_NAME_TO_CODE.values())

    for year in [2021, 2022, 2023, 2024]:
        for sc in state_codes:
            # Funding by source
            url = f"https://mchb.tvisdata.hrsa.gov/api/values/GetFinancialFundingBySourceData?year={year}&stateCode={sc}"
            try:
                result = subprocess.run(
                    ["curl", "-s", url],
                    capture_output=True, text=True, timeout=15,
                )
                if result.returncode != 0 or not result.stdout.strip():
                    continue
                data = json.loads(result.stdout)
                if not data.get("CallSuccess"):
                    continue
                pie = data.get("result", {}).get("pieResult", [])
                for item in pie:
                    val = item.get("Value")
                    if val is None:
                        continue
                    all_rows.append({
                        "state_code": sc,
                        "year": year,
                        "measure_type": "funding",
                        "category": str(item.get("Category", "")).strip(),
                        "value": float(val),
                    })
            except Exception:
                continue

            # Individuals served (Form 5a)
            url2 = f"https://mchb.tvisdata.hrsa.gov/api/values/GetForm5aDataByState?year={year}&stateCode={sc}"
            try:
                result = subprocess.run(
                    ["curl", "-s", url2],
                    capture_output=True, text=True, timeout=15,
                )
                if result.returncode != 0 or not result.stdout.strip():
                    continue
                data = json.loads(result.stdout)
                if not data.get("CallSuccess"):
                    continue
                r = data.get("result", {})
                cols = r.get("result1", r).get("columnResult5a", []) if isinstance(r.get("result1", r), dict) else []
                for item in cols:
                    val = item.get("Provided5aValue")
                    pct = item.get("Calculated5aPercent")
                    name = " ".join(str(item.get("Name", "")).split())
                    if val is not None:
                        all_rows.append({
                            "state_code": sc,
                            "year": year,
                            "measure_type": "individuals_served",
                            "category": name,
                            "value": float(val),
                        })
                    if pct is not None:
                        all_rows.append({
                            "state_code": sc,
                            "year": year,
                            "measure_type": "pct_population_served",
                            "category": name,
                            "value": float(pct),
                        })
            except Exception:
                continue

        print(f"  Year {year}: {len([r for r in all_rows if r['year'] == year])} records")

    if not all_rows:
        print("  SKIPPED -- no data from HRSA TVIS API")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_rows)
    df["source"] = "mchb.tvisdata.hrsa.gov"
    df["snapshot_date"] = SNAPSHOT_DATE

    con.execute("CREATE OR REPLACE TABLE _fact_tv AS SELECT * FROM df")

    count = write_parquet(con, "_fact_tv", _snapshot_path("title_v_mch"), dry_run)
    if count > 0:
        states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fact_tv").fetchone()[0]
        types = con.execute("SELECT DISTINCT measure_type FROM _fact_tv").fetchall()
        print(f"  {count:,} rows, {states} states, types: {[r[0] for r in types]}")
    con.execute("DROP TABLE IF EXISTS _fact_tv")
    return count


# ---------------------------------------------------------------------------
# Registry & Main
# ---------------------------------------------------------------------------
ALL_TABLES = {
    "fact_cdc_natality": build_cdc_natality,
    "fact_infant_mortality_state": build_infant_mortality_state,
    "fact_infant_mortality_quarterly": build_infant_mortality_quarterly,
    "fact_child_vaccination": build_child_vaccination,
    "fact_adolescent_vaccination": build_adolescent_vaccination,
    "fact_teen_birth_rate": build_teen_birth_rate,
    "fact_wic_nutrition": build_wic_nutrition,
    "fact_wic_participation": build_wic_participation,
    "fact_foster_care": build_foster_care,
    "fact_title_v_mch": build_title_v_mch,
}


def main():
    parser = argparse.ArgumentParser(description="Ingest maternal & child health data")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--table", type=str, help="Build a single table")
    args = parser.parse_args()

    RAW_DIR.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect()
    total_rows = 0
    tables_built = 0

    if args.table:
        if args.table not in ALL_TABLES:
            print(f"Unknown table: {args.table}")
            print(f"Available: {list(ALL_TABLES.keys())}")
            return
        count = ALL_TABLES[args.table](con, args.dry_run)
        total_rows += count
        tables_built += 1 if count > 0 else 0
    else:
        for name, builder in ALL_TABLES.items():
            try:
                count = builder(con, args.dry_run)
                total_rows += count
                tables_built += 1 if count > 0 else 0
            except Exception as e:
                print(f"  ERROR building {name}: {e}")

    con.close()
    print(f"\n{'='*60}")
    print(f"Maternal & Child Health ingestion complete.")
    print(f"  Tables built: {tables_built}")
    print(f"  Total rows:   {total_rows:,}")
    print(f"  Snapshot:     {SNAPSHOT_DATE}")
    if not args.dry_run:
        print(f"\nRemember to update db.py fact_names with:")
        for name in ALL_TABLES:
            print(f'    "{name}",')


if __name__ == "__main__":
    main()
