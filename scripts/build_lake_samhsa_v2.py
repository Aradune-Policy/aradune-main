#!/usr/bin/env python3
"""
build_lake_samhsa_v2.py — Ingest additional SAMHSA and behavioral health data.

Downloads and processes:
  1. TEDS-A 2023   — Treatment Episode Data Set, Admissions (1.6M records)
                     Aggregated to state × substance × demographics
  2. TEDS-D 2022   — Treatment Episode Data Set, Discharges (1.4M records)
                     Aggregated to state × substance × discharge reason
  3. NSDUH SAE 2024 — NSDUH Small Area Estimates, totals in thousands
                      41 measures × 51 states × 5 age groups
  4. CDC VSRR       — Provisional drug overdose death counts (updated)
                      81.9K rows, state × month × drug indicator

Tables built:
  fact_teds_admissions_detail  — TEDS 2023 admissions by state, substance, demographics
  fact_teds_discharges         — TEDS 2022 discharges by state, substance, outcome
  fact_nsduh_sae_totals_2024   — NSDUH 2024 state estimates in thousands
  fact_cdc_overdose_deaths     — CDC VSRR provisional overdose death counts (refreshed)

Usage:
  python3 scripts/build_lake_samhsa_v2.py
  python3 scripts/build_lake_samhsa_v2.py --dry-run
  python3 scripts/build_lake_samhsa_v2.py --only fact_teds_admissions_detail,fact_teds_discharges
"""

import argparse
import csv
import io
import os
import subprocess
import tempfile
import uuid
import zipfile
from datetime import date
from pathlib import Path

import duckdb

try:
    import openpyxl
except ImportError:
    print("WARNING: openpyxl not installed. NSDUH SAE Excel parsing will fail.")
    openpyxl = None

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
SAMHSA_DIR = RAW_DIR / "samhsa"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

# FIPS code -> state code mapping
STFIPS_TO_STATE = {
    "1": "AL", "2": "AK", "4": "AZ", "5": "AR", "6": "CA",
    "8": "CO", "9": "CT", "10": "DE", "11": "DC", "12": "FL",
    "13": "GA", "15": "HI", "16": "ID", "17": "IL", "18": "IN",
    "19": "IA", "20": "KS", "21": "KY", "22": "LA", "23": "ME",
    "24": "MD", "25": "MA", "26": "MI", "27": "MN", "28": "MS",
    "29": "MO", "30": "MT", "31": "NE", "32": "NV", "33": "NH",
    "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND",
    "39": "OH", "40": "OK", "41": "OR", "42": "PA", "44": "RI",
    "45": "SC", "46": "SD", "47": "TN", "48": "TX", "49": "UT",
    "50": "VT", "51": "VA", "53": "WA", "54": "WV", "55": "WI",
    "56": "WY", "72": "PR",
}

STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY", "Puerto Rico": "PR",
    "Total U.S.": "US", "Northeast": "NORTHEAST", "Midwest": "MIDWEST",
    "South": "SOUTH", "West": "WEST",
}

# TEDS substance codes
SUB_CODES = {
    "1": "None", "2": "Alcohol", "3": "Cocaine/crack", "4": "Marijuana/hashish",
    "5": "Heroin", "6": "Non-prescription methadone", "7": "Other opiates/synthetics",
    "8": "PCP", "9": "Hallucinogens", "10": "Methamphetamine/speed",
    "11": "Other amphetamines", "12": "Other stimulants", "13": "Benzodiazepines",
    "14": "Other tranquilizers", "15": "Barbiturates", "16": "Other sedatives/hypnotics",
    "17": "Inhalants", "18": "Over-the-counter medications", "19": "Other drugs",
    "-9": "Missing/unknown",
}

# TEDS service setting codes
SERVICE_CODES = {
    "1": "Detox, 24-hour, hospital inpatient",
    "2": "Detox, 24-hour, free-standing residential",
    "3": "Rehabilitation/residential, hospital (non-detox)",
    "4": "Rehabilitation/residential, short-term (30 days or fewer)",
    "5": "Rehabilitation/residential, long-term (more than 30 days)",
    "6": "Ambulatory, intensive outpatient",
    "7": "Ambulatory, non-intensive outpatient",
    "8": "Ambulatory, detoxification",
    "-9": "Missing/unknown",
}

# TEDS discharge reason codes (TEDS-D only)
REASON_CODES = {
    "1": "Treatment completed",
    "2": "Dropped out of treatment",
    "3": "Terminated by facility",
    "4": "Transferred to another treatment program or facility",
    "5": "Incarcerated",
    "6": "Death",
    "7": "Other",
    "-9": "Missing/unknown",
}

# Age group codes
AGE_CODES = {
    "1": "12-14", "2": "15-17", "3": "18-20", "4": "21-24",
    "5": "25-29", "6": "30-34", "7": "35-39", "8": "40-44",
    "9": "45-49", "10": "50-54", "11": "55-64", "12": "65+",
    "-9": "Missing",
}

# Race codes
RACE_CODES = {
    "1": "Alaska Native (Aleut, Eskimo, Indian)",
    "2": "American Indian (other than Alaska Native)",
    "3": "Asian/Pacific Islander",
    "4": "Black or African American",
    "5": "White",
    "9": "Other single race",
    "-9": "Missing/unknown",
}

# Sex/Gender codes
SEX_CODES = {
    "1": "Male", "2": "Female", "-9": "Missing/unknown",
}

# NSDUH table measure mapping (table number -> measure_id, measure_name)
NSDUH_TABLE_MAP = {
    1: ("illicit_drug_use_past_month", "Illicit Drug Use in the Past Month"),
    2: ("marijuana_use_past_year", "Marijuana Use in the Past Year"),
    3: ("marijuana_use_past_month", "Marijuana Use in the Past Month"),
    6: ("illicit_drug_non_marijuana_past_month", "Illicit Drug Use Other Than Marijuana in the Past Month"),
    7: ("cocaine_use_past_year", "Cocaine Use in the Past Year"),
    9: ("heroin_use_past_year", "Heroin Use in the Past Year"),
    11: ("hallucinogen_use_past_year", "Hallucinogen Use in the Past Year"),
    12: ("methamphetamine_use_past_year", "Methamphetamine Use in the Past Year"),
    13: ("rx_opioid_misuse_past_year", "Prescription Opioid Misuse in the Past Year"),
    14: ("opioid_misuse_past_year", "Opioid Misuse in the Past Year"),
    15: ("alcohol_use_past_month", "Alcohol Use in the Past Month"),
    16: ("binge_alcohol_past_month", "Binge Alcohol Use in the Past Month"),
    19: ("tobacco_use_past_month", "Tobacco Product Use in the Past Month"),
    20: ("cigarette_use_past_month", "Cigarette Use in the Past Month"),
    21: ("nicotine_vaping_past_month", "Nicotine Vaping in the Past Month"),
    24: ("substance_use_disorder_past_year", "Substance Use Disorder in the Past Year"),
    25: ("alcohol_use_disorder_past_year", "Alcohol Use Disorder in the Past Year"),
    27: ("drug_use_disorder_past_year", "Drug Use Disorder in the Past Year"),
    28: ("rx_opioid_use_disorder_past_year", "Prescription Opioid Use Disorder in the Past Year"),
    29: ("opioid_use_disorder_past_year", "Opioid Use Disorder in the Past Year"),
    30: ("received_su_treatment_past_year", "Received Substance Use Treatment in the Past Year"),
    31: ("needing_su_treatment_past_year", "Classified as Needing Substance Use Treatment in the Past Year"),
    32: ("unmet_su_treatment_need_past_year", "Did Not Receive Substance Use Treatment (Among Those Needing It)"),
    33: ("any_mental_illness_past_year", "Any Mental Illness in the Past Year"),
    34: ("serious_mental_illness_past_year", "Serious Mental Illness in the Past Year"),
    35: ("co_occurring_sud_ami_past_year", "Co-occurring SUD and Any Mental Illness in the Past Year"),
    36: ("co_occurring_sud_smi_past_year", "Co-occurring SUD and Serious Mental Illness in the Past Year"),
    37: ("received_mh_treatment_past_year", "Received Mental Health Treatment in the Past Year"),
    38: ("major_depressive_episode_past_year", "Major Depressive Episode in the Past Year"),
    39: ("serious_suicidal_thoughts_past_year", "Had Serious Thoughts of Suicide in the Past Year"),
    40: ("suicide_plans_past_year", "Made Any Suicide Plans in the Past Year"),
    41: ("suicide_attempt_past_year", "Attempted Suicide in the Past Year"),
}


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    """Write a DuckDB table to Parquet with ZSTD compression."""
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def _download_file(url: str, dest: Path, desc: str = "") -> bool:
    """Download a file using curl (urllib has issues on macOS)."""
    if dest.exists():
        print(f"  Already downloaded: {dest.name} ({dest.stat().st_size / 1e6:.1f} MB)")
        return True

    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"  Downloading {desc or dest.name}...")
    result = subprocess.run(
        ["curl", "-sL", "-o", str(dest), url],
        capture_output=True, text=True, timeout=300,
    )
    if result.returncode != 0:
        print(f"  ERROR downloading: {result.stderr}")
        return False

    if dest.stat().st_size < 1000:
        # Might be an HTML error page
        with open(dest, "rb") as f:
            head = f.read(200)
        if b"<!DOCTYPE" in head or b"<html" in head:
            print(f"  ERROR: Got HTML instead of data file")
            dest.unlink()
            return False

    print(f"  Downloaded {dest.name} ({dest.stat().st_size / 1e6:.1f} MB)")
    return True


# ──────────────────────────────────────────────────────────────────────────────
# Table 1: TEDS-A 2023 Admissions Detail (aggregated)
# ──────────────────────────────────────────────────────────────────────────────

def build_teds_admissions_detail(con, dry_run: bool) -> int:
    """
    Aggregate TEDS-A 2023 individual-level admissions to:
      state_code × primary_substance × service_setting × age_group × sex × race
    """
    print("\n[1/4] Building fact_teds_admissions_detail (TEDS-A 2023)...")

    zip_path = SAMHSA_DIR / "teds_a_2023.zip"
    csv_path = SAMHSA_DIR / "tedsa_puf_2023.csv"

    if not csv_path.exists():
        url = "https://www.samhsa.gov/data/system/files/media-puf-file/teds-a-2023-ds0001-bndl-data-csv_v1.zip"
        if not _download_file(url, zip_path, "TEDS-A 2023 CSV"):
            print("  SKIP: Could not download TEDS-A 2023")
            return 0

        print("  Extracting ZIP...")
        with zipfile.ZipFile(zip_path) as zf:
            for name in zf.namelist():
                if name.endswith(".csv"):
                    with zf.open(name) as src, open(csv_path, "wb") as dst:
                        dst.write(src.read())
                    break
        print(f"  Extracted: {csv_path.name} ({csv_path.stat().st_size / 1e6:.1f} MB)")

    # Read CSV and aggregate using Python for reliability
    print("  Parsing TEDS-A 2023 admissions...")
    agg = {}
    row_count = 0
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row_count += 1
            stfips = row.get("STFIPS", "")
            state_code = STFIPS_TO_STATE.get(stfips, "")
            if not state_code:
                continue

            sub1 = row.get("SUB1", "-9")
            services = row.get("SERVICES", "-9")
            age = row.get("AGE", "-9")
            sex = row.get("SEX", row.get("GENDER", "-9"))
            race = row.get("RACE", "-9")

            primary_substance = SUB_CODES.get(sub1, "Unknown")
            service_setting = SERVICE_CODES.get(services, "Unknown")
            age_group = AGE_CODES.get(age, "Unknown")
            sex_label = SEX_CODES.get(sex, "Unknown")
            race_label = RACE_CODES.get(race, "Unknown")

            # Substance flags
            alcflg = row.get("ALCFLG", "0") == "1"
            herflg = row.get("HERFLG", "0") == "1"
            cokeflg = row.get("COKEFLG", "0") == "1"
            marflg = row.get("MARFLG", "0") == "1"
            mthamflg = row.get("MTHAMFLG", "0") == "1"
            opsynflg = row.get("OPSYNFLG", "0") == "1"
            benzflg = row.get("BENZFLG", "0") == "1"

            methuse = row.get("METHUSE", "-9")
            mat_flag = methuse == "1"  # 1 = yes

            key = (state_code, primary_substance, service_setting, age_group, sex_label, race_label)
            if key not in agg:
                agg[key] = {
                    "admissions": 0,
                    "alcohol_involved": 0,
                    "heroin_involved": 0,
                    "cocaine_involved": 0,
                    "marijuana_involved": 0,
                    "meth_involved": 0,
                    "opioid_synth_involved": 0,
                    "benzo_involved": 0,
                    "mat_used": 0,
                }
            rec = agg[key]
            rec["admissions"] += 1
            if alcflg: rec["alcohol_involved"] += 1
            if herflg: rec["heroin_involved"] += 1
            if cokeflg: rec["cocaine_involved"] += 1
            if marflg: rec["marijuana_involved"] += 1
            if mthamflg: rec["meth_involved"] += 1
            if opsynflg: rec["opioid_synth_involved"] += 1
            if benzflg: rec["benzo_involved"] += 1
            if mat_flag: rec["mat_used"] += 1

    print(f"  Parsed {row_count:,} individual records -> {len(agg):,} aggregated rows")

    # Build DuckDB table
    rows = []
    for (state_code, primary_substance, service_setting, age_group, sex, race), vals in agg.items():
        rows.append({
            "state_code": state_code,
            "year": 2023,
            "primary_substance": primary_substance,
            "service_setting": service_setting,
            "age_group": age_group,
            "sex": sex,
            "race": race,
            "admissions": vals["admissions"],
            "alcohol_involved": vals["alcohol_involved"],
            "heroin_involved": vals["heroin_involved"],
            "cocaine_involved": vals["cocaine_involved"],
            "marijuana_involved": vals["marijuana_involved"],
            "meth_involved": vals["meth_involved"],
            "opioid_synth_involved": vals["opioid_synth_involved"],
            "benzo_involved": vals["benzo_involved"],
            "mat_used": vals["mat_used"],
            "source": "SAMHSA TEDS-A 2023 PUF",
            "snapshot": SNAPSHOT_DATE,
        })

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS fact_teds_admissions_detail")
    con.execute("CREATE TABLE fact_teds_admissions_detail AS SELECT * FROM df")

    return write_parquet(con, "fact_teds_admissions_detail", _snapshot_path("teds_admissions_detail"), dry_run)


# ──────────────────────────────────────────────────────────────────────────────
# Table 2: TEDS-D 2022 Discharges (aggregated)
# ──────────────────────────────────────────────────────────────────────────────

def build_teds_discharges(con, dry_run: bool) -> int:
    """
    Aggregate TEDS-D 2022 individual-level discharges to:
      state_code × primary_substance × service_setting × discharge_reason × age_group × sex
    """
    print("\n[2/4] Building fact_teds_discharges (TEDS-D 2022)...")

    zip_path = SAMHSA_DIR / "teds_d_2022.zip"
    csv_path = SAMHSA_DIR / "tedsd_puf_2022.csv"

    if not csv_path.exists():
        url = "https://www.samhsa.gov/data/system/files/media-puf-file/TEDS-D-2022-DS0001-bndl-data-csv_v1.zip"
        if not _download_file(url, zip_path, "TEDS-D 2022 CSV"):
            print("  SKIP: Could not download TEDS-D 2022")
            return 0

        print("  Extracting ZIP...")
        with zipfile.ZipFile(zip_path) as zf:
            for name in zf.namelist():
                if name.endswith(".csv"):
                    with zf.open(name) as src, open(csv_path, "wb") as dst:
                        dst.write(src.read())
                    break
        print(f"  Extracted: {csv_path.name} ({csv_path.stat().st_size / 1e6:.1f} MB)")

    print("  Parsing TEDS-D 2022 discharges...")
    agg = {}
    row_count = 0
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row_count += 1
            stfips = row.get("STFIPS", "")
            state_code = STFIPS_TO_STATE.get(stfips, "")
            if not state_code:
                continue

            sub1 = row.get("SUB1", "-9")
            services = row.get("SERVICES", "-9")
            reason = row.get("REASON", "-9")
            age = row.get("AGE", "-9")
            sex = row.get("GENDER", row.get("SEX", "-9"))

            primary_substance = SUB_CODES.get(sub1, "Unknown")
            service_setting = SERVICE_CODES.get(services, "Unknown")
            discharge_reason = REASON_CODES.get(reason, "Unknown")
            age_group = AGE_CODES.get(age, "Unknown")
            sex_label = SEX_CODES.get(sex, "Unknown")

            # Length of stay
            los_raw = row.get("LOS", "-9")
            try:
                los = int(los_raw) if los_raw not in ("-9", "") else None
            except ValueError:
                los = None

            # Substance flags
            alcflg = row.get("ALCFLG", "0") == "1"
            herflg = row.get("HERFLG", "0") == "1"
            mthamflg = row.get("MTHAMFLG", "0") == "1"
            opsynflg = row.get("OPSYNFLG", "0") == "1"

            key = (state_code, primary_substance, service_setting, discharge_reason, age_group, sex_label)
            if key not in agg:
                agg[key] = {
                    "discharges": 0,
                    "completed_treatment": 0,
                    "dropped_out": 0,
                    "alcohol_involved": 0,
                    "heroin_involved": 0,
                    "meth_involved": 0,
                    "opioid_synth_involved": 0,
                    "total_los_days": 0,
                    "los_count": 0,
                }
            rec = agg[key]
            rec["discharges"] += 1
            if reason == "1": rec["completed_treatment"] += 1
            if reason == "2": rec["dropped_out"] += 1
            if alcflg: rec["alcohol_involved"] += 1
            if herflg: rec["heroin_involved"] += 1
            if mthamflg: rec["meth_involved"] += 1
            if opsynflg: rec["opioid_synth_involved"] += 1
            if los is not None and los >= 0:
                rec["total_los_days"] += los
                rec["los_count"] += 1

    print(f"  Parsed {row_count:,} individual records -> {len(agg):,} aggregated rows")

    rows = []
    for (state_code, primary_substance, service_setting, discharge_reason, age_group, sex), vals in agg.items():
        avg_los = round(vals["total_los_days"] / vals["los_count"], 1) if vals["los_count"] > 0 else None
        rows.append({
            "state_code": state_code,
            "year": 2022,
            "primary_substance": primary_substance,
            "service_setting": service_setting,
            "discharge_reason": discharge_reason,
            "age_group": age_group,
            "sex": sex,
            "discharges": vals["discharges"],
            "completed_treatment": vals["completed_treatment"],
            "dropped_out": vals["dropped_out"],
            "alcohol_involved": vals["alcohol_involved"],
            "heroin_involved": vals["heroin_involved"],
            "meth_involved": vals["meth_involved"],
            "opioid_synth_involved": vals["opioid_synth_involved"],
            "avg_length_of_stay_days": avg_los,
            "source": "SAMHSA TEDS-D 2022 PUF",
            "snapshot": SNAPSHOT_DATE,
        })

    import pandas as pd
    df = pd.DataFrame(rows)
    con.execute("DROP TABLE IF EXISTS fact_teds_discharges")
    con.execute("CREATE TABLE fact_teds_discharges AS SELECT * FROM df")

    return write_parquet(con, "fact_teds_discharges", _snapshot_path("teds_discharges"), dry_run)


# ──────────────────────────────────────────────────────────────────────────────
# Table 3: NSDUH SAE 2024 Totals (in thousands)
# ──────────────────────────────────────────────────────────────────────────────

def build_nsduh_sae_totals(con, dry_run: bool) -> int:
    """
    Parse NSDUH 2024 Small Area Estimates totals Excel.
    41 tables, each with state × age group estimates in thousands.
    """
    print("\n[3/4] Building fact_nsduh_sae_totals_2024 (NSDUH SAE totals)...")

    if openpyxl is None:
        print("  SKIP: openpyxl not installed")
        return 0

    xlsx_path = SAMHSA_DIR / "nsduh_sae_totals_2024.xlsx"
    if not xlsx_path.exists():
        url = "https://www.samhsa.gov/data/sites/default/files/reports/rpt56987/2024-nsduh-sae-totals-tables-csvs/2024-nsduh-sae-totals-tables.xlsx"
        if not _download_file(url, xlsx_path, "NSDUH SAE 2024 totals Excel"):
            print("  SKIP: Could not download NSDUH SAE totals")
            return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Table "):
            continue
        table_num = int(sheet_name.replace("Table ", ""))
        if table_num not in NSDUH_TABLE_MAP:
            continue

        measure_id, measure_name = NSDUH_TABLE_MAP[table_num]
        ws = wb[sheet_name]

        # Find the header row (contains "Order" and "State")
        header_row = None
        header_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row and row[0] == "Order" and row[1] == "State":
                header_row = row
                header_idx = i
                break

        if header_row is None:
            continue

        # Parse column headers for age groups
        # Columns: Order, State, then groups of 3 (Estimate, CI Lower, CI Upper) per age group
        # Header format: "12+\nEstimate", "12+\n95% CI (Lower)", etc.
        age_groups = []
        col_idx = 2
        while col_idx < len(header_row) and header_row[col_idx] is not None:
            col_label = str(header_row[col_idx]).strip()
            # Extract age group from "12+\nEstimate" or "12-17\n95% CI (Lower)"
            if "\n" in col_label:
                age_part = col_label.split("\n")[0].strip()
                type_part = col_label.split("\n")[1].strip().lower()
                if "estimate" in type_part:
                    age_groups.append({"age_group": age_part, "est_col": col_idx})
                elif "lower" in type_part:
                    if age_groups:
                        age_groups[-1]["ci_lower_col"] = col_idx
                elif "upper" in type_part:
                    if age_groups:
                        age_groups[-1]["ci_upper_col"] = col_idx
            col_idx += 1

        # Read data rows
        data_started = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_idx:
                continue
            if not row or row[0] is None or row[1] is None:
                continue
            state_name = str(row[1]).strip()
            if not state_name or state_name.startswith("NOTE") or state_name.startswith("Source"):
                continue

            state_code = STATE_NAME_TO_CODE.get(state_name, "")
            if not state_code:
                continue

            for ag in age_groups:
                try:
                    estimate = row[ag["est_col"]]
                    ci_lower = row[ag.get("ci_lower_col", 0)] if "ci_lower_col" in ag else None
                    ci_upper = row[ag.get("ci_upper_col", 0)] if "ci_upper_col" in ag else None

                    if estimate is None:
                        continue
                    estimate = float(estimate)

                    all_rows.append({
                        "state_code": state_code,
                        "table_number": table_num,
                        "measure_id": measure_id,
                        "measure_name": measure_name,
                        "age_group": ag["age_group"],
                        "estimate_thousands": round(estimate, 1),
                        "ci_lower_thousands": round(float(ci_lower), 1) if ci_lower is not None else None,
                        "ci_upper_thousands": round(float(ci_upper), 1) if ci_upper is not None else None,
                        "data_period": "2023-2024",
                        "source": "SAMHSA NSDUH 2024 SAE Totals",
                        "snapshot": SNAPSHOT_DATE,
                    })
                except (TypeError, ValueError, IndexError):
                    continue

    wb.close()
    print(f"  Parsed {len(all_rows):,} rows from {len(NSDUH_TABLE_MAP)} measures")

    if not all_rows:
        print("  SKIP: No data parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_rows)
    con.execute("DROP TABLE IF EXISTS fact_nsduh_sae_totals_2024")
    con.execute("CREATE TABLE fact_nsduh_sae_totals_2024 AS SELECT * FROM df")

    return write_parquet(con, "fact_nsduh_sae_totals_2024", _snapshot_path("nsduh_sae_totals_2024"), dry_run)


# ──────────────────────────────────────────────────────────────────────────────
# Table 4: CDC VSRR Provisional Drug Overdose Deaths (refresh)
# ──────────────────────────────────────────────────────────────────────────────

def build_cdc_overdose_deaths(con, dry_run: bool) -> int:
    """
    Refresh CDC VSRR provisional drug overdose death counts.
    State × month × drug indicator, 2015-present.
    """
    print("\n[4/4] Building fact_cdc_overdose_deaths (CDC VSRR refresh)...")

    csv_path = SAMHSA_DIR / "vsrr_overdose_deaths.csv"
    if not csv_path.exists():
        url = "https://data.cdc.gov/api/views/xkb8-kh2a/rows.csv?accessType=DOWNLOAD"
        if not _download_file(url, csv_path, "CDC VSRR Overdose Deaths"):
            print("  SKIP: Could not download CDC VSRR data")
            return 0

    # Read directly with DuckDB
    con.execute("DROP TABLE IF EXISTS fact_cdc_overdose_deaths")
    con.execute(f"""
        CREATE TABLE fact_cdc_overdose_deaths AS
        SELECT
            "State" AS state_code,
            CAST("Year" AS INTEGER) AS year,
            "Month" AS month,
            "Period" AS period,
            "Indicator" AS indicator,
            TRY_CAST("Data Value" AS DOUBLE) AS data_value,
            TRY_CAST("Percent Complete" AS DOUBLE) AS pct_complete,
            TRY_CAST("Percent Pending Investigation" AS DOUBLE) AS pct_pending_investigation,
            "State Name" AS state_name,
            "Footnote" AS footnote,
            "Footnote Symbol" AS footnote_symbol,
            TRY_CAST("Predicted Value" AS DOUBLE) AS predicted_value,
            '{SNAPSHOT_DATE}' AS snapshot
        FROM read_csv_auto('{csv_path}', all_varchar=true)
    """)

    return write_parquet(con, "fact_cdc_overdose_deaths", _snapshot_path("cdc_overdose_deaths"), dry_run)


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

ALL_TABLES = {
    "fact_teds_admissions_detail": build_teds_admissions_detail,
    "fact_teds_discharges": build_teds_discharges,
    "fact_nsduh_sae_totals_2024": build_nsduh_sae_totals,
    "fact_cdc_overdose_deaths": build_cdc_overdose_deaths,
}


def main():
    parser = argparse.ArgumentParser(description="Build SAMHSA & BH lake tables")
    parser.add_argument("--dry-run", action="store_true", help="Parse but don't write Parquet")
    parser.add_argument("--only", type=str, default="", help="Comma-separated list of tables to build")
    args = parser.parse_args()

    SAMHSA_DIR.mkdir(parents=True, exist_ok=True)

    targets = set(args.only.split(",")) if args.only else set(ALL_TABLES.keys())

    con = duckdb.connect()
    results = {}

    for name, builder in ALL_TABLES.items():
        if name not in targets:
            continue
        try:
            count = builder(con, args.dry_run)
            results[name] = count
        except Exception as e:
            print(f"  ERROR building {name}: {e}")
            import traceback
            traceback.print_exc()
            results[name] = -1

    print("\n" + "=" * 60)
    print("SAMHSA v2 Build Summary")
    print("=" * 60)
    total = 0
    for name, count in results.items():
        status = f"{count:,} rows" if count > 0 else ("SKIP" if count == 0 else "ERROR")
        print(f"  {name:40s} {status}")
        if count > 0:
            total += count
    print(f"  {'TOTAL':40s} {total:,} rows")
    print(f"  Run ID: {RUN_ID}")
    print(f"  Snapshot: {SNAPSHOT_DATE}")

    con.close()


if __name__ == "__main__":
    main()
