#!/usr/bin/env python3
"""
build_lake_cms_program_stats.py — Ingest CMS Program Statistics utilization
and enrollment detail data into the Aradune data lake.

Sources (CY2023 ZIP files from data.cms.gov):
  - Medicare Inpatient Hospital Utilization (by state)
  - Medicare Skilled Nursing Facility Utilization (by state)
  - Medicare Home Health Agency Utilization (by state)
  - Medicare Hospice Utilization (by state)
  - Medicare-Medicaid Dual Enrollment Detail (by state, eligibility type)
  - Medicare Advantage Enrollment (by state, entitlement type)
  - Medicare Part D Enrollment (by state, plan type, LIS status)
  - Monthly MA/Part D Enrollment by Contract/Plan/State (CPSC, Dec 2023)

Tables built:
  fact_cps_inpatient_utilization   — Inpatient discharges, days, payments by state (CY2023)
  fact_cps_snf_utilization         — SNF admissions, days, payments by state (CY2023)
  fact_cps_hha_utilization         — HHA visits, episodes, payments by state (CY2023)
  fact_cps_hospice_utilization     — Hospice days, payments by state (CY2023)
  fact_cps_dual_enrollment_detail  — Dual enrollment by state and eligibility type (CY2023)
  fact_cps_ma_enrollment           — MA enrollment by state, entitlement, plan type (CY2023)
  fact_cps_part_d_enrollment       — Part D enrollment by state, plan type, LIS (CY2023)
  fact_ma_enrollment_plan          — MA/Part D enrollment by contract/plan/state (Dec 2023)

Usage:
  python3 scripts/build_lake_cms_program_stats.py
  python3 scripts/build_lake_cms_program_stats.py --dry-run
  python3 scripts/build_lake_cms_program_stats.py --only fact_cps_inpatient_utilization
"""

import argparse
import csv
import io
import json
import os
import re
import subprocess
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path

import duckdb

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl")
    raise

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "cms_program_stats"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

# ---------------------------------------------------------------------------
# Download URLs — CY2023 (most recent available as of March 2026)
# ---------------------------------------------------------------------------

DOWNLOADS = {
    "inpatient_2023.zip": "https://data.cms.gov/sites/default/files/2026-01/7b6c177c-76bc-45db-a26f-31640033ca85/MDCR%20INPT%20HOSP_CPS_05UIP_2023.zip",
    "snf_2023.zip": "https://data.cms.gov/sites/default/files/2026-01/7fb9cae0-38c4-4062-9915-ce7c6e9788da/MDCR%20SNF_CPS06USN_2023.zip",
    "hha_2023.zip": "https://data.cms.gov/sites/default/files/2026-01/MDCR%20HHA_CPS_07UHH_2023.zip",
    "hospice_2023.zip": "https://data.cms.gov/sites/default/files/2026-01/7d79ed87-18a1-445a-b9f2-a85c26989274/MDCR%20HOSPICE_CPS_08UHS_2023.zip",
    "dual_enrollment_2023.zip": "https://data.cms.gov/sites/default/files/2025-09/1104c73c-6cb7-422c-bb24-73236d1b5767/MDCR%20ENROLL%20AB%2040-48_CPS_02ENR_2023.zip",
    "ma_enrollment_2023.zip": "https://data.cms.gov/sites/default/files/2025-09/6723fcf5-4d03-476e-a72d-f382ef0cd856/MDCR%20ENROLL%20AB%2015-20_CPS_02ENR_2023.zip",
    "part_d_enrollment_2023.zip": "https://data.cms.gov/sites/default/files/2025-09/c0458b01-5731-43b4-a45c-16944a8253e7/MDCR%20ENROLL%20D_CPS_03EPD_2023.zip",
    "cpsc_enrollment_2023_12.zip": "https://www.cms.gov/files/zip/monthly-enrollment-cpsc-december-2023.zip",
}

# ---------------------------------------------------------------------------
# State name to code mapping
# ---------------------------------------------------------------------------

STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
    "U.S. Virgin Islands": "VI",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    """Write a DuckDB table to Parquet. Returns row count."""
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_kb = out_path.stat().st_size / 1024
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_kb:.1f} KB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def _download(url: str, dest: Path, label: str = "") -> bool:
    """Download a file using curl subprocess."""
    if dest.exists() and dest.stat().st_size > 100:
        print(f"  [cached] {label or dest.name}")
        return True
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"  Downloading {label or dest.name}...")
    result = subprocess.run(
        ["curl", "-sL", "-o", str(dest), url],
        capture_output=True, timeout=600
    )
    if result.returncode != 0 or not dest.exists() or dest.stat().st_size < 100:
        print(f"  FAILED to download {label or url}")
        return False
    size_mb = dest.stat().st_size / (1024 * 1024)
    print(f"  Downloaded {label or dest.name} ({size_mb:.1f} MB)")
    return True


def _extract_xlsx_from_zip(zip_path: Path) -> Path | None:
    """Extract the first xlsx from a ZIP file, return its path."""
    with zipfile.ZipFile(zip_path) as z:
        for name in z.namelist():
            if name.endswith(".xlsx"):
                z.extract(name, zip_path.parent)
                return zip_path.parent / name
    return None


def _clean_number(val):
    """Convert cell value to float, handling commas and CMS formatting."""
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("$", "").replace("%", "")
    if s in ("", "–", "-", "—", "N/A", "n/a", "*", ".", "..", "BLANK", "None"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _clean_state_name(name: str) -> str:
    """Strip footnote superscripts and whitespace from state/area names."""
    if not name:
        return ""
    return re.sub(r'[\d,*†‡§¹²³⁴⁵⁶⁷⁸⁹⁰˒]+$', '', str(name)).strip()


def _state_code_from_name(name: str) -> str | None:
    """Convert a state name to 2-letter code, handling CMS formatting quirks."""
    cleaned = _clean_state_name(name)
    # Try exact match
    code = STATE_NAME_TO_CODE.get(cleaned)
    if code:
        return code
    # Try case-insensitive
    for k, v in STATE_NAME_TO_CODE.items():
        if k.lower() == cleaned.lower():
            return v
    return None


def _is_state_row(area_name: str) -> bool:
    """Check if a row is a state-level data row (not header, total, blank, etc.)."""
    if not area_name or str(area_name).strip() == "":
        return False
    cleaned = _clean_state_name(str(area_name).strip())
    skip = {"All Areas", "United States", "BLANK", "Total", "",
            "Territories, Possessions, and Other", "Foreign Countries",
            "Unknown", "All Beneficiaries", "Age", "Sex", "Race",
            "Under 65 Years", "65 Years and Over", "Male", "Female",
            "White", "Black", "Hispanic", "Asian/Pacific Islander",
            "American Indian/Alaska Native", "Other", "Territories",
            "Possessions", "Demographic Characteristic", "Type of Plan and Demographic Characteristics"}
    if cleaned in skip:
        return False
    # Must be a recognizable state/territory
    return _state_code_from_name(cleaned) is not None


# ---------------------------------------------------------------------------
# INPATIENT UTILIZATION (Sheet 3: by state, CY2023)
# ---------------------------------------------------------------------------

def build_cps_inpatient_utilization(con, dry_run: bool) -> int:
    """Parse inpatient hospital utilization by state from CPS Excel."""
    print("Building fact_cps_inpatient_utilization...")

    zip_path = RAW_DIR / "inpatient_2023.zip"
    if not _download(DOWNLOADS["inpatient_2023.zip"], zip_path, "CPS Inpatient 2023"):
        return 0
    xlsx_path = _extract_xlsx_from_zip(zip_path)
    if not xlsx_path:
        print("  ERROR: No xlsx in ZIP")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    # Sheet 3 = by state, CY2023
    ws = wb[[s for s in wb.sheetnames if "HOSP 3" in s][0]]

    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        area = str(row[0]).strip() if row[0] else ""
        if not _is_state_row(area):
            continue
        state_code = _state_code_from_name(area)
        if not state_code:
            continue
        rows.append({
            "year": 2023,
            "state_code": state_code,
            "part_a_enrollees": _clean_number(row[1]),
            "persons_with_utilization": _clean_number(row[2]),
            "total_discharges": _clean_number(row[3]),
            "discharges_per_1000_enrollees": _clean_number(row[4]),
            "total_days_of_care": _clean_number(row[6]),
            "days_per_discharge": _clean_number(row[9]),
            "covered_days": _clean_number(row[10]),
            "covered_days_per_discharge": _clean_number(row[13]),
            "total_program_payments": _clean_number(row[14]),
            "payments_per_enrollee": _clean_number(row[15]),
            "payments_per_discharge": _clean_number(row[17]),
            "payments_per_covered_day": _clean_number(row[18]),
            "total_deductible": _clean_number(row[19]),
            "total_coinsurance": _clean_number(row[20]),
            "discharged_dead": _clean_number(row[21]),
        })
    wb.close()

    if not rows:
        print("  WARNING: No data rows parsed")
        return 0

    con.execute("DROP TABLE IF EXISTS _cps_inpatient")
    con.execute("""
        CREATE TABLE _cps_inpatient (
            year INTEGER, state_code VARCHAR, part_a_enrollees BIGINT,
            persons_with_utilization BIGINT, total_discharges BIGINT,
            discharges_per_1000_enrollees DECIMAL(8,1),
            total_days_of_care BIGINT, days_per_discharge DECIMAL(6,2),
            covered_days BIGINT, covered_days_per_discharge DECIMAL(6,2),
            total_program_payments BIGINT, payments_per_enrollee DECIMAL(10,2),
            payments_per_discharge DECIMAL(10,2), payments_per_covered_day DECIMAL(10,2),
            total_deductible BIGINT, total_coinsurance BIGINT,
            discharged_dead BIGINT
        )
    """)
    for r in rows:
        con.execute("""
            INSERT INTO _cps_inpatient VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [
            r["year"], r["state_code"], r["part_a_enrollees"],
            r["persons_with_utilization"], r["total_discharges"],
            r["discharges_per_1000_enrollees"],
            r["total_days_of_care"], r["days_per_discharge"],
            r["covered_days"], r["covered_days_per_discharge"],
            r["total_program_payments"], r["payments_per_enrollee"],
            r["payments_per_discharge"], r["payments_per_covered_day"],
            r["total_deductible"], r["total_coinsurance"],
            r["discharged_dead"],
        ])

    count = write_parquet(con, "_cps_inpatient", _snapshot_path("cps_inpatient_utilization"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cps_inpatient")
    return count


# ---------------------------------------------------------------------------
# SNF UTILIZATION (Sheet 3: by state, CY2023)
# ---------------------------------------------------------------------------

def build_cps_snf_utilization(con, dry_run: bool) -> int:
    """Parse SNF utilization by state from CPS Excel."""
    print("Building fact_cps_snf_utilization...")

    zip_path = RAW_DIR / "snf_2023.zip"
    if not _download(DOWNLOADS["snf_2023.zip"], zip_path, "CPS SNF 2023"):
        return 0
    xlsx_path = _extract_xlsx_from_zip(zip_path)
    if not xlsx_path:
        print("  ERROR: No xlsx in ZIP")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[[s for s in wb.sheetnames if "SNF 3" in s][0]]

    # Columns from row 4:
    # 0: Area of Residence
    # 1: Total Original Medicare Part A Enrollees
    # 2: Total Persons With Utilization
    # 3: Total Covered Admissions
    # 4: Covered Admissions Per Person With Utilization
    # 5: Covered Admissions Per 1,000 Enrollees
    # 6: Total Covered Days of Care
    # 7: Covered Days Per Admission
    # 8: Covered Days Per Person
    # 9: Covered Days Per 1,000 Enrollees
    # 10: Total Program Payments
    # 11: Payments Per Admission
    # 12: Payments Per Person
    # 13: Payments Per Covered Day
    # 14: Payments Per Enrollee

    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        area = str(row[0]).strip() if row[0] else ""
        if not _is_state_row(area):
            continue
        state_code = _state_code_from_name(area)
        if not state_code:
            continue
        rows.append({
            "year": 2023,
            "state_code": state_code,
            "part_a_enrollees": _clean_number(row[1]),
            "persons_with_utilization": _clean_number(row[2]),
            "total_covered_admissions": _clean_number(row[3]),
            "admissions_per_person": _clean_number(row[4]),
            "admissions_per_1000_enrollees": _clean_number(row[5]),
            "total_covered_days": _clean_number(row[6]),
            "covered_days_per_admission": _clean_number(row[7]),
            "covered_days_per_person": _clean_number(row[8]),
            "covered_days_per_1000_enrollees": _clean_number(row[9]),
            "total_program_payments": _clean_number(row[10]),
            "payments_per_admission": _clean_number(row[11]),
            "payments_per_person": _clean_number(row[12]),
            "payments_per_covered_day": _clean_number(row[13]),
            "payments_per_enrollee": _clean_number(row[14]),
        })
    wb.close()

    if not rows:
        print("  WARNING: No data rows parsed")
        return 0

    con.execute("DROP TABLE IF EXISTS _cps_snf")
    con.execute("""
        CREATE TABLE _cps_snf (
            year INTEGER, state_code VARCHAR, part_a_enrollees BIGINT,
            persons_with_utilization BIGINT, total_covered_admissions BIGINT,
            admissions_per_person DECIMAL(6,2), admissions_per_1000_enrollees DECIMAL(8,2),
            total_covered_days BIGINT, covered_days_per_admission DECIMAL(8,2),
            covered_days_per_person DECIMAL(8,2), covered_days_per_1000_enrollees DECIMAL(10,2),
            total_program_payments BIGINT, payments_per_admission DECIMAL(10,2),
            payments_per_person DECIMAL(10,2), payments_per_covered_day DECIMAL(10,2),
            payments_per_enrollee DECIMAL(10,2)
        )
    """)
    for r in rows:
        con.execute("""INSERT INTO _cps_snf VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [r["year"], r["state_code"], r["part_a_enrollees"],
             r["persons_with_utilization"], r["total_covered_admissions"],
             r["admissions_per_person"], r["admissions_per_1000_enrollees"],
             r["total_covered_days"], r["covered_days_per_admission"],
             r["covered_days_per_person"], r["covered_days_per_1000_enrollees"],
             r["total_program_payments"], r["payments_per_admission"],
             r["payments_per_person"], r["payments_per_covered_day"],
             r["payments_per_enrollee"]])

    count = write_parquet(con, "_cps_snf", _snapshot_path("cps_snf_utilization"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cps_snf")
    return count


# ---------------------------------------------------------------------------
# HHA UTILIZATION (Sheet 3: by state, CY2023)
# ---------------------------------------------------------------------------

def build_cps_hha_utilization(con, dry_run: bool) -> int:
    """Parse home health agency utilization by state from CPS Excel."""
    print("Building fact_cps_hha_utilization...")

    zip_path = RAW_DIR / "hha_2023.zip"
    if not _download(DOWNLOADS["hha_2023.zip"], zip_path, "CPS HHA 2023"):
        return 0
    xlsx_path = _extract_xlsx_from_zip(zip_path)
    if not xlsx_path:
        print("  ERROR: No xlsx in ZIP")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[[s for s in wb.sheetnames if "HHA 3" in s][0]]

    # Cols: Area, Enrollees, Persons, Visits, Visits/Person, Visits/1000,
    #       Episodes, Visits/Episode, Episodes/Person, Episodes/1000,
    #       Payments, Payments/Visit, Payments/Episode, Payments/Person, Payments/Enrollee

    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        area = str(row[0]).strip() if row[0] else ""
        if not _is_state_row(area):
            continue
        state_code = _state_code_from_name(area)
        if not state_code:
            continue
        rows.append({
            "year": 2023,
            "state_code": state_code,
            "total_enrollees": _clean_number(row[1]),
            "persons_with_utilization": _clean_number(row[2]),
            "total_service_visits": _clean_number(row[3]),
            "visits_per_person": _clean_number(row[4]),
            "visits_per_1000_enrollees": _clean_number(row[5]),
            "total_episodes": _clean_number(row[6]),
            "visits_per_episode": _clean_number(row[7]),
            "episodes_per_person": _clean_number(row[8]),
            "episodes_per_1000_enrollees": _clean_number(row[9]),
            "total_program_payments": _clean_number(row[10]),
            "payments_per_visit": _clean_number(row[11]),
            "payments_per_episode": _clean_number(row[12]),
            "payments_per_person": _clean_number(row[13]),
            "payments_per_enrollee": _clean_number(row[14]),
        })
    wb.close()

    if not rows:
        print("  WARNING: No data rows parsed")
        return 0

    con.execute("DROP TABLE IF EXISTS _cps_hha")
    con.execute("""
        CREATE TABLE _cps_hha (
            year INTEGER, state_code VARCHAR, total_enrollees BIGINT,
            persons_with_utilization BIGINT, total_service_visits BIGINT,
            visits_per_person DECIMAL(8,2), visits_per_1000_enrollees DECIMAL(10,2),
            total_episodes BIGINT, visits_per_episode DECIMAL(8,2),
            episodes_per_person DECIMAL(6,2), episodes_per_1000_enrollees DECIMAL(10,2),
            total_program_payments BIGINT, payments_per_visit DECIMAL(10,2),
            payments_per_episode DECIMAL(10,2), payments_per_person DECIMAL(10,2),
            payments_per_enrollee DECIMAL(10,2)
        )
    """)
    for r in rows:
        con.execute("""INSERT INTO _cps_hha VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [r["year"], r["state_code"], r["total_enrollees"],
             r["persons_with_utilization"], r["total_service_visits"],
             r["visits_per_person"], r["visits_per_1000_enrollees"],
             r["total_episodes"], r["visits_per_episode"],
             r["episodes_per_person"], r["episodes_per_1000_enrollees"],
             r["total_program_payments"], r["payments_per_visit"],
             r["payments_per_episode"], r["payments_per_person"],
             r["payments_per_enrollee"]])

    count = write_parquet(con, "_cps_hha", _snapshot_path("cps_hha_utilization"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cps_hha")
    return count


# ---------------------------------------------------------------------------
# HOSPICE UTILIZATION (Sheet 3: by state, CY2023)
# ---------------------------------------------------------------------------

def build_cps_hospice_utilization(con, dry_run: bool) -> int:
    """Parse hospice utilization by state from CPS Excel."""
    print("Building fact_cps_hospice_utilization...")

    zip_path = RAW_DIR / "hospice_2023.zip"
    if not _download(DOWNLOADS["hospice_2023.zip"], zip_path, "CPS Hospice 2023"):
        return 0
    xlsx_path = _extract_xlsx_from_zip(zip_path)
    if not xlsx_path:
        print("  ERROR: No xlsx in ZIP")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[[s for s in wb.sheetnames if "HOSPICE 3" in s][0]]

    # Cols: Area, Part A Enrollees, Persons, Covered Days, Days/Person,
    #       Days/1000 Enrollees, Payments, Payments/Person, Payments/Day,
    #       Payments/Enrollee, Discharged Dead

    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        area = str(row[0]).strip() if row[0] else ""
        if not _is_state_row(area):
            continue
        state_code = _state_code_from_name(area)
        if not state_code:
            continue
        rows.append({
            "year": 2023,
            "state_code": state_code,
            "part_a_enrollees": _clean_number(row[1]),
            "persons_with_utilization": _clean_number(row[2]),
            "total_covered_days": _clean_number(row[3]),
            "covered_days_per_person": _clean_number(row[4]),
            "covered_days_per_1000_enrollees": _clean_number(row[5]),
            "total_program_payments": _clean_number(row[6]),
            "payments_per_person": _clean_number(row[7]),
            "payments_per_covered_day": _clean_number(row[8]),
            "payments_per_enrollee": _clean_number(row[9]),
            "discharged_dead": _clean_number(row[10]),
        })
    wb.close()

    if not rows:
        print("  WARNING: No data rows parsed")
        return 0

    con.execute("DROP TABLE IF EXISTS _cps_hospice")
    con.execute("""
        CREATE TABLE _cps_hospice (
            year INTEGER, state_code VARCHAR, part_a_enrollees BIGINT,
            persons_with_utilization BIGINT, total_covered_days BIGINT,
            covered_days_per_person DECIMAL(8,2), covered_days_per_1000_enrollees DECIMAL(10,2),
            total_program_payments BIGINT, payments_per_person DECIMAL(10,2),
            payments_per_covered_day DECIMAL(10,2), payments_per_enrollee DECIMAL(10,2),
            discharged_dead BIGINT
        )
    """)
    for r in rows:
        con.execute("""INSERT INTO _cps_hospice VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [r["year"], r["state_code"], r["part_a_enrollees"],
             r["persons_with_utilization"], r["total_covered_days"],
             r["covered_days_per_person"], r["covered_days_per_1000_enrollees"],
             r["total_program_payments"], r["payments_per_person"],
             r["payments_per_covered_day"], r["payments_per_enrollee"],
             r["discharged_dead"]])

    count = write_parquet(con, "_cps_hospice", _snapshot_path("cps_hospice_utilization"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cps_hospice")
    return count


# ---------------------------------------------------------------------------
# DUAL ENROLLMENT DETAIL (Sheets 42 + 45 + 48: by state, CY2023)
# Tables 42 (Total MME by state), 45 (Original Medicare by state),
# 48 (MA by state) — all have the same eligibility type columns
# ---------------------------------------------------------------------------

def build_cps_dual_enrollment_detail(con, dry_run: bool) -> int:
    """Parse dual enrollment detail by state and eligibility type from CPS Excel."""
    print("Building fact_cps_dual_enrollment_detail...")

    zip_path = RAW_DIR / "dual_enrollment_2023.zip"
    if not _download(DOWNLOADS["dual_enrollment_2023.zip"], zip_path, "CPS Dual Enrollment 2023"):
        return 0
    xlsx_path = _extract_xlsx_from_zip(zip_path)
    if not xlsx_path:
        print("  ERROR: No xlsx in ZIP")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    all_rows = []

    # Parse three state-level sheets: Table 42 (total), 45 (original medicare), 48 (MA)
    sheet_configs = [
        ("42", "total"),
        ("45", "original_medicare"),
        ("48", "medicare_advantage"),
    ]

    for sheet_num, medicare_type in sheet_configs:
        target_sheet = None
        for sn in wb.sheetnames:
            if f"AB {sheet_num}" in sn:
                target_sheet = sn
                break
        if not target_sheet:
            print(f"  WARNING: Sheet {sheet_num} not found, skipping {medicare_type}")
            continue

        ws = wb[target_sheet]
        # Columns: Area, Total MMEs, Full-Benefit MMEs, QMBs Plus,
        #          SLMBs Plus, Other Full-Benefit, Partial-Benefit,
        #          QMBs, SLMBs, QDWIs & QI
        for row in ws.iter_rows(min_row=5, values_only=True):
            area = str(row[0]).strip() if row[0] else ""
            if not _is_state_row(area):
                continue
            state_code = _state_code_from_name(area)
            if not state_code:
                continue
            all_rows.append({
                "year": 2023,
                "state_code": state_code,
                "medicare_type": medicare_type,
                "total_mmes": _clean_number(row[1]),
                "full_benefit_mmes": _clean_number(row[2]),
                "qmb_plus": _clean_number(row[3]),
                "slmb_plus": _clean_number(row[4]),
                "other_full_benefit": _clean_number(row[5]),
                "partial_benefit_mmes": _clean_number(row[6]),
                "qmb_only": _clean_number(row[7]),
                "slmb_only": _clean_number(row[8]),
                "qdwi_qi": _clean_number(row[9]),
            })

    wb.close()

    if not all_rows:
        print("  WARNING: No data rows parsed")
        return 0

    con.execute("DROP TABLE IF EXISTS _cps_dual_detail")
    con.execute("""
        CREATE TABLE _cps_dual_detail (
            year INTEGER, state_code VARCHAR, medicare_type VARCHAR,
            total_mmes BIGINT, full_benefit_mmes BIGINT,
            qmb_plus BIGINT, slmb_plus BIGINT, other_full_benefit BIGINT,
            partial_benefit_mmes BIGINT, qmb_only BIGINT,
            slmb_only BIGINT, qdwi_qi BIGINT
        )
    """)
    for r in all_rows:
        con.execute("""INSERT INTO _cps_dual_detail VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [r["year"], r["state_code"], r["medicare_type"],
             r["total_mmes"], r["full_benefit_mmes"],
             r["qmb_plus"], r["slmb_plus"], r["other_full_benefit"],
             r["partial_benefit_mmes"], r["qmb_only"],
             r["slmb_only"], r["qdwi_qi"]])

    count = write_parquet(con, "_cps_dual_detail", _snapshot_path("cps_dual_enrollment_detail"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cps_dual_detail")
    return count


# ---------------------------------------------------------------------------
# MA ENROLLMENT (Sheets 19 + 20: by state, CY2023)
# Table 19: by state, total/aged/disabled
# Table 20: by state and entitlement type (with plan type breakdowns)
# ---------------------------------------------------------------------------

def build_cps_ma_enrollment(con, dry_run: bool) -> int:
    """Parse MA & Other Health Plan enrollment by state from CPS Excel."""
    print("Building fact_cps_ma_enrollment...")

    zip_path = RAW_DIR / "ma_enrollment_2023.zip"
    if not _download(DOWNLOADS["ma_enrollment_2023.zip"], zip_path, "CPS MA Enrollment 2023"):
        return 0
    xlsx_path = _extract_xlsx_from_zip(zip_path)
    if not xlsx_path:
        print("  ERROR: No xlsx in ZIP")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    all_rows = []

    # Sheet 19: by state — total, aged, disabled across Part A/B breakdowns
    ws19 = None
    for sn in wb.sheetnames:
        if "19" in sn:
            ws19 = wb[sn]
            break

    if ws19:
        # Get header to understand column layout
        header_row = None
        for row in ws19.iter_rows(min_row=4, max_row=4, values_only=True):
            header_row = [str(v).strip() if v else "" for v in row]

        for row in ws19.iter_rows(min_row=5, values_only=True):
            area = str(row[0]).strip() if row[0] else ""
            if not _is_state_row(area):
                continue
            state_code = _state_code_from_name(area)
            if not state_code:
                continue
            all_rows.append({
                "year": 2023,
                "state_code": state_code,
                "enrollment_type": "part_a_or_b",
                "total_enrollees": _clean_number(row[1]),
                "aged_enrollees": _clean_number(row[2]),
                "disabled_enrollees": _clean_number(row[3]),
            })
            # Part A and Part B
            all_rows.append({
                "year": 2023,
                "state_code": state_code,
                "enrollment_type": "part_a_and_b",
                "total_enrollees": _clean_number(row[4]),
                "aged_enrollees": _clean_number(row[5]),
                "disabled_enrollees": _clean_number(row[6]),
            })

    wb.close()

    if not all_rows:
        print("  WARNING: No data rows parsed")
        return 0

    con.execute("DROP TABLE IF EXISTS _cps_ma_enroll")
    con.execute("""
        CREATE TABLE _cps_ma_enroll (
            year INTEGER, state_code VARCHAR, enrollment_type VARCHAR,
            total_enrollees BIGINT, aged_enrollees BIGINT,
            disabled_enrollees BIGINT
        )
    """)
    for r in all_rows:
        con.execute("""INSERT INTO _cps_ma_enroll VALUES (?, ?, ?, ?, ?, ?)""",
            [r["year"], r["state_code"], r["enrollment_type"],
             r["total_enrollees"], r["aged_enrollees"],
             r["disabled_enrollees"]])

    count = write_parquet(con, "_cps_ma_enroll", _snapshot_path("cps_ma_enrollment"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cps_ma_enroll")
    return count


# ---------------------------------------------------------------------------
# PART D ENROLLMENT (Sheet 3: by state, CY2023)
# ---------------------------------------------------------------------------

def build_cps_part_d_enrollment(con, dry_run: bool) -> int:
    """Parse Part D enrollment by state from CPS Excel."""
    print("Building fact_cps_part_d_enrollment...")

    zip_path = RAW_DIR / "part_d_enrollment_2023.zip"
    if not _download(DOWNLOADS["part_d_enrollment_2023.zip"], zip_path, "CPS Part D Enrollment 2023"):
        return 0
    xlsx_path = _extract_xlsx_from_zip(zip_path)
    if not xlsx_path:
        print("  ERROR: No xlsx in ZIP")
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # Sheet 3: by state
    ws3 = None
    for sn in wb.sheetnames:
        if "D 3" in sn:
            ws3 = wb[sn]
            break

    if not ws3:
        print("  ERROR: Part D Sheet 3 not found")
        wb.close()
        return 0

    # Cols: Area, Total Medicare Enrollees, Total Part D, Stand-Alone PDP,
    #       MA-PD, No LIS, Full LIS, Partial LIS, Deemed LIS, EGWP,
    #       Retiree Drug Subsidy, No Part D

    rows = []
    for row in ws3.iter_rows(min_row=5, values_only=True):
        area = str(row[0]).strip() if row[0] else ""
        if not _is_state_row(area):
            continue
        state_code = _state_code_from_name(area)
        if not state_code:
            continue
        rows.append({
            "year": 2023,
            "state_code": state_code,
            "total_medicare_enrollees": _clean_number(row[1]),
            "total_part_d_enrollees": _clean_number(row[2]),
            "standalone_pdp_enrollees": _clean_number(row[3]),
            "ma_pdp_enrollees": _clean_number(row[4]),
            "no_lis_enrollees": _clean_number(row[5]),
            "full_lis_enrollees": _clean_number(row[6]),
            "partial_lis_enrollees": _clean_number(row[7]),
            "deemed_lis_enrollees": _clean_number(row[8]),
            "egwp_enrollees": _clean_number(row[9]),
            "retiree_drug_subsidy": _clean_number(row[10]),
            "no_part_d_no_rds": _clean_number(row[11]),
        })
    wb.close()

    if not rows:
        print("  WARNING: No data rows parsed")
        return 0

    con.execute("DROP TABLE IF EXISTS _cps_part_d")
    con.execute("""
        CREATE TABLE _cps_part_d (
            year INTEGER, state_code VARCHAR,
            total_medicare_enrollees BIGINT, total_part_d_enrollees BIGINT,
            standalone_pdp_enrollees BIGINT, ma_pdp_enrollees BIGINT,
            no_lis_enrollees BIGINT, full_lis_enrollees BIGINT,
            partial_lis_enrollees BIGINT, deemed_lis_enrollees BIGINT,
            egwp_enrollees BIGINT, retiree_drug_subsidy BIGINT,
            no_part_d_no_rds BIGINT
        )
    """)
    for r in rows:
        con.execute("""INSERT INTO _cps_part_d VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [r["year"], r["state_code"],
             r["total_medicare_enrollees"], r["total_part_d_enrollees"],
             r["standalone_pdp_enrollees"], r["ma_pdp_enrollees"],
             r["no_lis_enrollees"], r["full_lis_enrollees"],
             r["partial_lis_enrollees"], r["deemed_lis_enrollees"],
             r["egwp_enrollees"], r["retiree_drug_subsidy"],
             r["no_part_d_no_rds"]])

    count = write_parquet(con, "_cps_part_d", _snapshot_path("cps_part_d_enrollment"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cps_part_d")
    return count


# ---------------------------------------------------------------------------
# MA/PART D ENROLLMENT BY PLAN (CPSC monthly, Dec 2023)
# Aggregated to contract/plan/state level (from county-level detail)
# ---------------------------------------------------------------------------

def build_ma_enrollment_plan(con, dry_run: bool) -> int:
    """Parse monthly MA/Part D enrollment by contract/plan/state from CPSC ZIP."""
    print("Building fact_ma_enrollment_plan...")

    zip_path = RAW_DIR / "cpsc_enrollment_2023_12.zip"
    if not _download(DOWNLOADS["cpsc_enrollment_2023_12.zip"], zip_path, "CPSC Enrollment Dec 2023"):
        return 0

    # Extract CSVs
    with zipfile.ZipFile(zip_path) as z:
        z.extractall(RAW_DIR)

    enrollment_csv = RAW_DIR / "CPSC_Enrollment_2023_12" / "CPSC_Enrollment_Info_2023_12.csv"
    contract_csv = RAW_DIR / "CPSC_Enrollment_2023_12" / "CPSC_Contract_Info_2023_12.csv"

    if not enrollment_csv.exists() or not contract_csv.exists():
        print("  ERROR: CSV files not found after extraction")
        return 0

    # Load contract info into DuckDB
    con.execute("DROP TABLE IF EXISTS _cpsc_contract_raw")
    con.execute(f"""
        CREATE TABLE _cpsc_contract_raw AS
        SELECT * FROM read_csv('{contract_csv}', all_varchar=true, ignore_errors=true)
    """)

    # Load enrollment and aggregate to state level
    # The enrollment file has: Contract Number, Plan ID, SSA Code, FIPS Code, State, County, Enrollment
    # We aggregate to contract/plan/state, summing enrollment (treating * as 5 for suppressed)
    con.execute("DROP TABLE IF EXISTS _cpsc_enrollment_raw")
    con.execute(f"""
        CREATE TABLE _cpsc_enrollment_raw AS
        SELECT
            "Contract Number" AS contract_id,
            "Plan ID" AS plan_id,
            "State" AS state_code,
            "Enrollment" AS enrollment_raw
        FROM read_csv('{enrollment_csv}', all_varchar=true, ignore_errors=true)
        WHERE "State" IS NOT NULL AND "State" != ''
    """)

    # Aggregate to contract/plan/state and join with contract info
    con.execute("DROP TABLE IF EXISTS _ma_enrollment_plan")
    con.execute("""
        CREATE TABLE _ma_enrollment_plan AS
        SELECT
            2023 AS year,
            12 AS month,
            e.contract_id,
            e.plan_id,
            e.state_code,
            SUM(CASE WHEN e.enrollment_raw = '*' THEN 5
                     ELSE TRY_CAST(e.enrollment_raw AS INTEGER)
                END) AS enrollment,
            COUNT(*) AS county_count,
            c."Organization Type" AS organization_type,
            c."Plan Type" AS plan_type,
            c."Offers Part D" AS offers_part_d,
            c."SNP Plan" AS snp_plan,
            c."EGHP" AS eghp,
            c."Organization Marketing Name" AS organization_name,
            c."Plan Name" AS plan_name,
            c."Parent Organization" AS parent_organization
        FROM _cpsc_enrollment_raw e
        LEFT JOIN _cpsc_contract_raw c
            ON e.contract_id = c."Contract ID"
            AND e.plan_id = c."Plan ID"
        WHERE e.state_code IS NOT NULL
            AND LENGTH(e.state_code) = 2
        GROUP BY e.contract_id, e.plan_id, e.state_code,
                 c."Organization Type", c."Plan Type", c."Offers Part D",
                 c."SNP Plan", c."EGHP", c."Organization Marketing Name",
                 c."Plan Name", c."Parent Organization"
    """)

    count = write_parquet(con, "_ma_enrollment_plan", _snapshot_path("ma_enrollment_plan"), dry_run)
    con.execute("DROP TABLE IF EXISTS _cpsc_contract_raw")
    con.execute("DROP TABLE IF EXISTS _cpsc_enrollment_raw")
    con.execute("DROP TABLE IF EXISTS _ma_enrollment_plan")
    return count


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

ALL_BUILDERS = {
    "fact_cps_inpatient_utilization": build_cps_inpatient_utilization,
    "fact_cps_snf_utilization": build_cps_snf_utilization,
    "fact_cps_hha_utilization": build_cps_hha_utilization,
    "fact_cps_hospice_utilization": build_cps_hospice_utilization,
    "fact_cps_dual_enrollment_detail": build_cps_dual_enrollment_detail,
    "fact_cps_ma_enrollment": build_cps_ma_enrollment,
    "fact_cps_part_d_enrollment": build_cps_part_d_enrollment,
    "fact_ma_enrollment_plan": build_ma_enrollment_plan,
}


def main():
    parser = argparse.ArgumentParser(description="Build CMS Program Statistics tables")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    parser.add_argument("--only", type=str, help="Comma-separated list of tables to build")
    args = parser.parse_args()

    RAW_DIR.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect()

    targets = ALL_BUILDERS
    if args.only:
        selected = [t.strip() for t in args.only.split(",")]
        targets = {k: v for k, v in ALL_BUILDERS.items() if k in selected}

    manifest = {
        "run_id": RUN_ID,
        "snapshot_date": SNAPSHOT_DATE,
        "started_at": datetime.now().isoformat(),
        "tables": {},
    }

    total_rows = 0
    for table_name, builder_fn in targets.items():
        try:
            count = builder_fn(con, args.dry_run)
            manifest["tables"][table_name] = {"rows": count, "status": "ok"}
            total_rows += count
        except Exception as e:
            print(f"  ERROR building {table_name}: {e}")
            import traceback
            traceback.print_exc()
            manifest["tables"][table_name] = {"rows": 0, "status": f"error: {e}"}

    manifest["completed_at"] = datetime.now().isoformat()
    manifest["total_rows"] = total_rows

    if not args.dry_run:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest_path = META_DIR / f"cms_program_stats_{SNAPSHOT_DATE}.json"
        with open(manifest_path, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"\nManifest: {manifest_path.relative_to(LAKE_DIR)}")

    print(f"\n{'[DRY RUN] ' if args.dry_run else ''}Total: {total_rows:,} rows across {len(targets)} tables")
    con.close()


if __name__ == "__main__":
    main()
