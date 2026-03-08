#!/usr/bin/env python3
"""
build_lake_round2c.py — Ingest LTSS/HCBS and managed care datasets into the lake.

Sources:
  1. CMS-372 Waiver Data (2018-2019) — 1915(c) waiver expenditure & participants
  2. Managed Care Enrollment by Plan (includes PACE, 7,807 rows)
  3. MLTSS Enrollment (513 rows)
  4. Managed Care Enrollment by Population (515 rows)
  5. MA/PACE Plan Directory (CSV from CMS)

Tables built:
  fact_cms372_waiver         — 1915(c) waiver program data (participants, expenditures)
  fact_mc_enrollment_plan    — Managed care enrollment by plan (includes PACE orgs)
  fact_mltss_enrollment      — MLTSS enrollment by state (updated)
  fact_mc_enrollment_pop     — Managed care enrollment by program and population
  dim_pace_organization      — PACE plan directory from CMS

Usage:
  python3 scripts/build_lake_round2c.py
  python3 scripts/build_lake_round2c.py --dry-run
"""

import argparse
import json
import uuid
from datetime import date, datetime
from pathlib import Path

import duckdb

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
DIM_DIR = LAKE_DIR / "dimension"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())


def write_parquet(con, table_name: str, out_path: Path, dry_run: bool) -> int:
    count = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    if not dry_run and count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY {table_name} TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count:,} rows, {size_mb:.1f} MB)")
    elif dry_run:
        print(f"  [dry-run] {out_path.relative_to(LAKE_DIR)} ({count:,} rows)")
    return count


def _snapshot_path(fact_name: str) -> Path:
    return FACT_DIR / fact_name / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"


def build_cms372_waiver(con, dry_run: bool) -> int:
    """Build fact_cms372_waiver from CMS-372 annual report Excel files."""
    print("Building fact_cms372_waiver...")
    xlsx_path = RAW_DIR / "cms_372_waiver_2018_2019.xlsx"
    if not xlsx_path.exists():
        print("  SKIPPED — cms_372_waiver_2018_2019.xlsx not found")
        return 0

    try:
        import openpyxl
    except ImportError:
        print("  SKIPPED — openpyxl not installed")
        return 0

    all_rows = []
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    for sheet_name, year in [("B.1 Waiver Data 2019", 2019), ("B.2 Waiver Data 2018", 2018)]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))

        for row in rows_data[2:]:  # Skip title and header rows
            if not row or not row[0]:
                continue
            state = str(row[0]).strip()
            waiver_num = str(row[1]).strip() if row[1] else ""
            if waiver_num == "Total" or not state or state.startswith("Table") or state.startswith("Source"):
                continue

            # Parse numeric fields
            def _num(val):
                if val is None:
                    return None
                try:
                    return float(str(val).replace(",", "").strip())
                except (ValueError, TypeError):
                    return None

            all_rows.append({
                "state_name": state,
                "waiver_number": waiver_num,
                "waiver_name": str(row[2]).strip() if row[2] else "",
                "target_group": str(row[3]).strip() if row[3] else "",
                "subgroups": str(row[4]).strip() if row[4] else "",
                "total_participants": _num(row[5]),
                "total_days_of_service": _num(row[6]),
                "total_participant_months": _num(row[7]),
                "avg_participant_months": _num(row[8]),
                "total_expenditures": _num(row[9]),
                "year": year,
            })

    wb.close()

    if not all_rows:
        print("  No CMS-372 data parsed")
        return 0

    con.execute("""
        CREATE OR REPLACE TABLE _fact_372 (
            state_name VARCHAR, waiver_number VARCHAR, waiver_name VARCHAR,
            target_group VARCHAR, subgroups VARCHAR,
            total_participants DOUBLE, total_days_of_service DOUBLE,
            total_participant_months DOUBLE, avg_participant_months DOUBLE,
            total_expenditures DOUBLE, year INTEGER,
            source VARCHAR, snapshot_date DATE
        )
    """)
    for r in all_rows:
        con.execute("INSERT INTO _fact_372 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", [
            r["state_name"], r["waiver_number"], r["waiver_name"],
            r["target_group"], r["subgroups"],
            r["total_participants"], r["total_days_of_service"],
            r["total_participant_months"], r["avg_participant_months"],
            r["total_expenditures"], r["year"],
            "cms_372_annual_report", SNAPSHOT_DATE,
        ])

    count = write_parquet(con, "_fact_372", _snapshot_path("cms372_waiver"), dry_run)
    states = con.execute("SELECT COUNT(DISTINCT state_name) FROM _fact_372").fetchone()[0]
    waivers = con.execute("SELECT COUNT(DISTINCT waiver_number) FROM _fact_372").fetchone()[0]

    # Summary stats
    stats = con.execute("""
        SELECT year,
               COUNT(DISTINCT state_name) AS states,
               SUM(total_participants) AS participants,
               ROUND(SUM(total_expenditures) / 1e9, 2) AS expenditures_bn
        FROM _fact_372 GROUP BY year ORDER BY year
    """).fetchall()
    print(f"  {count:,} waiver program records, {states} states, {waivers} waivers")
    for s in stats:
        print(f"    {s[0]}: {s[1]} states, {s[2]:,.0f} participants, ${s[3]}B expenditures")

    # Top 5 by expenditure
    top = con.execute("""
        SELECT state_name, SUM(total_expenditures) AS exp
        FROM _fact_372 WHERE year = 2019
        GROUP BY state_name ORDER BY exp DESC LIMIT 5
    """).fetchall()
    print(f"  Top 5 states by 2019 waiver expenditure:")
    for t in top:
        print(f"    {t[0]}: ${t[1]/1e9:.2f}B")

    con.execute("DROP TABLE IF EXISTS _fact_372")
    return count


def build_mc_enrollment_plan(con, dry_run: bool) -> int:
    """Build fact_mc_enrollment_plan from managed care enrollment by plan."""
    print("Building fact_mc_enrollment_plan...")
    csv_path = RAW_DIR / "mc_enrollment_by_plan.csv"
    if not csv_path.exists():
        print("  SKIPPED — mc_enrollment_by_plan.csv not found")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_mc_plan AS
        SELECT
            State AS state_name,
            "Program Name" AS program_name,
            "Plan Name" AS plan_name,
            "Geographic Region" AS geographic_region,
            TRY_CAST(REPLACE("Medicaid-Only Enrollment", ',', '') AS INTEGER) AS medicaid_only_enrollment,
            TRY_CAST(REPLACE("Dual Enrollment", ',', '') AS INTEGER) AS dual_enrollment,
            TRY_CAST(REPLACE("Total Enrollment", ',', '') AS INTEGER) AS total_enrollment,
            TRY_CAST(Year AS INTEGER) AS year,
            "Parent Organization" AS parent_organization,
            CASE WHEN "Program Name" ILIKE '%PACE%' THEN TRUE ELSE FALSE END AS is_pace,
            'medicaid_gov_mc_enrollment_plan' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', header=true, all_varchar=true)
        WHERE State IS NOT NULL AND LENGTH(State) > 1
    """)

    count = write_parquet(con, "_fact_mc_plan", _snapshot_path("mc_enrollment_plan"), dry_run)
    years = con.execute("SELECT DISTINCT year FROM _fact_mc_plan ORDER BY year").fetchall()
    pace_count = con.execute("SELECT COUNT(*) FROM _fact_mc_plan WHERE is_pace").fetchone()[0]
    states = con.execute("SELECT COUNT(DISTINCT state_name) FROM _fact_mc_plan").fetchone()[0]
    print(f"  {count:,} plan enrollment records, {states} states, years: {[y[0] for y in years]}")
    print(f"  PACE plan records: {pace_count}")

    # Top PACE states
    pace_top = con.execute("""
        SELECT state_name, SUM(total_enrollment) AS enrollment
        FROM _fact_mc_plan WHERE is_pace AND year = (SELECT MAX(year) FROM _fact_mc_plan WHERE is_pace)
        GROUP BY state_name ORDER BY enrollment DESC LIMIT 5
    """).fetchall()
    if pace_top:
        print(f"  Top PACE states: {', '.join(f'{p[0]} ({p[1]:,})' for p in pace_top)}")

    con.execute("DROP TABLE IF EXISTS _fact_mc_plan")
    return count


def build_mltss_enrollment(con, dry_run: bool) -> int:
    """Build updated fact_mltss_enrollment from MLTSS enrollment CSV."""
    print("Building fact_mltss_enrollment (updated)...")
    csv_path = RAW_DIR / "mltss_enrollment_2024.csv"
    if not csv_path.exists():
        print("  SKIPPED — mltss_enrollment_2024.csv not found")
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact_mltss AS
        SELECT
            State AS state_name,
            TRY_CAST(REPLACE("Total Any Managed Care Enrollees", ',', '') AS INTEGER) AS total_mc_enrollees,
            "Comprehensive Managed Care LTSS Enrollees" AS comp_mc_ltss_enrollees,
            "Comprehensive Managed Care LTSS Percent" AS comp_mc_ltss_pct,
            "Managed LTSS Only Enrollees" AS mltss_only_enrollees,
            "Managed LTSS Only Percent" AS mltss_only_pct,
            TRY_CAST(Year AS INTEGER) AS year,
            'medicaid_gov_mltss_enrollment' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', header=true, all_varchar=true)
        WHERE State IS NOT NULL AND State != 'TOTALS'
          AND LENGTH(State) > 1
    """)

    count = write_parquet(con, "_fact_mltss", _snapshot_path("mltss_enrollment2"), dry_run)
    years = con.execute("SELECT DISTINCT year FROM _fact_mltss ORDER BY year").fetchall()
    states = con.execute("SELECT COUNT(DISTINCT state_name) FROM _fact_mltss").fetchone()[0]
    print(f"  {count} MLTSS enrollment records, {states} states, years: {[y[0] for y in years]}")

    con.execute("DROP TABLE IF EXISTS _fact_mltss")
    return count


def build_mc_enrollment_pop(con, dry_run: bool) -> int:
    """Build fact_mc_enrollment_pop (updated) from managed care enrollment by population."""
    print("Building fact_mc_enrollment_pop (updated)...")
    csv_path = RAW_DIR / "mc_enrollment_by_population.csv"
    if not csv_path.exists():
        print("  SKIPPED — mc_enrollment_by_population.csv not found")
        return 0

    # CSV has embedded commas in numbers and escaped quotes — use Python csv module
    import csv
    rows_data = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            state = row.get("State", "").strip()
            if not state or len(state) < 2:
                continue

            def _parse_int(val):
                if not val or val.strip() == "--":
                    return None
                try:
                    return int(val.replace(",", "").strip())
                except ValueError:
                    return None

            rows_data.append((
                state,
                _parse_int(row.get("Total Medicaid Enrollees", "")),
                _parse_int(row.get("Comprehensive MCO with or without MLTSS", "")),
                _parse_int(row.get("PCCM", "")),
                _parse_int(row.get("MLTSS only", "")),
                _parse_int(row.get("BHO (PIHP and/or PAHP)", "")),
                _parse_int(row.get("Dental", "")),
                _parse_int(row.get("Transportation", "")),
                _parse_int(row.get("PACE", "")),
                _parse_int(row.get("Year", "")),
            ))

    con.execute("""
        CREATE OR REPLACE TABLE _fact_mc_pop (
            state_name VARCHAR, total_medicaid_enrollees INTEGER,
            comprehensive_mco INTEGER, pccm INTEGER, mltss_only INTEGER,
            bho INTEGER, dental INTEGER, transportation INTEGER,
            pace INTEGER, year INTEGER,
            source VARCHAR, snapshot_date DATE
        )
    """)
    for r in rows_data:
        con.execute("INSERT INTO _fact_mc_pop VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                     list(r) + ["medicaid_gov_mc_enrollment_pop", SNAPSHOT_DATE])

    count = write_parquet(con, "_fact_mc_pop", _snapshot_path("mc_enrollment_pop2"), dry_run)
    years = con.execute("SELECT DISTINCT year FROM _fact_mc_pop ORDER BY year").fetchall()
    pace_states = con.execute("SELECT COUNT(*) FROM _fact_mc_pop WHERE pace > 0").fetchone()[0]
    print(f"  {count} MC enrollment by pop records, years: {[y[0] for y in years]}")
    print(f"  States with PACE enrollment: {pace_states}")

    con.execute("DROP TABLE IF EXISTS _fact_mc_pop")
    return count


def build_pace_directory(con, dry_run: bool) -> int:
    """Build dim_pace_organization from MA/PACE plan directory."""
    print("Building dim_pace_organization...")
    zip_path = RAW_DIR / "ma_pace_directory.zip"
    if not zip_path.exists():
        print("  SKIPPED — ma_pace_directory.zip not found")
        return 0

    import zipfile
    extract_dir = RAW_DIR / "ma_pace_directory"
    extract_dir.mkdir(exist_ok=True)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(extract_dir)

    # Find the CSV
    csv_files = list(extract_dir.glob("**/*.csv"))
    if not csv_files:
        print("  SKIPPED — no CSV found in zip")
        return 0

    csv_path = csv_files[0]
    print(f"  Reading {csv_path.name}")

    # Filter to PACE organizations only
    con.execute(f"""
        CREATE OR REPLACE TABLE _dim_pace AS
        SELECT *,
            'cms_ma_pace_directory' AS source,
            DATE '{SNAPSHOT_DATE}' AS snapshot_date
        FROM read_csv_auto('{csv_path}', header=true)
        WHERE "Organization Type" ILIKE '%PACE%'
    """)

    pace_count = con.execute("SELECT COUNT(*) FROM _dim_pace").fetchone()[0]
    if pace_count == 0:
        # Try alternate column name
        cols = con.execute(f"SELECT * FROM read_csv_auto('{csv_path}', header=true) LIMIT 0").description
        col_names = [c[0] for c in cols]
        print(f"  Columns: {col_names[:10]}")
        # Try with 'Plan Type' or similar
        for col in col_names:
            if 'type' in col.lower() or 'plan' in col.lower():
                vals = con.execute(f"SELECT DISTINCT \"{col}\" FROM read_csv_auto('{csv_path}', header=true) LIMIT 20").fetchall()
                print(f"  {col} values: {[v[0] for v in vals[:10]]}")

    out_path = DIM_DIR / "dim_pace_organization.parquet"
    if not dry_run and pace_count > 0:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        con.execute(f"COPY _dim_pace TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_mb = out_path.stat().st_size / (1024 * 1024)
        print(f"  -> dimension/dim_pace_organization.parquet ({pace_count} orgs, {size_mb:.1f} MB)")
    elif pace_count == 0:
        # Load all to dim if can't filter
        con.execute(f"""
            CREATE OR REPLACE TABLE _dim_pace AS
            SELECT *,
                'cms_ma_pace_directory' AS source,
                DATE '{SNAPSHOT_DATE}' AS snapshot_date
            FROM read_csv_auto('{csv_path}', header=true)
        """)
        pace_count = con.execute("SELECT COUNT(*) FROM _dim_pace").fetchone()[0]
        if not dry_run:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            con.execute(f"COPY _dim_pace TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
            size_mb = out_path.stat().st_size / (1024 * 1024)
            print(f"  -> dimension/dim_pace_organization.parquet ({pace_count} total orgs, {size_mb:.1f} MB)")
    else:
        print(f"  [dry-run] dim_pace_organization ({pace_count} orgs)")

    states = con.execute('SELECT COUNT(DISTINCT "Legal Entity State Code") FROM _dim_pace').fetchone()[0] if pace_count > 0 else 0
    print(f"  {pace_count} organizations across {states} states")

    con.execute("DROP TABLE IF EXISTS _dim_pace")
    return pace_count


ALL_TABLES = {
    "cms372": ("fact_cms372_waiver", build_cms372_waiver),
    "mc_plan": ("fact_mc_enrollment_plan", build_mc_enrollment_plan),
    "mltss": ("fact_mltss_enrollment2", build_mltss_enrollment),
    "mc_pop": ("fact_mc_enrollment_pop2", build_mc_enrollment_pop),
    "pace": ("dim_pace_organization", build_pace_directory),
}


def main():
    parser = argparse.ArgumentParser(description="Ingest LTSS/HCBS and MC datasets into Aradune lake")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--table", choices=list(ALL_TABLES.keys()) + ["all"], default="all")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"Round 2C — LTSS/HCBS & Managed Care — {SNAPSHOT_DATE}")
    print(f"{'='*60}")
    print(f"Run ID: {RUN_ID}\n")

    con = duckdb.connect()
    totals = {}

    tables_to_build = ALL_TABLES if args.table == "all" else {args.table: ALL_TABLES[args.table]}
    for key, (fact_name, builder) in tables_to_build.items():
        totals[fact_name] = builder(con, args.dry_run)
        print()

    con.close()

    print("=" * 60)
    print("ROUND 2C LAKE INGESTION COMPLETE")
    print("=" * 60)
    total_rows = sum(totals.values())
    for name, count in totals.items():
        status = "written" if not args.dry_run else "dry-run"
        print(f"  {name:40s} {count:>12,} rows  [{status}]")
    print(f"  {'TOTAL':40s} {total_rows:>12,} rows")

    if not args.dry_run and total_rows > 0:
        META_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "snapshot_date": SNAPSHOT_DATE,
            "pipeline_run_id": RUN_ID,
            "created_at": datetime.now().isoformat() + "Z",
            "tables": {name: {"rows": count} for name, count in totals.items()},
            "total_rows": total_rows,
        }
        manifest_file = META_DIR / f"manifest_round2c_{SNAPSHOT_DATE}.json"
        with open(manifest_file, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"\n  Manifest: {manifest_file}")

    print("\nDone.")


if __name__ == "__main__":
    main()
