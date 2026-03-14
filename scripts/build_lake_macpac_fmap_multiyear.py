#!/usr/bin/env python3
"""
build_lake_macpac_fmap_multiyear.py — Ingest MACPAC Exhibit 6 (FMAP + E-FMAP by state,
FYs 2023-2026) into the Aradune data lake.

Source: data/raw/state_fiscal/macpac_fmap_2023_2026.xlsx

Adds granular FMAP data including emergency FMAPs (PHE wind-down), E-FMAPs for CHIP,
and standard FMAPs for FY2024-2026. This complements the existing fmap_historical and
kff_fmap tables with MACPAC's authoritative quarterly breakdown.

Table built:
  macpac_fmap_multiyear — FMAP and E-FMAP by state, FY 2023 Q1 through FY 2026

Usage:
  python3 scripts/build_lake_macpac_fmap_multiyear.py
"""

import json
import re
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_PATH = PROJECT_ROOT / "data" / "raw" / "state_fiscal" / "macpac_fmap_2023_2026.xlsx"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
META_DIR = LAKE_DIR / "metadata"

SNAPSHOT_DATE = date.today().isoformat()
RUN_ID = str(uuid.uuid4())

STATE_MAP = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN",
    "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA",
    "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "District of Columbia": "DC",
    "Puerto Rico": "PR", "Guam": "GU", "Virgin Islands": "VI",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
}


def _parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if s in ("", "-", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _detect_fmap_columns(ws):
    """Dynamically detect FMAP and E-FMAP columns from Excel header rows.

    Scans rows 3-4 for headers containing fiscal year patterns (e.g., "FY 2023",
    "FY2024", "2025"). Returns (fmap_cols, efmap_cols, data_start_row) where each
    col list contains tuples of (col_idx, period_key, fiscal_year, period, is_emergency).

    Falls back to hardcoded positions if header detection fails.
    """
    # Hardcoded fallback (original column positions)
    FALLBACK_FMAP = [
        (2, "2023_q1q2", 2023, "Q1-Q2", True),
        (3, "2023_q3", 2023, "Q3", True),
        (4, "2023_q4", 2023, "Q4", True),
        (5, "2024", 2024, "Annual", False),
        (6, "2025", 2025, "Annual", False),
        (7, "2026", 2026, "Annual", False),
    ]
    FALLBACK_EFMAP = [
        (8, "2023_q1q2", 2023, "Q1-Q2", True),
        (9, "2023_q3", 2023, "Q3", True),
        (10, "2023_q4", 2023, "Q4", True),
        (11, "2024", 2024, "Annual", False),
    ]

    # Read header rows (rows 3-4 typically contain group headers + sub-headers)
    # Row 3: group labels ("FMAPs for Medicaid", "E-FMAPs for CHIP")
    # Row 4: period sub-headers ("FY 2023 Q1-2", "FY 2024", etc.)
    max_col = min(ws.max_column or 20, 30)  # Reasonable upper bound

    # Collect header text from rows 1-4 for each column
    col_headers = {}
    for col_idx in range(1, max_col + 1):
        texts = []
        for header_row in range(1, 5):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val and isinstance(cell_val, str):
                texts.append(cell_val.strip())
        col_headers[col_idx] = " | ".join(texts)

    # Detect which columns are FMAP vs E-FMAP by scanning row 3 for group headers
    fmap_group_start = None
    efmap_group_start = None
    for col_idx in range(2, max_col + 1):
        for header_row in range(1, 5):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val and isinstance(cell_val, str):
                text = cell_val.strip().upper()
                if "E-FMAP" in text or "ENHANCED" in text:
                    if efmap_group_start is None:
                        efmap_group_start = col_idx
                elif "FMAP" in text and "E-FMAP" not in text:
                    if fmap_group_start is None:
                        fmap_group_start = col_idx

    # Parse fiscal year + period from sub-headers (row 4, or merged into row 3)
    def _parse_fy_period(col_idx):
        """Extract (fiscal_year, period, is_emergency) from header text."""
        combined = col_headers.get(col_idx, "")
        if not combined:
            return None

        # Look for FY pattern: "FY 2023", "FY2024", or standalone "2024", "2025"
        fy_match = re.search(r'FY\s*(\d{4})', combined, re.IGNORECASE)
        if not fy_match:
            # Try standalone 4-digit year
            fy_match = re.search(r'\b(20\d{2})\b', combined)
        if not fy_match:
            return None

        fiscal_year = int(fy_match.group(1))

        # Detect quarter
        q_match = re.search(r'Q\s*(\d)[\s-]*(?:Q?\s*(\d))?', combined, re.IGNORECASE)
        if q_match:
            q1 = q_match.group(1)
            q2 = q_match.group(2)
            if q2:
                period = f"Q{q1}-Q{q2}"
            else:
                period = f"Q{q1}"
        else:
            period = "Annual"

        # Detect emergency FMAP
        is_emergency = bool(re.search(r'emerg', combined, re.IGNORECASE))

        return (fiscal_year, period, is_emergency)

    # Build column lists dynamically
    fmap_cols = []
    efmap_cols = []

    for col_idx in range(2, max_col + 1):
        parsed = _parse_fy_period(col_idx)
        if not parsed:
            continue

        fiscal_year, period, is_emergency = parsed
        period_key = f"{fiscal_year}"
        if period != "Annual":
            # Convert "Q1-Q2" -> "q1q2", "Q3" -> "q3", etc.
            period_key = f"{fiscal_year}_{period.lower().replace('-', '').replace('q', 'q')}"

        entry = (col_idx, period_key, fiscal_year, period, is_emergency)

        # Assign to FMAP or E-FMAP group based on detected group boundaries
        if efmap_group_start and col_idx >= efmap_group_start:
            efmap_cols.append(entry)
        elif fmap_group_start and col_idx >= fmap_group_start:
            fmap_cols.append(entry)
        else:
            # If no group headers detected, use column position heuristic
            fmap_cols.append(entry)

    # Determine data start row (first row with a state name after headers)
    data_start_row = 5  # default
    for row_num in range(3, 10):
        cell_val = ws.cell(row=row_num, column=1).value
        if cell_val and isinstance(cell_val, str):
            clean = re.sub(r'\d+$', '', cell_val.strip()).strip()
            if clean in STATE_MAP:
                data_start_row = row_num
                break

    # Validate: if dynamic detection found reasonable results, use them
    if fmap_cols and len(fmap_cols) >= 2:
        print(f"  Dynamic header detection: {len(fmap_cols)} FMAP cols, "
              f"{len(efmap_cols)} E-FMAP cols, data starts row {data_start_row}")
        for entry in fmap_cols:
            print(f"    FMAP  col {entry[0]}: FY{entry[2]} {entry[3]}"
                  f"{' (emergency)' if entry[4] else ''}")
        for entry in efmap_cols:
            print(f"    E-FMAP col {entry[0]}: FY{entry[2]} {entry[3]}"
                  f"{' (emergency)' if entry[4] else ''}")
        return fmap_cols, efmap_cols, data_start_row
    else:
        print("  Dynamic header detection failed, using hardcoded column positions")
        return FALLBACK_FMAP, FALLBACK_EFMAP, 5


def build():
    print(f"MACPAC FMAP Multi-Year ETL -- snapshot {SNAPSHOT_DATE}")
    print(f"  Source: {RAW_PATH.name}")

    if not RAW_PATH.exists():
        print("  ERROR: Source file not found")
        return

    wb = openpyxl.load_workbook(RAW_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Dynamically detect column layout from Excel headers.
    # Falls back to hardcoded positions if header detection fails.
    FMAP_COLS, EFMAP_COLS, data_start_row = _detect_fmap_columns(ws)

    rows = []
    for row_num in range(data_start_row, ws.max_row + 1):
        raw_state = ws.cell(row=row_num, column=1).value
        if not raw_state or not isinstance(raw_state, str):
            continue
        clean_state = re.sub(r"\d+$", "", raw_state.strip()).strip()
        state_code = STATE_MAP.get(clean_state)
        if not state_code:
            continue

        for col_idx, period_key, fy, quarter, is_emergency in FMAP_COLS:
            val = _parse_numeric(ws.cell(row=row_num, column=col_idx).value)
            if val is not None:
                rows.append({
                    "state_code": state_code,
                    "fiscal_year": fy,
                    "period": quarter,
                    "is_emergency_fmap": is_emergency,
                    "fmap_type": "medicaid",
                    "fmap_rate": round(val, 4),
                    "source": "macpac",
                    "snapshot_date": SNAPSHOT_DATE,
                })

        for col_idx, period_key, fy, quarter, is_emergency in EFMAP_COLS:
            val = _parse_numeric(ws.cell(row=row_num, column=col_idx).value)
            if val is not None:
                rows.append({
                    "state_code": state_code,
                    "fiscal_year": fy,
                    "period": quarter,
                    "is_emergency_fmap": is_emergency,
                    "fmap_type": "chip_enhanced",
                    "fmap_rate": round(val, 4),
                    "source": "macpac",
                    "snapshot_date": SNAPSHOT_DATE,
                })

    wb.close()

    if not rows:
        print("  No data parsed")
        return

    print(f"  Parsed {len(rows)} records")

    con = duckdb.connect()
    columns = list(rows[0].keys())
    col_defs = ", ".join(
        f"{c} {'INTEGER' if c == 'fiscal_year' else 'BOOLEAN' if c == 'is_emergency_fmap' else 'DOUBLE' if c == 'fmap_rate' else 'VARCHAR'}"
        for c in columns
    )
    con.execute(f"CREATE TABLE _fmap ({col_defs})")
    placeholders = ", ".join(["?"] * len(columns))
    con.executemany(
        f"INSERT INTO _fmap VALUES ({placeholders})",
        [tuple(r.get(c) for c in columns) for r in rows],
    )

    out_path = FACT_DIR / "macpac_fmap_multiyear" / f"snapshot={SNAPSHOT_DATE}" / "data.parquet"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    count = con.execute("SELECT COUNT(*) FROM _fmap").fetchone()[0]
    con.execute(f"COPY _fmap TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    size_mb = out_path.stat().st_size / (1024 * 1024)
    print(f"  -> {out_path.relative_to(LAKE_DIR)} ({count} rows, {size_mb:.2f} MB)")

    # Validation
    states = con.execute("SELECT COUNT(DISTINCT state_code) FROM _fmap").fetchone()[0]
    medicaid = con.execute("SELECT COUNT(*) FROM _fmap WHERE fmap_type = 'medicaid'").fetchone()[0]
    chip = con.execute("SELECT COUNT(*) FROM _fmap WHERE fmap_type = 'chip_enhanced'").fetchone()[0]
    print(f"  {count} rows total, {states} states/territories")
    print(f"  Medicaid FMAP records: {medicaid}")
    print(f"  CHIP E-FMAP records: {chip}")

    # Check FL
    fl = con.execute("SELECT fiscal_year, period, fmap_rate FROM _fmap WHERE state_code = 'FL' AND fmap_type = 'medicaid' ORDER BY fiscal_year, period").fetchall()
    print(f"  FL Medicaid FMAP:")
    for r in fl:
        print(f"    FY{r[0]} {r[1]}: {r[2]:.4f}")

    con.close()

    # Manifest
    manifest = {
        "pipeline_run": "macpac_fmap_multiyear",
        "run_id": RUN_ID,
        "snapshot_date": SNAPSHOT_DATE,
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "total_rows": count,
        "tables": ["macpac_fmap_multiyear"],
        "source_files": ["data/raw/state_fiscal/macpac_fmap_2023_2026.xlsx"],
        "source": "MACPAC MACStats Exhibit 6 (macpac.gov)",
        "notes": "FMAP rates as decimals (0-1). Includes emergency FMAPs for FY2023 Q1-Q4.",
    }
    manifest_path = META_DIR / f"manifest_macpac_fmap_multiyear_{SNAPSHOT_DATE}.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"  Manifest: {manifest_path.relative_to(LAKE_DIR)}")

    print(f"\n== Summary ==")
    print(f"  Table: macpac_fmap_multiyear ({count} rows)")
    print(f"  FMAP + E-FMAP for {states} states, FY2023-2026 (quarterly for FY2023)")


if __name__ == "__main__":
    build()
