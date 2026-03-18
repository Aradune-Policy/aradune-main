"""
Data import backend: upload CSV/XLSX/JSON files into DuckDB temp tables.

Provides session-based in-memory storage so uploaded data can be queried
by the Intelligence / NL2SQL endpoints alongside the main data lake.
"""

import csv
import io
import json
import time
import uuid
from typing import Any

from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel

from server.db import get_cursor
from server.utils.error_handler import safe_route

router = APIRouter(prefix="/api/import", tags=["import"])

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE = 50 * 1024 * 1024        # 50 MB per file
MAX_TOTAL_MEMORY = 500 * 1024 * 1024    # 500 MB total across all sessions
SESSION_TTL = 7200                       # 2 hours

# ---------------------------------------------------------------------------
# Session store
# ---------------------------------------------------------------------------

_sessions: dict[str, dict[str, Any]] = {}


def _total_memory() -> int:
    """Return total bytes stored across all sessions."""
    return sum(len(s["file_bytes"]) for s in _sessions.values())


def _evict_expired() -> None:
    """Remove sessions older than SESSION_TTL."""
    now = time.time()
    expired = [
        sid for sid, s in _sessions.items()
        if now - s["last_accessed"] > SESSION_TTL
    ]
    for sid in expired:
        _drop_table(sid)
        del _sessions[sid]


def _evict_lru() -> None:
    """Evict least-recently-used sessions until total memory is under cap."""
    while _total_memory() > MAX_TOTAL_MEMORY and _sessions:
        oldest_sid = min(_sessions, key=lambda sid: _sessions[sid]["last_accessed"])
        _drop_table(oldest_sid)
        del _sessions[oldest_sid]


def _drop_table(session_id: str) -> None:
    """Drop the DuckDB temp table for a session, if it exists."""
    sess = _sessions.get(session_id)
    if not sess:
        return
    table_name = sess["table_name"]
    try:
        with get_cursor() as cur:
            cur.execute(f"DROP TABLE IF EXISTS {table_name}")
    except Exception:
        pass


def _cleanup() -> None:
    """Run on each request: evict expired sessions first, then LRU if needed."""
    _evict_expired()
    _evict_lru()


def hydrate_session(session_id: str) -> None:
    """
    Re-hydrate a session's DuckDB temp table from stored bytes (sync version).
    Called by intelligence.py before queries to ensure user tables exist.
    """
    sess = _sessions.get(session_id)
    if not sess:
        return  # Session not found — silently skip

    sess["last_accessed"] = time.time()
    table_name = sess["table_name"]

    # Check if table already exists
    with get_cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM information_schema.tables WHERE table_name = ?",
            [table_name],
        )
        if cur.fetchone()[0] > 0:
            return  # Already loaded

    # Re-create from stored bytes
    filename = sess["filename"]
    if filename.lower().endswith((".xlsx", ".xls")):
        load_filename = filename.rsplit(".", 1)[0] + ".csv"
    else:
        load_filename = filename

    try:
        columns, row_count, _ = _load_into_duckdb(
            sess["file_bytes"], load_filename, table_name
        )
        sess["columns"] = columns
        sess["row_count"] = row_count
    except Exception:
        pass  # Best-effort — don't block Intelligence if hydration fails


# ---------------------------------------------------------------------------
# Helpers — parse uploaded files
# ---------------------------------------------------------------------------

def _parse_csv_bytes(raw: bytes) -> tuple[list[str], list[dict]]:
    """Parse CSV bytes, return (columns, rows_as_dicts)."""
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV file has no data rows")
    columns = list(rows[0].keys())
    return columns, rows


def _parse_json_bytes(raw: bytes) -> tuple[list[str], list[dict]]:
    """Parse JSON bytes (expects array of objects), return (columns, rows)."""
    data = json.loads(raw)
    if isinstance(data, dict):
        # Try common wrapper keys
        for key in ("data", "results", "rows", "records"):
            if key in data and isinstance(data[key], list):
                data = data[key]
                break
        else:
            raise ValueError("JSON must be an array of objects or contain a 'data'/'results' key")
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("JSON file has no data rows")
    if not isinstance(data[0], dict):
        raise ValueError("JSON array elements must be objects (key-value pairs)")
    columns = list(data[0].keys())
    return columns, data


def _parse_xlsx_bytes(raw: bytes) -> bytes:
    """Parse XLSX (first sheet) and return CSV bytes for DuckDB ingestion."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed on the server; XLSX uploads are unavailable",
        )

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("XLSX workbook has no sheets")

    output = io.StringIO()
    writer = csv.writer(output)
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)
        row_count += 1
    wb.close()

    if row_count < 2:  # header + at least 1 data row
        raise ValueError("XLSX file has no data rows")

    return output.getvalue().encode("utf-8")


def _short_id() -> str:
    """Return a short random hex string for table naming."""
    return uuid.uuid4().hex[:12]


# ---------------------------------------------------------------------------
# Load data into DuckDB
# ---------------------------------------------------------------------------

def _load_into_duckdb(
    file_bytes: bytes,
    filename: str,
    table_name: str,
) -> tuple[list[str], int, list[dict]]:
    """
    Load file bytes into a DuckDB table. Returns (columns, row_count, preview).

    For CSV (and XLSX converted to CSV), uses DuckDB's read_csv_auto for
    automatic type inference. For JSON, inserts from parsed dicts.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    with get_cursor() as cur:
        if ext in ("csv", "tsv", "xlsx", "xls"):
            # For XLSX, file_bytes is already the converted CSV bytes
            # Write to a temp DuckDB blob and use read_csv_auto
            cur.execute(f"DROP TABLE IF EXISTS {table_name}")
            cur.execute(
                f"""
                CREATE TABLE {table_name} AS
                SELECT * FROM read_csv_auto(
                    '{_write_temp_csv(cur, file_bytes, table_name)}',
                    header=true,
                    auto_detect=true,
                    sample_size=10000
                )
                """
            )
        elif ext == "json":
            columns, rows = _parse_json_bytes(file_bytes)
            if not rows:
                raise ValueError("No rows to load")
            # Build CREATE TABLE from first row, all VARCHAR, then let DuckDB cast
            col_defs = ", ".join(f'"{c}" VARCHAR' for c in columns)
            cur.execute(f"DROP TABLE IF EXISTS {table_name}")
            cur.execute(f"CREATE TABLE {table_name} ({col_defs})")
            # Insert in batches
            placeholders = ", ".join(["?"] * len(columns))
            col_names = ", ".join(f'"{c}"' for c in columns)
            batch_size = 1000
            for i in range(0, len(rows), batch_size):
                batch = rows[i : i + batch_size]
                values = [tuple(str(row.get(c, "")) if row.get(c) is not None else None for c in columns) for row in batch]
                for v in values:
                    cur.execute(
                        f"INSERT INTO {table_name} ({col_names}) VALUES ({placeholders})",
                        v,
                    )
        else:
            raise ValueError(f"Unsupported file type: .{ext}")

        # Get column info
        cur.execute(f"SELECT * FROM {table_name} LIMIT 0")
        columns = [desc[0] for desc in cur.description]

        # Get row count
        cur.execute(f"SELECT COUNT(*) FROM {table_name}")
        row_count = cur.fetchone()[0]

        # Get preview (first 10 rows)
        cur.execute(f"SELECT * FROM {table_name} LIMIT 10")
        preview_rows = cur.fetchall()
        preview = [dict(zip(columns, row)) for row in preview_rows]

    return columns, row_count, preview


def _write_temp_csv(cur: Any, csv_bytes: bytes, table_name: str) -> str:
    """
    Write CSV bytes to a temporary file path that DuckDB can read.
    Returns the path. Uses Python's tempfile module.
    """
    import tempfile
    import os

    tmp_dir = tempfile.gettempdir()
    path = os.path.join(tmp_dir, f"{table_name}.csv")
    with open(path, "wb") as f:
        f.write(csv_bytes)
    return path


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class QuarantineInfo(BaseModel):
    quarantine_count: int
    reasons: dict[str, int]  # reason_code -> count


class HydrateRequest(BaseModel):
    session_id: str


class ImportResponse(BaseModel):
    session_id: str
    table_name: str
    filename: str
    columns: list[str]
    row_count: int
    preview: list[dict]
    quarantine: QuarantineInfo | None = None


class SessionInfo(BaseModel):
    session_id: str
    filename: str
    table_name: str
    columns: list[str]
    row_count: int
    created_at: float
    last_accessed: float


# ---------------------------------------------------------------------------
# Quarantine pattern: validate loaded data, split invalid rows
# ---------------------------------------------------------------------------

# Columns that indicate rate data (triggers rate-specific validation)
_RATE_COLUMNS = {"rate", "medicaid_rate", "rate_nonfacility", "rate_facility", "amount"}
_CODE_COLUMNS = {"procedure_code", "cpt_hcpcs_code", "hcpcs_code", "cpt_code", "code"}
_STATE_COLUMNS = {"state_code", "state", "st"}


def _validate_and_quarantine(
    table_name: str, columns: list[str]
) -> QuarantineInfo | None:
    """
    Run validation checks on a loaded upload table. Move invalid rows to
    {table_name}_quarantine with a rejection_reason column. Returns
    QuarantineInfo if any rows were quarantined, else None.
    """
    col_set = {c.lower() for c in columns}
    checks: list[str] = []  # SQL CASE expressions for rejection reasons

    # Detect applicable columns
    state_col = next((c for c in columns if c.lower() in _STATE_COLUMNS), None)
    code_col = next((c for c in columns if c.lower() in _CODE_COLUMNS), None)
    rate_col = next((c for c in columns if c.lower() in _RATE_COLUMNS), None)

    # Build validation CASE expressions
    if state_col:
        checks.append(
            f"WHEN LENGTH(TRIM(\"{state_col}\")) != 2 "
            f"OR TRIM(\"{state_col}\") != UPPER(TRIM(\"{state_col}\")) "
            f"THEN 'invalid_state_code'"
        )

    if code_col:
        checks.append(
            f"WHEN \"{code_col}\" IS NULL OR TRIM(CAST(\"{code_col}\" AS VARCHAR)) = '' "
            f"THEN 'missing_code'"
        )

    if rate_col:
        checks.append(
            f'WHEN TRY_CAST("{rate_col}" AS DOUBLE) < 0 THEN \'negative_rate\''
        )
        checks.append(
            f'WHEN TRY_CAST("{rate_col}" AS DOUBLE) > 50000 THEN \'extreme_rate\''
        )

    if not checks:
        return None  # No applicable validation rules

    case_sql = "CASE " + " ".join(checks) + " ELSE NULL END"
    quarantine_table = f"{table_name}_quarantine"

    try:
        with get_cursor() as cur:
            # Add rejection_reason column via CASE
            cur.execute(f"""
                CREATE TABLE {quarantine_table} AS
                SELECT *, {case_sql} AS rejection_reason
                FROM {table_name}
                WHERE ({case_sql}) IS NOT NULL
            """)

            # Count quarantined rows
            cur.execute(f"SELECT COUNT(*) FROM {quarantine_table}")
            q_count = cur.fetchone()[0]

            if q_count == 0:
                cur.execute(f"DROP TABLE IF EXISTS {quarantine_table}")
                return None

            # Get reason breakdown
            cur.execute(f"""
                SELECT rejection_reason, COUNT(*) AS cnt
                FROM {quarantine_table}
                GROUP BY rejection_reason
                ORDER BY cnt DESC
            """)
            reasons = {row[0]: row[1] for row in cur.fetchall()}

            # Remove quarantined rows from main table
            cur.execute(f"""
                DELETE FROM {table_name}
                WHERE rowid IN (
                    SELECT rowid FROM {table_name}
                    WHERE ({case_sql}) IS NOT NULL
                )
            """)

            return QuarantineInfo(quarantine_count=q_count, reasons=reasons)

    except Exception:
        # Validation is best-effort; don't fail the upload
        try:
            with get_cursor() as cur:
                cur.execute(f"DROP TABLE IF EXISTS {quarantine_table}")
        except Exception:
            pass
        return None


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.post("", response_model=ImportResponse)
@safe_route(default_response={})
async def import_file(file: UploadFile = File(...)):
    """
    Upload a CSV, XLSX, or JSON file. Parses it, loads into a DuckDB temp
    table, and returns session metadata + a preview of the first 10 rows.
    """
    _cleanup()

    # --- Validate filename / extension ---
    filename = file.filename or "upload.csv"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ("csv", "tsv", "xlsx", "xls", "json"):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '.{ext}'. Accepted: CSV, XLSX, JSON.",
        )

    # --- Read bytes and check size ---
    raw = await file.read()
    if len(raw) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File exceeds maximum size of {MAX_FILE_SIZE // (1024 * 1024)} MB.",
        )
    if len(raw) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    # --- For XLSX, convert to CSV bytes first ---
    load_bytes = raw
    if ext in ("xlsx", "xls"):
        try:
            load_bytes = _parse_xlsx_bytes(raw)
            # Override ext so the loader treats it as CSV
            ext = "csv"
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {e}")

    # --- Generate session + table identifiers ---
    session_id = str(uuid.uuid4())
    short = _short_id()
    table_name = f"user_upload_{short}"

    # For CSV validation before DuckDB load (quick sanity check)
    if ext in ("csv", "tsv"):
        try:
            _parse_csv_bytes(load_bytes)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"CSV parse error: {e}")
    elif ext == "json":
        try:
            _parse_json_bytes(load_bytes)
        except (ValueError, json.JSONDecodeError) as e:
            raise HTTPException(status_code=400, detail=f"JSON parse error: {e}")

    # --- Load into DuckDB ---
    # Use the converted filename for XLSX so loader picks CSV path
    load_filename = filename if ext != "csv" or not filename.endswith((".xlsx", ".xls")) else filename
    # For xlsx that was converted, we need to signal CSV extension
    if filename.lower().endswith((".xlsx", ".xls")):
        load_filename = filename.rsplit(".", 1)[0] + ".csv"

    try:
        columns, row_count, preview = _load_into_duckdb(load_bytes, load_filename, table_name)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load data into DuckDB: {e}")

    # --- Run quarantine validation ---
    quarantine = _validate_and_quarantine(table_name, columns)

    # If rows were quarantined, update the row count
    if quarantine:
        with get_cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM {table_name}")
            row_count = cur.fetchone()[0]
            cur.execute(f"SELECT * FROM {table_name} LIMIT 10")
            preview = [dict(zip(columns, row)) for row in cur.fetchall()]

    # --- Store session ---
    now = time.time()
    _sessions[session_id] = {
        "session_id": session_id,
        "filename": filename,
        "file_bytes": load_bytes,
        "columns": columns,
        "row_count": row_count,
        "table_name": table_name,
        "created_at": now,
        "last_accessed": now,
    }

    # Evict LRU if over memory cap
    _evict_lru()

    return ImportResponse(
        session_id=session_id,
        table_name=table_name,
        filename=filename,
        columns=columns,
        row_count=row_count,
        preview=preview,
        quarantine=quarantine,
    )


@router.get("/sessions/{session_id}", response_model=SessionInfo)
@safe_route(default_response={})
async def get_session(session_id: str):
    """Return metadata for a given import session."""
    _cleanup()

    sess = _sessions.get(session_id)
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found or expired.")

    sess["last_accessed"] = time.time()

    return SessionInfo(
        session_id=sess["session_id"],
        filename=sess["filename"],
        table_name=sess["table_name"],
        columns=sess["columns"],
        row_count=sess["row_count"],
        created_at=sess["created_at"],
        last_accessed=sess["last_accessed"],
    )


@router.delete("/sessions/{session_id}")
@safe_route(default_response={})
async def delete_session(session_id: str):
    """Delete a session and drop its DuckDB temp table."""
    _cleanup()

    sess = _sessions.get(session_id)
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found or expired.")

    _drop_table(session_id)
    del _sessions[session_id]

    return {"status": "deleted", "session_id": session_id}


@router.get("/sessions/{session_id}/quarantine")
@safe_route(default_response={})
async def get_quarantine(session_id: str):
    """Return quarantined rows for a session, if any."""
    _cleanup()

    sess = _sessions.get(session_id)
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found or expired.")

    sess["last_accessed"] = time.time()
    quarantine_table = f"{sess['table_name']}_quarantine"

    try:
        with get_cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM information_schema.tables WHERE table_name = ?",
                [quarantine_table],
            )
            if cur.fetchone()[0] == 0:
                return {"quarantine_count": 0, "rows": []}

            cur.execute(f"SELECT * FROM {quarantine_table} LIMIT 100")
            columns = [desc[0] for desc in cur.description]
            rows = [dict(zip(columns, row)) for row in cur.fetchall()]
            cur.execute(f"SELECT COUNT(*) FROM {quarantine_table}")
            total = cur.fetchone()[0]

            return {"quarantine_count": total, "rows": rows}
    except Exception:
        return {"quarantine_count": 0, "rows": []}


@router.post("/hydrate")
@safe_route(default_response={})
async def hydrate_session(req: HydrateRequest):
    """
    Re-hydrate a session's DuckDB temp table from stored bytes.

    Called before Intelligence / NL2SQL queries if the table may have been
    dropped (e.g., after a DuckDB restart). If the table already exists,
    this is a no-op.
    """
    _cleanup()

    sess = _sessions.get(req.session_id)
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found or expired.")

    sess["last_accessed"] = time.time()
    table_name = sess["table_name"]

    # Check if table already exists in DuckDB
    with get_cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM information_schema.tables WHERE table_name = ?",
            [table_name],
        )
        exists = cur.fetchone()[0] > 0

    if exists:
        return {
            "status": "already_exists",
            "session_id": req.session_id,
            "table_name": table_name,
        }

    # Re-create table from stored bytes
    filename = sess["filename"]
    if filename.lower().endswith((".xlsx", ".xls")):
        load_filename = filename.rsplit(".", 1)[0] + ".csv"
    else:
        load_filename = filename

    try:
        columns, row_count, _ = _load_into_duckdb(
            sess["file_bytes"], load_filename, table_name
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to re-hydrate table: {e}"
        )

    # Update session metadata in case schema changed
    sess["columns"] = columns
    sess["row_count"] = row_count

    return {
        "status": "hydrated",
        "session_id": req.session_id,
        "table_name": table_name,
        "columns": columns,
        "row_count": row_count,
    }
