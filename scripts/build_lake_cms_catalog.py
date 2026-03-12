#!/usr/bin/env python3
"""
build_lake_cms_catalog.py — Ingest high-value CMS datasets from data.cms.gov catalog.

Downloads and processes datasets identified from the CMS data.json catalog that
fill gaps in the Aradune data lake. Focuses on CMS Program Statistics (ZIP/Excel)
and supplemental CSV datasets.

Tables produced:
  fact_cps_dual_enrollment        — Medicare-Medicaid dual enrollees by state, eligibility type (CY2018-2023)
  fact_cps_part_ab_summary        — Medicare Part A+B spending/utilization by state (CY2023)
  fact_cps_premiums               — Medicare premium beneficiaries and amounts by state (CY2023)
  fact_cps_providers              — Medicare certified providers by state and type (CY2023)
  fact_cps_part_d_utilization     — Part D drug fills and costs by state, plan type (CY2023)
  fact_medicare_covid_hosp        — Medicare COVID-19 hospitalization trends by state/demographics (2020-2023)
  fact_revalidation_due_date      — Provider revalidation due dates (2.9M providers)
  fact_innovation_participants    — CMS Innovation Center model participants (3.5K orgs)

Usage:
  python3 scripts/build_lake_cms_catalog.py
  python3 scripts/build_lake_cms_catalog.py --dry-run
  python3 scripts/build_lake_cms_catalog.py --only cps_dual_enrollment,medicare_covid_hosp
"""

import argparse
import json
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from datetime import date
from pathlib import Path

import duckdb
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "tmp"
LAKE_DIR = PROJECT_ROOT / "data" / "lake"
FACT_DIR = LAKE_DIR / "fact"
SNAP = str(date.today())

# State name to code mapping for CMS Program Statistics Excel files
STATE_NAME_TO_CODE = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "Puerto Rico": "PR", "Virgin Islands": "VI", "Guam": "GU",
    "American Samoa": "AS", "Northern Mariana Islands": "MP",
}


def _write_parquet(con, table: str, fact_name: str) -> int:
    """Write a DuckDB table to ZSTD Parquet in the lake."""
    out_dir = FACT_DIR / fact_name / f"snapshot={SNAP}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "data.parquet"
    n = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    if n > 0:
        con.execute(f"COPY {table} TO '{out}' (FORMAT PARQUET, COMPRESSION ZSTD)")
        size_kb = out.stat().st_size / 1024
        print(f"  -> {fact_name}: {n:,} rows ({size_kb:.1f} KB)")
    else:
        print(f"  -> {fact_name}: 0 rows (SKIPPED)")
    return n


def _download(url: str, dest: Path, label: str = "") -> bool:
    """Download a file via curl. Returns True if successful."""
    if dest.exists() and dest.stat().st_size > 100:
        print(f"  {label or dest.name}: already downloaded ({dest.stat().st_size:,} bytes)")
        return True
    print(f"  Downloading {label or dest.name}...")
    dest.parent.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        ["curl", "-s", "-L", "-o", str(dest), url],
        capture_output=True, timeout=300
    )
    if result.returncode != 0 or not dest.exists() or dest.stat().st_size < 100:
        print(f"  FAILED to download {label or dest.name}")
        return False
    print(f"  Downloaded: {dest.stat().st_size:,} bytes")
    return True


def _extract_zip_excel(zip_path: Path) -> Path:
    """Extract the first .xlsx from a ZIP, return the path."""
    with zipfile.ZipFile(zip_path, "r") as zf:
        xlsx_files = [n for n in zf.namelist() if n.endswith(".xlsx")]
        if not xlsx_files:
            raise ValueError(f"No .xlsx found in {zip_path}")
        zf.extract(xlsx_files[0], zip_path.parent)
        return zip_path.parent / xlsx_files[0]


def _parse_state_data(ws, header_row_idx=3, data_start_idx=5, first_col_name="Area of Residence"):
    """Parse a CMS Program Statistics Excel sheet with state-level data.

    Returns list of dicts with state_code and all numeric columns.
    header_row_idx: 0-based row index of the header row
    data_start_idx: 0-based row index where data starts
    """
    rows = list(ws.iter_rows(values_only=True))

    # Get headers (clean up newlines)
    headers = []
    for v in rows[header_row_idx]:
        if v is not None:
            h = str(v).replace("\n", " ").strip()
            headers.append(h)
        else:
            headers.append(None)

    # Parse data rows
    records = []
    for row in rows[data_start_idx:]:
        if row[0] is None:
            continue
        area = str(row[0]).strip()

        # Skip aggregate rows and blanks
        if area in ("BLANK", "", "All Areas", "All Areas ", "United States",
                     "Unknown", "Other Areas", "Abroad", "Foreign Countries"):
            continue

        # Map state name to code
        state_code = STATE_NAME_TO_CODE.get(area)
        if not state_code:
            # Try stripping footnote numbers
            clean = re.sub(r"\d+$", "", area).strip()
            state_code = STATE_NAME_TO_CODE.get(clean)
        if not state_code:
            continue

        record = {"state_code": state_code}
        for i, val in enumerate(row[1:], 1):
            if i < len(headers) and headers[i] is not None:
                col_name = headers[i]
                if val is not None and val != "*" and val != "BLANK":
                    try:
                        record[col_name] = float(val)
                    except (ValueError, TypeError):
                        record[col_name] = str(val)
                else:
                    record[col_name] = None
        records.append(record)

    return records


# ---------------------------------------------------------------------------
# 1. CPS Dual Enrollment
# ---------------------------------------------------------------------------

def build_cps_dual_enrollment(con, dry_run: bool) -> int:
    """CMS Program Statistics: Medicare-Medicaid dual enrollment by state and eligibility type."""
    zip_url = "https://data.cms.gov/sites/default/files/2025-09/1104c73c-6cb7-422c-bb24-73236d1b5767/MDCR%20ENROLL%20AB%2040-48_CPS_02ENR_2023.zip"
    zip_path = RAW_DIR / "cps_dual_2023.zip"

    print("Building cps_dual_enrollment...")
    if not _download(zip_url, zip_path, "CPS Dual Enrollment ZIP"):
        return 0

    xlsx_path = _extract_zip_excel(zip_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    all_records = []

    # Table 40: Trend data (2018-2023, national totals)
    ws = wb["MDCR ENROLL AB 40_CPS_02ENR"]
    rows = list(ws.iter_rows(values_only=True))
    headers_40 = [str(v).strip() if v else None for v in rows[3]]
    for row in rows[5:]:
        if row[0] is None or row[0] == "BLANK":
            continue
        year = row[0]
        if isinstance(year, (int, float)):
            year = int(year)
        else:
            continue
        record = {
            "year": year,
            "state_code": "US",
            "area_type": "national",
            "total_mmes": row[1],
            "full_benefit_mmes": row[2],
            "qmb_plus": row[3],
            "slmb_plus": row[4],
            "other_full_benefit": row[5],
            "partial_benefit_mmes": row[6],
            "qmb_only": row[7],
            "slmb_only": row[8],
            "qdwi_qi": row[9],
        }
        all_records.append(record)

    # Table 42: By state (CY2023)
    ws = wb["MDCR ENROLL AB 42_CPS_02ENR"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[5:]:
        if row[0] is None:
            continue
        area = str(row[0]).strip()
        if area in ("BLANK", "", "All Areas", "United States", "Unknown", "Other Areas", "Abroad"):
            continue
        state_code = STATE_NAME_TO_CODE.get(area) or STATE_NAME_TO_CODE.get(re.sub(r"\d+$", "", area).strip())
        if not state_code:
            continue

        record = {
            "year": 2023,
            "state_code": state_code,
            "area_type": "state",
            "total_mmes": row[1],
            "full_benefit_mmes": row[2],
            "qmb_plus": row[3],
            "slmb_plus": row[4],
            "other_full_benefit": row[5],
            "partial_benefit_mmes": row[6],
            "qmb_only": row[7],
            "slmb_only": row[8],
            "qdwi_qi": row[9],
        }
        all_records.append(record)

    wb.close()

    if not all_records:
        print("  No records parsed")
        return 0

    # Clean numeric fields
    num_cols = ["total_mmes", "full_benefit_mmes", "qmb_plus", "slmb_plus",
                "other_full_benefit", "partial_benefit_mmes", "qmb_only", "slmb_only", "qdwi_qi"]
    for r in all_records:
        for col in num_cols:
            v = r.get(col)
            if v is not None and v != "*":
                try:
                    r[col] = round(float(v))
                except (ValueError, TypeError):
                    r[col] = None
            else:
                r[col] = None

    import pandas as pd
    df = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _fact AS SELECT * FROM df")

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] cps_dual_enrollment: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "cps_dual_enrollment")


# ---------------------------------------------------------------------------
# 2. CPS Part A+B Summary
# ---------------------------------------------------------------------------

def build_cps_part_ab_summary(con, dry_run: bool) -> int:
    """CMS Program Statistics: Medicare Part A+B spending/utilization by state."""
    zip_url = "https://data.cms.gov/sites/default/files/2026-01/d6c794a4-6817-49ad-8d9a-932efb632d76/MDCR%20SUMMARY%20AB_CPS11SAB_2023.zip"
    zip_path = RAW_DIR / "cps_partab_2023.zip"

    print("Building cps_part_ab_summary...")
    if not _download(zip_url, zip_path, "CPS Part A+B Summary ZIP"):
        return 0

    xlsx_path = _extract_zip_excel(zip_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    all_records = []

    # Table 5: By state (CY2023) — the richest table
    ws = wb["MDCR SUMMARY AB 5_CPS_11SAB"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[3]  # header row

    col_map = {
        1: "total_enrollees",
        2: "persons_with_utilization",
        3: "total_program_payments",
        4: "payments_per_user",
        5: "payments_per_enrollee",
        6: "part_a_enrollees",
        7: "part_a_users",
        8: "part_a_payments",
        9: "part_a_payments_per_user",
        10: "part_a_payments_per_enrollee",
        11: "part_b_enrollees",
        12: "part_b_users",
        13: "part_b_payments",
        14: "part_b_payments_per_user",
        15: "part_b_payments_per_enrollee",
        16: "total_cost_sharing",
        17: "cost_sharing_per_user",
        18: "cost_sharing_per_enrollee",
        19: "discharged_dead",
    }

    for row in rows[5:]:
        if row[0] is None:
            continue
        area = str(row[0]).strip()
        if area in ("BLANK", "", "All Areas", "United States", "Unknown", "Other Areas", "Abroad"):
            continue
        state_code = STATE_NAME_TO_CODE.get(area) or STATE_NAME_TO_CODE.get(re.sub(r"\d+$", "", area).strip())
        if not state_code:
            continue

        record = {"year": 2023, "state_code": state_code}
        for idx, col_name in col_map.items():
            val = row[idx] if idx < len(row) else None
            if val is not None and val != "*" and val != "BLANK":
                try:
                    record[col_name] = float(val)
                except (ValueError, TypeError):
                    record[col_name] = None
            else:
                record[col_name] = None
        all_records.append(record)

    wb.close()

    if not all_records:
        print("  No records parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _fact AS SELECT * FROM df")

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] cps_part_ab_summary: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "cps_part_ab_summary")


# ---------------------------------------------------------------------------
# 3. CPS Premiums
# ---------------------------------------------------------------------------

def build_cps_premiums(con, dry_run: bool) -> int:
    """CMS Program Statistics: Medicare premiums by state (Part A, B, D)."""
    zip_url = "https://data.cms.gov/sites/default/files/2026-01/35b2cc41-1f3c-4ad1-ba60-7df16cad6939/MDCR%20PREMIUMS_CPS04PRM_2023.zip"
    zip_path = RAW_DIR / "cps_premiums_2023.zip"

    print("Building cps_premiums...")
    if not _download(zip_url, zip_path, "CPS Premiums ZIP"):
        return 0

    xlsx_path = _extract_zip_excel(zip_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    all_records = []

    # Parse Part A (Table 3), Part B (Table 6), Part D (Table 9) — all by state
    sheet_configs = [
        ("MDCR PREMIUMS 3_CPS_04PRM", "Part A", {
            1: "total_premium_benes",
            2: "total_premium_amount",
            3: "standard_base_benes",
            4: "standard_base_amount",
            5: "reduced_base_benes",
            6: "reduced_base_amount",
        }),
        ("MDCR PREMIUMS 6_CPS_04PRM", "Part B", {
            1: "total_premium_benes",
            2: "total_premium_amount",
            3: "standard_base_benes",
            4: "standard_base_amount",
            5: "irmaa_surcharge_benes",
            6: "irmaa_surcharge_amount",
        }),
        ("MDCR PREMIUMS 9_CPS_04PRM", "Part D", {
            1: "total_premium_benes",
            2: "total_premium_amount",
            3: "basic_benefit_benes",
            4: "basic_benefit_amount",
            5: "enhanced_benefit_benes",
            6: "enhanced_benefit_amount",
        }),
    ]

    for sheet_name, part, col_map in sheet_configs:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find the header row (has "Area of Residence")
        header_idx = None
        for i, row in enumerate(rows):
            if row[0] and "Area of Residence" in str(row[0]):
                header_idx = i
                break
        if header_idx is None:
            print(f"  WARNING: Could not find header in {sheet_name}")
            continue

        for row in rows[header_idx + 2:]:
            if row[0] is None:
                continue
            area = str(row[0]).strip()
            if area in ("BLANK", "", "All Areas", "All Areas ", "United States",
                        "Unknown", "Other Areas", "Abroad", "Foreign Countries"):
                continue
            state_code = STATE_NAME_TO_CODE.get(area) or STATE_NAME_TO_CODE.get(re.sub(r"\d+$", "", area).strip())
            if not state_code:
                continue

            record = {"year": 2023, "state_code": state_code, "part": part}
            for idx, col_name in col_map.items():
                val = row[idx] if idx < len(row) else None
                if val is not None and val != "*" and val != "BLANK":
                    try:
                        record[col_name] = float(val)
                    except (ValueError, TypeError):
                        record[col_name] = None
                else:
                    record[col_name] = None
            all_records.append(record)

    wb.close()

    if not all_records:
        print("  No records parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _fact AS SELECT * FROM df")

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] cps_premiums: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "cps_premiums")


# ---------------------------------------------------------------------------
# 4. CPS Providers
# ---------------------------------------------------------------------------

def build_cps_providers(con, dry_run: bool) -> int:
    """CMS Program Statistics: Medicare certified providers by state and type."""
    zip_url = "https://data.cms.gov/sites/default/files/2025-09/7e6dbbac-0873-4fe6-9d4a-4168b79e4556/MDCR%20PROVIDERS_CPS19PRV_2023.zip"
    zip_path = RAW_DIR / "cps_providers_2023.zip"

    print("Building cps_providers...")
    if not _download(zip_url, zip_path, "CPS Providers ZIP"):
        return 0

    xlsx_path = _extract_zip_excel(zip_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    all_records = []

    # Table 5: Providers by state and type (CY2023)
    ws = wb["MDCR PROVIDERS 5_CPS_19PRV"]
    rows = list(ws.iter_rows(values_only=True))

    # Get headers
    header_row = rows[2]
    provider_types = []
    for v in header_row[1:]:
        if v is not None:
            provider_types.append(str(v).replace("\n", " ").strip())

    for row in rows[4:]:
        if row[0] is None:
            continue
        area = str(row[0]).strip()
        if area in ("BLANK", "", "All Areas", "All Areas ", "United States",
                     "Unknown", "Other Areas", "Abroad"):
            continue
        state_code = STATE_NAME_TO_CODE.get(area) or STATE_NAME_TO_CODE.get(re.sub(r"\d+$", "", area).strip())
        if not state_code:
            continue

        for i, ptype in enumerate(provider_types):
            val = row[i + 1] if (i + 1) < len(row) else None
            if val is not None and val != "*" and val != "BLANK":
                try:
                    count = int(float(val))
                except (ValueError, TypeError):
                    count = None
            else:
                count = None

            all_records.append({
                "year": 2023,
                "state_code": state_code,
                "provider_type": ptype,
                "provider_count": count,
            })

    # Table 7: Non-institutional providers by state (trend 2019-2023)
    ws = wb["MDCR PROVIDERS 7_CPS_19PRV"]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] and "Provider State" in str(row[0]):
            header_idx = i
            break

    if header_idx is not None:
        # Years are in the header row, cols 1+
        year_cols = []
        for v in rows[header_idx][1:]:
            if v is not None:
                try:
                    year_cols.append(int(float(v)))
                except (ValueError, TypeError):
                    year_cols.append(str(v))

        for row in rows[header_idx + 2:]:
            if row[0] is None:
                continue
            area = str(row[0]).strip()
            if area in ("BLANK", "", "All Areas", "All Areas ", "United States",
                         "Unknown", "Other Areas", "Abroad"):
                continue
            state_code = STATE_NAME_TO_CODE.get(area) or STATE_NAME_TO_CODE.get(re.sub(r"\d+$", "", area).strip())
            if not state_code:
                continue

            for i, yr in enumerate(year_cols):
                if not isinstance(yr, int):
                    continue
                val = row[i + 1] if (i + 1) < len(row) else None
                if val is not None and val != "*" and val != "BLANK":
                    try:
                        count = int(float(val))
                    except (ValueError, TypeError):
                        count = None
                else:
                    count = None

                all_records.append({
                    "year": yr,
                    "state_code": state_code,
                    "provider_type": "Non-Institutional (Total)",
                    "provider_count": count,
                })

    wb.close()

    if not all_records:
        print("  No records parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _fact AS SELECT * FROM df")

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] cps_providers: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "cps_providers")


# ---------------------------------------------------------------------------
# 5. CPS Part D Utilization
# ---------------------------------------------------------------------------

def build_cps_part_d_utilization(con, dry_run: bool) -> int:
    """CMS Program Statistics: Part D drug utilization by state, plan type."""
    zip_url = "https://data.cms.gov/sites/default/files/2025-09/8580dab0-0060-4c42-a897-f2599e8a35fc/MDCR%20UTLZN%20D_CPS12UPD_2023.zip"
    zip_path = RAW_DIR / "cps_partd_2023.zip"

    print("Building cps_part_d_utilization...")
    if not _download(zip_url, zip_path, "CPS Part D Utilization ZIP"):
        return 0

    xlsx_path = _extract_zip_excel(zip_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    all_records = []

    # Table 6: By state — Drug fills and cost per enrollee by plan type
    ws = wb["MDCR UTLZN D 6_CPS_12UPD"]
    rows = list(ws.iter_rows(values_only=True))

    col_map = {
        1: "total_part_d_enrollees",
        2: "pdp_enrollees",
        3: "mapd_enrollees",
        4: "overall_avg_fills_per_enrollee",
        5: "pdp_avg_fills_per_enrollee",
        6: "mapd_avg_fills_per_enrollee",
        7: "overall_avg_cost_per_enrollee",
        8: "pdp_avg_cost_per_enrollee",
        9: "mapd_avg_cost_per_enrollee",
    }

    # Find first data row (after header + blank)
    for row in rows[4:]:
        if row[0] is None:
            continue
        area = str(row[0]).strip()
        if area in ("BLANK", "", "All Areas", "All Areas ", "United States",
                     "Unknown", "Other Areas", "Abroad"):
            continue
        state_code = STATE_NAME_TO_CODE.get(area) or STATE_NAME_TO_CODE.get(re.sub(r"\d+$", "", area).strip())
        if not state_code:
            continue

        record = {"year": 2023, "state_code": state_code}
        for idx, col_name in col_map.items():
            val = row[idx] if idx < len(row) else None
            if val is not None and val != "*" and val != "BLANK":
                try:
                    record[col_name] = float(val)
                except (ValueError, TypeError):
                    record[col_name] = None
            else:
                record[col_name] = None
        all_records.append(record)

    # Table 7: By state — Drug fills and cost per utilizer by plan type
    ws = wb["MDCR UTLZN D 7_CPS_12UPD"]
    rows = list(ws.iter_rows(values_only=True))

    utilizer_col_map = {
        1: "total_part_d_utilizers",
        2: "pdp_utilizers",
        3: "mapd_utilizers",
        4: "overall_avg_fills_per_utilizer",
        5: "pdp_avg_fills_per_utilizer",
        6: "mapd_avg_fills_per_utilizer",
        7: "overall_avg_cost_per_utilizer",
        8: "pdp_avg_cost_per_utilizer",
        9: "mapd_avg_cost_per_utilizer",
    }

    for row in rows[4:]:
        if row[0] is None:
            continue
        area = str(row[0]).strip()
        if area in ("BLANK", "", "All Areas", "All Areas ", "United States",
                     "Unknown", "Other Areas", "Abroad"):
            continue
        state_code = STATE_NAME_TO_CODE.get(area) or STATE_NAME_TO_CODE.get(re.sub(r"\d+$", "", area).strip())
        if not state_code:
            continue

        # Find matching enrollee record and merge
        for rec in all_records:
            if rec["state_code"] == state_code and rec["year"] == 2023:
                for idx, col_name in utilizer_col_map.items():
                    val = row[idx] if idx < len(row) else None
                    if val is not None and val != "*" and val != "BLANK":
                        try:
                            rec[col_name] = float(val)
                        except (ValueError, TypeError):
                            rec[col_name] = None
                    else:
                        rec[col_name] = None
                break

    wb.close()

    if not all_records:
        print("  No records parsed")
        return 0

    import pandas as pd
    df = pd.DataFrame(all_records)
    con.execute("CREATE OR REPLACE TABLE _fact AS SELECT * FROM df")

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] cps_part_d_utilization: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "cps_part_d_utilization")


# ---------------------------------------------------------------------------
# 6. Medicare COVID-19 Hospitalization Trends
# ---------------------------------------------------------------------------

def build_medicare_covid_hosp(con, dry_run: bool) -> int:
    """Medicare COVID-19 hospitalization trends by state, demographics."""
    csv_url = "https://data.cms.gov/sites/default/files/2024-01/4ae724cb-1224-496a-bcf5-3421817a9f03/COVID-19%20Hospitalization%20Trends%20Report%20Data%20File%20-%20Claims%20thru%2006.30.2023.csv"
    csv_path = RAW_DIR / "covid_hosp.csv"

    print("Building medicare_covid_hosp...")
    if not _download(csv_url, csv_path, "COVID-19 Hospitalization CSV"):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact AS
        SELECT
            TRY_CAST(Year AS INTEGER) AS year,
            TRY_CAST(Month AS INTEGER) AS month,
            Bene_Geo_Desc AS geo_desc,
            Bene_Mdcd_Mdcr_Enrl_Stus AS dual_status,
            Bene_Race_Desc AS race,
            Bene_Sex_Desc AS sex,
            Bene_Mdcr_Entlmt_Stus AS entitlement_status,
            Bene_Age_Desc AS age_group,
            Bene_RUCA_Desc AS rural_urban,
            TRY_CAST(Total_Hosp AS INTEGER) AS total_hospitalizations,
            TRY_CAST(Total_Enrl AS INTEGER) AS total_enrolled,
            TRY_CAST(Total_Hosp_Per100K AS DOUBLE) AS hosp_per_100k,
            TRY_CAST(Avg_LOS AS DOUBLE) AS avg_length_of_stay,
            TRY_CAST(Pct_Dschrg_SNF AS DOUBLE) AS pct_discharge_snf,
            TRY_CAST(Pct_Dschrg_Expired AS DOUBLE) AS pct_discharge_expired,
            TRY_CAST(Pct_Dschrg_Home AS DOUBLE) AS pct_discharge_home,
            TRY_CAST(Pct_Dschrg_Hspc AS DOUBLE) AS pct_discharge_hospice,
            TRY_CAST(Pct_Dschrg_HomeHealth AS DOUBLE) AS pct_discharge_home_health,
            TRY_CAST(Pct_Dschrg_Other AS DOUBLE) AS pct_discharge_other
        FROM read_csv_auto('{csv_path}', header=true, ignore_errors=true)
        WHERE Year IS NOT NULL
    """)

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] medicare_covid_hosp: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "medicare_covid_hosp")


# ---------------------------------------------------------------------------
# 7. Revalidation Due Date List
# ---------------------------------------------------------------------------

def build_revalidation_due_date(con, dry_run: bool) -> int:
    """Provider revalidation due dates — 2.9M providers with NPI, state, specialty."""
    csv_url = "https://data.cms.gov/sites/default/files/2026-02/aad0c65d-130c-4702-af81-78fde1a9ab08/revalidation_base.csv"
    csv_path = RAW_DIR / "revalidation_due.csv"

    print("Building revalidation_due_date...")
    if not _download(csv_url, csv_path, "Revalidation Due Date CSV"):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact AS
        SELECT
            "Enrollment ID" AS enrollment_id,
            "National Provider Identifier" AS npi,
            "First Name" AS first_name,
            "Last Name" AS last_name,
            "Organization Name" AS organization_name,
            "Enrollment State Code" AS state_code,
            "Enrollment Type" AS enrollment_type,
            "Provider Type Text" AS provider_type,
            "Enrollment Specialty" AS specialty,
            TRY_CAST("Revalidation Due Date" AS DATE) AS revalidation_due_date,
            TRY_CAST("Adjusted Due Date" AS DATE) AS adjusted_due_date,
            TRY_CAST("Individual Total Reassign To" AS INTEGER) AS total_reassign_to,
            "Receiving Benefits Reassignment" AS receiving_reassignment
        FROM read_csv_auto('{csv_path}', header=true, sample_size=20000, ignore_errors=true)
        WHERE "Enrollment State Code" IS NOT NULL
    """)

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] revalidation_due_date: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "revalidation_due_date")


# ---------------------------------------------------------------------------
# 8. Innovation Center Model Participants
# ---------------------------------------------------------------------------

def build_innovation_participants(con, dry_run: bool) -> int:
    """CMS Innovation Center model participants by organization."""
    csv_url = "https://data.cms.gov/sites/default/files/2026-02/d403f70e-e4d5-44f0-9758-477ace063253/Innovation_Center_Model_Participants-GUIDE_2_25_26.csv"
    csv_path = RAW_DIR / "innovation_participants.csv"

    print("Building innovation_participants...")
    if not _download(csv_url, csv_path, "Innovation Model Participants CSV"):
        return 0

    con.execute(f"""
        CREATE OR REPLACE TABLE _fact AS
        SELECT
            "Name of Initiative" AS model_name,
            "Organization Name" AS organization_name,
            Notes AS notes,
            "Street Address" AS street_address,
            City AS city,
            State AS state_code,
            "State Based" AS state_based,
            Website AS website,
            Category AS category,
            MSA_Name AS msa_name,
            "Unique ID" AS unique_id
        FROM read_csv_auto('{csv_path}', header=true, ignore_errors=true)
        WHERE "Name of Initiative" IS NOT NULL
    """)

    if dry_run:
        n = con.execute("SELECT COUNT(*) FROM _fact").fetchone()[0]
        print(f"  [dry-run] innovation_participants: {n:,} rows")
        return n
    return _write_parquet(con, "_fact", "innovation_participants")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

BUILDERS = {
    "cps_dual_enrollment": build_cps_dual_enrollment,
    "cps_part_ab_summary": build_cps_part_ab_summary,
    "cps_premiums": build_cps_premiums,
    "cps_providers": build_cps_providers,
    "cps_part_d_utilization": build_cps_part_d_utilization,
    "medicare_covid_hosp": build_medicare_covid_hosp,
    "revalidation_due_date": build_revalidation_due_date,
    "innovation_participants": build_innovation_participants,
}


def main():
    parser = argparse.ArgumentParser(description="Ingest CMS catalog datasets into the data lake")
    parser.add_argument("--dry-run", action="store_true", help="Parse but don't write parquet")
    parser.add_argument("--only", type=str, default=None,
                        help="Comma-separated list of tables to build")
    args = parser.parse_args()

    # Ensure tmp dir exists
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect()

    targets = BUILDERS
    if args.only:
        only_set = set(args.only.split(","))
        targets = {k: v for k, v in BUILDERS.items() if k in only_set}

    total_rows = 0
    results = {}

    print(f"\n{'='*60}")
    print(f"  CMS Catalog Ingestion — {len(targets)} tables")
    print(f"  Snapshot: {SNAP}")
    print(f"{'='*60}\n")

    for name, builder in targets.items():
        try:
            n = builder(con, args.dry_run)
            results[name] = n
            total_rows += n
            print()
        except Exception as e:
            print(f"  ERROR building {name}: {e}")
            import traceback
            traceback.print_exc()
            results[name] = -1
            print()

    print(f"\n{'='*60}")
    print(f"  Summary")
    print(f"{'='*60}")
    for name, n in results.items():
        status = f"{n:>12,} rows" if n >= 0 else "     FAILED"
        print(f"  {name:<35s} {status}")
    print(f"  {'─'*48}")
    print(f"  {'TOTAL':<35s} {total_rows:>12,} rows")
    print(f"{'='*60}\n")

    # Write manifest
    if not args.dry_run:
        manifest = {
            "snapshot": SNAP,
            "tables": {k: v for k, v in results.items() if v > 0},
            "total_rows": total_rows,
        }
        manifest_path = LAKE_DIR / "metadata" / f"cms_catalog_{SNAP}.json"
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        with open(manifest_path, "w") as f:
            json.dump(manifest, f, indent=2)
        print(f"  Manifest: {manifest_path}")

    con.close()


if __name__ == "__main__":
    main()
